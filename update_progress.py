#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""更新Excel开发管理表中的任务状态
用法: python3 update_progress.py <任务编号> <状态>
示例: python3 update_progress.py A1.01 完成
      python3 update_progress.py A1.01,A1.02,A1.03 完成
"""
import sys
from openpyxl import load_workbook

EXCEL_PATH = '/workspace/宗门论道_Godot开发管理工具包.xlsx'
MAIN_SHEET = '2.开发主表'
STATUS_COL = 15  # O列

# 状态颜色（与Excel样式一致）
STATUS_COLORS = {
    '待办': 'FCE4D6',
    '进行中': 'FFF2CC',
    '完成': 'C6EFCE',
}

def update_status(codes, status):
    """更新指定任务编号的状态"""
    if status not in STATUS_COLORS:
        print(f'错误: 状态必须是 {list(STATUS_COLORS.keys())} 之一')
        return False
    
    wb = load_workbook(EXCEL_PATH)
    ws = wb[MAIN_SHEET]
    
    found = {code: False for code in codes}
    updated = 0
    
    for row in range(3, ws.max_row + 1):
        code = ws.cell(row=row, column=2).value  # B列=编号
        if code in codes:
            found[code] = True
            cell = ws.cell(row=row, column=STATUS_COL)
            old_status = cell.value
            cell.value = status
            # 更新背景色
            from openpyxl.styles import PatternFill
            cell.fill = PatternFill('solid', fgColor=STATUS_COLORS[status])
            print(f'  {code}: {old_status} → {status}')
            updated += 1
    
    # 检查未找到的任务
    for code, was_found in found.items():
        if not was_found:
            print(f'  ⚠ {code}: 未找到（请检查编号）')
    
    if updated > 0:
        # 更新进度统计表（openpyxl不自动计算公式，需手动写入数值）
        refresh_stats(wb)
        wb.save(EXCEL_PATH)
        print(f'\n✓ 已更新 {updated} 个任务，Excel已保存')

        # 打印当前进度统计
        print_progress(wb)
        return True
    else:
        print('\n✗ 没有任务被更新')
        return False


# 阶段 → 统计表行号映射
PHASE_ROWS = {
    'P1': 3, 'P2': 4, 'P3': 5, 'P4': 6, 'P5': 7,
    'P6': 8, 'P7': 9, 'P8': 10,
    'P9': 11, 'P10': 11, 'P11': 11, 'P12': 11, 'P13': 11, 'P14': 11,
    'P16': 12, 'P17': 12, 'P18': 12,
}

def refresh_stats(wb):
    """根据开发主表数据，重新计算并填充进度统计表的数值"""
    ws_main = wb[MAIN_SHEET]
    ws_stats = wb['3.进度统计']

    # 按阶段统计L3任务
    stats = {}  # {行号: {total, done, in_progress, todo}}
    for row in range(3, ws_main.max_row + 1):
        lvl = ws_main.cell(row=row, column=1).value
        if not lvl or lvl.strip() != 'L3':
            continue
        phase = ws_main.cell(row=row, column=3).value  # C列=阶段
        status = ws_main.cell(row=row, column=STATUS_COL).value
        stat_row = PHASE_ROWS.get(phase)
        if stat_row is None:
            continue
        if stat_row not in stats:
            stats[stat_row] = {'total': 0, 'done': 0, 'in_progress': 0, 'todo': 0}
        stats[stat_row]['total'] += 1
        if status == '完成':
            stats[stat_row]['done'] += 1
        elif status == '进行中':
            stats[stat_row]['in_progress'] += 1
        else:
            stats[stat_row]['todo'] += 1

    # 写入统计表（B=总数 C=完成 D=进行中 E=待办 F=完成率）
    grand_total = grand_done = grand_inprog = grand_todo = 0
    for row_num in range(3, 12):  # 行3-11为各阶段，行12为PvP，行13为合计
        if row_num in stats:
            s = stats[row_num]
        else:
            s = {'total': 0, 'done': 0, 'in_progress': 0, 'todo': 0}
        ws_stats.cell(row=row_num, column=2).value = s['total']
        ws_stats.cell(row=row_num, column=3).value = s['done']
        ws_stats.cell(row=row_num, column=4).value = s['in_progress']
        ws_stats.cell(row=row_num, column=5).value = s['todo']
        pct = s['done'] / s['total'] if s['total'] > 0 else 0
        ws_stats.cell(row=row_num, column=6).value = round(pct, 4)
        # 设置百分比格式
        ws_stats.cell(row=row_num, column=6).number_format = '0.0%'
        grand_total += s['total']
        grand_done += s['done']
        grand_inprog += s['in_progress']
        grand_todo += s['todo']

    # 第12行（P16-18 PvP）可能也在stats中
    if 12 in stats:
        s = stats[12]
        ws_stats.cell(row=12, column=2).value = s['total']
        ws_stats.cell(row=12, column=3).value = s['done']
        ws_stats.cell(row=12, column=4).value = s['in_progress']
        ws_stats.cell(row=12, column=5).value = s['todo']
        pct = s['done'] / s['total'] if s['total'] > 0 else 0
        ws_stats.cell(row=12, column=6).value = round(pct, 4)
        ws_stats.cell(row=12, column=6).number_format = '0.0%'
        grand_total += s['total']
        grand_done += s['done']
        grand_inprog += s['in_progress']
        grand_todo += s['todo']
    else:
        for col in range(2, 6):
            ws_stats.cell(row=12, column=col).value = 0
        ws_stats.cell(row=12, column=6).value = 0
        ws_stats.cell(row=12, column=6).number_format = '0.0%'

    # 合计行（第13行）
    ws_stats.cell(row=13, column=2).value = grand_total
    ws_stats.cell(row=13, column=3).value = grand_done
    ws_stats.cell(row=13, column=4).value = grand_inprog
    ws_stats.cell(row=13, column=5).value = grand_todo
    grand_pct = grand_done / grand_total if grand_total > 0 else 0
    ws_stats.cell(row=13, column=6).value = round(grand_pct, 4)
    ws_stats.cell(row=13, column=6).number_format = '0.0%'


def print_progress(wb):
    """打印当前进度统计"""
    ws_main = wb[MAIN_SHEET]
    total = done = 0
    for row in range(3, ws_main.max_row + 1):
        lvl = ws_main.cell(row=row, column=1).value
        if lvl and lvl.strip() == 'L3':
            total += 1
            if ws_main.cell(row=row, column=STATUS_COL).value == '完成':
                done += 1
    pct = done / total * 100 if total > 0 else 0
    print(f'\n📊 当前进度: {done}/{total} ({pct:.1f}%)')

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('用法: python3 update_progress.py <任务编号> <状态>')
        print('示例: python3 update_progress.py A1.01 完成')
        print('      python3 update_progress.py A1.01,A1.02 完成')
        sys.exit(1)
    
    codes = sys.argv[1].split(',')
    status = sys.argv[2]
    update_status(codes, status)
