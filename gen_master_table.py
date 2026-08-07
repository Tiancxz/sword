#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
重新生成《宗门论道》开发管理工具包 v2
核心改动：用一张「开发主表」替代原来混乱的3张表(路线图+功能清单+任务拆解)
7层嵌套：系统大类 > 模块 > 子功能 > 实现原理 > 对应函数 > 输入/输出 > 开发状态
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = '/workspace/宗门论道_开发管理工具包.xlsx'
wb = openpyxl.load_workbook(PATH)

# 删除旧的重叠Sheet
for name in ['2.开发路线图', '3.功能清单', '4.任务拆解', '8.进度看板']:
    if name in wb.sheetnames:
        del wb[name]

# ===== 样式 =====
F_TITLE   = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
F_HEADER  = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
F_L1      = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
F_L2      = Font(name='微软雅黑', size=11, bold=True, color='1F4E79')
F_L3      = Font(name='微软雅黑', size=10, bold=True, color='333333')
F_CELL    = Font(name='微软雅黑', size=10)
F_CODE    = Font(name='Consolas', size=10, color='333333')
F_BOLD    = Font(name='微软雅黑', size=10, bold=True)
F_NOTE    = Font(name='微软雅黑', size=9, color='888888', italic=True)

BG_TITLE  = PatternFill('solid', fgColor='2F5496')
BG_HEADER = PatternFill('solid', fgColor='4472C4')
BG_L1_A   = PatternFill('solid', fgColor='1F4E79')  # 系统大类
BG_L1_B   = PatternFill('solid', fgColor='2E75B6')
BG_L1_C   = PatternFill('solid', fgColor='5B9BD5')
BG_L1_D   = PatternFill('solid', fgColor='70AD47')
BG_L1_E   = PatternFill('solid', fgColor='C55A11')
BG_L2     = PatternFill('solid', fgColor='D6E4F0')  # 模块
BG_L3     = PatternFill('solid', fgColor='F2F2F2')  # 子功能
BG_TODO   = PatternFill('solid', fgColor='FFC7CE')
BG_DOING  = PatternFill('solid', fgColor='FFEB9C')
BG_DONE   = PatternFill('solid', fgColor='C6EFCE')

A_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
A_L = Alignment(horizontal='left', vertical='center', wrap_text=True)
A_LT= Alignment(horizontal='left', vertical='top', wrap_text=True)

BD = Border(
    left=Side(style='thin', color='B4C7E7'),
    right=Side(style='thin', color='B4C7E7'),
    top=Side(style='thin', color='B4C7E7'),
    bottom=Side(style='thin', color='B4C7E7')
)

def set_w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title(ws, text, cols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_TITLE; c.fill = BG_TITLE; c.alignment = A_C
    ws.row_dimensions[row].height = 36

def header(ws, headers, row):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = F_HEADER; c.fill = BG_HEADER; c.alignment = A_C; c.border = BD
    ws.row_dimensions[row].height = 28

# ================================================================
# 新 Sheet 2: 开发主表（★唯一核心表★）
# ================================================================
ws = wb.create_sheet('2.开发主表(唯一)', 1)  # 插入到第2位
set_w(ws, [6, 16, 20, 28, 42, 28, 24, 8, 8])
title(ws, '★ 开发主表 — 唯一核心表：从系统大类到函数级，逐层展开 ★', 9)
header(ws, ['编号', '系统大类', '模块', '子功能', '实现原理（怎么实现的）', '对应函数 / 方法', '输入 → 输出', '版本', '状态'], 3)

# 数据：7层嵌套
# (编号, 系统, 模块, 子功能, 实现原理, 函数, 输入→输出, 版本, 状态)
# 系统=第一层(色块) / 模块=第二层 / 子功能=第三层 / 原理=第四层 / 函数=第五层 / 输入输出=第六层 / 状态=第七层
D = []

# ============ A. 引擎框架 ============
S = 'A.引擎框架'
M = 'A1.渲染循环'
D += [
    ('A1.01', S, M, '主循环驱动', 'requestAnimationFrame循环，每帧计算dt=(now-lastTime)/1000，dt上限0.033防跳帧，调用update(dt)再render()', 'Director.loop(timestamp)', 'timestamp → void', 'V1', '待办'),
    ('A1.02', S, M, '逻辑更新', '每帧调用当前场景的onUpdate(dt)，场景内部驱动游戏逻辑', 'Director.update(dt)', 'dt → void', 'V1', '待办'),
    ('A1.03', S, M, '画面渲染', '每帧清屏→调用当前场景onRender()→场景内部绘制所有元素', 'Director.render()', 'void → void', 'V1', '待办'),
    ('A1.04', S, M, '引擎初始化', '获取canvas 2D context→创建SceneManager→创建Input→绑定触摸→启动循环', 'Director.init()', 'void → void', 'V1', '待办'),
]
M = 'A2.输入系统'
D += [
    ('A2.01', S, M, '触摸监听', 'wx.onTouchStart/Move/End绑定，获取触摸坐标', 'Input.init()', 'void → void', 'V1', '待办'),
    ('A2.02', S, M, '点击区域注册', '场景/UI向Input注册矩形区域{id,x,y,w,h,cb}，存入touchRegions数组', 'Input.register(id,x,y,w,h,cb)', '区域参数 → void', 'V1', '待办'),
    ('A2.03', S, M, '命中检测', '触摸时遍历touchRegions，判断点是否在矩形内(x<tx<x+w且y<ty<y+h)，命中则调用cb', 'Input.onTouchStart(tx,ty)', '触摸坐标 → 触发cb', 'V1', '待办'),
    ('A2.04', S, M, '区域清理', '场景切换时清空所有注册区域，防止旧场景残留响应', 'Input.clear()', 'void → void', 'V1', '待办'),
]
M = 'A3.事件总线'
D += [
    ('A3.01', S, M, '事件注册', '模块通过on(event,cb)注册监听，存入listeners Map', 'EventSystem.on(event,cb)', '事件名+回调 → void', 'V1', '待办'),
    ('A3.02', S, M, '事件触发', '通过emit(event,data)触发，遍历该事件的所有cb并调用', 'EventSystem.emit(event,data)', '事件名+数据 → void', 'V1', '待办'),
    ('A3.03', S, M, '事件移除', '模块销毁时off(event,cb)移除监听，防止内存泄漏', 'EventSystem.off(event,cb)', '事件名+回调 → void', 'V1', '待办'),
]
M = 'A4.场景管理'
D += [
    ('A4.01', S, M, '场景注册', '注册场景名→场景实例到Map', 'SceneManager.register(name,scene)', '名称+场景 → void', 'V1', '待办'),
    ('A4.02', S, M, '场景切换', '调用旧场景onExit()→清理→切换currentScene→调用新场景onEnter()', 'SceneManager.switch(name)', '场景名 → void', 'V1', '待办'),
]
M = 'A5.资源加载'
D += [
    ('A5.01', S, M, '图片预加载', 'wx.createImage()→src=url→onload回调→计数→全部完成触发onComplete', 'ResourceLoader.loadImages(urls,cb)', 'URL数组+回调 → void', 'V1', '待办'),
    ('A5.02', S, M, '加载进度', '已加载数/总数×100%，回调通知UI显示进度条', 'ResourceLoader.getProgress()', 'void → 百分比', 'V1', '待办'),
]
M = 'A6.渲染器'
D += [
    ('A6.01', S, M, '矩形绘制', 'ctx.fillRect绘制色块(单位/阵法占位)', 'Renderer.rect(x,y,w,h,color)', '坐标+颜色 → void', 'V1', '待办'),
    ('A6.02', S, M, '文字绘制', 'ctx.fillText绘制名称/血量/灵力', 'Renderer.text(text,x,y,color,size)', '文本+坐标 → void', 'V1', '待办'),
    ('A6.03', S, M, '圆形绘制', 'ctx.arc绘制粒子/特效', 'Renderer.circle(x,y,r,color)', '坐标+半径 → void', 'V1', '待办'),
    ('A6.04', S, M, '血条绘制', '按比例绘制红绿渐变血条(背景红+前景绿)', 'Renderer.healthBar(x,y,w,h,ratio)', '坐标+比例 → void', 'V1', '待办'),
]
M = 'A7.粒子系统'
D += [
    ('A7.01', S, M, '粒子发射', '在指定位置生成N个粒子{x,y,vx,vy,life,color,size}，初始速度随机方向', 'ParticleSystem.emit(x,y,count,color)', '坐标+数量+颜色 → void', 'V1', '待办'),
    ('A7.02', S, M, '粒子更新', '每帧x+=vx*dt, y+=vy*dt, life-=dt，life<=0则移除(复用对象池)', 'ParticleSystem.update(dt)', 'dt → void', 'V1', '待办'),
    ('A7.03', S, M, '粒子渲染', '遍历存活粒子，按life/maxLife算透明度，绘制圆形', 'ParticleSystem.render()', 'void → void', 'V1', '待办'),
]
M = 'A8.音频管理'
D += [
    ('A8.01', S, M, 'BGM播放', 'wx.getBackgroundAudioManager→设置src→play，循环播放', 'AudioManager.playBGM(url)', 'URL → void', 'V1', '待办'),
    ('A8.02', S, M, 'SFX播放', 'wx.createInnerAudioContext→设置src→play，播完自动销毁(用对象池复用)', 'AudioManager.playSFX(url)', 'URL → void', 'V1', '待办'),
    ('A8.03', S, M, '音量控制', '存储bgmVol/sfxVol到localStorage，设置时更新所有音频实例', 'AudioManager.setVolume(type,vol)', '类型+音量 → void', 'V1', '待办'),
]

# ============ B. 游戏数据 ============
S = 'B.游戏数据'
M = 'B1.常量配置'
D += [
    ('B1.01', S, M, '全局常量', '棋盘9格/灵力2.8s/大殿30血/手牌4张等所有硬编码值，集中在一个文件，代码中只引用不写死', 'Constants.js', '文件 → 所有模块引用', 'V1', '待办'),
]
M = 'B2.卡牌数据表'
D += [
    ('B2.01', S, M, '卡牌属性配置', '19张卡牌完整属性{id,name,type,cost,hp,atk,speed,range,interval,traits}，配置驱动，加卡牌只加配置不改逻辑', 'Cards.js', '文件 → 所有模块引用', 'V1', '待办'),
    ('B2.02', S, M, '卡牌查表', '通过cardId从Cards表读取属性，用于创建Unit/Formation实例', 'Cards.get(cardId)', '卡牌ID → 卡牌属性Object', 'V1', '待办'),
]
M = 'B3.卡组手牌'
D += [
    ('B3.01', S, M, '卡组初始化', '传入8张卡牌ID数组→复制到drawPile→Fisher-Yates洗牌→抽3张到手牌', 'Deck.init(cardIds)', '卡牌ID数组 → void', 'V1', '待办'),
    ('B3.02', S, M, '洗牌算法', '从后往前遍历，随机交换位置(Fisher-Yates算法，O(n)均匀洗牌)', 'Deck.shuffle()', 'void → void', 'V1', '待办'),
    ('B3.03', S, M, '抽牌', '从drawPile末尾pop一张→push到hand→手牌满4张或牌堆空时不抽', 'Deck.draw()', 'void → 卡牌ID或null', 'V1', '待办'),
    ('B3.04', S, M, '打出后补牌', '打出牌后启动drawTimer=2秒，倒计时到0时调用draw()补1张', 'Deck.update(dt)', 'dt → void', 'V1', '待办'),
    ('B3.05', S, M, '出牌检查', '检查手牌index对应的卡牌费用是否<=当前灵力', 'Deck.canPlay(index,energy)', '手牌index+灵力 → bool', 'V1', '待办'),
    ('B3.06', S, M, '打出牌', '从hand数组移除指定index的牌→启动drawTimer', 'Deck.playCard(index)', '手牌index → 卡牌ID', 'V1', '待办'),
]
M = 'B4.预设卡组'
D += [
    ('B4.01', S, M, '3套预设卡组', '快攻型(低费体修为主)/控制型(阵法+法术)/坦克型(御兽+重岳)各8张', 'DeckPresets.js', '文件 → Deck.init使用', 'V1', '待办'),
]

# ============ C. 战斗逻辑 ============
S = 'C.战斗逻辑'
M = 'C1.战斗主循环'
D += [
    ('C1.01', S, M, '每帧总入口', '按顺序调用: 灵力回复→抽牌→单位更新→阵法更新→长老技能→AI决策→清理死亡→检查胜负→倒计时', 'BattleLogic.update(dt)', 'dt → void', 'V1', '待办'),
]
M = 'C2.单位实体'
D += [
    ('C2.01', S, M, '从卡牌创建单位', '读Cards表属性→创建Unit对象→设置owner/x/y/facing(owner=0向上facing=1, owner=1向下facing=-1)→设置state=walking', 'Unit.fromCard(cardId,owner,x,y)', '卡牌ID+归属+坐标 → Unit实例', 'V1', '待办'),
    ('C2.02', S, M, 'buff系统', 'buff={type,value,duration,elapsed}，每帧elapsed+=dt，超时移除。类型: slow(减速)/speed(加速)/stun(定身)/shield(护盾)/atkBoost(攻击加成)', 'Unit.addBuff(type,value,duration)', '类型+数值+持续 → void', 'V1', '待办'),
    ('C2.03', S, M, 'buff更新', '每帧遍历buffs→elapsed+=dt→超则移除→重算speed(slow:-0.5, speed:+0.5, stun:0)', 'Unit.updateBuffs(dt)', 'dt → void', 'V1', '待办'),
    ('C2.04', S, M, '实际移速计算', 'baseSpeed + 所有buff的value之和(slow为负, speed为正, stun直接返回0)', 'Unit.getEffectiveSpeed()', 'void → 移速数值', 'V1', '待办'),
    ('C2.05', S, M, '受到伤害', 'hp-=amount→if hp<=0: state=dead→触发死亡回调(自爆等)→如果attacker有反震则attacker也受伤', 'Unit.takeDamage(amount,attacker)', '伤害量+攻击者 → 实际伤害', 'V1', '待办'),
]
M = 'C3.移动系统'
D += [
    ('C3.01', S, M, '★单位移动★', 'if state==walking: 先找目标→if有目标且在攻击范围内: state=fighting→else: y += effectiveSpeed * dt * facing→检查是否到大殿', 'MovementSystem.moveUnit(unit,dt)', '单位+dt → void', 'V1', '待办'),
    ('C3.02', S, M, '★目标选择★', '获取所有敌方单位+敌方阵法→过滤同列(x相同或阵法相邻)→按y距离排序→返回最近的→如果最近的不在攻击范围内返回null(继续走)', 'MovementSystem.findTarget(unit,model)', '单位+状态 → 目标或null', 'V1', '待办'),
    ('C3.03', S, M, '碰撞检测', '计算单位与目标的y轴距离，如果<attackRange(近战1格/远程3格)则返回true(可攻击)', 'MovementSystem.checkCollision(unit,target)', '单位+目标 → bool', 'V1', '待办'),
    ('C3.04', S, M, '到达大殿检测', 'if owner==0且y>=8(到顶): return true→if owner==1且y<=0(到底): return true', 'MovementSystem.checkHallReach(unit,model)', '单位+状态 → bool', 'V1', '待办'),
]
M = 'C4.战斗系统'
D += [
    ('C4.01', S, M, '攻击主逻辑', 'if now - lastAttackTime >= attackInterval: 根据traits选择近战/远程/范围→执行→重置lastAttackTime', 'CombatSystem.attack(attacker,target,dt)', '攻击者+目标+dt → void', 'V1', '待办'),
    ('C4.02', S, M, '近战互扣', 'target.takeDamage(attacker.atk, attacker) → attacker.takeDamage(target.atk, target) 双方同时掉血', 'CombatSystem.meleeAttack(attacker,target)', '攻击者+目标 → void', 'V1', '待办'),
    ('C4.03', S, M, '远程单方', 'target.takeDamage(attacker.atk) 只扣对方，攻击者不掉血(射程3格)', 'CombatSystem.rangedAttack(attacker,target)', '攻击者+目标 → void', 'V1', '待办'),
    ('C4.04', S, M, '范围伤害', '对目标数组每个target执行takeDamage(damage)', 'CombatSystem.aoeAttack(source,targets,damage)', '来源+目标数组+伤害 → void', 'V1', '待办'),
    ('C4.05', S, M, '自爆', '对相邻3格内所有敌方造成damage→单位消失(state=dead)', 'CombatSystem.kamikaze(unit,targets)', '单位+目标数组 → void', 'V1', '待办'),
    ('C4.06', S, M, '★击杀后继续推进★', '当target.isDead()时: killer.state=walking → killer.target=null → 继续移动(不消失，核心修复)', 'CombatSystem.handleKill(killer,victim)', '击杀者+被杀者 → void', 'V1', '待办'),
    ('C4.07', S, M, '大殿受伤', 'if hallShield>0: 免疫不扣血 → else: hallHp-=amount → if hallHp<=0: 游戏结束', 'CombatSystem.damageHall(player,amount)', '玩家+伤害量 → void', 'V1', '待办'),
]
M = 'C5.阵法系统'
D += [
    ('C5.01', S, M, '布阵', '检查格子是否在冷却中→if冷却: return false → 创建Formation→加入formations→扣灵力→打出手牌', 'BattleLogic.placeFormation(player,cardId,gx,gy)', '玩家+卡牌+格子坐标 → bool', 'V1', '待办'),
    ('C5.02', S, M, '阵法攻击', '每帧遍历阵法→if isActive: 检查范围内敌方单位→有则按interval攻击(调用CombatSystem)', 'Formation.update(dt,model)', 'dt+状态 → void', 'V1', '待办'),
    ('C5.03', S, M, '阵法冷却', '阵法被毁时记录该格子坐标到formationCooldowns，8秒后过期，期间不能在此格布阵', 'Player.checkCooldown(gx,gy)', '格子坐标 → bool(可否布阵)', 'V1', '待办'),
    ('C5.04', S, M, '阵法被禁', '镇魂符使阵法isActive=false，3秒内不攻击不拦截，计时后恢复', 'Formation.setSilence(duration)', '持续秒 → void', 'V1', '待办'),
]
M = 'C6.法术系统'
D += [
    ('C6.01', S, M, '法术入口', '根据cardId switch分支，调用对应cast方法，扣灵力，打出手牌', 'SpellSystem.cast(cardId,caster,target,model)', '卡牌ID+施法方+目标+状态 → void', 'V1', '待办'),
    ('C6.02', S, M, '万剑归宗', '遍历己方所有单位→atk+=1→addBuff("speed",0.3,5)→持续5秒', 'SpellSystem.castWanJian(caster)', '施法方 → void', 'V1', '待办'),
    ('C6.03', S, M, '五雷正法', '获取目标位置3格内所有敌方单位+阵法→各takeDamage(4)', 'SpellSystem.castWuLei(target,model)', '目标位置+状态 → void', 'V1', '待办'),
    ('C6.04', S, M, '御风诀', 'target.addBuff("speed",0.5,5)→移速+0.5持续5秒', 'SpellSystem.castYuFeng(target)', '目标单位 → void', 'V1', '待办'),
    ('C6.05', S, M, '镇魂符', 'target.setSilence(3)→阵法失效3秒', 'SpellSystem.castZhenHun(target)', '目标阵法 → void', 'V1', '待办'),
    ('C6.06', S, M, '金钟罩', 'caster.hallShield=3→大殿3秒免疫伤害', 'SpellSystem.castJinZhong(caster)', '施法方 → void', 'V1', '待办'),
    ('C6.07', S, M, '移山倒海', '获取目标区域单位→每个y-=2*facing(推后)→takeDamage(1)', 'SpellSystem.castYiShan(targets)', '目标数组 → void', 'V1', '待办'),
    ('C6.08', S, M, '困仙索', 'target.addBuff("stun",0,2)→定身2秒(speed=0)', 'SpellSystem.castKunXian(target)', '目标单位 → void', 'V1', '待办'),
    ('C6.09', S, M, '天雷诀', '获取目标3格内敌方单位→各takeDamage(4)(清兵用)', 'SpellSystem.castTianLei(targets)', '目标数组 → void', 'V1', '待办'),
]
M = 'C7.长老技能'
D += [
    ('C7.01', S, M, '技能计时', '长老每帧elderTimer+=dt→if>=5秒: triggerRandom()→重置timer', 'ElderSkillSystem.update(elder,dt,model)', '长老+dt+状态 → void', 'V1', '待办'),
    ('C7.02', S, M, '★随机分支★', 'branches=[飞剑,丹药,符箓,御兽]→index=Math.floor(Math.random()*4)→调用对应方法', 'ElderSkillSystem.triggerRandom(elder,model)', '长老+状态 → void', 'V1', '待办'),
    ('C7.03', S, M, '飞剑分支', '获取长老周围3格内敌方单位→各takeDamage(3)(范围伤害)', 'ElderSkillSystem.castFlyingSword(elder,model)', '长老+状态 → void', 'V1', '待办'),
    ('C7.04', S, M, '丹药分支', '长老自身hp+=3(不超maxHp)→周围己方单位各hp+=2(不超maxHp)', 'ElderSkillSystem.castPill(elder,model)', '长老+状态 → void', 'V1', '待办'),
    ('C7.05', S, M, '符箓分支', '获取长老前方敌方单位→调用castTianLei(天雷诀清兵)', 'ElderSkillSystem.castTalisman(elder,model)', '长老+状态 → void', 'V1', '待办'),
    ('C7.06', S, M, '御兽分支', '在长老位置召唤1个护山灵兽Unit(owner同长老，向敌方推进)', 'ElderSkillSystem.castBeast(elder,model)', '长老+状态 → void', 'V1', '待办'),
]
M = 'C8.灵力系统'
D += [
    ('C8.01', S, M, '实时回复', '每帧energyTimer+=dt→if>=2.8秒: energy++(不超上限)→重置timer', 'Player.updateEnergy(dt)', 'dt → void', 'V1', '待办'),
    ('C8.02', S, M, '上限增长', 'elapsedTime每30秒→energyMax++(封顶10)→上限提升', 'Player.updateEnergyMax(dt)', 'dt → void', 'V1', '待办'),
    ('C8.03', S, M, '加时倍率', 'if state==overtime: energyTimer增长×1.5(加速回复)', 'Player.updateEnergy(dt)', 'dt → void', 'V1', '待办'),
]
M = 'C9.出牌执行'
D += [
    ('C9.01', S, M, '出兵', '检查灵力→Unit.fromCard创建实例→设置出生y(owner=0在y=0, owner=1在y=8)→加入units→扣灵力→打出手牌', 'BattleLogic.spawnUnit(player,cardId,x)', '玩家+卡牌+列x → Unit', 'V1', '待办'),
    ('C9.02', S, M, '施法', '检查灵力→SpellSystem.cast()→扣灵力→打出手牌', 'BattleLogic.castSpell(player,cardId,target)', '玩家+卡牌+目标 → void', 'V1', '待办'),
]
M = 'C10.胜负判定'
D += [
    ('C10.01', S, M, '大殿摧毁检测', '每帧检查双方hallHp→if任一<=0: 设置winner=对方→state=ended', 'BattleChecker.checkHall(model)', '状态 → 0/1/null', 'V1', '待办'),
    ('C10.02', S, M, '时限检测', 'if time<=0: if双方血量不等→winner=血量高者→else进入加时', 'BattleChecker.checkTime(model)', '状态 → 0/1/null', 'V1', '待办'),
    ('C10.03', S, M, '加时结算', '加时60秒后→if仍平→winner=灵力多者', 'BattleChecker.checkOvertime(model)', '状态 → 0/1/null', 'V1', '待办'),
]
M = 'C11.死亡清理'
D += [
    ('C11.01', S, M, '清理死亡单位', '遍历所有units→if state==dead: 触发死亡效果(自爆)→从数组移除', 'BattleLogic.removeDeadUnits()', 'void → void', 'V1', '待办'),
    ('C11.02', S, M, '清理被毁阵法', '遍历formations→if hp<=0: 记录格子冷却→从数组移除', 'BattleLogic.removeDeadFormations()', 'void → void', 'V1', '待办'),
]

# ============ D. AI系统 ============
S = 'D.AI系统'
M = 'D1.AI决策'
D += [
    ('D1.01', S, M, 'AI主循环', 'thinkTimer+=dt→if>=thinkInterval: think()→重置timer。间隔: easy=3.5s/normal=2.5s/hard=1.5s', 'AI.update(dt,model)', 'dt+状态 → void', 'V1', '待办'),
    ('D1.02', S, M, '★决策核心★', '读取灵力→decideAttackRatio算攻守比→pickCard选牌→if有牌: 根据类型执行spawn/place/cast', 'AI.think(model)', '状态 → void', 'V1', '待办'),
    ('D1.03', S, M, '攻守比计算', 'hpPercent=hallHp/hallMaxHp→if>0.6: return 0.7(偏攻)→if<0.3: return 0.2(偏守)→else: return 0.4', 'AI.decideAttackRatio(model)', '状态 → 0~1', 'V1', '待办'),
    ('D1.04', S, M, '选牌逻辑', '过滤手牌中灵力够的→if ratio>0.5: 优先攻方单位(体修/剑修/御兽)→else: 优先阵法→随机选一张', 'AI.pickCard(energy,ratio)', '灵力+攻守比 → 卡牌index或null', 'V1', '待办'),
    ('D1.05', S, M, '布阵位置选择', 'easy: 随机选阵法区格子→normal: 优先大殿前2~3格→hard: 根据敌方单位位置针对性布阵(高速单位前方布寒霜阵)', 'AI.pickFormationPos(model)', '状态 → {gx,gy}', 'V1', '待办'),
]
M = 'D2.AI难度'
D += [
    ('D2.01', S, M, '简单AI', 'thinkInterval=3.5s→60%出兵40%布阵→随机选牌→随机位置', 'AI.init(playerId,"easy")', 'void → void', 'V1', '待办'),
    ('D2.02', S, M, '普通AI', 'thinkInterval=2.5s→根据血量调攻守比→优先大殿前布阵', 'AI.init(playerId,"normal")', 'void → void', 'V1', '待办'),
    ('D2.03', S, M, '困难AI', 'thinkInterval=1.5s→根据局势选牌→针对性布阵→反应型施法', 'AI.init(playerId,"hard")', 'void → void', 'V1', '待办'),
]

# ============ E. 渲染与UI ============
S = 'E.渲染与UI'
M = 'E1.战斗场景'
D += [
    ('E1.01', S, M, '场景进入', '创建GameModel→创建BattleLogic→注册Input区域→加载资源', 'BattleScene.onEnter()', 'void → void', 'V1', '待办'),
    ('E1.02', S, M, '场景更新', '调用battleLogic.update(dt)驱动游戏逻辑', 'BattleScene.onUpdate(dt)', 'dt → void', 'V1', '待办'),
    ('E1.03', S, M, '场景渲染(分层)', '按顺序: 背景山道→大殿→阵法→单位(y排序)→特效→UI', 'BattleScene.onRender()', 'void → void', 'V1', '待办'),
    ('E1.04', S, M, '场景退出', '清理Input注册→销毁资源→停止逻辑', 'BattleScene.onExit()', 'void → void', 'V1', '待办'),
]
M = 'E2.背景渲染'
D += [
    ('E2.01', S, M, '山道背景', 'V1: 纯色渐变(深绿→浅绿)模拟山道→V1.5: 替换为正式山道图', 'BattleScene.renderBackground()', 'void → void', 'V1', '待办'),
    ('E2.02', S, M, '棋盘格子线', '绘制主路(中间列)+两侧阵法区格子边框线', 'BattleScene.renderGrid()', 'void → void', 'V1', '待办'),
]
M = 'E3.大殿渲染'
D += [
    ('E3.01', S, M, '大殿色块', '顶部/底部各绘制大色块(金色)+宗门名文字', 'BattleScene.renderHalls()', 'void → void', 'V1', '待办'),
    ('E3.02', S, M, '大殿血条', '大殿上方绘制血条(背景红+前景绿，按比例)', 'BattleScene.renderHallHp()', 'void → void', 'V1', '待办'),
    ('E3.03', S, M, '受击特效', '大殿受击时: 屏幕震动(offset随机±3px)→红色闪烁(0.2s)', 'BattleScene.renderHallHit()', 'void → void', 'V1', '待办'),
]
M = 'E4.单位渲染'
D += [
    ('E4.01', S, M, '单位色块', '攻方红/守方蓝色块+名称首字→V1.5替换为Spine动画', 'BattleScene.renderUnits()', 'void → void', 'V1', '待办'),
    ('E4.02', S, M, 'y轴深度排序', '所有单位按y坐标排序后绘制(y小的先画=远的在后，y大的后画=近的在前)', 'BattleScene.renderUnits()', 'void → void', 'V1', '待办'),
    ('E4.03', S, M, '单位血条', '单位上方绘制小血条(宽=单位宽,高=4px)', 'BattleScene.renderUnitHp()', 'void → void', 'V1', '待办'),
    ('E4.04', S, M, '状态指示', 'fighting时单位闪烁→有buff时显示图标(减速蓝/加速绿/定身灰)', 'BattleScene.renderUnitState()', 'void → void', 'V1', '待办'),
]
M = 'E5.阵法渲染'
D += [
    ('E5.01', S, M, '阵法光阵', '阵法区格子绘制半透明青色光阵+阵法名', 'BattleScene.renderFormations()', 'void → void', 'V1', '待办'),
    ('E5.02', S, M, '阵法状态', '正常: 完整光阵→受击: 闪烁→被禁: 灰暗→冷却: 半透明+倒计时数字', 'BattleScene.renderFormationState()', 'void → void', 'V1', '待办'),
]
M = 'E6.特效渲染'
D += [
    ('E6.01', S, M, '命中粒子', '攻击命中时在目标位置emit 4~6个粒子(0.3s寿命)', 'ParticleSystem.emit(x,y,4,"#FF6B6B")', '坐标+数量+颜色 → void', 'V1', '待办'),
    ('E6.02', S, M, '死亡碎裂', '单位死亡时emit 8个粒子向四周散开(0.5s)', 'ParticleSystem.emit(x,y,8,color)', '坐标+数量+颜色 → void', 'V1', '待办'),
    ('E6.03', S, M, '法术光效', '法术释放时在目标区域绘制全屏光效(0.8s渐隐)', 'BattleScene.renderSpellFx()', 'void → void', 'V1', '待办'),
]
M = 'E7.UI-手牌栏'
D += [
    ('E7.01', S, M, '手牌显示', '底部横排4张牌: 色块+卡名+费用→选中时上浮20px+发光边框', 'HandBar.render()', 'void → void', 'V1', '待办'),
    ('E7.02', S, M, '灰显不可出', '灵力<费用时: 手牌灰度+不可点击', 'HandBar.updateGrey(energy)', '灵力 → void', 'V1', '待办'),
    ('E7.03', S, M, '点击选中', '点击手牌→selectedCard=index→上浮→高亮可放置区域', 'HandBar.onTap(index)', '手牌index → void', 'V1', '待办'),
    ('E7.04', S, M, '出牌动画', '出牌后: 手牌飞向战场→手牌位置留空→2秒后新牌滑入补位', 'HandBar.playCardAnim(index)', '手牌index → void', 'V1', '待办'),
]
M = 'E8.UI-灵力条'
D += [
    ('E8.01', S, M, '灵力显示', '手牌上方: 紫色液条+数字(如5/7)→液条宽度按energy/energyMax比例', 'EnergyBar.render()', 'void → void', 'V1', '待办'),
    ('E8.02', S, M, '回复动画', '灵力+1时: 液条有上升动画+数字跳动', 'EnergyBar.playGainAnim()', 'void → void', 'V1', '待办'),
]
M = 'E9.UI-顶部HUD'
D += [
    ('E9.01', S, M, '双方血条', '顶部左(我方)/右(敌方)各一条大殿血条', 'HUD.renderHallBars()', 'void → void', 'V1', '待办'),
    ('E9.02', S, M, '计时器', '顶部中央: 倒计时数字(180→0)→最后30秒变红色', 'HUD.renderTimer()', 'void → void', 'V1', '待办'),
    ('E9.03', S, M, '加时提示', '进入加时时: 中央显示"加时赛"+60秒倒计时', 'HUD.renderOvertime()', 'void → void', 'V1', '待办'),
]
M = 'E10.出牌交互'
D += [
    ('E10.01', S, M, '手牌→选目标', '选中手牌后→根据卡牌类型高亮不同区域(单位:主路/阵法:阵法区/法术:全屏)→点击目标执行', 'BattleScene.onCardTap(index)', '手牌index → void', 'V1', '待办'),
    ('E10.02', S, M, '格子点击', '点击格子→if selectedCard!=-1: 根据卡牌类型执行spawn/place/cast→清除选中', 'BattleScene.onGridTap(gx,gy)', '格子坐标 → void', 'V1', '待办'),
    ('E10.03', S, M, '取消选择', '再次点击已选中手牌→取消选中→清除高亮', 'BattleScene.onCardTap(index)', '手牌index → void', 'V1', '待办'),
]
M = 'E11.结算场景'
D += [
    ('E11.01', S, M, '胜负展示', '大字体"胜利"/"失败"+动画(0.5s缩放)', 'ResultScene.renderResult()', 'void → void', 'V1', '待办'),
    ('E11.02', S, M, '摧毁度', '显示双方摧毁度百分比(我方XX% vs 敌方XX%)', 'ResultScene.renderDestroyRate()', 'void → void', 'V1', '待办'),
    ('E11.03', S, M, '再来一局', '按钮→点击后切换回BattleScene重新初始化', 'ResultScene.onReplay()', 'void → void', 'V1', '待办'),
]
M = 'E12.新手引导'
D += [
    ('E12.01', S, M, '引导1-出牌', '高亮第一张手牌+箭头→"点击出兵"→玩家点击后进入下一步', 'TutorialGuide.step1()', 'void → void', 'V1', '待办'),
    ('E12.02', S, M, '引导2-灵力', '高亮灵力条→"灵力不够时等待回复"→等3秒后下一步', 'TutorialGuide.step2()', 'void → void', 'V1', '待办'),
    ('E12.03', S, M, '引导3-布阵', '高亮阵法区格子→"点击布阵拦截"→玩家布阵后下一步', 'TutorialGuide.step3()', 'void → void', 'V1', '待办'),
    ('E12.04', S, M, '引导4-目标', '箭头指向敌方大殿→"摧毁大殿获胜"→2秒后结束引导', 'TutorialGuide.step4()', 'void → void', 'V1', '待办'),
]

# ============ F. 社交(V1.5) ============
S = 'F.社交系统(V1.5)'
M = 'F1.登录存档'
D += [
    ('F1.01', S, M, '微信登录', 'wx.login()→获取code→云函数换取openid→存到本地', 'CloudManager.login()', 'void → openid', 'V1.5', '待办'),
    ('F1.02', S, M, '云存档', 'wx.cloud.database()→存用户档案{uid,nickname,灵石,仙玉,卡牌,卡组,战绩}', 'CloudManager.save(data)', '存档数据 → void', 'V1.5', '待办'),
    ('F1.03', S, M, '读档', '云数据库where openid→get→返回用户档案', 'CloudManager.load()', 'void → 存档数据', 'V1.5', '待办'),
]
M = 'F2.异步PvP'
D += [
    ('F2.01', S, M, '布阵上传', '玩家布阵完成后→将阵法配置上传到云数据库→生成挑战链接', 'AsyncPvP.uploadLayout(layout)', '布阵数据 → 挑战ID', 'V1.5', '待办'),
    ('F2.02', S, M, '加载对手', '通过挑战ID→云数据库读取对手布阵→在本地模拟AI攻打', 'AsyncPvP.loadOpponent(challengeId)', '挑战ID → 布阵数据', 'V1.5', '待办'),
    ('F2.03', S, M, '回放', '记录对战过程→生成回放数据→对方可观看', 'AsyncPvP.saveReplay(battleLog)', '战斗日志 → 回放ID', 'V1.5', '待办'),
]
M = 'F3.排行榜分享'
D += [
    ('F3.01', S, M, '好友排行', 'wx.getOpenDataContext()→开放数据域→按胜场/段位排序', 'RankManager.renderFriendRank()', 'void → void', 'V1.5', '待办'),
    ('F3.02', S, M, '分享', 'wx.shareAppMessage→设置title/imagePath→玩家分享到聊天', 'ShareManager.share(title,img)', '标题+图片 → void', 'V1.5', '待办'),
]

# ============ G. 养成变现(V1.5) ============
S = 'G.养成变现(V1.5)'
M = 'G1.养成系统'
D += [
    ('G1.01', S, M, '弟子升级', '消耗灵石→level++→hp/atk按成长系数提升', 'ProgressionSystem.levelUp(unitId)', '单位ID → void', 'V1.5', '待办'),
    ('G1.02', S, M, '弟子升星', '消耗剑魄→突破境界(练气→筑基→金丹→元婴)→解锁更高上限+新技能', 'ProgressionSystem.starUp(unitId)', '单位ID → void', 'V1.5', '待办'),
    ('G1.03', S, M, '卡牌收集', '宝箱随机掉落卡牌碎片→集齐N个碎片→合成新卡', 'ProgressionSystem.mergeCard(cardId)', '卡牌ID → void', 'V1.5', '待办'),
]
M = 'G2.变现系统'
D += [
    ('G2.01', S, M, '激励广告', 'wx.createRewardedVideoAd→show→onClose回调发奖励(灵石/宝箱)', 'AdManager.showReward(cb)', '回调 → void', 'V1.5', '待办'),
    ('G2.02', S, M, '内购', 'wx.requestPayment→下单→支付→回调发仙玉', 'IAPManager.buy(productId,cb)', '商品ID+回调 → void', 'V1.5', '待办'),
    ('G2.03', S, M, '通行证', '赛季任务进度→免费/付费双线奖励→付费需内购解锁', 'BattlePass.update(taskId)', '任务ID → void', 'V1.5', '待办'),
]

# ============ H. 实时PvP(V2) ============
S = 'H.实时PvP(V2)'
M = 'H1.网络通信'
D += [
    ('H1.01', S, M, 'WebSocket连接', 'wx.connectSocket→onOpen/onMessage/onClose→心跳保活(每10s ping)', 'NetworkClient.connect(url)', 'URL → void', 'V2', '待办'),
    ('H1.02', S, M, '断线重连', 'onClose时记录游戏状态→自动重连→恢复对局', 'NetworkClient.reconnect()', 'void → void', 'V2', '待办'),
]
M = 'H2.状态同步'
D += [
    ('H2.01', S, M, '操作上报', '玩家出牌/布阵/施法→封装操作→send到服务端', 'SyncManager.sendAction(action)', '操作数据 → void', 'V2', '待办'),
    ('H2.02', S, M, '状态接收', '服务端广播状态→本地应用→渲染', 'SyncManager.onState(state)', '状态数据 → void', 'V2', '待办'),
]
M = 'H3.匹配段位'
D += [
    ('H3.01', S, M, '匹配', '上报ELO→服务端匹配相近对手→返回房间ID', 'MatchMaker.match(elo)', 'ELO → 房间ID', 'V2', '待办'),
    ('H3.02', S, M, '段位赛', '青铜→白银→黄金→...→宗师，胜负加减分', 'RankSystem.updateRank(win)', '是否胜利 → void', 'V2', '待办'),
]

# ===== 写入数据 =====
r = 4
current_sys = ''
current_mod = ''
sys_fills = {
    'A.引擎框架': BG_L1_A,
    'B.游戏数据': BG_L1_B,
    'C.战斗逻辑': BG_L1_C,
    'D.AI系统': BG_L1_D,
    'E.渲染与UI': BG_L1_E,
}

for row_data in D:
    num, sys, mod, func, principle, fn, io, ver, status = row_data

    # 写入
    vals = [num, sys, mod, func, principle, fn, io, ver, status]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.border = BD
        c.alignment = A_LT
        # 编号列
        if i == 1:
            c.font = F_CODE; c.alignment = A_C
        # 系统大类列
        elif i == 2:
            c.font = F_L2
        # 模块列
        elif i == 3:
            c.font = F_L3
        # 子功能列
        elif i == 4:
            c.font = F_BOLD
        # 实现原理列
        elif i == 5:
            c.font = F_CELL
        # 函数列
        elif i == 6:
            c.font = F_CODE
        # 输入输出列
        elif i == 7:
            c.font = F_CODE
        # 版本列
        elif i == 8:
            c.font = F_CELL; c.alignment = A_C
        # 状态列
        elif i == 9:
            c.font = F_BOLD; c.alignment = A_C
            if v == '待办': c.fill = BG_TODO
            elif v == '进行中': c.fill = BG_DOING
            elif v == '已完成': c.fill = BG_DONE

    # 系统大类颜色条
    if sys != current_sys:
        current_sys = sys
    ws.cell(row=r, column=2).fill = sys_fills.get(sys, BG_L2)
    ws.cell(row=r, column=2).font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    ws.cell(row=r, column=2).alignment = A_C

    # 行高
    ws.row_dimensions[r].height = 42
    r += 1

# 冻结表头
ws.freeze_panes = 'A4'

# 条件格式：状态列
from openpyxl.formatting.rule import FormulaRule
ws.conditional_formatting.add(f'I4:I{r-1}',
    FormulaRule(formula=['$I4="已完成"'], fill=BG_DONE))
ws.conditional_formatting.add(f'I4:I{r-1}',
    FormulaRule(formula=['$I4="进行中"'], fill=BG_DOING))

# 底部统计
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
c = ws.cell(row=r, column=1, value=f'共 {len(D)} 个开发点 | 5个系统大类 + V1.5/V2扩展 | 按编号顺序从上往下做，做完一个把"状态"改为"已完成"')
c.font = F_NOTE; c.alignment = A_L

# ================================================================
# 新 Sheet: 进度统计（自动计算）
# ================================================================
ws2 = wb.create_sheet('3.进度统计')
set_w(ws2, [16, 10, 10, 10, 10, 12])
title(ws2, '进度统计 — 自动汇总', 6)
header(ws2, ['系统大类', '总点数', '已完成', '进行中', '待办', '完成率'], 3)

# 统计
from collections import OrderedDict
stats = OrderedDict()
for row_data in D:
    sys = row_data[1]
    if sys not in stats:
        stats[sys] = {'total': 0, 'done': 0, 'doing': 0, 'todo': 0}
    stats[sys]['total'] += 1
    s = row_data[8]
    if s == '已完成': stats[sys]['done'] += 1
    elif s == '进行中': stats[sys]['doing'] += 1
    else: stats[sys]['todo'] += 1

r = 4
for sys, s in stats.items():
    rate = f"{s['done']}/{s['total']} ({s['done']*100//s['total']}%)"
    for i, v in enumerate([sys, s['total'], s['done'], s['doing'], s['todo'], rate], 1):
        c = ws2.cell(row=r, column=i, value=v)
        c.font = F_CELL; c.alignment = A_C; c.border = BD
    ws2.cell(row=r, column=1).fill = sys_fills.get(sys, BG_L2)
    ws2.cell(row=r, column=1).font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    ws2.row_dimensions[r].height = 24
    r += 1

# 总计
t = sum(s['total'] for s in stats.values())
d = sum(s['done'] for s in stats.values())
di = sum(s['doing'] for s in stats.values())
td = sum(s['todo'] for s in stats.values())
rate = f"{d}/{t} ({d*100//t if t else 0}%)"
for i, v in enumerate(['总计', t, d, di, td, rate], 1):
    c = ws2.cell(row=r, column=i, value=v)
    c.font = F_HEADER; c.fill = BG_HEADER; c.alignment = A_C; c.border = BD
ws2.row_dimensions[r].height = 26

# 保存
wb.save(PATH)
print(f'Excel已更新: {PATH}')
print(f'总Sheet数: {len(wb.sheetnames)}')
print(f'Sheet列表: {wb.sheetnames}')
print(f'开发主表: {len(D)}个开发点')
sys_count = len(set(row[1] for row in D))
mod_count = len(set((row[1], row[2]) for row in D))
print(f'系统大类: {sys_count}个 | 模块: {mod_count}个')
for sys, s in stats.items():
    print(f'  {sys}: {s["total"]}点 (已完成{s["done"]}/进行中{s["doing"]}/待办{s["todo"]})')
