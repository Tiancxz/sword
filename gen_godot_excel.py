#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""《宗门论道》Godot版 — Excel生成脚本 v2（优化美观版）
10个Sheet：1.阶段规划 2.开发主表 3.进度统计 4.卡牌数据 5.数值配置
6.AI提问指南 7.资源清单 8.决策日志 9.跨平台导出 10.设计思路"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== 加载数据 =====
with open('/workspace/_data_godot.json') as f:
    D = json.load(f)
L3_ROWS = [d for d in D if d[0] == 'L3']
print(f'加载 {len(L3_ROWS)} 个L3任务')

# ===== 配色方案（统一蓝色系） =====
C_TITLE_BG = '1F3864'      # 标题深蓝
C_HEAD_BG = '2E5C8A'       # 表头中蓝
C_L1_BG = '1F4E79'         # L1深蓝
C_L2_BG = 'BDD7EE'         # L2浅蓝
C_L3_BG = 'FFFFFF'         # L3白
C_L3_ALT_BG = 'F2F7FC'     # L3浅灰蓝交替
C_CODE_BG = 'FFF8E1'       # 代码浅黄
C_CRIT_BG = 'FFF3E0'       # 完成标准浅橙
C_PHASE_BG = 'E8EEF4'      # 阶段行浅灰蓝
C_DONE = 'C6EFCE'
C_TODO = 'FCE4D6'
C_DOING = 'FFF2CC'
C_WHITE = 'FFFFFF'
C_DARK = '1F3864'
C_TEXT = '333333'

FONT = '微软雅黑'
MONO = 'Consolas'

# ===== 样式工厂 =====
def mk_font(size=10, bold=False, color=C_TEXT, name=FONT):
    return Font(name=name, size=size, bold=bold, color=color)

def mk_fill(color):
    return PatternFill('solid', fgColor=color)

def mk_align(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

thin = Side(style='thin', color='C0C0C0')
medium = Side(style='medium', color=C_HEAD_BG)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_TOP_THICK = Border(left=thin, right=thin, top=medium, bottom=thin)

def style_cell(cell, font=None, fill=None, align=None, border=BORDER):
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border

# ===== 行高计算 =====
def calc_row_height(texts, col_widths):
    """根据文本内容和列宽计算所需行高"""
    max_lines = 1
    for text, width in zip(texts, col_widths):
        if not text: continue
        text = str(text)
        # 每行能容纳的字符数（中文算2，英文算1，粗略按width*1.5）
        chars_per_line = max(1, int(width * 1.5))
        # 计算需要的行数（考虑显式换行）
        lines = text.split('\n')
        total = 0
        for line in lines:
            total += max(1, (len(line) + chars_per_line - 1) // chars_per_line)
        max_lines = max(max_lines, total)
    return max(22, min(120, max_lines * 16 + 6))

wb = Workbook()

# ============================================================
# Sheet 1: 阶段规划
# ============================================================
ws1 = wb.active; ws1.title = '1.阶段规划'
ws1.merge_cells('A1:H1')
style_cell(ws1['A1'], mk_font(16, True, C_WHITE), mk_fill(C_TITLE_BG), mk_align('center', 'center'))
ws1['A1'] = '《宗门论道》开发阶段规划（Godot引擎版）'
ws1.row_dimensions[1].height = 38

headers = ['阶段', '名称', '目标', '包含模块', '核心产出', '前置条件', '工时h', '里程碑']
for i, h in enumerate(headers, 1):
    style_cell(ws1.cell(row=2, column=i, value=h),
               mk_font(11, True, C_WHITE), mk_fill(C_HEAD_BG), mk_align('center', 'center'))
ws1.row_dimensions[2].height = 28

phases = [
    ('P0', '环境搭建', '安装Godot+插件+导出工具链', 'Godot 4.x\ngodot_for_minigame\nGodotSteam', '开发环境就绪', '—', 1.0, 'M0 环境就绪'),
    ('P1', '项目骨架', '搭建Godot骨架和autoload单例', 'A1主场景\nA2输入 A3事件总线\nA4场景 A5渲染辅助', 'Godot项目可运行', 'P0', 2.0, 'M1 空项目跑通'),
    ('P2', '游戏数据', '配置驱动的卡牌/卡组数据', 'B1常量\nB2卡牌 B3卡组\nB4预设', '卡牌数据可查可抽', 'P1', 3.0, 'M2 数据层完成'),
    ('P3', '核心战斗', '实时双向推进+碰撞+战斗', 'C1循环 C2单位\nC3移动 C4战斗', '单位能走能打', 'P2', 4.0, 'M3 核心玩法可玩'),
    ('P4', '扩展战斗', '阵法/法术/长老/灵力/胜负', 'C5-C11 阵法法术\nD1 AI决策', '完整战斗流程', 'P3', 6.0, 'M4 完整对战'),
    ('P5', 'AI系统', 'AI自动出牌布阵', 'D1 AI决策', 'PvE可玩', 'P4', 2.3, 'M5 AI对战可玩'),
    ('P6', '渲染与UI', '完整画面+手牌+HUD+引导', 'E1-E11 渲染UI', '游戏视觉完整', 'P4', 6.5, 'M6 视觉完整'),
    ('P8', '跨平台导出', '微信/Steam/安卓三平台', 'I1抽象层\nI2微信 I3Steam\nI4安卓', '三平台可导出', 'P6', 10.0, 'M7 多平台导出'),
    ('P9-P14', '社交养成变现', '登录/存档/排行/广告/内购 (V1.5)', 'F1登录 F2异步PvP\nF3排行 G1养成 G2变现', 'V1.5上线', 'P8', 8.0, 'M8 V1.5上线'),
    ('P16-P18', '实时PvP (V2)', 'WebSocket+同步+匹配+段位', 'H1网络 H2同步\nH3匹配段位', 'V2上线', 'P9', 6.0, 'M9 V2上线'),
]
col_widths_1 = [10, 16, 30, 24, 22, 12, 9, 18]
for i, w in enumerate(col_widths_1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

r = 3
for p in phases:
    for c, v in enumerate(p, 1):
        cell = ws1.cell(row=r, column=c, value=v)
        if c <= 2:
            style_cell(cell, mk_font(11, True, C_DARK), mk_fill(C_PHASE_BG), mk_align('center', 'center'))
        elif c == 7:
            style_cell(cell, mk_font(11, True), mk_fill(C_L3_BG), mk_align('center', 'center'))
        else:
            style_cell(cell, mk_font(10), mk_fill(C_L3_BG if r % 2 else C_L3_ALT_BG), mk_align('left', 'center'))
    ws1.row_dimensions[r].height = calc_row_height(list(p), col_widths_1)
    r += 1
# 合计行
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
style_cell(ws1.cell(row=r, column=1, value='合计'), mk_font(12, True, C_WHITE), mk_fill(C_L1_BG), mk_align('center', 'center'))
style_cell(ws1.cell(row=r, column=7, value=f'=SUM(G3:G{r-1})'), mk_font(12, True, C_WHITE), mk_fill(C_L1_BG), mk_align('center', 'center'))
ws1.merge_cells(start_row=r, start_column=8, end_row=r, end_column=8)
style_cell(ws1.cell(row=r, column=8, value='总工时'), mk_font(11, True, C_WHITE), mk_fill(C_L1_BG), mk_align('center', 'center'))
ws1.row_dimensions[r].height = 30
ws1.freeze_panes = 'A3'

# ============================================================
# Sheet 2: 开发主表
# ============================================================
MAIN = '2.开发主表'
ws2 = wb.create_sheet(MAIN)
ws2.merge_cells('A1:O1')
style_cell(ws2['A1'], mk_font(16, True, C_WHITE), mk_fill(C_TITLE_BG), mk_align('center', 'center'))
ws2['A1'] = '《宗门论道》开发主表（Godot GDScript版）'
ws2.row_dimensions[1].height = 38

main_headers = ['层级', '编号', '阶段', '系统', '模块', '任务名称', '实现原理', '函数签名', '文件路径', '输入输出', '前置依赖', '工时h', '完成标准', '版本', '状态']
for i, h in enumerate(main_headers, 1):
    style_cell(ws2.cell(row=2, column=i, value=h),
               mk_font(10, True, C_WHITE), mk_fill(C_HEAD_BG), mk_align('center', 'center'))
ws2.row_dimensions[2].height = 32

# 完成标准字典
CRITERIA = {
    'A1.01':'①Godot项目能打开 ②_process每帧打印 ③运行无报错',
    'A1.02':'①场景切换后on_update调用 ②delta值合理(0.016)',
    'A1.03':'①_draw绘制矩形 ②queue_redraw触发重绘',
    'A1.04':'①_ready初始化autoload ②自动进入初始场景',
    'A2.01':'①InputEventScreenTouch响应 ②print坐标 ③多平台触屏OK',
    'A2.02':'①register后点击回调 ②Rect2命中准确',
    'A2.03':'①只触发命中的cb ②坐标传递正确',
    'A2.04':'①clear后旧cb不触发 ②场景切换无残留',
    'A3.01':'①emit信号能被connect收到 ②参数传递正确',
    'A3.02':'①任何节点可触发 ②监听者收到事件',
    'A3.03':'①_ready中connect ②_exit_tree中断开 ③无重复连接',
    'A4.01':'①register后switch能切换 ②路径正确',
    'A4.02':'①切换时_exit_tree→_enter_tree ②旧场景释放',
    'A5.01':'①draw_rect填充正确 ②颜色可配',
    'A5.02':'①draw_string渲染文字 ②字号颜色正确',
    'A5.03':'①血条比例正确 ②绿底红前',
    'A6.01':'①CPUParticles2D发射 ②数量颜色正确',
    'A6.02':'①emit后粒子播放 ②自动停止',
    'A7.01':'①BGM循环播放 ②音量可调',
    'A7.02':'①SFX播放一次 ②不阻塞主线程',
    'B1.01':'①Const可全局引用 ②数值与文档一致',
    'B2.01':'①20张卡牌都有cardId ②字段完整 ③static可查',
    'B2.02':'①存在返回字典 ②不存在返回{}+报错',
    'B3.01':'①shuffle后顺序随机 ②初始手牌3张',
    'B3.02':'①多次shuffle分布均匀',
    'B3.03':'①手牌满返回"" ②牌库空返回"" ③正常返回cardId',
    'B3.04':'①2秒后自动抽牌 ②手牌不超过4',
    'B3.05':'①灵力不足返回false ②足够返回true',
    'B3.06':'①打出后从手牌移除 ②2秒后补牌',
    'B4.01':'①3套卡组各8张 ②卡组内卡牌都存在',
    'C1.01':'①每帧所有子系统更新 ②time递减 ③无漏调',
    'C2.01':'①从cardId创建单位 ②属性正确 ③位置正确',
    'C2.02':'①buff添加成功 ②duration倒计时 ③到期移除',
    'C2.03':'①slow减0.5 ②stun返回0 ③最低为0',
    'C2.04':'①扣血正确 ②hp到0标记dead ③返回实际伤害',
    'C3.01':'①walking时y递增 ②遇敌变fighting ③到殿变dead',
    'C3.02':'①优先同列最近敌 ②无敌返回null',
    'C3.03':'①距离<=range命中 ②跨列不命中',
    'C3.04':'①攻方到顶 ②守方到底 ③触发大殿伤害',
    'C4.01':'①fighting时定期攻击 ②间隔正确 ③击杀后推进',
    'C4.02':'①双方同时掉血 ②伤害值正确',
    'C4.03':'①只有目标掉血 ②攻击者无损',
    'C4.04':'①范围内全部受伤 ②伤害值正确',
    'C4.05':'①击杀者state→walking ②target→null ③继续前进',
    'C4.06':'①护盾优先 ②hp扣减 ③到0触发结束',
    'C5.01':'①检查冷却 ②扣灵力 ③Area2D创建',
    'C5.02':'①敌人进入触发攻击 ②owner判断正确',
    'C5.03':'①8秒冷却 ②同位置不能立即再布',
    'C5.04':'①禁用期间不攻击 ②到期恢复',
    'C6.01':'①match分支正确 ②参数传递OK',
    'C6.02':'①全员+1攻 ②加速5秒',
    'C6.03':'①范围3格 ②伤害4',
    'C6.04':'①目标+0.5速 ②持续5秒',
    'C6.05':'①目标禁用3秒',
    'C6.06':'①护盾+3 ②大殿不受伤',
    'C6.07':'①目标后退2格 ②受1伤',
    'C6.08':'①目标stun 2秒 ②不移动不攻击',
    'C6.09':'①范围内伤害4',
    'C7.01':'①每5秒触发 ②非长老不触发',
    'C7.02':'①四分支等概率 ②每次随机',
    'C7.03':'①范围3格 ②AOE伤害3',
    'C7.04':'①自身+3血 ②附近友军+2血 ③不超上限',
    'C7.05':'①前方3格 ②天雷伤害4',
    'C7.06':'①召唤兽 ②位置正确 ③属于同方',
    'C8.01':'①2.8秒+1灵力 ②不超上限 ③加时加快',
    'C8.02':'①每30秒+1上限 ②封顶10',
    'C9.01':'①扣灵力 ②创建单位 ③加入units',
    'C9.02':'①扣灵力 ②调用SpellSystem ③移除手牌',
    'C10.01':'①hp到0返回对手 ②否则-1',
    'C10.02':'①时间到判血量 ②平局进加时',
    'C10.03':'①加时到判灵力 ②返回胜者',
    'C11.01':'①dead单位移除 ②kamikaze触发 ③queue_free',
    'C11.02':'①hp=0阵法移除 ②8秒冷却',
    'D1.01':'①每3秒思考 ②不卡帧',
    'D1.02':'①能出兵/布阵/施法 ②灵力不足时等待',
    'D1.03':'①血多偏攻 ②血少偏守 ③比例正确',
    'D1.04':'①从手牌选可出 ②有单位优先',
    'D1.05':'①easy随机 ②normal守家 ③hard堵路',
    'E1.01':'①场景初始化 ②注册输入 ③模型就绪',
    'E1.02':'①_process驱动 ②子节点更新',
    'E1.03':'①分层渲染正确 ②YSort排序OK',
    'E1.04':'①清理输入 ②释放模型 ③断开信号',
    'E2.01':'①渐变背景 ②颜色正确',
    'E2.02':'①格子线 ②间距正确',
    'E3.01':'①Sprite显示 ②位置正确',
    'E3.02':'①血条比例 ②颜色区分',
    'E3.03':'①Tween抖动 ②0.05秒',
    'E4.01':'①Sprite+YSort ②颜色区分阵营',
    'E4.02':'①血条上方 ②比例正确',
    'E5.01':'①半透明 ②禁用时变暗',
    'E6.01':'①4张牌显示 ②不可出变灰',
    'E6.02':'①点击高亮 ②再点取消',
    'E7.01':'①灵力数实时 ②比例正确',
    'E8.01':'①双方血条 ②比例正确',
    'E8.02':'①倒计时 ②30秒变红',
    'E9.01':'①选牌 ②灵力不足不可选',
    'E9.02':'①选目标执行 ②执行后取消选中',
    'E10.01':'①胜负文字 ②颜色区分',
    'E10.02':'①摧毁度百分比 ②双方对比',
    'E10.03':'①点击重开 ②场景重置',
    'E11.01':'①高亮第一张 ②点击后进step2',
    'E11.02':'①await 3秒 ③进step3',
    'E11.03':'①高亮阵法区 ②布阵后进step4',
    'E11.04':'①画箭头 ②2秒后引导结束',
    'F1.01':'①微信登录 ②Steam登录 ③安卓登录 都能拿openid',
    'F1.02':'①微信云存档 ②Steam云 ③安卓云 统一接口',
    'F2.01':'①布局上传 ②返回challengeId',
    'F2.02':'①下载布局 ②本地AI模拟对战',
    'F3.01':'①好友排行显示 ②分数正确',
    'F3.02':'①分享链接 ②图片正确',
    'G1.01':'①升级属性增长 ②数值正确',
    'G1.02':'①升星 ②境界名称变化',
    'G2.01':'①广告看完回调 ②跨平台统一',
    'G2.02':'①支付成功回调 ②跨平台统一',
    'H1.01':'①WebSocket连接 ②能收发消息',
    'H1.02':'①断线保存 ②自动重连 ③状态恢复',
    'H2.01':'①操作上报 ②服务端收到',
    'H2.02':'①状态接收 ②应用同步 ③无冲突',
    'H3.01':'①匹配成功 ②返回房间号',
    'H3.02':'①elo增减 ②段位更新',
    'I1.01':'①检测平台正确 ②各平台返回对应枚举',
    'I1.02':'①微信存 ②Steam存 ③安卓存 都OK',
    'I1.03':'①微信广告 ②安卓广告 OK ③Steam无广告返回false',
    'I1.04':'①微信支付 ②Steam支付 ③安卓支付 都OK',
    'I2.01':'①插件启用 ②Export面板出现',
    'I2.02':'①导出预设 ②资源筛选正确',
    'I2.03':'①wx.* API可用 ②编译无错',
    'I2.04':'①主包<4MB ②分包加载成功',
    'I2.05':'①微信工具打开 ②真机预览 ③可提审',
    'I3.01':'①Steamworks加载 ②API可调',
    'I3.02':'①成就解锁 ②Steam后台显示',
    'I3.03':'①云存档写入 ②Steam云同步',
    'I3.04':'①.exe导出 ②可运行',
    'I4.01':'①.aab导出 ②签名正确',
    'I4.02':'①触屏OK ②安全区适配',
    'I4.03':'①谷歌支付 ②回调正确',
    'I4.04':'①上架成功 ②可下载',
}

col_widths_2 = [5, 7, 5, 5, 14, 16, 38, 26, 24, 14, 12, 6, 30, 6, 7]
for i, w in enumerate(col_widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

r = 3
alt = False
for d in D:
    if d[0] == 'L1':
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
        cell = ws2.cell(row=r, column=1, value=f'  {d[4]}')
        style_cell(cell, mk_font(13, True, C_WHITE), mk_fill(C_L1_BG), mk_align('left', 'center'))
        for c in range(2, 16):
            ws2.cell(row=r, column=c).border = BORDER
            ws2.cell(row=r, column=c).fill = mk_fill(C_L1_BG)
        ws2.row_dimensions[r].height = 28
        r += 1; alt = False; continue
    if d[0] == 'L2':
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
        cell = ws2.cell(row=r, column=1, value=f'    ▸ {d[5]}  [{d[2]}]')
        style_cell(cell, mk_font(11, True, C_DARK), mk_fill(C_L2_BG), mk_align('left', 'center'))
        for c in range(2, 16):
            ws2.cell(row=r, column=c).border = BORDER
            ws2.cell(row=r, column=c).fill = mk_fill(C_L2_BG)
        ws2.row_dimensions[r].height = 24
        r += 1; alt = False; continue
    # L3
    code = d[1]
    criteria = CRITERIA.get(code, '')
    values = [d[0], d[1], d[2], d[3], d[5], d[6], d[7], d[8], d[9], d[10], d[11], d[12], criteria, d[13], d[14]]
    bg = C_L3_ALT_BG if alt else C_L3_BG
    for c, v in enumerate(values, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        if c in (7, 8, 9):  # 实现原理/函数签名/文件路径
            style_cell(cell, mk_font(9, name=MONO), mk_fill(C_CODE_BG), mk_align('left', 'center'))
        elif c == 13:  # 完成标准
            style_cell(cell, mk_font(9, color='7F6000'), mk_fill(C_CRIT_BG), mk_align('left', 'center'))
        elif c in (1, 2, 3, 12, 14, 15):  # 居中列
            style_cell(cell, mk_font(10), mk_fill(bg), mk_align('center', 'center'))
        else:
            style_cell(cell, mk_font(10), mk_fill(bg), mk_align('left', 'center'))
        if c == 15:  # 状态着色
            if v == '待办': cell.fill = mk_fill(C_TODO)
            elif v == '进行中': cell.fill = mk_fill(C_DOING)
            elif v == '完成': cell.fill = mk_fill(C_DONE)
    ws2.row_dimensions[r].height = calc_row_height(values, col_widths_2)
    r += 1; alt = not alt
last_row = r - 1
ws2.freeze_panes = 'G3'
ws2.auto_filter.ref = f'A2:O{last_row}'

# ============================================================
# Sheet 3: 进度统计
# ============================================================
ws3 = wb.create_sheet('3.进度统计')
ws3.merge_cells('A1:F1')
style_cell(ws3['A1'], mk_font(16, True, C_WHITE), mk_fill(C_TITLE_BG), mk_align('center', 'center'))
ws3['A1'] = '进度统计（自动计算）'
ws3.row_dimensions[1].height = 38

for i, h in enumerate(['阶段', '总任务数', '完成数', '进行中', '待办', '完成率'], 1):
    style_cell(ws3.cell(row=2, column=i, value=h),
               mk_font(11, True, C_WHITE), mk_fill(C_HEAD_BG), mk_align('center', 'center'))
ws3.row_dimensions[2].height = 28

phase_map = [
    ('P1', 'P1 项目骨架'), ('P2', 'P2 游戏数据'), ('P3', 'P3 核心战斗'), ('P4', 'P4 扩展战斗'),
    ('P5', 'P5 AI系统'), ('P6', 'P6 渲染UI'), ('P7', 'P0 环境搭建'), ('P8', 'P8 跨平台导出'),
    ('P9', 'P9-14 社交养成'), ('P16', 'P16-18 实时PvP'),
]
phase_codes = {
    'P1': ['P1'], 'P2': ['P2'], 'P3': ['P3'], 'P4': ['P4'], 'P5': ['P5'],
    'P6': ['P6'], 'P7': ['P7'], 'P8': ['P8'],
    'P9': ['P9','P10','P11','P12','P13','P14'], 'P16': ['P16','P17','P18'],
}
col_widths_3 = [22, 11, 11, 11, 11, 11]
for i, w in enumerate(col_widths_3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

r = 3
for pid, pname in phase_map:
    codes = phase_codes[pid]
    code_str = ','.join([f'"{c}"' for c in codes])
    total_f = f'=SUMPRODUCT(--ISNUMBER(MATCH({MAIN}!C3:C{last_row},{{{code_str}}},0)))'
    done_f = f'=SUMPRODUCT(--ISNUMBER(MATCH({MAIN}!C3:C{last_row},{{{code_str}}},0))*({MAIN}!O3:O{last_row}="完成"))'
    doing_f = f'=SUMPRODUCT(--ISNUMBER(MATCH({MAIN}!C3:C{last_row},{{{code_str}}},0))*({MAIN}!O3:O{last_row}="进行中"))'
    todo_f = f'=SUMPRODUCT(--ISNUMBER(MATCH({MAIN}!C3:C{last_row},{{{code_str}}},0))*({MAIN}!O3:O{last_row}="待办"))'
    pct_f = f'=IF(B{r}>0,C{r}/B{r},0)'
    bg = C_L3_ALT_BG if r % 2 == 0 else C_L3_BG
    for c, v in enumerate([pname, total_f, done_f, doing_f, todo_f, pct_f], 1):
        cell = ws3.cell(row=r, column=c, value=v)
        if c == 1:
            style_cell(cell, mk_font(11, True, C_DARK), mk_fill(C_PHASE_BG), mk_align('left', 'center'))
        else:
            style_cell(cell, mk_font(11), mk_fill(bg), mk_align('center', 'center'))
        if c == 6: cell.number_format = '0.0%'
    ws3.row_dimensions[r].height = 26
    r += 1
# 合计行
for c, v in enumerate(['合计', f'=SUM(B3:B{r-1})', f'=SUM(C3:C{r-1})', f'=SUM(D3:D{r-1})', f'=SUM(E3:E{r-1})', f'=IF(C{r}>0,C{r}/B{r},0)'], 1):
    cell = ws3.cell(row=r, column=c, value=v)
    style_cell(cell, mk_font(12, True, C_WHITE), mk_fill(C_L1_BG), mk_align('center', 'center'))
    if c == 6: cell.number_format = '0.0%'
ws3.row_dimensions[r].height = 30
ws3.freeze_panes = 'A3'

# ============================================================
# Sheet 4: 卡牌数据
# ============================================================
ws4 = wb.create_sheet('4.卡牌数据')
ws4.merge_cells('A1:K1')
style_cell(ws4['A1'], mk_font(16, True, C_WHITE), mk_fill(C_TITLE_BG), mk_align('center', 'center'))
ws4['A1'] = '卡牌数据表（cardId 与代码一致）'
ws4.row_dimensions[1].height = 38

card_headers = ['cardId', '名称', '阵营', '类型', '费用', '血量', '攻击', '移速', '范围', '特性', '稀有度']
for i, h in enumerate(card_headers, 1):
    style_cell(ws4.cell(row=2, column=i, value=h),
               mk_font(10, True, C_WHITE), mk_fill(C_HEAD_BG), mk_align('center', 'center'))
ws4.row_dimensions[2].height = 28

cards = [
    ('beast_disciple', '宗门御兽弟子', '攻方', '单位', 3, 5, 1, 0.8, 0, '控制兽当肉盾', '凡品'),
    ('body_disciple', '宗门体修弟子', '攻方', '单位', 2, 4, 2, 1.0, 0, '普通近战', '凡品'),
    ('sword_disciple', '宗门剑修弟子', '攻方', '单位', 3, 2, 3, 1.0, 3, '远程飞剑', '凡品'),
    ('elder_jindan', '金丹期长老', '攻方', '精英', 6, 8, 4, 0.8, 1, '5秒随机技能', '宝品'),
    ('wan_jian', '万剑归宗', '攻方', '法术', 4, 0, 0, 0, 0, '全员+1攻+加速', '灵品'),
    ('wu_lei', '五雷正法', '攻方', '法术', 4, 0, 0, 0, 3, '范围AOE伤害4', '灵品'),
    ('yu_feng', '御风诀', '攻方', '法术', 2, 0, 0, 0, 0, '单体+0.5速5秒', '凡品'),
    ('zhen_hun', '镇魂符', '攻方', '法术', 3, 0, 0, 0, 0, '禁用阵法3秒', '凡品'),
    ('jin_zhong', '金钟罩', '攻方', '法术', 3, 0, 0, 0, 0, '大殿护盾+3', '灵品'),
    ('yi_shan', '移山倒海', '攻方', '法术', 5, 0, 0, 0, 0, '推后2格+1伤', '宝品'),
    ('kun_xian', '困仙索', '攻方', '法术', 3, 0, 0, 0, 0, '单体stun2秒', '灵品'),
    ('tian_lei', '天雷诀', '攻方', '法术', 4, 0, 0, 0, 3, '范围伤害4', '灵品'),
    ('jiemai_formation', '截脉阵', '守方', '阵法', 2, 4, 2, 0, 1, '基础拦截', '凡品'),
    ('hanshuang_formation', '寒霜阵', '守方', '阵法', 3, 3, 1, 0, 1, '命中减速1回合', '凡品'),
    ('wanren_formation', '万刃阵', '守方', '阵法', 4, 5, 3, 0, 1, '高输出拦截', '灵品'),
    ('fanzhen_formation', '反震阵', '守方', '阵法', 3, 3, 0, 0, 1, '反伤50%', '灵品'),
    ('tianluo_formation', '天罗阵', '守方', '阵法', 5, 6, 2, 0, 3, '范围拦截', '宝品'),
    ('ying_jian', '影剑', '守方', '单位', 3, 3, 3, 1.5, 1, '冲向最近敌', '凡品'),
    ('huti_jianling', '护体剑灵', '守方', '单位', 4, 5, 2, 0, 1, '大殿临时护盾', '灵品'),
    ('guardian_beast', '守门灵兽', '守方', '单位', 3, 6, 1, 0.6, 1, '高血量肉盾', '凡品'),
    ('elder_flying_sword', '长老·飞剑诀', '通用', '技能', 0, 0, 0, 0, 3, '长老随机技-AOE', '宝品'),
    ('elder_pill', '长老·丹药', '通用', '技能', 0, 0, 0, 0, 0, '长老随机技-治疗', '宝品'),
    ('elder_talisman', '长老·符箓', '通用', '技能', 0, 0, 0, 0, 3, '长老随机技-天雷', '宝品'),
]
# 阵营颜色
SIDE_COLORS = {'攻方': 'FDE9D9', '守方': 'D6EAF8', '通用': 'E8DAEF'}
col_widths_4 = [20, 16, 6, 6, 6, 6, 6, 6, 6, 20, 8]
for i, w in enumerate(col_widths_4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

r = 3
for card in cards:
    side = card[2]
    for c, v in enumerate(card, 1):
        cell = ws4.cell(row=r, column=c, value=v)
        bg = C_L3_ALT_BG if r % 2 == 0 else C_L3_BG
        if c == 1:  # cardId代码列
            style_cell(cell, mk_font(9, name=MONO), mk_fill(C_CODE_BG), mk_align('left', 'center'))
        elif c == 3:  # 阵营列着色
            style_cell(cell, mk_font(10, True), mk_fill(SIDE_COLORS.get(side, bg)), mk_align('center', 'center'))
        elif c in (4, 5, 6, 7, 8, 9, 11):
            style_cell(cell, mk_font(10), mk_fill(bg), mk_align('center', 'center'))
        else:
            style_cell(cell, mk_font(10), mk_fill(bg), mk_align('left', 'center'))
    ws4.row_dimensions[r].height = 22
    r += 1
ws4.freeze_panes = 'A3'

# ============================================================
# Sheet 5: 数值配置
# ============================================================
ws5 = wb.create_sheet('5.数值配置')
ws5.merge_cells('A1:D1')
style_cell(ws5['A1'], mk_font(16, True, C_WHITE), mk_fill(C_TITLE_BG), mk_align('center', 'center'))
ws5['A1'] = '核心数值配置（GDScript Const.gd）'
ws5.row_dimensions[1].height = 38

for i, h in enumerate(['参数', '数值', '说明', 'GDScript常量名'], 1):
    style_cell(ws5.cell(row=2, column=i, value=h),
               mk_font(11, True, C_WHITE), mk_fill(C_HEAD_BG), mk_align('center', 'center'))
ws5.row_dimensions[2].height = 28

configs = [
    ('棋盘长度', '9格', '双方大殿间隔', 'BOARD_LENGTH'),
    ('大殿血量', '30', '到0即败', 'HALL_HP'),
    ('初始灵力', '5', '开局值', 'ENERGY_START'),
    ('灵力回复速率', '2.8秒/点', '实时回复', 'ENERGY_REGEN'),
    ('灵力上限', '5→10', '每30秒+1', 'ENERGY_MAX_CAP'),
    ('对局时长', '120秒', '到时比血量', 'BATTLE_TIME'),
    ('加时赛', '60秒', '平局加时比灵力', 'OVERTIME_TIME'),
    ('手牌上限', '4张', '满不抽', 'HAND_MAX'),
    ('出牌后补牌', '2秒', '自动补', 'DRAW_DELAY'),
    ('长老技能间隔', '5秒', '随机触发', 'ELDER_SKILL_INTERVAL'),
    ('阵法冷却', '8秒', '同位置不能再布', 'FORMATION_COOLDOWN'),
    ('移速基准', '1.0格/秒', '体修/剑修', 'BASE_SPEED'),
    ('减速效果', '-0.5格/秒', '寒霜阵/困仙', 'SLOW_AMOUNT'),
    ('攻击间隔', '1秒', '近战/远程', 'ATTACK_INTERVAL'),
    ('AI思考间隔', '3秒', 'AI出牌频率', 'AI_THINK_INTERVAL'),
    ('护盾最大', '3层', '金钟罩', 'MAX_SHIELD'),
]
col_widths_5 = [18, 14, 24, 24]
for i, w in enumerate(col_widths_5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

r = 3
for cfg in configs:
    bg = C_L3_ALT_BG if r % 2 == 0 else C_L3_BG
    for c, v in enumerate(cfg, 1):
        cell = ws5.cell(row=r, column=c, value=v)
        if c == 4:
            style_cell(cell, mk_font(9, name=MONO), mk_fill(C_CODE_BG), mk_align('left', 'center'))
        elif c == 2:
            style_cell(cell, mk_font(11, True, C_DARK), mk_fill(bg), mk_align('center', 'center'))
        else:
            style_cell(cell, mk_font(10), mk_fill(bg), mk_align('left', 'center'))
    ws5.row_dimensions[r].height = 24
    r += 1
ws5.freeze_panes = 'A3'

# ============================================================
# Sheet 6: AI提问指南
# ============================================================
ws6 = wb.create_sheet('6.AI提问指南')
ws6.merge_cells('A1:B1')
style_cell(ws6['A1'], mk_font(16, True, C_WHITE), mk_fill(C_TITLE_BG), mk_align('center', 'center'))
ws6['A1'] = '如何向AI提问开发（Godot版）'
ws6.row_dimensions[1].height = 38

guide = [
    ('使用方法', '把下表内容完整复制发给AI，让AI按规范实现'),
    ('提问模板', '请实现以下Godot功能：\n【编号】{编号}\n【任务】{任务名称}\n【实现原理】{实现原理}\n【函数签名】{函数签名}\n【文件路径】{文件路径}\n【输入输出】{输入输出}\n【前置依赖】{前置依赖}\n【完成标准】{完成标准}\n请用GDScript实现，遵循Godot 4.x规范。'),
    ('示例-移动系统', '请实现以下Godot功能：\n【编号】C3.01\n【任务】★单位移动★\n【实现原理】if state!="walking": return。target=find_target(unit)。if target and check_collision: state="fighting"。else: position.y+=get_effective_speed()*delta*facing。\n【函数签名】static func move_unit(unit: Unit, delta: float, model: Dictionary) -> void:\n【文件路径】scripts/game/MovementSystem.gd\n【输入输出】Unit+delta+model → void\n【前置依赖】C2.03,C3.02,C3.04\n【完成标准】①walking时y递增 ②遇敌变fighting ③到殿变dead\n请用GDScript实现，遵循Godot 4.x规范。'),
    ('验证方式', 'AI写完后，用以下方式验证：\n1. godot --headless --check-script scripts/game/MovementSystem.gd  语法检查\n2. 在编辑器中运行，观察控制台无报错\n3. 对照"完成标准"逐条验证\n4. 修改"2.开发主表"中该任务的状态为"完成"'),
    ('每日流程', '1. 打开"3.进度统计"找完成率<100%的阶段\n2. 到"2.开发主表"筛选该阶段\n3. 找第一个"待办"的L3行\n4. 核对"前置依赖"是否都已完成\n5. 复制"实现原理+函数签名+文件路径+完成标准"发给AI\n6. AI实现后验证\n7. 改状态列，进度自动更新'),
    ('Godot特性提示', '1. GDScript用缩进（Tab）不用大括号\n2. 静态变量/函数用 static 前缀\n3. 信号用 signal 关键字定义\n4. autoload单例在Project Settings配置\n5. 节点用 @onready var 获取引用\n6. 用 await 等待异步，不用 callback\n7. 场景切换用 get_tree().change_scene_to_file()\n8. 粒子用 CPUParticles2D 节点\n9. Tween动画用 create_tween()'),
]
r = 2
for title, content in guide:
    style_cell(ws6.cell(row=r, column=1, value=title), mk_font(11, True, C_DARK), mk_fill(C_L2_BG), mk_align('center', 'center'))
    cell = ws6.cell(row=r, column=2, value=content)
    style_cell(cell, mk_font(10), mk_fill(C_L3_BG), mk_align('left', 'top'))
    ws6.row_dimensions[r].height = max(30, content.count('\n') * 16 + 24)
    r += 1
ws6.column_dimensions['A'].width = 18
ws6.column_dimensions['B'].width = 88

# ============================================================
# Sheet 7: 资源清单
# ============================================================
ws7 = wb.create_sheet('7.资源清单')
ws7.merge_cells('A1:H1')
style_cell(ws7['A1'], mk_font(16, True, C_WHITE), mk_fill(C_TITLE_BG), mk_align('center', 'center'))
ws7['A1'] = '美术资源清单（Godot格式）'
ws7.row_dimensions[1].height = 38

for i, h in enumerate(['资源ID', '名称', '类型', '尺寸px', '帧数', '格式', 'Godot路径', '说明'], 1):
    style_cell(ws7.cell(row=2, column=i, value=h),
               mk_font(10, True, C_WHITE), mk_fill(C_HEAD_BG), mk_align('center', 'center'))
ws7.row_dimensions[2].height = 28

resources = [
    ('RES001', '山道背景', '背景', '720x1280', 1, 'PNG', 'res://assets/bg/mountain_path.png', '渐变绿'),
    ('RES002', '棋盘格子', 'UI', '720x200', 1, 'PNG', 'res://assets/ui/grid.png', '9列'),
    ('RES003', '攻方大殿', '建筑', '120x140', 1, 'PNG', 'res://assets/buildings/hall_attack.png', '底部'),
    ('RES004', '守方大殿', '建筑', '120x140', 1, 'PNG', 'res://assets/buildings/hall_defend.png', '顶部'),
    ('RES005', '御兽弟子', '单位', '48x64', 4, 'PNG', 'res://assets/units/beast_disciple.png', 'Sprite帧动画'),
    ('RES006', '体修弟子', '单位', '48x64', 4, 'PNG', 'res://assets/units/body_disciple.png', 'Sprite帧动画'),
    ('RES007', '剑修弟子', '单位', '48x64', 4, 'PNG', 'res://assets/units/sword_disciple.png', 'Sprite帧动画'),
    ('RES008', '金丹长老', '精英', '64x80', 6, 'PNG', 'res://assets/units/elder_jindan.png', 'Sprite帧动画'),
    ('RES009', '截脉阵', '阵法', '64x64', 1, 'PNG', 'res://assets/formations/jiemai.png', '半透明'),
    ('RES010', '寒霜阵', '阵法', '64x64', 1, 'PNG', 'res://assets/formations/hanshuang.png', '蓝色调'),
    ('RES011', '万刃阵', '阵法', '64x64', 1, 'PNG', 'res://assets/formations/wanren.png', '红色调'),
    ('RES012', '反震阵', '阵法', '64x64', 1, 'PNG', 'res://assets/formations/fanzhen.png', '紫色调'),
    ('RES013', '天罗阵', '阵法', '96x64', 1, 'PNG', 'res://assets/formations/tianluo.png', '范围3格'),
    ('RES014', '影剑', '单位', '48x64', 4, 'PNG', 'res://assets/units/ying_jian.png', '守方'),
    ('RES015', '护体剑灵', '单位', '48x64', 4, 'PNG', 'res://assets/units/huti_jianling.png', '守方'),
    ('RES016', '守门灵兽', '单位', '56x64', 4, 'PNG', 'res://assets/units/guardian_beast.png', '守方'),
    ('RES017', '万剑归宗特效', '特效', '128x128', 8, 'PNG', 'res://assets/effects/wan_jian.png', 'CPUParticles2D'),
    ('RES018', '五雷正法特效', '特效', '128x128', 8, 'PNG', 'res://assets/effects/wu_lei.png', 'CPUParticles2D'),
    ('RES019', '天雷诀特效', '特效', '128x128', 6, 'PNG', 'res://assets/effects/tian_lei.png', 'CPUParticles2D'),
    ('RES020', '大殿受击特效', '特效', '120x140', 4, 'PNG', 'res://assets/effects/hall_hit.png', 'Tween抖动'),
    ('RES021', '卡牌背面', 'UI', '80x110', 1, 'PNG', 'res://assets/ui/card_back.png', '手牌槽'),
    ('RES022', '灵力条', 'UI', '200x24', 1, 'PNG', 'res://assets/ui/energy_bar.png', 'TextureProgress'),
    ('RES023', '血条', 'UI', '60x8', 1, 'PNG', 'res://assets/ui/health_bar.png', 'TextureProgress'),
    ('RES024', 'BGM-战斗', '音频', '—', '—', 'OGG', 'res://assets/audio/bgm_battle.ogg', 'AudioStreamPlayer循环'),
    ('RES025', 'SFX-出兵', '音频', '—', '—', 'WAV', 'res://assets/audio/sfx_spawn.wav', '短音效'),
    ('RES026', 'SFX-攻击', '音频', '—', '—', 'WAV', 'res://assets/audio/sfx_attack.wav', '短音效'),
    ('RES027', 'SFX-阵法', '音频', '—', '—', 'WAV', 'res://assets/audio/sfx_formation.wav', '短音效'),
    ('RES028', 'SFX-法术', '音频', '—', '—', 'WAV', 'res://assets/audio/sfx_spell.wav', '短音效'),
    ('RES029', 'SFX-大殿受击', '音频', '—', '—', 'WAV', 'res://assets/audio/sfx_hall_hit.wav', '短音效'),
    ('RES030', 'SFX-胜利', '音频', '—', '—', 'WAV', 'res://assets/audio/sfx_win.wav', '短音效'),
    ('RES031', 'SFX-失败', '音频', '—', '—', 'WAV', 'res://assets/audio/sfx_lose.wav', '短音效'),
    ('RES032', '引导箭头', 'UI', '32x32', 1, 'PNG', 'res://assets/ui/arrow.png', '新手引导'),
    ('RES033', '引导高亮框', 'UI', '80x110', 1, 'PNG', 'res://assets/ui/highlight.png', '新手引导'),
]
col_widths_7 = [8, 16, 8, 12, 6, 6, 42, 16]
for i, w in enumerate(col_widths_7, 1):
    ws7.column_dimensions[get_column_letter(i)].width = w

r = 3
for res in resources:
    bg = C_L3_ALT_BG if r % 2 == 0 else C_L3_BG
    for c, v in enumerate(res, 1):
        cell = ws7.cell(row=r, column=c, value=v)
        if c == 7:  # Godot路径代码列
            style_cell(cell, mk_font(9, name=MONO), mk_fill(C_CODE_BG), mk_align('left', 'center'))
        elif c in (1, 3, 5, 6):
            style_cell(cell, mk_font(10), mk_fill(bg), mk_align('center', 'center'))
        else:
            style_cell(cell, mk_font(10), mk_fill(bg), mk_align('left', 'center'))
    ws7.row_dimensions[r].height = 22
    r += 1
ws7.freeze_panes = 'A3'

# ============================================================
# Sheet 8: 决策日志与问题
# ============================================================
ws8 = wb.create_sheet('8.决策日志与问题')
ws8.merge_cells('A1:G1')
style_cell(ws8['A1'], mk_font(16, True, C_WHITE), mk_fill(C_TITLE_BG), mk_align('center', 'center'))
ws8['A1'] = '决策日志与已知问题'
ws8.row_dimensions[1].height = 38

for i, h in enumerate(['类型', '编号', '日期', '标题', '决策/描述', '状态', '优先级'], 1):
    style_cell(ws8.cell(row=2, column=i, value=h),
               mk_font(10, True, C_WHITE), mk_fill(C_HEAD_BG), mk_align('center', 'center'))
ws8.row_dimensions[2].height = 28

decisions = [
    ('决策', 'D001', '2026-08-01', '游戏题材', '修仙宗门大战，非飞剑/现代战争', '已执行', '高'),
    ('决策', 'D002', '2026-08-01', '玩法模式', '实时策略卡牌，双向推进', '已执行', '高'),
    ('决策', 'D003', '2026-08-02', '地图视角', '折中方案：俯视斜坡，两边空缺放阵法', '已执行', '中'),
    ('决策', 'D004', '2026-08-03', '攻方卡牌', '人物卡4种+技能卡8种，金丹长老随机技能', '已执行', '高'),
    ('决策', 'D005', '2026-08-03', '守方设计', '阵法5种+拦截单位3种，阵法有冷却', '已执行', '高'),
    ('决策', 'D006', '2026-08-04', '平衡机制', '灵力实时回复+击杀后继续推进+大殿护盾', '已执行', '高'),
    ('决策', 'D007', '2026-08-05', '数值框架', '120秒对局+60秒加时+灵力5→10', '已执行', '中'),
    ('决策', 'D008', '2026-08-05', 'AI难度', 'easy/normal/hard三档，布阵策略不同', '已执行', '中'),
    ('决策', 'D009', '2026-08-06', '变现设计', '混合变现：激励广告+内购(皮肤/通行证)', '已执行', '中'),
    ('决策', 'D010', '2026-08-06', '版本规划', 'V1单机PvE→V1.5社交养成→V2实时PvP', '已执行', '高'),
    ('决策', 'D011', '2026-08-08', '★引擎选择★', '从原生JS改为Godot 4.x，理由：跨平台导出(微信/Steam/谷歌)原生支持', '已执行', '高'),
    ('决策', 'D012', '2026-08-08', '微信导出方案', '使用godot_for_minigame插件，非Godot官方但活跃维护', '已执行', '高'),
    ('决策', 'D013', '2026-08-08', '平台抽象层', 'PlatformAdapter autoload统一接口，业务层不直接调平台API', '已执行', '高'),
]
issues = [
    ('问题', 'I001', '2026-08-08', '微信包体限制', 'Godot 4.x WASM较大，主包4MB限制需分包加载', '待解决', '高'),
    ('问题', 'I002', '2026-08-08', '插件版本兼容', 'godot_for_minigame认证Godot 4.6.1，其他版本可能不兼容', '待解决', '高'),
    ('问题', 'I003', '2026-08-08', 'GDScript学习曲线', '需学GDScript(类Python)，估计1~2周上手', '待解决', '中'),
    ('问题', 'I004', '2026-08-08', 'Steam SDK集成', 'GodotSteam需编译GDExtension，有编译门槛', '待解决', '中'),
    ('问题', 'I005', '2026-08-08', '谷歌支付审核', 'Google Play Billing审核严格，需测试账号', '待解决', '中'),
    ('问题', 'I006', '2026-08-08', '跨平台输入差异', 'PC鼠标 vs 手机触屏 vs 微信touch，需统一InputEvent处理', '待解决', '中'),
    ('问题', 'I007', '2026-08-08', '纹理压缩格式', '微信ETC2/SteamS3TC/安卓ASTC，需按平台导出不同格式', '待解决', '低'),
    ('问题', 'I008', '2026-08-08', '苹果商店', 'iOS导出需Mac+iTunes Connect，暂不列入V1', '待解决', '低'),
]
col_widths_8 = [8, 8, 12, 16, 42, 10, 8]
for i, w in enumerate(col_widths_8, 1):
    ws8.column_dimensions[get_column_letter(i)].width = w

# 决策区域标题行
r = 3
ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
style_cell(ws8.cell(row=r, column=1, value='  ◆ 设计决策'), mk_font(12, True, C_WHITE), mk_fill(C_L1_BG), mk_align('left', 'center'))
ws8.row_dimensions[r].height = 26
r += 1
for d in decisions:
    bg = C_L3_ALT_BG if r % 2 == 0 else C_L3_BG
    for c, v in enumerate(d, 1):
        cell = ws8.cell(row=r, column=c, value=v)
        if c == 6:  # 状态
            if v == '已执行': style_cell(cell, mk_font(10, True, '006100'), mk_fill(C_DONE), mk_align('center', 'center'))
            elif v == '待解决': style_cell(cell, mk_font(10, True, '9C0006'), mk_fill(C_TODO), mk_align('center', 'center'))
        elif c == 7:  # 优先级
            color = '9C0006' if v == '高' else ('9C6500' if v == '中' else '375623')
            style_cell(cell, mk_font(10, True, color), mk_fill(bg), mk_align('center', 'center'))
        elif c in (1, 2, 3):
            style_cell(cell, mk_font(10), mk_fill(bg), mk_align('center', 'center'))
        else:
            style_cell(cell, mk_font(10), mk_fill(bg), mk_align('left', 'center'))
    ws8.row_dimensions[r].height = 24
    r += 1
# 问题区域标题行
ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
style_cell(ws8.cell(row=r, column=1, value='  ◆ 已知问题'), mk_font(12, True, C_WHITE), mk_fill(C_L1_BG), mk_align('left', 'center'))
ws8.row_dimensions[r].height = 26
r += 1
for d in issues:
    bg = C_L3_ALT_BG if r % 2 == 0 else C_L3_BG
    for c, v in enumerate(d, 1):
        cell = ws8.cell(row=r, column=c, value=v)
        if c == 6:
            if v == '已执行' or v == '已解决': style_cell(cell, mk_font(10, True, '006100'), mk_fill(C_DONE), mk_align('center', 'center'))
            elif v == '待解决': style_cell(cell, mk_font(10, True, '9C0006'), mk_fill(C_TODO), mk_align('center', 'center'))
        elif c == 7:
            color = '9C0006' if v == '高' else ('9C6500' if v == '中' else '375623')
            style_cell(cell, mk_font(10, True, color), mk_fill(bg), mk_align('center', 'center'))
        elif c in (1, 2, 3):
            style_cell(cell, mk_font(10), mk_fill(bg), mk_align('center', 'center'))
        else:
            style_cell(cell, mk_font(10), mk_fill(bg), mk_align('left', 'center'))
    ws8.row_dimensions[r].height = 24
    r += 1
ws8.freeze_panes = 'A3'

# ============================================================
# Sheet 9: 跨平台导出
# ============================================================
ws9 = wb.create_sheet('9.跨平台导出')
ws9.merge_cells('A1:F1')
style_cell(ws9['A1'], mk_font(16, True, C_WHITE), mk_fill(C_TITLE_BG), mk_align('center', 'center'))
ws9['A1'] = '跨平台导出方案对比'
ws9.row_dimensions[1].height = 38

for i, h in enumerate(['平台', '导出方式', '关键插件/SDK', '审核要求', '包体限制', '注意事项'], 1):
    style_cell(ws9.cell(row=2, column=i, value=h),
               mk_font(11, True, C_WHITE), mk_fill(C_HEAD_BG), mk_align('center', 'center'))
ws9.row_dimensions[2].height = 28

platforms = [
    ('微信小游戏', 'godot_for_minigame插件', 'godot_for_minigame', '微信小游戏审核\n(内容+资质)', '主包4MB\n分包20MB', '①需企业资质\n②WASM分包\n③wx.* API通过PlatformAdapter\n④触屏优先'),
    ('Steam', 'Godot原生\nDesktop导出', 'GodotSteam\n(GDExtension)', 'Steamworks审核\n(内容+年龄分级)', '无硬限制\n(建议<2GB)', '①需Steamworks账号\n②.exe+.pck\n③成就/云存档\n④支持Mac/Linux'),
    ('Google Play', 'Godot原生\nAndroid导出', 'GooglePlayBilling', '谷歌商店审核\n(内容+隐私政策)', '.aab格式', '①需Google Play开发者账号\n②keystore签名\n③目标API等级要求\n④触屏适配'),
    ('iOS', 'Godot原生iOS导出\n(需Mac)', '—', 'App Store审核\n(严格)', '4GB', '①需Mac+Xcode\n②Apple开发者账号\n③暂不列入V1'),
    ('Web', 'Godot原生\nWeb导出', '—', '无需审核', '建议<50MB', '①HTML5导出\n②不支持部分功能\n③可用于Demo'),
]
col_widths_9 = [14, 20, 20, 20, 14, 38]
for i, w in enumerate(col_widths_9, 1):
    ws9.column_dimensions[get_column_letter(i)].width = w

r = 3
for p in platforms:
    bg = C_L3_ALT_BG if r % 2 == 0 else C_L3_BG
    for c, v in enumerate(p, 1):
        cell = ws9.cell(row=r, column=c, value=v)
        if c == 1:
            style_cell(cell, mk_font(11, True, C_DARK), mk_fill(C_PHASE_BG), mk_align('center', 'center'))
        elif c in (2, 3):
            style_cell(cell, mk_font(9, name=MONO), mk_fill(C_CODE_BG), mk_align('left', 'center'))
        else:
            style_cell(cell, mk_font(10), mk_fill(bg), mk_align('left', 'center'))
    ws9.row_dimensions[r].height = calc_row_height(list(p), col_widths_9)
    r += 1
ws9.freeze_panes = 'A3'

# ============================================================
# Sheet 10: 设计思路
# ============================================================
ws10 = wb.create_sheet('10.设计思路')
ws10.merge_cells('A1:B1')
style_cell(ws10['A1'], mk_font(16, True, C_WHITE), mk_fill(C_TITLE_BG), mk_align('center', 'center'))
ws10['A1'] = '游戏设计思路'
ws10.row_dimensions[1].height = 38

design = [
    ('一句话简介', '修仙版皇室战争：实时双向推进，攻方破阵摧殿，守方布阵拦截，120秒决胜负'),
    ('核心循环', '开局→灵力回复→出兵/布阵/施法→单位自动推进→碰撞战斗→击杀继续推进→摧毁大殿/时限到→结算'),
    ('攻方策略', '①低费快攻(体修+御风诀)  ②远程消耗(剑修+万剑归宗)  ③精英一波(金丹长老+金钟罩)'),
    ('守方策略', '①层层拦截(截脉+万刃)  ②减速消耗(寒霜+困仙)  ③反伤坦克(反震+灵兽)'),
    ('平衡机制', '①灵力实时回复(不卡手)  ②击杀后继续推进(进攻有收益)  ③大殿护盾(防守有保障)  ④阵法冷却(不能无脑布阵)  ⑤加时赛(避免平局)'),
    ('单局节奏', '0-30秒: 试探出牌  30-60秒: 中期拉锯  60-90秒: 决战期  90-120秒: 收官/加时'),
    ('深度来源', '①20+卡牌组合  ②金丹长老随机技能(每局不同)  ③阵法位置策略  ④灵力管理  ⑤攻守时机'),
    ('Godot优势', '①跨平台一次开发  ②节点系统简化渲染  ③信号系统解耦  ④粒子/Tween内置  ⑤编辑器可视化UI'),
    ('版本路线', 'V1(P1-P8): 单机PvE+跨平台导出 → V1.5(P9-P14): 社交+养成+变现 → V2(P16-P18): 实时PvP+段位赛'),
    ('开发优先级', 'P0环境→P1骨架→P2数据→P3核心战斗→P4扩展战斗→P5 AI→P6 UI→P8跨平台导出→上线验证→P9+扩展'),
]
r = 2
for title, content in design:
    style_cell(ws10.cell(row=r, column=1, value=title), mk_font(11, True, C_WHITE), mk_fill(C_L2_BG), mk_align('center', 'center'))
    cell = ws10.cell(row=r, column=2, value=content)
    style_cell(cell, mk_font(10), mk_fill(C_L3_BG), mk_align('left', 'center'))
    ws10.row_dimensions[r].height = max(28, len(content) // 45 * 16 + 24)
    r += 1
ws10.column_dimensions['A'].width = 16
ws10.column_dimensions['B'].width = 88

# ===== 保存 =====
output = '/workspace/宗门论道_Godot开发管理工具包.xlsx'
wb.save(output)
print(f'✓ 已生成: {output}')
print(f'  Sheet数: {len(wb.sheetnames)}')
print(f'  L3任务数: {len(L3_ROWS)}')
print(f'  主表行数: {last_row}')
