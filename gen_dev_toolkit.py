#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
《宗门论道》完整游戏开发管理工具包
不是一份"设计文档"，而是一个你每天用的项目管理工具。
10个Sheet覆盖：使用说明 → 路线图 → 功能清单 → 任务拆解 → 卡牌数据 → 数值配置 → AI提问指南 → 进度看板 → Bug跟踪 → 开发日志
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()

# ===== 样式库 =====
F_TITLE   = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
F_HEADER  = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
F_SECTION = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')
F_CELL    = Font(name='微软雅黑', size=10)
F_BOLD    = Font(name='微软雅黑', size=10, bold=True)
F_NOTE    = Font(name='微软雅黑', size=10, color='666666', italic=True)

BG_TITLE   = PatternFill('solid', fgColor='2F5496')
BG_HEADER  = PatternFill('solid', fgColor='4472C4')
BG_SECTION = PatternFill('solid', fgColor='D6E4F0')
BG_ATTACK  = PatternFill('solid', fgColor='FCE4D6')
BG_DEFEND  = PatternFill('solid', fgColor='D6E4F0')
BG_SPELL   = PatternFill('solid', fgColor='E2EFDA')
BG_ARRAY   = PatternFill('solid', fgColor='FFF2CC')
BG_DONE    = PatternFill('solid', fgColor='C6EFCE')
BG_DOING   = PatternFill('solid', fgColor='FFEB9C')
BG_TODO    = PatternFill('solid', fgColor='FFC7CE')
BG_HIGH    = PatternFill('solid', fgColor='F8CBAD')
BG_MED     = PatternFill('solid', fgColor='FFF2CC')
BG_LOW     = PatternFill('solid', fgColor='E2EFDA')

A_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
A_LEFT   = Alignment(horizontal='left', vertical='center', wrap_text=True)
A_LEFTT  = Alignment(horizontal='left', vertical='top', wrap_text=True)

BD = Border(
    left=Side(style='thin', color='B4C7E7'),
    right=Side(style='thin', color='B4C7E7'),
    top=Side(style='thin', color='B4C7E7'),
    bottom=Side(style='thin', color='B4C7E7')
)

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_row(ws, text, cols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_TITLE; c.fill = BG_TITLE; c.alignment = A_CENTER
    ws.row_dimensions[row].height = 32

def header_row(ws, headers, row):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = F_HEADER; c.fill = BG_HEADER; c.alignment = A_CENTER; c.border = BD
    ws.row_dimensions[row].height = 24

def data_row(ws, values, row, fill=None, bold_first=False):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = F_BOLD if (bold_first and i == 1) else F_CELL
        c.alignment = A_LEFTT if i == len(values) else A_CENTER
        c.border = BD
        if fill: c.fill = fill

def section_row(ws, text, cols, row):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_SECTION; c.fill = BG_SECTION; c.alignment = A_LEFT
    ws.row_dimensions[row].height = 22

# ================================================================
# Sheet 1: 使用说明（开发方法论）
# ================================================================
ws = wb.active
ws.title = '1.使用说明'
set_widths(ws, [22, 65])
title_row(ws, '《宗门论道》开发管理工具包 — 使用说明', 2)

r = 3
section_row(ws, '一、这个Excel是什么', 2, r); r += 1
intro = [
    ('定位', '这不是一份"读一遍就放着"的设计文档，而是一个你【每天打开用】的项目管理工具。'),
    ('解决什么问题', '你不知道全貌→看「功能清单」；你不知道下一步→看「任务拆解」；你不知道怎么让AI干活→看「AI提问指南」；你不知道进度→看「进度看板」。'),
    ('怎么用', '每天开发前：打开「任务拆解」找下一个「待办」任务→看「完成标准」→用「AI提问指南」的模板让AI帮你做→做完标记「已完成」→记录到「开发日志」。'),
]
for k, v in intro:
    ws.cell(row=r, column=1, value=k).font = F_BOLD
    ws.cell(row=r, column=1).fill = BG_SECTION; ws.cell(row=r, column=1).border = BD; ws.cell(row=r, column=1).alignment = A_CENTER
    ws.cell(row=r, column=2, value=v).font = F_CELL
    ws.cell(row=r, column=2).alignment = A_LEFTT; ws.cell(row=r, column=2).border = BD
    ws.row_dimensions[r].height = 40; r += 1

r += 1
section_row(ws, '二、10个Sheet各自干什么', 2, r); r += 1
sheets_info = [
    ('1.使用说明', '你现在看的这个。讲清楚开发方法论和工具怎么用。'),
    ('2.开发路线图', '鸟瞰全局：分3个大版本，每个版本有几个阶段，每个阶段做什么。让你知道"现在在哪、还要走多远"。'),
    ('3.功能清单', '完整功能列表：游戏要做哪些功能，按模块分类，标了优先级和状态。让你知道"总共要做多少东西"。'),
    ('4.任务拆解', '★核心Sheet★：把每个功能拆成具体的开发任务，带依赖关系和完成标准。你每天就照着这个表从上往下做。'),
    ('5.卡牌数据', '所有卡牌的数值表：攻方弟子/守方单位/阵法/法术/长老分支。改数值直接在这改。'),
    ('6.数值配置', '所有可调参数（灵力回复/大殿血量/移速等）集中管理。平衡调优时在这改。'),
    ('7.AI提问指南', '★核心Sheet★：教你如何向AI提问，附带可直接复制使用的提示词模板。这是让你"不用自己写代码"的关键。'),
    ('8.进度看板', '自动统计各阶段完成率，看一眼就知道整体进度。'),
    ('9.Bug跟踪', '发现Bug记录在这，追踪修复状态。'),
    ('10.开发日志', '每天记录：做了什么/遇到什么问题/下一步计划。积累开发经验。'),
]
header_row(ws, ['Sheet名', '干什么用的'], r); r += 1
for name, desc in sheets_info:
    data_row(ws, [name, desc], r); ws.row_dimensions[r].height = 36; r += 1

r += 1
section_row(ws, '三、单人+AI的开发方法论（重要！）', 2, r); r += 1
method = [
    ('核心原则', '你不需要自己写代码。你的角色是"产品经理+测试员"，AI的角色是"程序员"。你负责：定需求、验结果、提Bug。AI负责：写代码、修Bug、做设计。'),
    ('开发循环', '每个任务都走这个循环：①看任务拆解表找下一个任务 → ②用AI提问指南的模板把任务发给AI → ③AI完成后你验证（跑模拟/看效果）→ ④通过则标记完成，不通过则把问题发给AI修 → ⑤记录到开发日志。'),
    ('任务粒度', '任务拆解表里的每个任务都是"1次AI对话能完成"的粒度。不要一次让AI做10个任务，它会做不好。一个一个来，做完验证再做下一个。'),
    ('验证意识', '每个任务都有「完成标准」。AI说做完了不算完成，你按标准验证通过才算完成。验证方法：跑模拟/看渲染/检查语法。'),
    ('版本意识', 'V1只做MVP（最小可玩版本），不要贪多。V1验证玩法，V1.5加社交美术，V2做PvP。每个版本有明确的"做完标准"。'),
    ('进度管理', '每完成一个任务，在「任务拆解」表把状态改为「已完成」。每周看一次「进度看板」，确认整体进度健康。'),
]
for k, v in method:
    ws.cell(row=r, column=1, value=k).font = F_BOLD
    ws.cell(row=r, column=1).fill = BG_SECTION; ws.cell(row=r, column=1).border = BD; ws.cell(row=r, column=1).alignment = A_CENTER
    ws.cell(row=r, column=2, value=v).font = F_CELL
    ws.cell(row=r, column=2).alignment = A_LEFTT; ws.cell(row=r, column=2).border = BD
    ws.row_dimensions[r].height = 55; r += 1

r += 1
section_row(ws, '四、给AI下指令的黄金公式', 2, r); r += 1
formula = [
    ('公式', '【角色】+【任务】+【上下文】+【完成标准】+【自主决策授权】'),
    ('示例', '"你是微信小游戏开发专家。请实现单位的实时移动逻辑：单位沿主路向敌方大殿推进，移速=配置值，遇敌方单位/阵法停下交战，击杀后继续推进。参考卡牌数据见Cards.js。完成标准：①代码能通过node --check语法检查 ②AI vs AI模拟能看到单位移动 ③单位到大殿后对大殿造成伤害。遇到设计细节未明确的按合理默认自行决定，做完统一汇报，不要中途问我。"'),
    ('关键', '①给足上下文（卡牌数据/已有代码/设计文档）②明确完成标准（怎么算做完）③授权自主决策（别让它停下来问你）④一次一个任务（别一次塞太多）'),
]
for k, v in formula:
    ws.cell(row=r, column=1, value=k).font = F_BOLD
    ws.cell(row=r, column=1).fill = BG_SECTION; ws.cell(row=r, column=1).border = BD; ws.cell(row=r, column=1).alignment = A_CENTER
    ws.cell(row=r, column=2, value=v).font = F_CELL
    ws.cell(row=r, column=2).alignment = A_LEFTT; ws.cell(row=r, column=2).border = BD
    ws.row_dimensions[r].height = 70; r += 1

# ================================================================
# Sheet 2: 开发路线图
# ================================================================
ws = wb.create_sheet('2.开发路线图')
set_widths(ws, [10, 16, 35, 12, 12])
title_row(ws, '开发路线图 — 三个版本全景', 5)
header_row(ws, ['版本', '阶段', '目标', '预计周数', '状态'], 3)

phases = [
    ('V1.0', 'P1-工程骨架', '创建小游戏工程+渲染循环+输入系统+事件系统', '1周', '待办'),
    ('V1.0', 'P2-核心数据', '卡牌数据表+游戏状态模型+卡组手牌系统', '1周', '待办'),
    ('V1.0', 'P3-战斗逻辑', '单位移动+碰撞+交战+击杀推进+阵法+法术+长老技能', '2周', '待办'),
    ('V1.0', 'P4-AI系统', 'AI灵力管理+出牌决策+布阵+难度分级', '1周', '待办'),
    ('V1.0', 'P5-渲染场景', '背景+大殿+单位+阵法+特效+UI渲染', '1.5周', '待办'),
    ('V1.0', 'P6-系统整合', '灵力回复+抽牌+胜负+结算+新手引导', '1周', '待办'),
    ('V1.0', 'P7-测试平衡', '模拟器+平衡调优+性能测试+真机测试', '1周', '待办'),
    ('V1.0', 'P8-提交审核', '修Bug+提交微信审核+上线', '1周', '待办'),
    ('V1.5', 'P9-Spine美术', 'Spine骨骼动画接入+资源管线+全部单位美术', '2周', '待办'),
    ('V1.5', 'P10-云开发', '微信登录+云开发存档+用户档案', '1周', '待办'),
    ('V1.5', 'P11-异步PvP', '异步对战(布阵+回放)+匹配系统', '2周', '待办'),
    ('V1.5', 'P12-社交系统', '排行榜+分享+好友挑战', '1周', '待办'),
    ('V1.5', 'P13-养成系统', '弟子升级+升星+卡牌收集+宝箱', '2周', '待办'),
    ('V1.5', 'P14-变现系统', '激励广告+通行证+月卡+内购', '1.5周', '待办'),
    ('V1.5', 'P15-测试上线', '测试+修Bug+提交审核', '1.5周', '待办'),
    ('V2.0', 'P16-实时通信', 'WebSocket框架+状态同步方案', '2周', '待办'),
    ('V2.0', 'P17-同步逻辑', '帧同步/状态同步+断线重连+延迟补偿', '3周', '待办'),
    ('V2.0', 'P18-匹配段位', '匹配系统+段位赛+ELO评分', '2周', '待办'),
    ('V2.0', 'P19-赛季活动', '赛季系统+锦标赛+排行榜重置', '2周', '待办'),
    ('V2.0', 'P20-内容扩充', '新卡牌+新流派+大平衡调整', '2周', '待办'),
    ('V2.0', 'P21-压测上线', '压力测试+修Bug+上线', '2周', '待办'),
]
r = 4
for p in phases:
    v = p[0]
    fill = BG_ATTACK if v == 'V1.0' else (BG_DEFEND if v == 'V1.5' else BG_SPELL)
    data_row(ws, list(p), r, fill=fill)
    ws.row_dimensions[r].height = 28; r += 1

r += 1
section_row(ws, '里程碑验收标准', 5, r); r += 1
milestones = [
    ('V1.0上线', 'AI vs AI模拟10局：双方摧毁度20%~70%，胜率~50%，无0%死局。真机60fps。新手引导完成率>80%。无崩溃。'),
    ('V1.5上线', 'Spine美术全部接入。云存档正常。异步PvP可对战。排行榜可用。养成系统可升级。广告+内购可支付。次留35%+。'),
    ('V2.0上线', '实时PvP延迟<200ms。匹配<10秒。段位赛可排名。赛季可重置。压测1000人并发不崩。实时PvP活跃占比60%+。'),
]
for name, criteria in milestones:
    ws.cell(row=r, column=1, value=name).font = F_BOLD
    ws.cell(row=r, column=1).fill = BG_SECTION; ws.cell(row=r, column=1).border = BD; ws.cell(row=r, column=1).alignment = A_CENTER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.cell(row=r, column=2, value=criteria).font = F_CELL
    ws.cell(row=r, column=2).alignment = A_LEFTT; ws.cell(row=r, column=2).border = BD
    ws.row_dimensions[r].height = 50; r += 1

# ================================================================
# Sheet 3: 功能清单
# ================================================================
ws = wb.create_sheet('3.功能清单')
set_widths(ws, [14, 28, 40, 8, 10, 10])
title_row(ws, '功能清单 — 游戏要做哪些功能（完整版）', 6)
header_row(ws, ['模块', '功能点', '详细描述', '优先级', '版本', '状态'], 3)

features = [
    # 核心引擎
    ('核心引擎', '渲染循环', 'requestAnimationFrame 60fps循环，update+render分离', '高', 'V1', '待办'),
    ('核心引擎', '输入系统', '触摸事件处理，点击区域注册与命中检测', '高', 'V1', '待办'),
    ('核心引擎', '事件总线', '发布订阅模式，模块间解耦通信', '高', 'V1', '待办'),
    ('核心引擎', '资源加载器', '图片/音频预加载，加载进度回调', '中', 'V1', '待办'),
    ('核心引擎', '场景管理', '场景切换（菜单/战斗/结算）', '高', 'V1', '待办'),
    # 数据层
    ('数据层', '卡牌数据表', '所有卡牌属性配置（攻方/守方/阵法/法术）', '高', 'V1', '待办'),
    ('数据层', '游戏状态模型', '双方大殿血量/灵力/单位列表/时间', '高', 'V1', '待办'),
    ('数据层', '卡组系统', '8张卡组配置，可重复携带', '高', 'V1', '待办'),
    ('数据层', '手牌系统', '4张手牌，打出后2秒抽牌补位', '高', 'V1', '待办'),
    ('数据层', '抽牌系统', '从牌库随机抽牌，手牌满时不抽', '高', 'V1', '待办'),
    # 战斗逻辑
    ('战斗逻辑', '单位移动', '沿主路向敌方大殿推进，移速×dt', '高', 'V1', '待办'),
    ('战斗逻辑', '碰撞检测', 'AABB矩形碰撞，检测同列最近敌方目标', '高', 'V1', '待办'),
    ('战斗逻辑', '近战交战', '进入攻击范围停下，按间隔互相扣血', '高', 'V1', '待办'),
    ('战斗逻辑', '远程攻击', '剑修弟子射程3格，只打人不掉血', '高', 'V1', '待办'),
    ('战斗逻辑', '击杀后继续推进', '★核心★杀掉目标后不消失，继续前进', '高', 'V1', '待办'),
    ('战斗逻辑', '到达大殿伤害', '单位到大殿格→对大殿造成攻击力伤害→消失', '高', 'V1', '待办'),
    ('战斗逻辑', '阵法机制', '布在阵法区，不动有血量，主动攻击经过敌人', '高', 'V1', '待办'),
    ('战斗逻辑', '阵法冷却', '阵法被毁格子8秒内不能再布阵', '高', 'V1', '待办'),
    ('战斗逻辑', '范围伤害', '天罗阵/五雷正法/天雷诀打相邻格所有敌人', '高', 'V1', '待办'),
    ('战斗逻辑', '减速机制', '寒霜阵命中后移速-0.5持续2秒', '中', 'V1', '待办'),
    ('战斗逻辑', '反震机制', '反震阵被攻击时反弹50%伤害', '中', 'V1', '待办'),
    ('战斗逻辑', '定身机制', '困仙索使敌人2秒无法移动', '中', 'V1', '待办'),
    ('战斗逻辑', '推后机制', '移山倒海使区域敌人后退2格+1伤', '中', 'V1', '待办'),
    ('战斗逻辑', '护盾机制', '金钟罩使大殿3秒免疫伤害', '中', 'V1', '待办'),
    ('战斗逻辑', '加速机制', '御风诀/万剑归宗提升移速持续X秒', '中', 'V1', '待办'),
    ('战斗逻辑', '禁阵机制', '镇魂符使阵法3秒失效', '中', 'V1', '待办'),
    ('战斗逻辑', '自爆机制', '护山灵兽死亡对相邻格造成1伤', '低', 'V1', '待办'),
    ('战斗逻辑', '长老技能-飞剑', '金丹长老随机释放万剑齐发(范围伤)', '中', 'V1', '待办'),
    ('战斗逻辑', '长老技能-丹药', '金丹长老随机释放回血', '中', 'V1', '待办'),
    ('战斗逻辑', '长老技能-符箓', '金丹长老随机释放天雷诀(清兵)', '中', 'V1', '待办'),
    ('战斗逻辑', '长老技能-御兽', '金丹长老随机召唤灵兽', '中', 'V1', '待办'),
    ('战斗逻辑', '护法长老技能', '护法长老随机释放4分支(对标金丹长老)', '中', 'V1', '待办'),
    # AI系统
    ('AI系统', 'AI灵力管理', '根据难度决定攒灵力还是即时出牌', '高', 'V1', '待办'),
    ('AI系统', 'AI出牌决策', '根据大殿血量决定攻守比例', '高', 'V1', '待办'),
    ('AI系统', 'AI布阵位置', '根据难度选择布阵格子(随机/大殿前/针对性)', '中', 'V1', '待办'),
    ('AI系统', 'AI难度分级', '简单/普通/困难三档，决策间隔与策略不同', '高', 'V1', '待办'),
    ('AI系统', 'AI思考间隔', '简单3.5s/普通2.5s/困难1.5s', '中', 'V1', '待办'),
    # 渲染
    ('渲染', '背景渲染', '山道蜿蜒+云雾效果(占位色块)', '中', 'V1', '待办'),
    ('渲染', '大殿渲染', '顶部/底部大殿+血量条+受击特效', '高', 'V1', '待办'),
    ('渲染', '单位渲染', '单位精灵+y排序深度+状态动画(占位色块)', '高', 'V1', '待办'),
    ('渲染', '阵法渲染', '阵法光阵+状态(正常/受击/禁阵/冷却)', '中', 'V1', '待办'),
    ('渲染', '特效系统', '命中粒子/死亡碎裂/法术光效/大殿震动', '中', 'V1', '待办'),
    ('渲染', 'UI-手牌栏', '底部4张牌横排，选中上浮发光，费用不足灰显', '高', 'V1', '待办'),
    ('渲染', 'UI-灵力条', '灵力数字+液条+回复动画', '高', 'V1', '待办'),
    ('渲染', 'UI-血量计时', '顶部双方血条+中央计时器', '高', 'V1', '待办'),
    # 系统功能
    ('系统功能', '灵力实时回复', '每2.8秒+1，上限随时间5→10', '高', 'V1', '待办'),
    ('系统功能', '胜负判定', '大殿血量归零/时限到比血量/加时比灵力', '高', 'V1', '待办'),
    ('系统功能', '加时赛', '180秒平局→加时60秒，灵力×1.5', '中', 'V1', '待办'),
    ('系统功能', '结算页面', '胜负展示+摧毁度+奖励+再来一局', '高', 'V1', '待办'),
    ('系统功能', '新手引导', '4步强制教程(出牌/灵力/布阵/目标)', '高', 'V1', '待办'),
    ('系统功能', '音效系统', 'BGM+音效播放+音量控制', '低', 'V1', '待办'),
    ('系统功能', '卡组预设', '3套预设卡组(快攻/控制/坦克)', '中', 'V1', '待办'),
    # V1.5
    ('美术(V1.5)', 'Spine动画接入', 'Spine Runtime集成+骨骼动画播放', '高', 'V1.5', '待办'),
    ('美术(V1.5)', '正式美术资源', '全部单位/阵法/大殿/UI正式美术', '高', 'V1.5', '待办'),
    ('社交(V1.5)', '微信登录', 'wx.login+用户授权+获取openid', '高', 'V1.5', '待办'),
    ('社交(V1.5)', '云开发存档', '云数据库存用户档案/卡组/战绩', '高', 'V1.5', '待办'),
    ('社交(V1.5)', '异步PvP', '挑战好友布阵+回放对战', '高', 'V1.5', '待办'),
    ('社交(V1.5)', '排行榜', '开放数据域好友/群排行', '中', 'V1.5', '待办'),
    ('社交(V1.5)', '分享系统', '对局分享/求助/炫耀', '中', 'V1.5', '待办'),
    ('养成(V1.5)', '弟子升级', '消耗灵石提升等级增加属性', '中', 'V1.5', '待办'),
    ('养成(V1.5)', '弟子升星', '消耗剑魄突破境界(练气→筑基→金丹→元婴)', '中', 'V1.5', '待办'),
    ('养成(V1.5)', '卡牌收集', '宝箱/碎片合成获取新卡', '中', 'V1.5', '待办'),
    ('变现(V1.5)', '激励广告', '每日宝箱/双倍奖励/续战复活/体力', '中', 'V1.5', '待办'),
    ('变现(V1.5)', '通行证', '赛季任务+免费/付费双线奖励', '中', 'V1.5', '待办'),
    ('变现(V1.5)', '内购支付', '仙玉/月卡/皮肤/礼包', '中', 'V1.5', '待办'),
    # V2
    ('PvP(V2)', 'WebSocket通信', '实时双向通信框架', '高', 'V2', '待办'),
    ('PvP(V2)', '状态同步', '服务端权威+客户端预测+延迟补偿', '高', 'V2', '待办'),
    ('PvP(V2)', '断线重连', '断线后可重连恢复对局', '高', 'V2', '待办'),
    ('PvP(V2)', '匹配系统', '按段位/ELO匹配对手', '高', 'V2', '待办'),
    ('PvP(V2)', '段位赛', '青铜→宗师段位+升降级', '中', 'V2', '待办'),
    ('PvP(V2)', '赛季系统', '4周赛季+重置+奖励', '中', 'V2', '待办'),
    ('PvP(V2)', '锦标赛', '淘汰赛/积分赛模式', '低', 'V2', '待办'),
]
r = 4
for f in features:
    pri = f[3]
    fill = BG_HIGH if pri == '高' else (BG_MED if pri == '中' else BG_LOW)
    data_row(ws, list(f), r, fill=fill)
    ws.row_dimensions[r].height = 24; r += 1

# ================================================================
# Sheet 4: 任务拆解（★核心★）
# ================================================================
ws = wb.create_sheet('4.任务拆解')
set_widths(ws, [8, 10, 30, 40, 35, 10, 10])
title_row(ws, '★ 任务拆解 — 每天照着这个表从上往下做 ★', 7)
header_row(ws, ['任务ID', '阶段', '任务名称', '任务描述', '完成标准（怎么算做完）', '依赖', '状态'], 3)

tasks = [
    # P1 工程骨架
    ('T1.1', 'P1', '创建工程配置文件', '创建game.json/game.js/project.config.json/sitemap.json/package.json，compileType=game', '微信开发者工具能打开不报错', '-', '待办'),
    ('T1.2', 'P1', '实现渲染循环Director', 'requestAnimationFrame循环，update(dt)+render()分离，dt上限33ms', '能60fps跑空循环，console能看到帧率', 'T1.1', '待办'),
    ('T1.3', 'P1', '实现输入系统Input', 'wx触摸事件→坐标→点击区域注册→命中检测→回调', '点屏幕能触发注册的回调', 'T1.1', '待办'),
    ('T1.4', 'P1', '实现事件系统EventSystem', '发布订阅模式on/emit/off，模块间解耦', 'emit事件后on的回调能收到', 'T1.1', '待办'),
    ('T1.5', 'P1', '实现场景管理SceneManager', '场景注册/切换/生命周期(onEnter/onExit)', '能从空场景A切换到空场景B', 'T1.2', '待办'),
    # P2 核心数据
    ('T2.1', 'P2', '定义全局常量Constants', '棋盘尺寸/灵力参数/时间/坐标等所有常量', '常量定义完整，无硬编码', 'T1.1', '待办'),
    ('T2.2', 'P2', '配置卡牌数据Cards.js', '所有卡牌属性（攻方3弟子+金丹长老+守方3单位+护法长老+5阵法+8法术）', '19张卡牌数据完整，字段无缺', 'T2.1', '待办'),
    ('T2.3', 'P2', '定义Unit数据结构', 'id/cardId/owner/x/y/hp/atk/speed/range/interval/state/traits/buffs', '结构完整，能从卡牌数据生成Unit实例', 'T2.2', '待办'),
    ('T2.4', 'P2', '定义GameModel游戏状态', 'time/players[2]/{hallHp/energy/deck/hand/units/formations}/state', '能初始化一局完整游戏状态', 'T2.3', '待办'),
    ('T2.5', 'P2', '实现卡组手牌系统Deck', '8张卡组→洗牌→抽3张起手→打出后2秒抽1张→手牌上限4', '能完整模拟抽牌流程', 'T2.2', '待办'),
    # P3 战斗逻辑
    ('T3.1', 'P3', '实现单位移动', '单位沿主路y轴移动，speed×dt×direction，方向朝敌方大殿', '模拟能看到单位y坐标变化', 'T2.4', '待办'),
    ('T3.2', 'P3', '实现碰撞检测', '检测同列最近敌方单位/阵法，距离<attackRange时返回目标', '单位接近时能检测到目标', 'T3.1', '待办'),
    ('T3.3', 'P3', '实现近战交战', '进入范围→停下→按interval互相扣血→血量归零标记死亡', '两个单位能互打至一方死亡', 'T3.2', '待办'),
    ('T3.4', 'P3', '实现击杀后继续推进', '★核心★击杀目标后state=walk，继续移动不消失', '单位杀掉敌人后继续前进（关键验证点）', 'T3.3', '待办'),
    ('T3.5', 'P3', '实现到达大殿伤害', '单位y到达大殿格→对大殿造成atk伤害→单位消失', '单位到大殿后大殿血量减少', 'T3.4', '待办'),
    ('T3.6', 'P3', '实现远程攻击', '剑修弟子射程3格，只打敌人不掉血', '远程单位能隔格打人自己不掉血', 'T3.3', '待办'),
    ('T3.7', 'P3', '实现阵法布设与机制', '布在阵法区格子，不动有血量，主动攻击经过敌人', '阵法能拦截并攻击经过的敌方单位', 'T3.3', '待办'),
    ('T3.8', 'P3', '实现阵法冷却', '阵法被毁格子8秒内不能再布阵', '格子被毁后8秒内布阵被拒绝', 'T3.7', '待办'),
    ('T3.9', 'P3', '实现范围伤害', '天罗阵/五雷正法/天雷诀打相邻格所有敌人', '范围攻击能同时打多个目标', 'T3.3', '待办'),
    ('T3.10', 'P3', '实现减速机制', '寒霜阵命中后buff: speed-0.5持续2秒', '被寒霜阵打后移速降低2秒', 'T3.7', '待办'),
    ('T3.11', 'P3', '实现反震机制', '反震阵被攻击时反弹50%伤害给攻击者', '打反震阵的攻击者自己也掉血', 'T3.7', '待办'),
    ('T3.12', 'P3', '实现定身/推后/护盾/加速/禁阵', '困仙索定身2s/移山倒海推后2格/金钟罩免疫3s/御风诀加速/镇魂符禁阵', '5种法术效果各自生效', 'T3.3', '待办'),
    ('T3.13', 'P3', '实现长老随机技能', '金丹长老/护法长老每5秒随机释放4分支之一', '长老每5秒释放一个随机技能', 'T3.3', '待办'),
    # P4 AI
    ('T4.1', 'P4', 'AI灵力管理', '简单够就出/普通攒到4+/困难根据局势', 'AI能按难度管理灵力', 'T2.5', '待办'),
    ('T4.2', 'P4', 'AI出牌决策', '根据大殿血量决定攻守比例+选牌', 'AI能合理出牌不是纯随机', 'T4.1', '待办'),
    ('T4.3', 'P4', 'AI布阵位置', '简单随机/普通大殿前/困难针对性', 'AI布阵位置合理', 'T4.2', '待办'),
    ('T4.4', 'P4', 'AI难度三档', '简单3.5s/普通2.5s/困难1.5s+策略差异', '三档AI行为有明显差异', 'T4.3', '待办'),
    # P5 渲染
    ('T5.1', 'P5', '背景+棋盘渲染', '山道背景(色块)+主路+阵法区格子线', '屏幕能看到棋盘结构', 'T1.5', '待办'),
    ('T5.2', 'P5', '大殿渲染', '顶部/底部大殿(色块)+血量条+受击闪烁', '能看到双方大殿和血量', 'T5.1', '待办'),
    ('T5.3', 'P5', '单位渲染', '单位色块+名称+y排序深度+简单动画', '能看到单位在棋盘上移动', 'T5.1', '待办'),
    ('T5.4', 'P5', '阵法渲染', '阵法色块+状态(正常/受击/禁阵/冷却倒计时)', '能看到阵法及其状态', 'T5.1', '待办'),
    ('T5.5', 'P5', '特效渲染', '命中粒子/死亡碎裂/法术光效(简单版)', '战斗时有视觉反馈', 'T5.3', '待办'),
    ('T5.6', 'P5', 'UI-手牌栏渲染', '底部4张牌+费用+选中上浮+灰显', '能看到手牌并能选中', 'T5.1', '待办'),
    ('T5.7', 'P5', 'UI-灵力条渲染', '灵力数字+液条+上限', '能看到灵力实时变化', 'T5.6', '待办'),
    ('T5.8', 'P5', 'UI-血量计时渲染', '顶部双方血条+中央计时器', '能看到血量和倒计时', 'T5.2', '待办'),
    ('T5.9', 'P5', '出牌交互', '点手牌→高亮→点目标→出牌→灵力扣除→手牌补位', '能完整走通出牌流程', 'T5.6,T5.7', '待办'),
    # P6 系统整合
    ('T6.1', 'P6', '灵力实时回复', '每2.8秒+1，上限随时间5→10', '灵力按时间自动回复', 'T2.4', '待办'),
    ('T6.2', 'P6', '胜负判定', '大殿归零/时限到比血量/加时比灵力', '能正确判定胜负', 'T3.5', '待办'),
    ('T6.3', 'P6', '加时赛', '180秒平局→加时60秒灵力×1.5', '平局时进入加时赛', 'T6.2', '待办'),
    ('T6.4', 'P6', '结算页面', '胜负展示+摧毁度+再来一局按钮', '战斗结束后显示结算', 'T6.2', '待办'),
    ('T6.5', 'P6', '新手引导(4步)', '引导出牌/灵力/布阵/目标，高亮+箭头', '新手能跟着引导完成首局', 'T5.9', '待办'),
    ('T6.6', 'P6', '音效系统', 'BGM+SFX播放+音量控制', '有声音反馈', 'T5.9', '待办'),
    ('T6.7', 'P6', '卡组预设(3套)', '快攻型/控制型/坦克型预设卡组', '能选择不同预设卡组', 'T2.5', '待办'),
    # P7 测试平衡
    ('T7.1', 'P7', '纯逻辑模拟器', 'AI vs AI跑10局，输出摧毁度/胜率/时长', '模拟能跑通输出数据', 'T4.4', '待办'),
    ('T7.2', 'P7', '平衡性调优', '根据模拟数据调参数，目标:摧毁度20~70%/胜率~50%', '模拟10局达标', 'T7.1', '待办'),
    ('T7.3', 'P7', '性能测试', '同屏20单位60fps，内存<150MB', '真机不卡', 'T5.9', '待办'),
    ('T7.4', 'P7', '真机测试', '微信开发者工具+真机预览，无崩溃', '真机能正常游玩', 'T7.3', '待办'),
    ('T7.5', 'P7', '提交审核', '修Bug+提交微信小游戏审核', '审核通过', 'T7.4', '待办'),
]
r = 4
for t in tasks:
    data_row(ws, list(t), r)
    ws.row_dimensions[r].height = 36; r += 1

# 条件格式：状态列着色
from openpyxl.formatting.rule import FormulaRule
ws.conditional_formatting.add(f'G4:G{r-1}',
    FormulaRule(formula=['$G4="已完成"'], fill=BG_DONE))
ws.conditional_formatting.add(f'G4:G{r-1}',
    FormulaRule(formula=['$G4="进行中"'], fill=BG_DOING))
ws.conditional_formatting.add(f'G4:G{r-1}',
    FormulaRule(formula=['$G4="待办"'], fill=BG_TODO))

# ================================================================
# Sheet 5: 卡牌数据
# ================================================================
ws = wb.create_sheet('5.卡牌数据')
set_widths(ws, [14, 10, 8, 8, 8, 8, 8, 30, 8])
title_row(ws, '卡牌数据表 — 改数值直接在这改', 9)

# 攻方人物
r = 3
section_row(ws, '攻方·人物卡（出场即向敌方大殿推进）', 9, r); r += 1
header_row(ws, ['名称','类型','费用','血量','攻击','移速','间隔(s)','特性','稀有度'], r); r += 1
attack_cards = [
    ('宗门体修弟子','普通弟子',2,4,2,1.0,1.0,'普通近战士兵，基础兵','凡品'),
    ('宗门剑修弟子','普通弟子',3,3,3,0.9,1.3,'远程(射程3)，用飞剑，不掉血','凡品'),
    ('宗门御兽弟子','普通弟子',3,6,2,0.8,1.2,'控兽当肉盾，高血低速','凡品'),
    ('金丹期长老','精英长老',6,10,3,0.8,1.5,'每5s随机释放4分支技能','宝品'),
]
for c in attack_cards:
    data_row(ws, list(c), r, fill=BG_ATTACK); r += 1

# 守方人物
r += 1
section_row(ws, '守方·人物卡（布在阵法区/大殿前，不推进或半动）', 9, r); r += 1
header_row(ws, ['名称','类型','费用','血量','攻击','射程','间隔(s)','特性','稀有度'], r); r += 1
defend_cards = [
    ('护山傀儡','防守单位',3,6,2,1,1.2,'缓慢移动的肉盾守卫，拦截','凡品'),
    ('护山灵兽','防守单位',4,8,2,1,1.0,'高血拦截，死亡时自爆1伤','灵品'),
    ('护法长老','精英长老',6,10,3,1,1.5,'镇守大殿前，每5s随机释放4分支','宝品'),
]
for c in defend_cards:
    data_row(ws, list(c), r, fill=BG_DEFEND); r += 1

# 阵法
r += 1
section_row(ws, '技能卡·阵法（双方通用，布在阵法区，不动有血量）', 9, r); r += 1
header_row(ws, ['名称','类型','费用','血量','攻击','射程','间隔(s)','特性','稀有度'], r); r += 1
array_cards = [
    ('截脉阵','阵法',2,4,2,1,1.0,'基础拦截，便宜','凡品'),
    ('寒霜阵','阵法',3,3,1,1,1.0,'命中后敌人移速-0.5(2s)','凡品'),
    ('万刃阵','阵法',4,5,3,1,1.0,'高输出拦截','灵品'),
    ('反震阵','阵法',3,3,0,'-',1.0,'反伤50%','灵品'),
    ('天罗阵','阵法',5,6,2,1,1.0,'范围(打相邻所有敌人)','宝品'),
]
for c in array_cards:
    data_row(ws, list(c), r, fill=BG_ARRAY); r += 1

# 法术
r += 1
section_row(ws, '技能卡·法术（即时效果，需选目标）', 9, r); r += 1
header_row(ws, ['名称','类型','费用','效果','','','','偏向','稀有度'], r); r += 1
spell_cards = [
    ('万剑归宗','法术',5,'全己方单位+1攻、移速+0.3(5s)','','','','攻方','宝品'),
    ('五雷正法','法术',4,'区域3格内敌方受4伤','','','','攻方/通用','灵品'),
    ('御风诀','法术',2,'指定己方单位移速+0.5(5s)','','','','通用','凡品'),
    ('镇魂符','法术',3,'指定敌方阵法失效3秒','','','','通用','凡品'),
    ('金钟罩','法术',3,'己方大殿免疫伤害3秒','','','','守方','灵品'),
    ('移山倒海','法术',4,'区域敌人推后2格+1伤','','','','守方','灵品'),
    ('困仙索','法术',2,'指定敌人定身2秒','','','','守方','凡品'),
    ('天雷诀','法术',4,'范围3格内敌方单位受4伤(清兵)','','','','守方/通用','灵品'),
]
for c in spell_cards:
    data_row(ws, list(c), r, fill=BG_SPELL); r += 1

# 长老分支
r += 1
section_row(ws, '长老技能分支（每5秒随机释放1个）', 9, r); r += 1
header_row(ws, ['长老','分支','流派','效果','','','','触发',''], r); r += 1
elder_skills = [
    ('金丹期长老','飞剑分支','剑修','万剑齐发(范围伤害)','','','','随机(每5s)',''),
    ('金丹期长老','丹药分支','丹修','自身+周围己方回血','','','','随机(每5s)',''),
    ('金丹期长老','符箓分支','符修','天雷诀(范围清兵)','','','','随机(每5s)',''),
    ('金丹期长老','御兽分支','御兽','召唤灵兽助战','','','','随机(每5s)',''),
    ('护法长老','阵法分支','阵修','自动补截脉阵','','','','随机(每5s)',''),
    ('护法长老','符箓分支','符修','天雷诀(清兵)','','','','随机(每5s)',''),
    ('护法长老','丹药分支','丹修','阵法/大殿回血','','','','随机(每5s)',''),
    ('护法长老','御兽分支','御兽','召唤护山灵兽','','','','随机(每5s)',''),
]
for c in elder_skills:
    fill = BG_ATTACK if '金丹' in c[0] else BG_DEFEND
    data_row(ws, list(c), r, fill=fill); r += 1

# ================================================================
# Sheet 6: 数值配置
# ================================================================
ws = wb.create_sheet('6.数值配置')
set_widths(ws, [16, 14, 35, 14, 14])
title_row(ws, '数值配置 — 所有可调参数集中管理', 5)
header_row(ws, ['参数名', '当前值', '说明', '调高偏向', '调低偏向'], 3)

params = [
    ('灵力回复间隔', '2.8秒', '每多少秒回复1点灵力', '攻方(出兵多)', '守方(布阵少)'),
    ('灵力初始上限', '5', '开局灵力上限', '攻方', '守方'),
    ('灵力上限增长间隔', '30秒', '每多少秒上限+1', '攻方', '守方'),
    ('灵力上限封顶', '10', '灵力上限最大值', '攻方', '守方'),
    ('大殿血量', '30', '大殿初始血量', '守方', '攻方'),
    ('单局时长', '180秒', '正常对战时间', '守方', '攻方'),
    ('加时时长', '60秒', '平局后加时时间', '中立', '中立'),
    ('加时灵力倍率', '1.5', '加时赛灵力回复倍率', '攻方', '守方'),
    ('手牌数量', '4', '同时持有手牌数', '攻方(选择多)', '守方'),
    ('卡组数量', '8', '卡组总卡牌数', '中立', '中立'),
    ('抽牌延迟', '2秒', '打出后多久抽新牌', '守方', '攻方'),
    ('阵法冷却时间', '8秒', '阵法被毁后格子冷却时间', '攻方', '守方'),
    ('长老技能间隔', '5秒', '长老多久释放一次技能', '守方', '攻方'),
    ('棋盘长度', '9格', '大殿到大殿的格子数(含大殿)', '守方', '攻方'),
    ('同屏单位上限', '30', '最多同时存在单位数', '中立(性能)', '中立(性能)'),
    ('远程单位射程', '3格', '剑修弟子攻击射程', '攻方', '守方'),
    ('减速幅度', '0.5', '寒霜阵/麻痹减速值', '守方', '攻方'),
    ('减速持续', '2秒', '减速持续时间', '守方', '攻方'),
    ('反震比例', '50%', '反震阵反弹伤害比例', '守方', '攻方'),
    ('护盾持续', '3秒', '金钟罩大殿免疫时间', '守方', '攻方'),
    ('定身持续', '2秒', '困仙索定身时间', '守方', '攻方'),
    ('推后格数', '2格', '移山倒海推后距离', '守方', '攻方'),
    ('加速持续', '5秒', '御风诀/万剑归宗加速时间', '攻方', '守方'),
    ('禁阵持续', '3秒', '镇魂符禁阵时间', '攻方', '守方'),
    ('AI简单思考间隔', '3.5秒', '简单AI决策间隔', 'AI强', 'AI弱'),
    ('AI普通思考间隔', '2.5秒', '普通AI决策间隔', 'AI强', 'AI弱'),
    ('AI困难思考间隔', '1.5秒', '困难AI决策间隔', 'AI强', 'AI弱'),
    ('目标帧率', '60fps', '渲染目标帧率', '-', '-'),
    ('dt上限', '33ms', '单帧最大时间步(防跳帧)', '-', '-'),
]
r = 4
for p in params:
    data_row(ws, list(p), r)
    ws.row_dimensions[r].height = 22; r += 1

# 平衡验证标准
r += 1
section_row(ws, '平衡验证标准（模拟10局AI vs AI）', 5, r); r += 1
verify = [
    '双方摧毁度都在 20%~70% 区间',
    '胜率：攻守方各 ~50%（40%~60%）',
    '0% 摧毁度对局 < 10%',
    '90% 对局在 150~210 秒内结束',
    '双方大殿都被打到（有来有回）',
]
for v in verify:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value='• ' + v)
    c.font = F_CELL; c.alignment = A_LEFT; c.border = BD
    r += 1

# ================================================================
# Sheet 7: AI提问指南（★核心★）
# ================================================================
ws = wb.create_sheet('7.AI提问指南')
set_widths(ws, [18, 75])
title_row(ws, '★ AI提问指南 — 如何让AI替你干活 ★', 2)

r = 3
section_row(ws, '一、黄金公式：每次给AI下指令都按这个结构', 2, r); r += 1
ws.cell(row=r, column=1, value='公式').font = F_BOLD
ws.cell(row=r, column=1).fill = BG_SECTION; ws.cell(row=r, column=1).border = BD; ws.cell(row=r, column=1).alignment = A_CENTER
ws.cell(row=r, column=2, value='【角色定位】+【具体任务】+【上下文/参考文件】+【完成标准】+【自主决策授权】').font = F_CELL
ws.cell(row=r, column=2).alignment = A_LEFTT; ws.cell(row=r, column=2).border = BD
ws.row_dimensions[r].height = 30; r += 1

ws.cell(row=r, column=1, value='示例').font = F_BOLD
ws.cell(row=r, column=1).fill = BG_SECTION; ws.cell(row=r, column=1).border = BD; ws.cell(row=r, column=1).alignment = A_CENTER
ws.cell(row=r, column=2, value='你是微信小游戏开发专家。\n请实现单位的实时移动逻辑：单位沿主路向敌方大殿推进，移速=卡牌配置值，遇敌方单位/阵法停下交战，击杀后继续推进。\n参考卡牌数据见Cards.js，参考设计文档见「御剑对决_游戏设计方案.txt」。\n完成标准：①代码能通过node --check语法检查 ②AI vs AI模拟能看到单位移动 ③单位到大殿后对大殿造成伤害。\n遇到设计细节未明确的按合理默认自行决定，做完统一汇报，不要中途停下来问我。').font = F_CELL
ws.cell(row=r, column=2).alignment = A_LEFTT; ws.cell(row=r, column=2).border = BD
ws.row_dimensions[r].height = 120; r += 2

section_row(ws, '二、6类场景的提示词模板（直接复制使用）', 2, r); r += 1
templates = [
    ('场景1：让AI实现功能',
     '你是微信小游戏开发专家。\n请实现【功能名】：【功能描述，参考任务拆解表的"任务描述"列】。\n参考文件：【列出相关文件路径，如js/game/BattleLogic.js, js/config/Cards.js】。\n完成标准：【参考任务拆解表的"完成标准"列】。\n遇到设计细节未明确的按合理默认自行决定，做完统一汇报，不要中途问我。'),

    ('场景2：让AI修Bug',
     '我遇到了一个Bug：\n【现象描述：什么情况下发生了什么】\n【错误信息：贴上报错日志】\n【相关文件：【文件路径】】\n请分析原因并修复。修复后说明改了什么、为什么这样改。'),

    ('场景3：让AI做平衡调优',
     '请运行AI vs AI模拟10局，输出每局的：双方摧毁度/胜方/对局时长。\n当前参数见「数值配置」表。\n如果发现平衡问题（如摧毁度0%或一边倒），请调整参数并重新模拟，直到达到验证标准：双方摧毁度20~70%，胜率~50%。\n每次调参请记录：调了什么参数、从多少调到多少、为什么。'),

    ('场景4：让AI审查代码',
     '请审查以下文件的代码质量：\n【文件路径】\n检查项：①是否有逻辑Bug ②是否有性能问题 ③是否有边界情况未处理 ④是否符合设计文档。\n列出所有发现的问题，按严重程度排序，每个问题给出修复建议。'),

    ('场景5：让AI设计新内容',
     '请设计【内容名，如"新流派/新卡牌/新关卡"】：\n要求：①符合现有卡牌体系的数值模型（参考卡牌数据表） ②与现有内容不重复 ③标注稀有度和费用 ④给出完整的属性数值表。\n参考现有设计文档和卡牌数据。设计完请说明设计思路和平衡考虑。'),

    ('场景6：让AI自主连续开发',
     '以下是V1阶段的任务清单（按顺序）：\n【粘贴任务拆解表中连续的3~5个任务】\n请按顺序逐个完成，每个任务完成后：\n①标记该任务为"已完成"\n②简要记录做了什么\n③立即开始下一个任务\n全部做完后统一汇报。遇到设计细节未明确的按合理默认自行决定，不要中途问我。\n每个任务都必须通过其"完成标准"验证才算完成。'),
]
for name, template in templates:
    ws.cell(row=r, column=1, value=name).font = F_BOLD
    ws.cell(row=r, column=1).fill = BG_SECTION; ws.cell(row=r, column=1).border = BD; ws.cell(row=r, column=1).alignment = A_CENTER
    ws.cell(row=r, column=2, value=template).font = F_CELL
    ws.cell(row=r, column=2).alignment = A_LEFTT; ws.cell(row=r, column=2).border = BD
    ws.row_dimensions[r].height = 110; r += 1

r += 1
section_row(ws, '三、常见错误与纠正', 2, r); r += 1
mistakes = [
    ('❌ 太模糊', '"做个战斗系统"→AI不知道从哪开始。✅ 改为："实现单位移动：沿y轴推进，speed×dt，遇敌停下"'),
    ('❌ 太大', '"把整个游戏做完"→AI会做不好。✅ 改为：一次一个任务（参考任务拆解表粒度）'),
    ('❌ 没标准', '"帮我写代码"→不知道做完没。✅ 改为：给完成标准"能通过node --check + 模拟能跑"'),
    ('❌ 没授权', 'AI每步都停下来问你怎么做。✅ 改为："遇到细节自行决定，做完汇报，不要中途问我"'),
    ('❌ 没验证', 'AI说做完了你就信了。✅ 改为：按完成标准自己验证（跑模拟/看效果/查语法）'),
    ('❌ 不给上下文', 'AI不知道已有的代码和设计。✅ 改为：告诉它参考文件路径和设计文档'),
]
for wrong, right in mistakes:
    ws.cell(row=r, column=1, value=wrong).font = F_CELL
    ws.cell(row=r, column=1).border = BD; ws.cell(row=r, column=1).alignment = A_LEFTT
    ws.cell(row=r, column=2, value=right).font = F_CELL
    ws.cell(row=r, column=2).border = BD; ws.cell(row=r, column=2).alignment = A_LEFTT
    ws.row_dimensions[r].height = 36; r += 1

r += 1
section_row(ws, '四、每日开发流程（照着做）', 2, r); r += 1
daily = [
    ('第1步', '打开「任务拆解」表，找到第一个状态为"待办"的任务'),
    ('第2步', '阅读该任务的"任务描述"和"完成标准"'),
    ('第3步', '打开「AI提问指南」，用「场景1」模板，把任务描述和完成标准填进去，发给AI'),
    ('第4步', 'AI完成后，你按"完成标准"验证（跑模拟/看效果/查语法）'),
    ('第5步', '验证通过→在「任务拆解」表标记"已完成"→在「开发日志」记录今天做了什么'),
    ('第6步', '验证不通过→用「场景2」模板把问题发给AI修→重复第4步'),
    ('第7步', '重复第1~6步，做下一个任务。每天做2~4个任务为宜'),
]
for step, desc in daily:
    ws.cell(row=r, column=1, value=step).font = F_BOLD
    ws.cell(row=r, column=1).fill = BG_SECTION; ws.cell(row=r, column=1).border = BD; ws.cell(row=r, column=1).alignment = A_CENTER
    ws.cell(row=r, column=2, value=desc).font = F_CELL
    ws.cell(row=r, column=2).alignment = A_LEFTT; ws.cell(row=r, column=2).border = BD
    ws.row_dimensions[r].height = 30; r += 1

# ================================================================
# Sheet 8: 进度看板
# ================================================================
ws = wb.create_sheet('8.进度看板')
set_widths(ws, [12, 10, 10, 10, 10, 12])
title_row(ws, '进度看板 — 一眼看全局', 6)
header_row(ws, ['阶段', '总任务数', '已完成', '进行中', '待办', '完成率'], 3)

# 统计各阶段
from collections import Counter
phase_stats = {}
for t in tasks:
    phase = t[1]
    if phase not in phase_stats:
        phase_stats[phase] = {'total': 0, 'done': 0, 'doing': 0, 'todo': 0}
    phase_stats[phase]['total'] += 1
    status = t[6]
    if status == '已完成':
        phase_stats[phase]['done'] += 1
    elif status == '进行中':
        phase_stats[phase]['doing'] += 1
    else:
        phase_stats[phase]['todo'] += 1

r = 4
for phase in sorted(phase_stats.keys()):
    s = phase_stats[phase]
    rate = f"{s['done']}/{s['total']} ({s['done']*100//s['total']}%)"
    data_row(ws, [phase, s['total'], s['done'], s['doing'], s['todo'], rate], r)
    ws.row_dimensions[r].height = 24; r += 1

# 总计
total_all = sum(s['total'] for s in phase_stats.values())
done_all = sum(s['done'] for s in phase_stats.values())
doing_all = sum(s['doing'] for s in phase_stats.values())
todo_all = sum(s['todo'] for s in phase_stats.values())
rate_all = f"{done_all}/{total_all} ({done_all*100//total_all if total_all else 0}%)"
ws.cell(row=r, column=1, value='总计').font = F_BOLD
ws.cell(row=r, column=1).fill = BG_HEADER; ws.cell(row=r, column=1).font = F_HEADER
for i, v in enumerate(['', total_all, done_all, doing_all, todo_all, rate_all], 1):
    c = ws.cell(row=r, column=i, value=v)
    c.font = F_HEADER; c.fill = BG_HEADER; c.alignment = A_CENTER; c.border = BD
ws.row_dimensions[r].height = 26; r += 2

# 使用说明
section_row(ws, '使用说明', 6, r); r += 1
notes = [
    '完成一个任务后，去「任务拆解」表把该任务的"状态"列改为"已完成"',
    '开始做一个任务时，把"状态"改为"进行中"',
    '本表的数字会自动更新（需重新打开Excel刷新公式，或手动更新）',
    '每周看一次整体完成率，确认进度健康',
]
for n in notes:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(row=r, column=1, value='• ' + n)
    c.font = F_CELL; c.alignment = A_LEFT; c.border = BD
    r += 1

# ================================================================
# Sheet 9: Bug跟踪
# ================================================================
ws = wb.create_sheet('9.Bug跟踪')
set_widths(ws, [8, 12, 35, 10, 10, 30, 12])
title_row(ws, 'Bug与问题跟踪', 7)
header_row(ws, ['编号', '发现日期', '问题描述', '严重度', '状态', '修复方案', '修复日期'], 3)
# 留空行供填写
for i in range(4, 24):
    for c in range(1, 8):
        ws.cell(row=i, column=c).border = BD
        ws.cell(row=i, column=c).font = F_CELL
        ws.cell(row=i, column=c).alignment = A_LEFTT
    ws.row_dimensions[i].height = 24

ws.conditional_formatting.add('D4:D23',
    FormulaRule(formula=['$D4="严重"'], fill=BG_TODO))
ws.conditional_formatting.add('D4:D23',
    FormulaRule(formula=['$D4="一般"'], fill=BG_DOING))
ws.conditional_formatting.add('D4:D23',
    FormulaRule(formula=['$D4="轻微"'], fill=BG_DONE))

ws.conditional_formatting.add('E4:E23',
    FormulaRule(formula=['$E4="已修复"'], fill=BG_DONE))
ws.conditional_formatting.add('E4:E23',
    FormulaRule(formula=['$E4="修复中"'], fill=BG_DOING))
ws.conditional_formatting.add('E4:E23',
    FormulaRule(formula=['$E4="待修复"'], fill=BG_TODO))

r = 25
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
c = ws.cell(row=r, column=1, value='严重度：严重(红/阻塞开发) / 一般(黄/影响体验) / 轻微(绿/不影响主流程）    状态：待修复 / 修复中 / 已修复')
c.font = F_NOTE; c.alignment = A_LEFT

# ================================================================
# Sheet 10: 开发日志
# ================================================================
ws = wb.create_sheet('10.开发日志')
set_widths(ws, [12, 10, 30, 30, 30])
title_row(ws, '开发日志 — 每天记录', 5)
header_row(ws, ['日期', '完成任务', '做了什么', '遇到的问题', '下一步计划'], 3)
for i in range(4, 54):
    for c in range(1, 6):
        ws.cell(row=i, column=c).border = BD
        ws.cell(row=i, column=c).font = F_CELL
        ws.cell(row=i, column=c).alignment = A_LEFTT
    ws.row_dimensions[i].height = 40

r = 55
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws.cell(row=r, column=1, value='每天开发结束前花2分钟填写。记录"做了什么"和"遇到的问题"能帮你积累经验，下次遇到类似问题能快速回忆。')
c.font = F_NOTE; c.alignment = A_LEFT

# ===== 保存 =====
output = '/workspace/宗门论道_开发管理工具包.xlsx'
wb.save(output)
print(f'Excel已生成: {output}')
print(f'共{len(wb.sheetnames)}个Sheet: {wb.sheetnames}')
print(f'任务拆解: {len(tasks)}个任务')
print(f'功能清单: {len(features)}个功能点')
print(f'卡牌数据: {len(attack_cards)+len(defend_cards)+len(array_cards)+len(spell_cards)+len(elder_skills)}张卡牌')
print(f'数值配置: {len(params)}个参数')
print(f'AI提问模板: {len(templates)}个')
