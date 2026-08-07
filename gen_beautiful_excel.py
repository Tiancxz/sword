#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
《宗门论道》开发管理工具包 v3 — 美化版
全新设计：美观配色 + 清晰层级标注 + 游戏设计思路 + 玩法说明
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

PATH = '/workspace/宗门论道_开发管理工具包.xlsx'
wb = openpyxl.Workbook()

# ===== 美观配色方案（修仙主题：墨蓝+金+青） =====
C_INK    = '2C3E50'  # 墨蓝（主标题）
C_GOLD   = 'B7950B'  # 金色（修仙强调）
C_BLUE   = '2874A6'  # 蓝
C_TEAL   = '117A65'  # 青
C_PURPLE = '6C3483'  # 紫
C_RED    = 'C0392B'  # 红
C_GREEN  = '27AE60'  # 绿

# 层级渐变色（从深到浅）
L1_FILLS = {
    'A': ('1A5276', 'FFFFFF'),  # 引擎-深蓝
    'B': ('2874A6', 'FFFFFF'),  # 数据-蓝
    'C': ('117A65', 'FFFFFF'),  # 战斗-青
    'D': ('B7950B', 'FFFFFF'),  # AI-金
    'E': ('8E44AD', 'FFFFFF'),  # 渲染-紫
    'F': ('C0392B', 'FFFFFF'),  # 社交-红
    'G': ('D35400', 'FFFFFF'),  # 变现-橙
    'H': ('34495E', 'FFFFFF'),  # PvP-灰蓝
}
L2_FILL = 'D6EAF8'  # 模块层-浅蓝
L3_FILL = 'FDF2E9'  # 子功能层-浅橙
L4_FILL = 'F4ECF7'  # 原理层-浅紫
L5_FILL = 'E8F8F5'  # 函数层-浅青

# 状态色
ST_TODO  = 'FADBD8'  # 待办-浅红
ST_DOING = 'FEF9E7'  # 进行中-浅黄
ST_DONE  = 'D5F5E3'  # 已完成-浅绿

# 字体
F_BIG    = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
F_TITLE  = Font(name='微软雅黑', size=13, bold=True, color='FFFFFF')
F_HEADER = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
F_L1     = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
F_L2     = Font(name='微软雅黑', size=11, bold=True, color=C_INK)
F_L3     = Font(name='微软雅黑', size=10, bold=True, color='333333')
F_CELL   = Font(name='微软雅黑', size=10, color='2C3E50')
F_CODE   = Font(name='Consolas', size=10, color='1A5276')
F_SMALL  = Font(name='微软雅黑', size=9, color='7F8C8D')
F_TAG    = Font(name='微软雅黑', size=8, bold=True, color='FFFFFF')
F_GOLD   = Font(name='微软雅黑', size=11, bold=True, color=C_GOLD)

# 对齐
A_C  = Alignment(horizontal='center', vertical='center', wrap_text=True)
A_L  = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
A_LT = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=1)
A_LM = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 边框（细浅色）
BD_THIN = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC')
)
BD_NONE = Border()

def set_w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def big_title(ws, text, cols, row=1, color=C_INK):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_BIG
    c.fill = PatternFill('solid', fgColor=color)
    c.alignment = A_C
    ws.row_dimensions[row].height = 44

def sub_title(ws, text, cols, row, color=C_GOLD):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_TITLE
    c.fill = PatternFill('solid', fgColor=color)
    c.alignment = A_L
    ws.row_dimensions[row].height = 30

def header_row(ws, headers, row, color=C_INK):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = F_HEADER
        c.fill = PatternFill('solid', fgColor=color)
        c.alignment = A_C
        c.border = BD_THIN
    ws.row_dimensions[row].height = 30

# ================================================================
# Sheet 1: 游戏设计思路与玩法
# ================================================================
ws = wb.active
ws.title = '1.设计思路与玩法'
ws.sheet_view.showGridLines = False
set_w(ws, [4, 20, 20, 20, 20, 20, 4])

r = 1
big_title(ws, '《宗门论道》 — 游戏设计思路与玩法', 7, r, C_INK); r += 1
# 空行
ws.row_dimensions[r].height = 8; r += 1

# === 一、游戏概述 ===
sub_title(ws, '一、游戏概述', 7, r, C_GOLD); r += 1
overview = [
    ('游戏名称', '宗门论道（暂定）'),
    ('游戏类型', '实时策略卡牌（RTS-lite + 卡牌费用制）'),
    ('游戏题材', '东方修仙 / 宗门大战'),
    ('运行平台', '微信小游戏（Canvas 2D 渲染）'),
    ('单局时长', '3~5 分钟（180秒正赛 + 60秒加时）'),
    ('目标用户', '18~35岁，喜欢轻度策略、修仙题材、碎片化对战的玩家'),
    ('一句话简介', '皇室战争式的实时卡牌攻防，修仙皮——你与对手各据山门一座，实时派弟子出征、布阵拦截、长老施法，摧毁对方宗门大殿者胜。'),
]
for k, v in overview:
    ws.cell(row=r, column=2, value=k).font = F_L2
    ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=L2_FILL)
    ws.cell(row=r, column=2).alignment = A_C; ws.cell(row=r, column=2).border = BD_THIN
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.cell(row=r, column=3, value=v).font = F_CELL
    ws.cell(row=r, column=3).alignment = A_LM; ws.cell(row=r, column=3).border = BD_THIN
    ws.row_dimensions[r].height = 28; r += 1

r += 1
# === 二、核心玩法 ===
sub_title(ws, '二、核心玩法 — 实时双线互推', 7, r, C_BLUE); r += 1

# 玩法说明
play_intro = [
    '你和对手各据一座宗门大殿（30血），实时同时出兵互攻，谁先拆掉对方大殿谁赢，3分钟时限。',
    '',
    '【你做什么】',
    '  1. 出兵：花灵力出弟子（体修/剑修/御兽），单位自动沿山道向敌方大殿推进，遇敌自动交战，杀完继续冲',
    '  2. 布阵：在主路两侧阵法区布阵法（不动，拦截+攻击经过的敌人），剑阵双方通用',
    '  3. 施法：放法术（五雷正法范围炸、万剑归宗全队加速、金钟罩大殿免疫等）',
    '  4. 出长老：花6费出金丹长老，每5秒随机释放飞剑/丹药/符箓/御兽分支技能',
    '',
    '【灵力系统】',
    '  实时回复：每2.8秒+1，开局上限5随时间涨到10。灵力不够出不了牌——攒大单位还是铺小单位，是核心决策。',
    '',
    '【地图视角】折中方案：逻辑上下直推 + 视觉山道攻山 + 两侧阵法区',
]
for line in play_intro:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(row=r, column=2, value=line)
    if line.startswith('【'):
        c.font = F_GOLD
    else:
        c.font = F_CELL
    c.alignment = A_LM; c.border = BD_THIN
    ws.row_dimensions[r].height = 24 if line else 8; r += 1

r += 1
# 地图示意
sub_title(ws, '地图布局示意', 7, r, C_TEAL); r += 1
map_lines = [
    '     ┌─────────────────────────┐',
    '     │      [敌方大殿] 30血     │  ← 你要拆的（顶端）',
    '     │  阵法区│主路│阵法区      │  ← 两侧可布阵拦截',
    '     │       │    │            │',
    '     │  阵法区│主路│阵法区      │  ← 中段，单位在此交战',
    '     │       │    │            │',
    '     │  阵法区│主路│阵法区      │  ← 下段',
    '     │      [我方大殿] 30血     │  ← 你要守的（底端）',
    '     └─────────────────────────┘',
    '     单位沿主路向上推进（2朝向：上/下），背景做成山道攻山感',
]
for line in map_lines:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(row=r, column=2, value=line)
    c.font = Font(name='Consolas', size=11, color=C_INK)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = BD_THIN
    ws.row_dimensions[r].height = 20; r += 1

r += 1
# === 三、设计思路 ===
sub_title(ws, '三、设计思路 — 为什么这么设计', 7, r, C_PURPLE); r += 1

design_thoughts = [
    ('为什么做实时不做回合制？',
     '回合制的痛点：等对方出招很无聊，节奏慢。实时制双方同时操作互不等待，每一秒都有决策压力，紧张感是核心体验。参考皇室战争成功验证了这一点。'),
    ('为什么攻方"击杀后继续推进"？',
     '★这是最核心的设计修复★。旧版"遇阵即停打完即没"导致攻方永远过不去，0%摧毁度。改成击杀后不消失继续冲，攻方才有突破感，守方才有"阵被破"的危机感。这是整个游戏好不好玩的命门。'),
    ('为什么阵法有8秒冷却？',
     '防止守方无限补阵堵路。阵被破后该格8秒不能再布，攻方有突破窗口。这创造了"攻守博弈"的节奏：守方要判断哪格该补、哪格该放弃。'),
    ('为什么用折中视角不用斜45度？',
     '纯斜45度等距视角美术成本3~4倍（每单位需8方向vs2方向），碰撞复杂（菱形vs矩形）。折中方案：逻辑上下直推（实现简单），视觉做成山道攻山（保留修仙味），两全其美。'),
    ('为什么金丹长老是随机技能？',
     '固定技能会变成"最优解"——玩家算出哪个分支最强就只用那个。随机性增加不可预测性，每次出长老都有惊喜/惊吓，增加对局变化和观赏性。'),
    ('为什么剑阵双方通用？',
     '如果阵法只能防守用，攻方缺乏掩护手段。双方通用后，攻方可铺阵掩护推进（阵法跟着部队走），守方布阵拦截，策略空间更大。'),
    ('为什么选修仙题材？',
     '①国内过审零风险（法术不血腥）②修仙认知度高（飞剑/阵法/灵兽玩家秒懂）③"攻山"主题与推塔玩法天然契合④市场有验证（一念逍遥等）。现代战争过审风险极高，科幻需重设题材。'),
    ('为什么V1先做3个流派？',
     'MVP原则：用最少的验证核心玩法。剑修（突进）+符修（远程）+阵修（防守）覆盖了攻防基本循环。验证好玩后再加丹修/傀儡/御兽，避免过度设计。'),
]
for q, a in design_thoughts:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
    ws.cell(row=r, column=2, value='问').font = F_TAG
    ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=C_PURPLE)
    ws.cell(row=r, column=2).alignment = A_C; ws.cell(row=r, column=2).border = BD_THIN
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.cell(row=r, column=3, value=q).font = F_L2
    ws.cell(row=r, column=3).alignment = A_LM; ws.cell(row=r, column=3).border = BD_THIN
    ws.cell(row=r, column=3).fill = PatternFill('solid', fgColor=L4_FILL)
    ws.row_dimensions[r].height = 26; r += 1

    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
    ws.cell(row=r, column=2, value='答').font = F_TAG
    ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=C_TEAL)
    ws.cell(row=r, column=2).alignment = A_C; ws.cell(row=r, column=2).border = BD_THIN
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.cell(row=r, column=3, value=a).font = F_CELL
    ws.cell(row=r, column=3).alignment = A_LT; ws.cell(row=r, column=3).border = BD_THIN
    ws.row_dimensions[r].height = 50; r += 1

r += 1
# === 四、攻守平衡设计 ===
sub_title(ws, '四、攻守平衡设计 — 怎么保证攻得进去也防得下来', 7, r, C_RED); r += 1

balance = [
    ('攻方保障（能攻进去）', C_RED, [
        '1. 单位击杀后继续推进（不消失）——攻方有突破感',
        '2. 阵法被毁格子8秒冷却——守方不能无限堵路',
        '3. 攻方有破阵工具：体修扛伤、五雷正法清阵、爆裂符换阵',
        '4. 灵力实时回复——攻方可持续出兵施压',
    ]),
    ('守方保障（能防下来）', C_BLUE, [
        '1. 阵法便宜性价比高（截脉阵费2血4，换攻方更多时间）',
        '2. 多阵叠加+特种阵法：寒霜减速、反震反伤、天罗范围',
        '3. 反制法术兜底：金钟罩免疫、移山倒海推后、困仙索定身',
        '4. 护山傀儡/灵兽提供肉盾拦截',
    ]),
    ('平衡验证标准', C_GOLD, [
        '模拟10局AI vs AI：双方摧毁度都在20%~70%区间',
        '胜率：攻守方各~50%（40%~60%）',
        '0%摧毁度对局<10%（不再出现打不进去的死局）',
        '90%对局在150~210秒内结束',
    ]),
]
for title_text, color, items in balance:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(row=r, column=2, value=title_text)
    c.font = F_TITLE; c.fill = PatternFill('solid', fgColor=color)
    c.alignment = A_L; c.border = BD_THIN
    ws.row_dimensions[r].height = 28; r += 1
    for item in items:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c = ws.cell(row=r, column=2, value='  ' + item)
        c.font = F_CELL; c.alignment = A_LM; c.border = BD_THIN
        ws.row_dimensions[r].height = 22; r += 1
    r += 1

r += 1
# === 五、卡牌体系 ===
sub_title(ws, '五、卡牌体系总览', 7, r, C_GOLD); r += 1
cards_overview = [
    ('攻方·普通弟子', C_RED, '出场即向敌方大殿推进', '体修弟子(近战/费2) | 剑修弟子(远程/费3) | 御兽弟子(肉盾/费3)'),
    ('攻方·精英长老', C_RED, '每5秒随机释放4分支技能', '金丹期长老(费6/血10) → 飞剑/丹药/符箓/御兽 随机'),
    ('守方·防守单位', C_BLUE, '布在阵法区/大殿前，不推进', '护山傀儡(肉盾/费3) | 护山灵兽(高血自爆/费4) | 护法长老(费6)'),
    ('阵法（双方通用）', C_TEAL, '布在阵法区格子，不动有血量', '截脉阵/寒霜阵/万刃阵/反震阵/天罗阵'),
    ('法术（即时效果）', C_PURPLE, '需选目标，即时生效', '万剑归宗/五雷正法/御风诀/镇魂符/金钟罩/移山倒海/困仙索/天雷诀'),
]
for name, color, desc, detail in cards_overview:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
    c = ws.cell(row=r, column=2, value=name)
    c.font = F_L1; c.fill = PatternFill('solid', fgColor=color)
    c.alignment = A_C; c.border = BD_THIN
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    c = ws.cell(row=r, column=3, value=desc)
    c.font = F_CELL; c.alignment = A_LM; c.border = BD_THIN
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    c = ws.cell(row=r, column=5, value=detail)
    c.font = F_CODE; c.alignment = A_LM; c.border = BD_THIN
    ws.row_dimensions[r].height = 30; r += 1

r += 1
# === 六、版本规划 ===
sub_title(ws, '六、版本规划', 7, r, C_INK); r += 1
versions = [
    ('V1.0', 'MVP实时制PvE', '9~10周', '实时核心循环+3弟子+金丹长老+5阵法+8法术+AI对手+新手引导\n验收：模拟10局摧毁度20~70%，真机60fps'),
    ('V1.5', '社交+养成+美术', '11周', 'Spine美术+云存档+异步PvP+排行榜+养成系统+变现\n验收：次留35%+，付费率3%+'),
    ('V2.0', '实时PvP', '13周', 'WebSocket+状态同步+匹配段位+赛季+内容扩充\n验收：实时PvP延迟<200ms，活跃占比60%+'),
]
header_row(ws, ['版本', '主题', '周期', '内容与验收标准'], r, C_INK)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=3)
ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=4)
ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
r += 1
for ver, theme, period, content in versions:
    ws.cell(row=r, column=2, value=ver).font = F_L1
    ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=C_GOLD)
    ws.cell(row=r, column=2).alignment = A_C; ws.cell(row=r, column=2).border = BD_THIN
    ws.cell(row=r, column=3, value=theme).font = F_L2
    ws.cell(row=r, column=3).alignment = A_C; ws.cell(row=r, column=3).border = BD_THIN
    ws.cell(row=r, column=4, value=period).font = F_CELL
    ws.cell(row=r, column=4).alignment = A_C; ws.cell(row=r, column=4).border = BD_THIN
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    ws.cell(row=r, column=5, value=content).font = F_CELL
    ws.cell(row=r, column=5).alignment = A_LT; ws.cell(row=r, column=5).border = BD_THIN
    ws.row_dimensions[r].height = 48; r += 1

# ================================================================
# Sheet 2: 开发主表（美化版 + 层级标注）
# ================================================================
ws = wb.create_sheet('2.开发主表')
ws.sheet_view.showGridLines = False
# 列: 层级标签 | 编号 | L1系统 | L2模块 | L3子功能 | L4实现原理 | L5函数 | 输入→输出 | 版本 | 状态
set_w(ws, [5, 8, 12, 14, 20, 42, 26, 22, 6, 8])

big_title(ws, '开发主表 — 从系统到函数，逐层展开', 10, 1, C_INK)
header_row(ws, ['层级', '编号', 'L1 系统', 'L2 模块', 'L3 子功能', 'L4 实现原理', 'L5 函数', '输入 → 输出', '版本', '状态'], 2, C_INK)
ws.freeze_panes = 'A3'

# 数据（复用之前的D数据，精简版）
D = []

# ============ A. 引擎框架 ============
S, SN = 'A', 'A.引擎框架'
M = 'A1.渲染循环'
D += [
    ('L1', 'A1', S, SN, M, '—', '—', '—', '—', '—'),
    ('L2', 'A1', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'A1.01', S, '', M, '主循环驱动', 'requestAnimationFrame循环，每帧dt=(now-lastTime)/1000，上限0.033防跳帧，update(dt)→render()', 'Director.loop(ts)', 'timestamp → void', 'V1', '待办'),
    ('L3', 'A1.02', S, '', M, '逻辑更新', '每帧调用当前场景onUpdate(dt)，场景内驱动游戏逻辑', 'Director.update(dt)', 'dt → void', 'V1', '待办'),
    ('L3', 'A1.03', S, '', M, '画面渲染', '每帧清屏→调用当前场景onRender()', 'Director.render()', 'void → void', 'V1', '待办'),
    ('L3', 'A1.04', S, '', M, '引擎初始化', '获取canvas→创建SceneManager→创建Input→绑定触摸→启动循环', 'Director.init()', 'void → void', 'V1', '待办'),
]
M = 'A2.输入系统'
D += [
    ('L2', 'A2', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'A2.01', S, '', M, '触摸监听', 'wx.onTouchStart/Move/End绑定，获取触摸坐标', 'Input.init()', 'void → void', 'V1', '待办'),
    ('L3', 'A2.02', S, '', M, '点击区域注册', '场景/UI向Input注册矩形区域{id,x,y,w,h,cb}，存入数组', 'Input.register(id,x,y,w,h,cb)', '区域参数 → void', 'V1', '待办'),
    ('L3', 'A2.03', S, '', M, '命中检测', '触摸时遍历区域，判断点在矩形内(x<tx<x+w且y<ty<y+h)，命中调cb', 'Input.onTouchStart(tx,ty)', '坐标 → 触发cb', 'V1', '待办'),
    ('L3', 'A2.04', S, '', M, '区域清理', '场景切换时清空注册区域，防旧场景残留', 'Input.clear()', 'void → void', 'V1', '待办'),
]
M = 'A3.事件总线'
D += [
    ('L2', 'A3', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'A3.01', S, '', M, '事件注册', 'on(event,cb)注册监听，存入listeners Map', 'EventSystem.on(event,cb)', '事件名+回调 → void', 'V1', '待办'),
    ('L3', 'A3.02', S, '', M, '事件触发', 'emit(event,data)遍历该事件所有cb并调用', 'EventSystem.emit(event,data)', '事件名+数据 → void', 'V1', '待办'),
    ('L3', 'A3.03', S, '', M, '事件移除', 'off(event,cb)移除监听，防内存泄漏', 'EventSystem.off(event,cb)', '事件名+回调 → void', 'V1', '待办'),
]
M = 'A4.场景管理'
D += [
    ('L2', 'A4', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'A4.01', S, '', M, '场景注册', '注册场景名→实例到Map', 'SceneManager.register(name,scene)', '名称+场景 → void', 'V1', '待办'),
    ('L3', 'A4.02', S, '', M, '场景切换', '旧场景onExit()→清理→切换→新场景onEnter()', 'SceneManager.switch(name)', '场景名 → void', 'V1', '待办'),
]
M = 'A5.渲染器'
D += [
    ('L2', 'A5', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'A5.01', S, '', M, '矩形绘制', 'ctx.fillRect绘制色块(单位/阵法占位)', 'Renderer.rect(x,y,w,h,color)', '坐标+颜色 → void', 'V1', '待办'),
    ('L3', 'A5.02', S, '', M, '文字绘制', 'ctx.fillText绘制名称/血量/灵力', 'Renderer.text(text,x,y,color,size)', '文本+坐标 → void', 'V1', '待办'),
    ('L3', 'A5.03', S, '', M, '血条绘制', '按比例绘制红绿渐变血条(背景红+前景绿)', 'Renderer.healthBar(x,y,w,h,ratio)', '坐标+比例 → void', 'V1', '待办'),
]
M = 'A6.粒子系统'
D += [
    ('L2', 'A6', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'A6.01', S, '', M, '粒子发射', '在指定位置生成N个粒子{x,y,vx,vy,life,color,size}，随机方向', 'ParticleSystem.emit(x,y,count,color)', '坐标+数量 → void', 'V1', '待办'),
    ('L3', 'A6.02', S, '', M, '粒子更新', '每帧x+=vx*dt,y+=vy*dt,life-=dt，life<=0移除(复用对象池)', 'ParticleSystem.update(dt)', 'dt → void', 'V1', '待办'),
    ('L3', 'A6.03', S, '', M, '粒子渲染', '遍历存活粒子，按life/maxLife算透明度绘制', 'ParticleSystem.render()', 'void → void', 'V1', '待办'),
]
M = 'A7.音频管理'
D += [
    ('L2', 'A7', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'A7.01', S, '', M, 'BGM播放', 'wx.getBackgroundAudioManager→src→play循环', 'AudioManager.playBGM(url)', 'URL → void', 'V1', '待办'),
    ('L3', 'A7.02', S, '', M, 'SFX播放', 'wx.createInnerAudioContext→src→play(用对象池复用)', 'AudioManager.playSFX(url)', 'URL → void', 'V1', '待办'),
]

# ============ B. 游戏数据 ============
S, SN = 'B', 'B.游戏数据'
M = 'B1.常量配置'
D += [
    ('L1', 'B', S, SN, '—', '—', '—', '—', '—', '—'),
    ('L2', 'B1', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'B1.01', S, '', M, '全局常量', '棋盘9格/灵力2.8s/大殿30血/手牌4张等所有硬编码值集中管理', 'Constants.js', '文件 → 引用', 'V1', '待办'),
]
M = 'B2.卡牌数据表'
D += [
    ('L2', 'B2', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'B2.01', S, '', M, '卡牌配置', '19张卡牌完整属性配置驱动，加卡牌只加配置不改逻辑', 'Cards.js', '文件 → 引用', 'V1', '待办'),
    ('L3', 'B2.02', S, '', M, '卡牌查表', '通过cardId从Cards表读属性，创建Unit/Formation', 'Cards.get(cardId)', 'ID → 属性Object', 'V1', '待办'),
]
M = 'B3.卡组手牌'
D += [
    ('L2', 'B3', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'B3.01', S, '', M, '卡组初始化', '传入8张ID→复制到drawPile→Fisher-Yates洗牌→抽3张起手', 'Deck.init(cardIds)', 'ID数组 → void', 'V1', '待办'),
    ('L3', 'B3.02', S, '', M, '洗牌算法', '从后往前遍历随机交换(Fisher-Yates O(n)均匀)', 'Deck.shuffle()', 'void → void', 'V1', '待办'),
    ('L3', 'B3.03', S, '', M, '抽牌', 'drawPile末尾pop→push到hand→满4张或牌空不抽', 'Deck.draw()', 'void → 卡牌ID/null', 'V1', '待办'),
    ('L3', 'B3.04', S, '', M, '打出后补牌', '打出牌后drawTimer=2秒→倒计时到0抽1张', 'Deck.update(dt)', 'dt → void', 'V1', '待办'),
    ('L3', 'B3.05', S, '', M, '出牌检查', '检查手牌index对应费用<=灵力', 'Deck.canPlay(idx,energy)', 'index+灵力 → bool', 'V1', '待办'),
    ('L3', 'B3.06', S, '', M, '打出牌', '从hand移除指定index→启动drawTimer', 'Deck.playCard(idx)', 'index → 卡牌ID', 'V1', '待办'),
]

# ============ C. 战斗逻辑 ============
S, SN = 'C', 'C.战斗逻辑'
M = 'C1.战斗主循环'
D += [
    ('L1', 'C', S, SN, '—', '—', '—', '—', '—', '—'),
    ('L2', 'C1', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'C1.01', S, '', M, '每帧总入口', '顺序: 灵力回复→抽牌→单位更新→阵法更新→长老技能→AI→清理死亡→检查胜负→倒计时', 'BattleLogic.update(dt)', 'dt → void', 'V1', '待办'),
]
M = 'C2.单位实体'
D += [
    ('L2', 'C2', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'C2.01', S, '', M, '从卡牌创建单位', '读Cards→创建Unit→设置facing(0向上=1,1向下=-1)→state=walking', 'Unit.fromCard(cardId,owner,x,y)', 'ID+归属+坐标 → Unit', 'V1', '待办'),
    ('L3', 'C2.02', S, '', M, 'buff系统', 'buff={type,value,duration}，每帧倒计时超时移除', 'Unit.addBuff(type,value,dur)', '类型+数值+持续 → void', 'V1', '待办'),
    ('L3', 'C2.03', S, '', M, '实际移速', 'baseSpeed+buff之和(slow:-0.5/speed:+0.5/stun:0)', 'Unit.getEffectiveSpeed()', 'void → 移速', 'V1', '待办'),
    ('L3', 'C2.04', S, '', M, '受到伤害', 'hp-=amount→if<=0:state=dead→触发死亡效果(自爆)', 'Unit.takeDamage(amount,attacker)', '伤害+攻击者 → 实际伤害', 'V1', '待办'),
]
M = 'C3.移动系统'
D += [
    ('L2', 'C3', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'C3.01', S, '', M, '★单位移动★', 'if walking: 找目标→有目标且在范围: fighting→else: y+=speed*dt*facing→检查大殿', 'MovementSystem.moveUnit(unit,dt)', '单位+dt → void', 'V1', '待办'),
    ('L3', 'C3.02', S, '', M, '★目标选择★', '获取敌方单位+阵法→过滤同列→按y距离排序→返回最近的→不在范围返回null', 'MovementSystem.findTarget(unit,model)', '单位+状态 → 目标/null', 'V1', '待办'),
    ('L3', 'C3.03', S, '', M, '碰撞检测', 'y轴距离<attackRange(近战1/远程3)返回true', 'MovementSystem.checkCollision(u,t)', '单位+目标 → bool', 'V1', '待办'),
    ('L3', 'C3.04', S, '', M, '到达大殿', 'owner==0且y>=8 或 owner==1且y<=0 → true', 'MovementSystem.checkHallReach(u,m)', '单位+状态 → bool', 'V1', '待办'),
]
M = 'C4.战斗系统'
D += [
    ('L2', 'C4', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'C4.01', S, '', M, '攻击主逻辑', 'if now-lastAttack>=interval: 选近战/远程/范围→执行→重置时间', 'CombatSystem.attack(attacker,target,dt)', '攻击者+目标+dt → void', 'V1', '待办'),
    ('L3', 'C4.02', S, '', M, '近战互扣', 'target.takeDamage(atk)→attacker.takeDamage(target.atk)双方掉血', 'CombatSystem.meleeAttack(a,t)', '攻击者+目标 → void', 'V1', '待办'),
    ('L3', 'C4.03', S, '', M, '远程单方', 'target.takeDamage(atk)只扣对方，攻击者不掉', 'CombatSystem.rangedAttack(a,t)', '攻击者+目标 → void', 'V1', '待办'),
    ('L3', 'C4.04', S, '', M, '范围伤害', '对目标数组每个takeDamage(damage)', 'CombatSystem.aoeAttack(src,targets,dmg)', '来源+数组+伤害 → void', 'V1', '待办'),
    ('L3', 'C4.05', S, '', M, '★击杀后继续推进★', 'target.isDead()时: killer.state=walking→target=null→继续走(核心修复)', 'CombatSystem.handleKill(killer,victim)', '击杀者+被杀者 → void', 'V1', '待办'),
    ('L3', 'C4.06', S, '', M, '大殿受伤', 'if hallShield>0: 免疫→else: hallHp-=amount→if<=0: 结束', 'CombatSystem.damageHall(player,amount)', '玩家+伤害 → void', 'V1', '待办'),
]
M = 'C5.阵法系统'
D += [
    ('L2', 'C5', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'C5.01', S, '', M, '布阵', '检查格子冷却→创建Formation→加入数组→扣灵力→打出手牌', 'BattleLogic.placeFormation(p,cardId,gx,gy)', '玩家+卡牌+坐标 → bool', 'V1', '待办'),
    ('L3', 'C5.02', S, '', M, '阵法攻击', '每帧if isActive: 检查范围敌方→有则按interval攻击', 'Formation.update(dt,model)', 'dt+状态 → void', 'V1', '待办'),
    ('L3', 'C5.03', S, '', M, '阵法冷却', '被毁时记录格子坐标，8秒后过期，期间不能在此布', 'Player.checkCooldown(gx,gy)', '坐标 → bool', 'V1', '待办'),
    ('L3', 'C5.04', S, '', M, '阵法被禁', '镇魂符使isActive=false，3秒不攻击，计时后恢复', 'Formation.setSilence(duration)', '秒 → void', 'V1', '待办'),
]
M = 'C6.法术系统'
D += [
    ('L2', 'C6', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'C6.01', S, '', M, '法术入口', 'switch(cardId)分支调用对应cast方法→扣灵力→打出手牌', 'SpellSystem.cast(cardId,caster,target,model)', 'ID+施法方+目标 → void', 'V1', '待办'),
    ('L3', 'C6.02', S, '', M, '万剑归宗', '遍历己方单位→atk+=1→addBuff("speed",0.3,5)', 'SpellSystem.castWanJian(caster)', '施法方 → void', 'V1', '待办'),
    ('L3', 'C6.03', S, '', M, '五雷正法', '目标3格内敌方+阵法各受4伤', 'SpellSystem.castWuLei(target,model)', '目标+状态 → void', 'V1', '待办'),
    ('L3', 'C6.04', S, '', M, '御风诀', 'target.addBuff("speed",0.5,5)移速+0.5持续5s', 'SpellSystem.castYuFeng(target)', '目标 → void', 'V1', '待办'),
    ('L3', 'C6.05', S, '', M, '镇魂符', 'target.setSilence(3)阵法失效3s', 'SpellSystem.castZhenHun(target)', '目标阵法 → void', 'V1', '待办'),
    ('L3', 'C6.06', S, '', M, '金钟罩', 'caster.hallShield=3大殿免疫3s', 'SpellSystem.castJinZhong(caster)', '施法方 → void', 'V1', '待办'),
    ('L3', 'C6.07', S, '', M, '移山倒海', 'targets每个y-=2*facing(推后)→takeDamage(1)', 'SpellSystem.castYiShan(targets)', '目标数组 → void', 'V1', '待办'),
    ('L3', 'C6.08', S, '', M, '困仙索', 'target.addBuff("stun",0,2)定身2s(speed=0)', 'SpellSystem.castKunXian(target)', '目标 → void', 'V1', '待办'),
    ('L3', 'C6.09', S, '', M, '天雷诀', '目标3格内敌方各受4伤(清兵)', 'SpellSystem.castTianLei(targets)', '目标数组 → void', 'V1', '待办'),
]
M = 'C7.长老技能'
D += [
    ('L2', 'C7', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'C7.01', S, '', M, '技能计时', 'elderTimer+=dt→if>=5s: triggerRandom()→重置', 'ElderSkillSystem.update(elder,dt,model)', '长老+dt+状态 → void', 'V1', '待办'),
    ('L3', 'C7.02', S, '', M, '★随机分支★', 'branches=[飞剑,丹药,符箓,御兽]→random选1→调用对应方法', 'ElderSkillSystem.triggerRandom(elder,model)', '长老+状态 → void', 'V1', '待办'),
    ('L3', 'C7.03', S, '', M, '飞剑分支', '长老周围3格内敌方各受3伤(范围)', 'ElderSkillSystem.castFlyingSword(elder,model)', '长老+状态 → void', 'V1', '待办'),
    ('L3', 'C7.04', S, '', M, '丹药分支', '长老hp+3(不超max)→周围己方各hp+2', 'ElderSkillSystem.castPill(elder,model)', '长老+状态 → void', 'V1', '待办'),
    ('L3', 'C7.05', S, '', M, '符箓分支', '长老前方敌方→castTianLei(清兵)', 'ElderSkillSystem.castTalisman(elder,model)', '长老+状态 → void', 'V1', '待办'),
    ('L3', 'C7.06', S, '', M, '御兽分支', '长老位置召唤1个灵兽Unit(同owner)', 'ElderSkillSystem.castBeast(elder,model)', '长老+状态 → void', 'V1', '待办'),
]
M = 'C8.灵力系统'
D += [
    ('L2', 'C8', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'C8.01', S, '', M, '实时回复', 'energyTimer+=dt→if>=2.8s: energy++(不超上限)→重置', 'Player.updateEnergy(dt)', 'dt → void', 'V1', '待办'),
    ('L3', 'C8.02', S, '', M, '上限增长', 'elapsedTime每30s→energyMax++(封顶10)', 'Player.updateEnergyMax(dt)', 'dt → void', 'V1', '待办'),
]
M = 'C9.出牌执行'
D += [
    ('L2', 'C9', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'C9.01', S, '', M, '出兵', '检查灵力→Unit.fromCard→设置出生y→加入units→扣灵力→打出', 'BattleLogic.spawnUnit(p,cardId,x)', '玩家+卡牌+列 → Unit', 'V1', '待办'),
    ('L3', 'C9.02', S, '', M, '施法', '检查灵力→SpellSystem.cast()→扣灵力→打出', 'BattleLogic.castSpell(p,cardId,target)', '玩家+卡牌+目标 → void', 'V1', '待办'),
]
M = 'C10.胜负判定'
D += [
    ('L2', 'C10', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'C10.01', S, '', M, '大殿摧毁', '每帧检查hallHp→if任一<=0: winner=对方→ended', 'BattleChecker.checkHall(model)', '状态 → 0/1/null', 'V1', '待办'),
    ('L3', 'C10.02', S, '', M, '时限检测', 'if time<=0: 血量不等→血量高胜→else进加时', 'BattleChecker.checkTime(model)', '状态 → 0/1/null', 'V1', '待办'),
    ('L3', 'C10.03', S, '', M, '加时结算', '加时60s后→if仍平→winner=灵力多者', 'BattleChecker.checkOvertime(model)', '状态 → 0/1/null', 'V1', '待办'),
]
M = 'C11.死亡清理'
D += [
    ('L2', 'C11', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'C11.01', S, '', M, '清理死亡单位', '遍历units→if dead: 触发死亡效果(自爆)→移除', 'BattleLogic.removeDeadUnits()', 'void → void', 'V1', '待办'),
    ('L3', 'C11.02', S, '', M, '清理被毁阵法', '遍历formations→if hp<=0: 记录冷却→移除', 'BattleLogic.removeDeadFormations()', 'void → void', 'V1', '待办'),
]

# ============ D. AI系统 ============
S, SN = 'D', 'D.AI系统'
M = 'D1.AI决策'
D += [
    ('L1', 'D', S, SN, '—', '—', '—', '—', '—', '—'),
    ('L2', 'D1', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'D1.01', S, '', M, 'AI主循环', 'thinkTimer+=dt→if>=interval: think()→重置。间隔easy3.5/normal2.5/hard1.5', 'AI.update(dt,model)', 'dt+状态 → void', 'V1', '待办'),
    ('L3', 'D1.02', S, '', M, '★决策核心★', '读灵力→decideAttackRatio→pickCard→if有牌: spawn/place/cast', 'AI.think(model)', '状态 → void', 'V1', '待办'),
    ('L3', 'D1.03', S, '', M, '攻守比计算', 'hp%>0.6: return 0.7(偏攻)→<0.3: 0.2(偏守)→else: 0.4', 'AI.decideAttackRatio(model)', '状态 → 0~1', 'V1', '待办'),
    ('L3', 'D1.04', S, '', M, '选牌逻辑', '过滤灵力够的→ratio>0.5优先攻方→else优先阵法→随机选', 'AI.pickCard(energy,ratio)', '灵力+攻守比 → index/null', 'V1', '待办'),
    ('L3', 'D1.05', S, '', M, '布阵位置', 'easy随机→normal大殿前2~3格→hard针对性(高速前布寒霜)', 'AI.pickFormationPos(model)', '状态 → {gx,gy}', 'V1', '待办'),
]

# ============ E. 渲染与UI ============
S, SN = 'E', 'E.渲染与UI'
M = 'E1.战斗场景'
D += [
    ('L1', 'E', S, SN, '—', '—', '—', '—', '—', '—'),
    ('L2', 'E1', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'E1.01', S, '', M, '场景进入', '创建GameModel→创建BattleLogic→注册Input→加载资源', 'BattleScene.onEnter()', 'void → void', 'V1', '待办'),
    ('L3', 'E1.02', S, '', M, '场景更新', '调用battleLogic.update(dt)驱动逻辑', 'BattleScene.onUpdate(dt)', 'dt → void', 'V1', '待办'),
    ('L3', 'E1.03', S, '', M, '场景渲染(分层)', '顺序: 背景→大殿→阵法→单位(y排序)→特效→UI', 'BattleScene.onRender()', 'void → void', 'V1', '待办'),
    ('L3', 'E1.04', S, '', M, '场景退出', '清理Input→销毁资源→停止逻辑', 'BattleScene.onExit()', 'void → void', 'V1', '待办'),
]
M = 'E2.背景渲染'
D += [
    ('L2', 'E2', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'E2.01', S, '', M, '山道背景', 'V1: 纯色渐变(深绿→浅绿)→V1.5: 替换正式山道图', 'BattleScene.renderBackground()', 'void → void', 'V1', '待办'),
    ('L3', 'E2.02', S, '', M, '棋盘格子线', '绘制主路(中间列)+两侧阵法区边框线', 'BattleScene.renderGrid()', 'void → void', 'V1', '待办'),
]
M = 'E3.大殿渲染'
D += [
    ('L2', 'E3', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'E3.01', S, '', M, '大殿色块', '顶部/底部各绘制金色色块+宗门名', 'BattleScene.renderHalls()', 'void → void', 'V1', '待办'),
    ('L3', 'E3.02', S, '', M, '大殿血条', '大殿上方血条(背景红+前景绿按比例)', 'BattleScene.renderHallHp()', 'void → void', 'V1', '待办'),
    ('L3', 'E3.03', S, '', M, '受击特效', '受击时屏幕震动(offset±3px)+红色闪烁0.2s', 'BattleScene.renderHallHit()', 'void → void', 'V1', '待办'),
]
M = 'E4.单位渲染'
D += [
    ('L2', 'E4', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'E4.01', S, '', M, '单位色块', '攻方红/守方蓝色块+名称→V1.5替换Spine', 'BattleScene.renderUnits()', 'void → void', 'V1', '待办'),
    ('L3', 'E4.02', S, '', M, 'y轴深度排序', '按y排序后绘制(y小先画=远在后,y大后画=近在前)', 'BattleScene.renderUnits()', 'void → void', 'V1', '待办'),
    ('L3', 'E4.03', S, '', M, '单位血条', '单位上方小血条(宽=单位宽,高=4px)', 'BattleScene.renderUnitHp()', 'void → void', 'V1', '待办'),
]
M = 'E5.阵法渲染'
D += [
    ('L2', 'E5', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'E5.01', S, '', M, '阵法光阵', '阵法区格子半透明青色光阵+阵法名', 'BattleScene.renderFormations()', 'void → void', 'V1', '待办'),
    ('L3', 'E5.02', S, '', M, '阵法状态', '正常完整→受击闪烁→被禁灰暗→冷却半透明+倒计时', 'BattleScene.renderFormationState()', 'void → void', 'V1', '待办'),
]
M = 'E6.UI-手牌栏'
D += [
    ('L2', 'E6', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'E6.01', S, '', M, '手牌显示', '底部横排4张: 色块+卡名+费用→选中上浮20px+发光', 'HandBar.render()', 'void → void', 'V1', '待办'),
    ('L3', 'E6.02', S, '', M, '灰显不可出', '灵力<费用: 灰度+不可点', 'HandBar.updateGrey(energy)', '灵力 → void', 'V1', '待办'),
    ('L3', 'E6.03', S, '', M, '点击选中', '点击→selectedCard=index→上浮→高亮可放置区', 'HandBar.onTap(index)', 'index → void', 'V1', '待办'),
]
M = 'E7.UI-灵力条'
D += [
    ('L2', 'E7', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'E7.01', S, '', M, '灵力显示', '手牌上方紫色液条+数字(5/7)按比例', 'EnergyBar.render()', 'void → void', 'V1', '待办'),
]
M = 'E8.UI-顶部HUD'
D += [
    ('L2', 'E8', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'E8.01', S, '', M, '双方血条', '顶部左(我方)/右(敌方)大殿血条', 'HUD.renderHallBars()', 'void → void', 'V1', '待办'),
    ('L3', 'E8.02', S, '', M, '计时器', '顶部中央倒计时→最后30秒变红', 'HUD.renderTimer()', 'void → void', 'V1', '待办'),
]
M = 'E9.出牌交互'
D += [
    ('L2', 'E9', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'E9.01', S, '', M, '手牌→选目标', '选中后高亮区域(单位:主路/阵法:阵法区/法术:全屏)→点目标执行', 'BattleScene.onCardTap(index)', 'index → void', 'V1', '待办'),
    ('L3', 'E9.02', S, '', M, '格子点击', '点击格子→if有选中: spawn/place/cast→清除选中', 'BattleScene.onGridTap(gx,gy)', '坐标 → void', 'V1', '待办'),
]
M = 'E10.结算场景'
D += [
    ('L2', 'E10', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'E10.01', S, '', M, '胜负展示', '大字"胜利"/"失败"+缩放动画0.5s', 'ResultScene.renderResult()', 'void → void', 'V1', '待办'),
    ('L3', 'E10.02', S, '', M, '摧毁度', '显示双方摧毁度(我方XX% vs 敌方XX%)', 'ResultScene.renderDestroyRate()', 'void → void', 'V1', '待办'),
    ('L3', 'E10.03', S, '', M, '再来一局', '按钮→点击切回BattleScene重新初始化', 'ResultScene.onReplay()', 'void → void', 'V1', '待办'),
]
M = 'E11.新手引导'
D += [
    ('L2', 'E11', S, '', M, '—', '—', '—', 'V1', ''),
    ('L3', 'E11.01', S, '', M, '引导1-出牌', '高亮第一张手牌+箭头→"点击出兵"→点击后下一步', 'TutorialGuide.step1()', 'void → void', 'V1', '待办'),
    ('L3', 'E11.02', S, '', M, '引导2-灵力', '高亮灵力条→"灵力不够时等待回复"→等3s', 'TutorialGuide.step2()', 'void → void', 'V1', '待办'),
    ('L3', 'E11.03', S, '', M, '引导3-布阵', '高亮阵法区→"点击布阵拦截"→布阵后下一步', 'TutorialGuide.step3()', 'void → void', 'V1', '待办'),
    ('L3', 'E11.04', S, '', M, '引导4-目标', '箭头指向敌方大殿→"摧毁大殿获胜"→2s后结束', 'TutorialGuide.step4()', 'void → void', 'V1', '待办'),
]

# ============ F/G/H (V1.5/V2 简版) ============
S, SN = 'F', 'F.社交系统(V1.5)'
D += [
    ('L1', 'F', S, SN, '—', '—', '—', '—', 'V1.5', '—'),
    ('L3', 'F1.01', S, '', 'F1.登录', '微信登录', 'wx.login→code→云函数换openid', 'CloudManager.login()', 'void → openid', 'V1.5', '待办'),
    ('L3', 'F1.02', S, '', 'F1.存档', '云存档', 'wx.cloud.database存用户档案', 'CloudManager.save(data)', '数据 → void', 'V1.5', '待办'),
    ('L3', 'F2.01', S, '', 'F2.异步PvP', '布阵上传', '布阵完成→上传云数据库→生成挑战链接', 'AsyncPvP.uploadLayout(layout)', '布阵 → 挑战ID', 'V1.5', '待办'),
    ('L3', 'F2.02', S, '', 'F2.异步PvP', '加载对手', '通过挑战ID读对手布阵→本地AI模拟攻打', 'AsyncPvP.loadOpponent(id)', 'ID → 布阵数据', 'V1.5', '待办'),
    ('L3', 'F3.01', S, '', 'F3.排行分享', '好友排行', 'wx.getOpenDataContext→开放数据域排序', 'RankManager.renderFriendRank()', 'void → void', 'V1.5', '待办'),
    ('L3', 'F3.02', S, '', 'F3.排行分享', '分享', 'wx.shareAppMessage→设置title/image', 'ShareManager.share(title,img)', '标题+图 → void', 'V1.5', '待办'),
]
S, SN = 'G', 'G.养成变现(V1.5)'
D += [
    ('L1', 'G', S, SN, '—', '—', '—', '—', 'V1.5', '—'),
    ('L3', 'G1.01', S, '', 'G1.养成', '弟子升级', '消耗灵石→level++→hp/atk按成长系数提升', 'ProgressionSystem.levelUp(unitId)', 'ID → void', 'V1.5', '待办'),
    ('L3', 'G1.02', S, '', 'G1.养成', '弟子升星', '消耗剑魄→突破境界(练气→筑基→金丹→元婴)', 'ProgressionSystem.starUp(unitId)', 'ID → void', 'V1.5', '待办'),
    ('L3', 'G2.01', S, '', 'G2.变现', '激励广告', 'wx.createRewardedVideoAd→show→onClose发奖', 'AdManager.showReward(cb)', '回调 → void', 'V1.5', '待办'),
    ('L3', 'G2.02', S, '', 'G2.变现', '内购', 'wx.requestPayment→下单→支付→发仙玉', 'IAPManager.buy(productId,cb)', '商品+回调 → void', 'V1.5', '待办'),
]
S, SN = 'H', 'H.实时PvP(V2)'
D += [
    ('L1', 'H', S, SN, '—', '—', '—', '—', 'V2', '—'),
    ('L3', 'H1.01', S, '', 'H1.网络', 'WebSocket', 'wx.connectSocket→onOpen/Message/Close→心跳10s', 'NetworkClient.connect(url)', 'URL → void', 'V2', '待办'),
    ('L3', 'H1.02', S, '', 'H1.网络', '断线重连', 'onClose记录状态→自动重连→恢复对局', 'NetworkClient.reconnect()', 'void → void', 'V2', '待办'),
    ('L3', 'H2.01', S, '', 'H2.同步', '操作上报', '出牌/布阵/施法→封装→send到服务端', 'SyncManager.sendAction(action)', '操作 → void', 'V2', '待办'),
    ('L3', 'H2.02', S, '', 'H2.同步', '状态接收', '服务端广播→本地应用→渲染', 'SyncManager.onState(state)', '状态 → void', 'V2', '待办'),
    ('L3', 'H3.01', S, '', 'H3.匹配', '匹配系统', '上报ELO→服务端匹配相近对手→返回房间ID', 'MatchMaker.match(elo)', 'ELO → 房间ID', 'V2', '待办'),
    ('L3', 'H3.02', S, '', 'H3.匹配', '段位赛', '青铜→宗师，胜负加减分', 'RankSystem.updateRank(win)', '是否胜 → void', 'V2', '待办'),
]

# 写入数据
r = 3
for row_data in D:
    row_list = list(row_data)
    while len(row_list) < 11:
        row_list.append('')
    level, num, sys_code, sys_name, mod, func, principle, fn, io, ver, status = row_list
    vals = [level, num, sys_name if level=='L1' else '', mod, func, principle, fn, io, ver, status]

    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.border = BD_THIN
        c.alignment = A_LT

    col = r  # row number

    # 层级标签列
    lc = ws.cell(row=r, column=1)
    lc.font = F_TAG; lc.alignment = A_C
    if level == 'L1':
        bg, fg = L1_FILLS.get(sys_code, ('333333','FFFFFF'))
        lc.fill = PatternFill('solid', fgColor=bg)
        lc.value = 'L1\n系统'
        ws.row_dimensions[r].height = 32
    elif level == 'L2':
        lc.fill = PatternFill('solid', fgColor=L2_FILL)
        lc.value = 'L2\n模块'
        ws.row_dimensions[r].height = 24
    elif level == 'L3':
        lc.fill = PatternFill('solid', fgColor=L3_FILL)
        lc.value = 'L3\n功能'
        ws.row_dimensions[r].height = 40

    # 编号列
    ws.cell(row=r, column=2).font = F_CODE
    ws.cell(row=r, column=2).alignment = A_C

    # L1系统列（只有L1行显示）
    if level == 'L1':
        bg, fg = L1_FILLS.get(sys_code, ('333333','FFFFFF'))
        c = ws.cell(row=r, column=3)
        c.fill = PatternFill('solid', fgColor=bg)
        c.font = F_L1; c.alignment = A_C

    # L2模块列
    if level == 'L2':
        c = ws.cell(row=r, column=4)
        c.fill = PatternFill('solid', fgColor=L2_FILL)
        c.font = F_L2; c.alignment = A_L

    # L3子功能列
    if level == 'L3':
        c = ws.cell(row=r, column=5)
        c.fill = PatternFill('solid', fgColor=L3_FILL)
        c.font = F_CELL
        # 关键功能标星加粗
        if '★' in func:
            c.font = Font(name='微软雅黑', size=10, bold=True, color=C_RED)

    # L4原理列
    if level == 'L3':
        c = ws.cell(row=r, column=6)
        c.fill = PatternFill('solid', fgColor=L4_FILL)
        c.font = F_CELL

    # L5函数列
    if level == 'L3':
        c = ws.cell(row=r, column=7)
        c.fill = PatternFill('solid', fgColor=L5_FILL)
        c.font = F_CODE

    # 输入输出列
    if level == 'L3':
        ws.cell(row=r, column=8).font = F_CODE

    # 版本列
    vc = ws.cell(row=r, column=9)
    vc.font = F_CELL; vc.alignment = A_C
    if level == 'L3':
        v = ver
        if v == 'V1': vc.fill = PatternFill('solid', fgColor='D6EAF8')
        elif v == 'V1.5': vc.fill = PatternFill('solid', fgColor='FADBD8')
        elif v == 'V2': vc.fill = PatternFill('solid', fgColor='D5D8DC')

    # 状态列
    sc = ws.cell(row=r, column=10)
    sc.font = Font(name='微软雅黑', size=10, bold=True)
    sc.alignment = A_C
    if level == 'L3':
        if status == '待办': sc.fill = PatternFill('solid', fgColor=ST_TODO)
        elif status == '进行中': sc.fill = PatternFill('solid', fgColor=ST_DOING)
        elif status == '已完成': sc.fill = PatternFill('solid', fgColor=ST_DONE)
    else:
        sc.value = ''

    r += 1

# 条件格式
ws.conditional_formatting.add(f'J3:J{r-1}',
    FormulaRule(formula=['$J3="已完成"'], fill=PatternFill('solid', fgColor=ST_DONE)))
ws.conditional_formatting.add(f'J3:J{r-1}',
    FormulaRule(formula=['$J3="进行中"'], fill=PatternFill('solid', fgColor=ST_DOING)))
ws.conditional_formatting.add(f'J3:J{r-1}',
    FormulaRule(formula=['$J3="待办"'], fill=PatternFill('solid', fgColor=ST_TODO)))

# 底部说明
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
c = ws.cell(row=r, column=1, value='层级说明：L1系统大类(色块) → L2模块(浅蓝) → L3子功能(浅橙) → L4实现原理(浅紫) → L5函数(浅青) | 做完一个把"状态"改为"已完成"')
c.font = F_SMALL; c.alignment = A_L

# ================================================================
# Sheet 3: 进度统计（卡片式）
# ================================================================
ws = wb.create_sheet('3.进度统计')
ws.sheet_view.showGridLines = False
set_w(ws, [4, 22, 10, 10, 10, 10, 14, 4])

big_title(ws, '进度统计', 8, 1, C_INK)

# 统计
from collections import OrderedDict
stats = OrderedDict()
for row_data in D:
    if row_data[0] != 'L3': continue
    sys_code = row_data[2]
    sys_name = row_data[3]
    if sys_code not in stats:
        stats[sys_code] = {'name': sys_name, 'total': 0, 'done': 0, 'doing': 0, 'todo': 0}
    stats[sys_code]['total'] += 1
    s = row_data[10]
    if s == '已完成': stats[sys_code]['done'] += 1
    elif s == '进行中': stats[sys_code]['doing'] += 1
    else: stats[sys_code]['todo'] += 1

header_row(ws, ['系统大类', '总点数', '已完成', '进行中', '待办', '完成率', '进度条'], 3, C_INK)
# 调整列
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=1)
ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=7)

r = 4
for sys_code, s in stats.items():
    bg, fg = L1_FILLS.get(sys_code, ('333333','FFFFFF'))
    rate = s['done'] / s['total'] if s['total'] else 0
    rate_str = f"{s['done']}/{s['total']} ({int(rate*100)}%)"
    bar_len = int(rate * 20)
    bar = '█' * bar_len + '░' * (20 - bar_len)

    vals = ['', s['name'], s['total'], s['done'], s['doing'], s['todo'], rate_str, bar]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.border = BD_THIN
        c.alignment = A_C if i > 2 else A_L
        c.font = F_CELL

    # 系统名色块
    c = ws.cell(row=r, column=2)
    c.fill = PatternFill('solid', fgColor=bg)
    c.font = F_L1; c.alignment = A_C

    # 进度条
    c = ws.cell(row=r, column=8)
    c.font = Font(name='Consolas', size=11, color=C_GREEN if rate > 0.5 else (C_GOLD if rate > 0.2 else C_RED))

    ws.row_dimensions[r].height = 30
    r += 1

# 总计
t = sum(s['total'] for s in stats.values())
d = sum(s['done'] for s in stats.values())
di = sum(s['doing'] for s in stats.values())
td = sum(s['todo'] for s in stats.values())
rate = d / t if t else 0
rate_str = f"{d}/{t} ({int(rate*100)}%)"
bar_len = int(rate * 20)
bar = '█' * bar_len + '░' * (20 - bar_len)

vals = ['', '总计', t, d, di, td, rate_str, bar]
for i, v in enumerate(vals, 1):
    c = ws.cell(row=r, column=i, value=v)
    c.font = F_HEADER; c.fill = PatternFill('solid', fgColor=C_INK)
    c.alignment = A_C if i > 2 else A_L
    c.border = BD_THIN
ws.row_dimensions[r].height = 34

# 保存
wb.save(PATH)
print(f'Excel已重新生成: {PATH}')
print(f'Sheet: {wb.sheetnames}')
print(f'开发主表: {len([d for d in D if d[0]=="L3"])}个L3功能点')
print(f'系统大类: {len([d for d in D if d[0]=="L1"])}个')
