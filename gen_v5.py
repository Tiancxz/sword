#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""《宗门论道》开发管理工具包 v5 — 全部8项改进"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

PATH = '/workspace/宗门论道_开发管理工具包.xlsx'
wb = openpyxl.Workbook()

C_INK='2C3E50'; C_GOLD='B7950B'; C_BLUE='2874A6'; C_TEAL='117A65'
C_RED='C0392B'; C_PURPLE='8E44AD'; C_GREEN='27AE60'; C_ORANGE='D35400'

L1_FILLS={'A':('1A5276','FFFFFF'),'B':('2874A6','FFFFFF'),'C':('117A65','FFFFFF'),'D':('B7950B','FFFFFF'),'E':('8E44AD','FFFFFF'),'F':('C0392B','FFFFFF'),'G':('D35400','FFFFFF'),'H':('34495E','FFFFFF')}
L2_FILL='D6EAF8'; L3_FILL='FDF2E9'; L4_FILL='F4ECF7'; L5_FILL='E8F8F5'
ST_TODO='FADBD8'; ST_DOING='FEF9E7'; ST_DONE='D5F5E3'
PHASE_COLORS={'P1':'1A5276','P2':'2874A6','P3':'117A65','P4':'B7950B','P5':'8E44AD','P6':'C0392B','P7':'D35400','P8':'27AE60','P9':'2E86C1','P10':'85C1E9','P11':'F8C471','P12':'F0B27A','P13':'BB8FCE','P14':'F1948A','P16':'5D6D7E','P17':'85929E','P18':'AAB7B8'}

F_BIG=Font(name='微软雅黑',size=16,bold=True,color='FFFFFF')
F_TITLE=Font(name='微软雅黑',size=13,bold=True,color='FFFFFF')
F_HEADER=Font(name='微软雅黑',size=11,bold=True,color='FFFFFF')
F_L1=Font(name='微软雅黑',size=11,bold=True,color='FFFFFF')
F_L2=Font(name='微软雅黑',size=11,bold=True,color=C_INK)
F_CELL=Font(name='微软雅黑',size=10,color='2C3E50')
F_BOLD=Font(name='微软雅黑',size=10,bold=True,color='2C3E50')
F_CODE=Font(name='Consolas',size=10,color='1A5276')
F_SMALL=Font(name='微软雅黑',size=9,color='7F8C8D')
F_TAG=Font(name='微软雅黑',size=8,bold=True,color='FFFFFF')
F_GOLD=Font(name='微软雅黑',size=11,bold=True,color=C_GOLD)
F_RESULT=Font(name='微软雅黑',size=10,bold=True,color=C_TEAL)
F_PHASE=Font(name='微软雅黑',size=12,bold=True,color='FFFFFF')

A_C=Alignment(horizontal='center',vertical='center',wrap_text=True)
A_L=Alignment(horizontal='left',vertical='center',wrap_text=True,indent=1)
A_LT=Alignment(horizontal='left',vertical='top',wrap_text=True,indent=1)
A_LM=Alignment(horizontal='left',vertical='center',wrap_text=True)
BD=Border(left=Side(style='thin',color='D5D8DC'),right=Side(style='thin',color='D5D8DC'),top=Side(style='thin',color='D5D8DC'),bottom=Side(style='thin',color='D5D8DC'))

def set_w(ws,widths):
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
def big_title(ws,text,cols,row=1,color=C_INK):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=cols)
    c=ws.cell(row=row,column=1,value=text); c.font=F_BIG; c.fill=PatternFill('solid',fgColor=color); c.alignment=A_C
    ws.row_dimensions[row].height=44
def sub_title(ws,text,cols,row,color=C_GOLD):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=cols)
    c=ws.cell(row=row,column=1,value=text); c.font=F_TITLE; c.fill=PatternFill('solid',fgColor=color); c.alignment=A_L
    ws.row_dimensions[row].height=30
def header_row(ws,headers,row,color=C_INK):
    for i,h in enumerate(headers,1):
        c=ws.cell(row=row,column=i,value=h); c.font=F_HEADER; c.fill=PatternFill('solid',fgColor=color); c.alignment=A_C; c.border=BD
    ws.row_dimensions[row].height=30
def calc_row_height(text,col_w=40):
    if not text: return 24
    text=str(text); wrapped=sum(max(1,(len(line)+col_w-1)//col_w) for line in text.split('\n'))
    return max(24,min(120,wrapped*16+8))

# ===== 主表数据 =====
D=[]
# A.引擎框架
SC,SN='A','A.引擎框架'
for mod,items in [
    ('A1.渲染循环',[
        ('L3','A1.01','P1','主循环驱动','requestAnimationFrame((ts)=>{ dt=(ts-lastTime)/1000; dt=Math.min(dt,0.033); this.update(dt); this.render(); this.lastTime=ts; })。lastTime初始=0，首帧dt钳制。running=false停止。','Director.loop(ts)','js/core/Director.js','ts:number → void','—',0.5),
        ('L3','A1.02','P1','逻辑更新','let scene=this.sceneManager.current; if(scene) scene.onUpdate(dt)。scene为null跳过。','Director.update(dt)','js/core/Director.js','dt:number → void','A1.01',0.3),
        ('L3','A1.03','P1','画面渲染','let ctx=this.canvas.getContext("2d"); ctx.clearRect(0,0,W,H); let scene=...; if(scene) scene.onRender(ctx)。','Director.render()','js/core/Director.js','void → void','A1.01',0.3),
        ('L3','A1.04','P1','引擎初始化','this.canvas=wx.createCanvas(); this.ctx=canvas.getContext("2d"); wx.getSystemInfoSync()获取屏幕尺寸; this.sceneManager=new SceneManager(); this.input=new Input(); this.input.init(); this.start()。','Director.init()','js/core/Director.js','void → void','A1.01,A2.01,A3.01,A4.01',0.5),
    ]),
    ('A2.输入系统',[
        ('L3','A2.01','P1','触摸监听','wx.onTouchStart((e)=>this.onTouchStart(e.touches[0].clientX,e.touches[0].clientY)); 同理onTouchEnd。','Input.init()','js/core/Input.js','void → void','—',0.3),
        ('L3','A2.02','P1','点击区域注册','this.regions.push({id,x,y,w,h,cb})。场景onEnter注册，onExit时clear。','Input.register(id,x,y,w,h,cb)','js/core/Input.js','id+坐标+回调 → void','A2.01',0.3),
        ('L3','A2.03','P1','命中检测','for(let r of this.regions){ if(tx>=r.x&&tx<=r.x+r.w&&ty>=r.y&&ty<=r.y+r.h){ r.cb(tx,ty); break } }','Input.onTouchStart(tx,ty)','js/core/Input.js','坐标 → 触发cb','A2.02',0.3),
        ('L3','A2.04','P1','区域清理','this.regions=[]。场景切换onExit必须调用。','Input.clear()','js/core/Input.js','void → void','A2.02',0.2),
    ]),
    ('A3.事件总线',[
        ('L3','A3.01','P1','事件注册','if(!this.listeners.has(event)) this.listeners.set(event,[]); this.listeners.get(event).push(cb)。','EventSystem.on(event,cb)','js/core/EventSystem.js','事件名+回调 → void','—',0.3),
        ('L3','A3.02','P1','事件触发','let cbs=this.listeners.get(event); if(cbs) for(let cb of cbs) cb(data)。','EventSystem.emit(event,data)','js/core/EventSystem.js','事件名+数据 → void','A3.01',0.2),
        ('L3','A3.03','P1','事件移除','let cbs=this.listeners.get(event); if(cbs){ let i=cbs.indexOf(cb); if(i>=0) cbs.splice(i,1) }。','EventSystem.off(event,cb)','js/core/EventSystem.js','事件名+回调 → void','A3.01',0.2),
    ]),
    ('A4.场景管理',[
        ('L3','A4.01','P1','场景注册','this.scenes.set(name,scene)。','SceneManager.register(name,scene)','js/core/SceneManager.js','名称+场景 → void','—',0.2),
        ('L3','A4.02','P1','场景切换','if(this.current) this.current.onExit(); this.current=this.scenes.get(name); if(this.current) this.current.onEnter()。','SceneManager.switch(name)','js/core/SceneManager.js','场景名 → void','A4.01',0.3),
    ]),
    ('A5.渲染器',[
        ('L3','A5.01','P1','矩形绘制','ctx.fillStyle=color; ctx.fillRect(x,y,w,h)。','Renderer.rect(ctx,x,y,w,h,color)','js/core/Renderer.js','ctx+坐标+颜色 → void','—',0.2),
        ('L3','A5.02','P1','文字绘制','ctx.fillStyle=color; ctx.font=size+"px 微软雅黑"; ctx.fillText(text,x,y)。','Renderer.text(ctx,text,x,y,color,size)','js/core/Renderer.js','ctx+文本+坐标 → void','—',0.2),
        ('L3','A5.03','P1','血条绘制','ctx.fillStyle="#E74C3C"; ctx.fillRect(x,y,w,h); ctx.fillStyle="#27AE60"; ctx.fillRect(x,y,w*ratio,h)。','Renderer.healthBar(ctx,x,y,w,h,ratio)','js/core/Renderer.js','ctx+坐标+比例 → void','A5.01',0.3),
    ]),
    ('A6.粒子系统',[
        ('L3','A6.01','P6','粒子发射','for(let i=0;i<count;i++){ angle=random*PI*2; speed=20+random*40; push({x,y,vx,vy,life,maxLife,color,size}) }','ParticleSystem.emit(x,y,count,color)','js/core/ParticleSystem.js','坐标+数量+颜色 → void','—',0.5),
        ('L3','A6.02','P6','粒子更新','for(p of particles){ p.x+=p.vx*dt; p.y+=p.vy*dt; p.life-=dt } 过滤life<=0。','ParticleSystem.update(dt)','js/core/ParticleSystem.js','dt → void','A6.01',0.3),
        ('L3','A6.03','P6','粒子渲染','for(p of particles){ ctx.globalAlpha=p.life/p.maxLife; draw } ctx.globalAlpha=1。','ParticleSystem.render(ctx)','js/core/ParticleSystem.js','ctx → void','A6.01',0.3),
    ]),
    ('A7.音频管理',[
        ('L3','A7.01','P6','BGM播放','this.bgm=wx.getBackgroundAudioManager(); this.bgm.src=url; this.bgm.loop=true; this.bgm.play()。','AudioManager.playBGM(url)','js/core/AudioManager.js','URL → void','—',0.3),
        ('L3','A7.02','P6','SFX播放','let audio=wx.createInnerAudioContext(); audio.src=url; audio.play(); audio.onEnded(()=>audio.destroy())。','AudioManager.playSFX(url)','js/core/AudioManager.js','URL → void','—',0.3),
    ]),
]:
    D.append(('L1','','',SC,SN,'','','','','','','','','',''))
    D.append(('L2',mod.split('.')[0],'P1' if 'A6' not in mod and 'A7' not in mod else 'P6',SC,'',mod,'','','','','','','','V1',''))
    for it in items:
        D.append((it[0],it[1],it[2],SC,'',mod,it[3],it[4],it[5],it[6],it[7],it[8],'V1','待办'))

# B.游戏数据
SC,SN='B','B.游戏数据'
for mod,items in [
    ('B1.常量配置',[('L3','B1.01','P2','全局常量','module.exports={BOARD_LENGTH:9,HALL_HP:30,ENERGY_START:5,ENERGY_REGEN:2.8,...}。','Constants.js','js/config/Constants.js','文件 → 引用','—',0.5)]),
    ('B2.卡牌数据表',[
        ('L3','B2.01','P2','卡牌配置','module.exports={"body_disciple":{name,type,faction,cost,hp,atk,...},...共19张}','Cards.js','js/config/Cards.js','文件 → 引用','B1.01',1.0),
        ('L3','B2.02','P2','卡牌查表','let card=this.data[cardId]; if(!card) console.error(); return card。','Cards.get(cardId)','js/config/Cards.js','cardId → Card','B2.01',0.3),
    ]),
    ('B3.卡组手牌',[
        ('L3','B3.01','P2','卡组初始化','this.drawPile=[...cardIds]; this.shuffle(); this.hand=[]; for(i=0;i<3;i++) this.draw()。','Deck.init(cardIds)','js/game/Deck.js','cardIds[] → void','B2.01',0.5),
        ('L3','B3.02','P2','洗牌算法','for(i=drawPile.length-1;i>0;i--){ j=floor(random*(i+1)); swap }','Deck.shuffle()','js/game/Deck.js','void → void','B3.01',0.3),
        ('L3','B3.03','P2','抽牌','if(hand.length>=4) return null; if(drawPile.length==0) return null; let c=drawPile.pop(); hand.push(c); return c','Deck.draw()','js/game/Deck.js','void → cardId|null','B3.01',0.3),
        ('L3','B3.04','P2','打出后补牌','if(drawTimer>0){ drawTimer-=dt; if(drawTimer<=0){ draw(); drawTimer=0 } }','Deck.update(dt)','js/game/Deck.js','dt → void','B3.03',0.3),
        ('L3','B3.05','P2','出牌检查','let card=Cards.get(hand[index]); return energy>=card.cost。','Deck.canPlay(index,energy)','js/game/Deck.js','index+energy → bool','B2.02',0.2),
        ('L3','B3.06','P2','打出牌','let cardId=hand.splice(index,1)[0]; drawTimer=2.0; return cardId。','Deck.playCard(index)','js/game/Deck.js','index → cardId','B3.05',0.2),
    ]),
    ('B4.预设卡组',[('L3','B4.01','P2','3套预设卡组','module.exports={rush:[...8张],control:[...],tank:[...]}','DeckPresets.js','js/config/DeckPresets.js','文件 → Deck.init','B2.01',0.5)]),
]:
    D.append(('L1','','',SC,SN,'','','','','','','','','',''))
    D.append(('L2',mod.split('.')[0],'P2',SC,'',mod,'','','','','','','','V1',''))
    for it in items:
        D.append((it[0],it[1],it[2],SC,'',mod,it[3],it[4],it[5],it[6],it[7],it[8],'V1','待办'))

# C.战斗逻辑
SC,SN='C','C.战斗逻辑'
for mod,items in [
    ('C1.战斗主循环',[('L3','C1.01','P3','每帧总入口','updateEnergy(dt); updateDraw(dt); updateUnits(dt); updateFormations(dt); updateElders(dt); updateAI(dt); removeDead(); checkEnd(); model.time-=dt; model.elapsedTime+=dt','BattleLogic.update(dt)','js/game/BattleLogic.js','dt:number → void','C3.01,C4.01,C5.02,C7.01,C8.01',1.0)]),
    ('C2.单位实体',[
        ('L3','C2.01','P3','从卡牌创建单位','let c=Cards.get(cardId); return {id,cardId,owner,x,y,hp:c.hp,maxHp:c.hp,atk:c.atk,speed:c.speed,baseSpeed:c.speed,attackRange:c.range||1,attackInterval:c.interval,lastAttackTime:0,target:null,state:"walking",traits:c.traits||[],buffs:[],facing:owner===0?1:-1,isElder:c.type==="elder",elderTimer:0}','Unit.fromCard(cardId,owner,x,y)','js/game/Unit.js','cardId+owner+x+y → Unit','B2.02',0.5),
        ('L3','C2.02','P3','buff系统','this.buffs.push({type,value,duration,elapsed:0})。类型:slow/speed/stun/shield/atkBoost。','Unit.addBuff(type,value,duration)','js/game/Unit.js','type+value+duration → void','C2.01',0.5),
        ('L3','C2.03','P3','实际移速计算','let spd=baseSpeed; for(b of buffs){ if(slow) spd-=0.5; if(speed) spd+=0.5; if(stun) return 0 } return max(0,spd)','Unit.getEffectiveSpeed()','js/game/Unit.js','void → number','C2.02',0.3),
        ('L3','C2.04','P3','受到伤害','if(hp<=0) return 0; hp-=amount; if(hp<=0){ hp=0; state="dead" } return amount','Unit.takeDamage(amount,attacker)','js/game/Unit.js','amount+attacker → int','C2.01',0.3),
    ]),
    ('C3.移动系统',[
        ('L3','C3.01','P3','★单位移动★','if(state!="walking") return; target=findTarget(unit,model); if(target&&checkCollision){ state="fighting"; target=target } else { y+=getEffectiveSpeed()*dt*facing; if(checkHallReach){ state="dead"; damageHall() } }','MovementSystem.moveUnit(unit,dt,model)','js/game/MovementSystem.js','unit+dt+model → void','C2.03,C3.02,C3.04',1.0),
        ('L3','C3.02','P3','★目标选择★','enemies=getEnemyUnits(); formations=getEnemyFormations(); candidates=[...enemies,...formations].filter(e=>e.x===unit.x); if(empty) return null; sort by |y-unit.y|; if(nearest dist<=attackRange) return nearest; return null','MovementSystem.findTarget(unit,model)','js/game/MovementSystem.js','unit+model → Unit|Formation|null','C2.01',0.8),
        ('L3','C3.03','P3','碰撞检测','let dist=Math.abs(target.y-unit.y); return dist<=unit.attackRange;','MovementSystem.checkCollision(unit,target)','js/game/MovementSystem.js','unit+target → bool','C3.02',0.2),
        ('L3','C3.04','P3','到达大殿检测','if(owner===0 && y>=BOARD_LENGTH-1) return true; if(owner===1 && y<=0) return true; return false','MovementSystem.checkHallReach(unit,model)','js/game/MovementSystem.js','unit+model → bool','B1.01',0.2),
    ]),
    ('C4.战斗系统',[
        ('L3','C4.01','P3','攻击主逻辑','if(state!="fighting"||!target) return; if(now-lastAttackTime>=interval){ if(ranged) rangedAttack else meleeAttack; lastAttackTime=now; if(target.hp<=0) handleKill() }','CombatSystem.attack(unit,dt,model)','js/game/CombatSystem.js','unit+dt+model → void','C2.04,C4.02,C4.05',1.0),
        ('L3','C4.02','P3','近战互扣','target.takeDamage(attacker.atk,attacker); attacker.takeDamage(target.atk,target); //双方同时掉血','CombatSystem.meleeAttack(attacker,target)','js/game/CombatSystem.js','attacker+target → void','C2.04',0.3),
        ('L3','C4.03','P3','远程单方','target.takeDamage(attacker.atk); //攻击者不掉血','CombatSystem.rangedAttack(attacker,target)','js/game/CombatSystem.js','attacker+target → void','C2.04',0.2),
        ('L3','C4.04','P4','范围伤害','for(t of targets) t.takeDamage(damage)','CombatSystem.aoeAttack(source,targets,damage)','js/game/CombatSystem.js','source+targets[]+damage → void','C2.04',0.3),
        ('L3','C4.05','P3','★击杀后继续推进★','killer.state="walking"; killer.target=null; //不消失继续走！核心修复','CombatSystem.handleKill(killer,victim)','js/game/CombatSystem.js','killer+victim → void','C4.01',0.3),
        ('L3','C4.06','P3','大殿受伤','if(player.hallShield>0) return; player.hallHp-=amount; if(hallHp<=0){ hallHp=0; model.state="ended"; model.winner=1-player.id }','CombatSystem.damageHall(player,amount,model)','js/game/CombatSystem.js','player+amount+model → void','C3.04',0.3),
    ]),
    ('C5.阵法系统',[
        ('L3','C5.01','P4','布阵','let key=gx+","+gy; if(cooldowns[key]&&elapsedTime<cooldowns[key]) return false; let f=Formation.fromCard(); formations.push(f); spendEnergy(); playCard(); return true','BattleLogic.placeFormation(player,cardId,gx,gy,model)','js/game/BattleLogic.js','player+cardId+gx+gy+model → bool','B3.06,C5.03',0.8),
        ('L3','C5.02','P4','阵法攻击','if(!f.isActive) return; enemies=getEnemyUnits(); for(e of enemies){ if(e.x===f.gridX&&|e.y-f.gridY|<=range){ if(now-lastAttack>=interval){ rangedAttack(f,e); lastAttack=now; break } } }','Formation.update(f,dt,model)','js/game/Formation.js','f+dt+model → void','C4.03',0.5),
        ('L3','C5.03','P4','阵法冷却','let key=gx+","+gy; return !cooldowns[key]||elapsedTime>=cooldowns[key]; //true=可布阵','Player.checkCooldown(gx,gy,model)','js/game/Player.js','gx+gy+model → bool','—',0.3),
        ('L3','C5.04','P4','阵法被禁','f.isActive=false; f.silenceTimer=duration; //update中倒数到0恢复','Formation.setSilence(duration)','js/game/Formation.js','duration:number → void','C5.02',0.2),
    ]),
    ('C6.法术系统',[
        ('L3','C6.01','P4','法术入口','switch(cardId){ case "wan_jian": castWanJian(); break; case "wu_lei": castWuLei(); break; ... }','SpellSystem.cast(cardId,caster,target,model)','js/game/SpellSystem.js','cardId+caster+target+model → void','—',0.5),
        ('L3','C6.02','P4','万剑归宗','let units=players[caster].units; for(u of units){ u.atk+=1; u.addBuff("speed",0.3,5) }','SpellSystem.castWanJian(casterId,model)','js/game/SpellSystem.js','casterId+model → void','C2.02',0.3),
        ('L3','C6.03','P4','五雷正法','enemies=getEnemyUnits(caster).concat(getEnemyFormations(caster)); inRange=enemies.filter(e=>|e.y-target.y|<=3&&|e.x-target.x|<=1); for(e of inRange) e.takeDamage(4)','SpellSystem.castWuLei(target,casterId,model)','js/game/SpellSystem.js','target+casterId+model → void','C2.04,C4.04',0.5),
        ('L3','C6.04','P4','御风诀','target.addBuff("speed",0.5,5)','SpellSystem.castYuFeng(target)','js/game/SpellSystem.js','target:Unit → void','C2.02',0.2),
        ('L3','C6.05','P4','镇魂符','target.setSilence(3)','SpellSystem.castZhenHun(target)','js/game/SpellSystem.js','target:Formation → void','C5.04',0.2),
        ('L3','C6.06','P4','金钟罩','players[casterId].hallShield=3','SpellSystem.castJinZhong(casterId,model)','js/game/SpellSystem.js','casterId+model → void','—',0.2),
        ('L3','C6.07','P4','移山倒海','for(t of targets){ t.y-=2*t.facing; t.takeDamage(1) }','SpellSystem.castYiShan(targets)','js/game/SpellSystem.js','targets:Unit[] → void','C2.04',0.3),
        ('L3','C6.08','P4','困仙索','target.addBuff("stun",0,2)','SpellSystem.castKunXian(target)','js/game/SpellSystem.js','target:Unit → void','C2.02',0.2),
        ('L3','C6.09','P4','天雷诀','for(t of targets) t.takeDamage(4)','SpellSystem.castTianLei(targets)','js/game/SpellSystem.js','targets:Unit[] → void','C2.04',0.2),
    ]),
    ('C7.长老技能',[
        ('L3','C7.01','P4','技能计时','if(!unit.isElder) return; unit.elderTimer+=dt; if(elderTimer>=5.0){ elderTimer=0; triggerRandom() }','ElderSkillSystem.update(unit,dt,model)','js/game/ElderSkillSystem.js','unit+dt+model → void','C2.01',0.3),
        ('L3','C7.02','P4','★随机分支★','let branches=["flyingSword","pill","talisman","beast"]; let idx=floor(random*4); this[branches[idx]](elder,model);','ElderSkillSystem.triggerRandom(elder,model)','js/game/ElderSkillSystem.js','elder+model → void','C7.01',0.3),
        ('L3','C7.03','P4','飞剑分支','enemies=getEnemyUnits(elder.owner); nearby=enemies.filter(|y-elder.y|<=3); aoeAttack(elder,nearby,3)','ElderSkillSystem.flyingSword(elder,model)','js/game/ElderSkillSystem.js','elder+model → void','C4.04',0.3),
        ('L3','C7.04','P4','丹药分支','elder.hp=min(maxHp,hp+3); allies=units; for(a of allies){ if(a!==elder&&|a.y-elder.y|<=2) a.hp=min(maxHp,a.hp+2) }','ElderSkillSystem.pill(elder,model)','js/game/ElderSkillSystem.js','elder+model → void','—',0.3),
        ('L3','C7.05','P4','符箓分支','enemies=getEnemyUnits(); front=enemies.filter(|y-elder.y|<=3); castTianLei(front)','ElderSkillSystem.talisman(elder,model)','js/game/ElderSkillSystem.js','elder+model → void','C6.09',0.3),
        ('L3','C7.06','P4','御兽分支','let beast=Unit.fromCard("spirit_beast",elder.owner,elder.x,elder.y); units.push(beast)','ElderSkillSystem.beast(elder,model)','js/game/ElderSkillSystem.js','elder+model → void','C2.01',0.3),
    ]),
    ('C8.灵力系统',[
        ('L3','C8.01','P4','实时回复','regenRate=overtime?REGEN/1.5:REGEN; energyTimer+=dt; if(energyTimer>=regenRate){ energyTimer-=regenRate; energy=min(energyMax,energy+1) }','Player.updateEnergy(dt,model)','js/game/Player.js','dt+model → void','B1.01',0.3),
        ('L3','C8.02','P4','上限增长','let newMax=min(MAX_CAP, START+floor(elapsedTime/30)); energyMax=newMax','Player.updateEnergyMax(model)','js/game/Player.js','model → void','B1.01',0.2),
    ]),
    ('C9.出牌执行',[
        ('L3','C9.01','P4','出兵','let card=Cards.get(cardId); if(!spendEnergy(cost)) return null; let unit=Unit.fromCard(cardId,player.id,x,spawnY); units.push(unit); playCard(handIndex); return unit','BattleLogic.spawnUnit(player,cardId,x,model)','js/game/BattleLogic.js','player+cardId+x+model → Unit','B3.06,C2.01',0.5),
        ('L3','C9.02','P4','施法','let card=Cards.get(cardId); if(!spendEnergy(cost)) return; SpellSystem.cast(cardId,player.id,target,model); playCard(handIndex)','BattleLogic.castSpell(player,cardId,target,model)','js/game/BattleLogic.js','player+cardId+target+model → void','B3.06,C6.01',0.5),
    ]),
    ('C10.胜负判定',[
        ('L3','C10.01','P4','大殿摧毁','if(players[0].hallHp<=0) return 1; if(players[1].hallHp<=0) return 0; return null','BattleChecker.checkHall(model)','js/game/BattleChecker.js','model → 0|1|null','C4.06',0.2),
        ('L3','C10.02','P4','时限检测','if(time>0) return null; if(p0.hp!=p1.hp) return p0>p1?0:1; state="overtime"; time=60; return null','BattleChecker.checkTime(model)','js/game/BattleChecker.js','model → 0|1|null','C8.01',0.3),
        ('L3','C10.03','P4','加时结算','if(time>0) return null; return p0.energy>p1.energy?0:1','BattleChecker.checkOvertime(model)','js/game/BattleChecker.js','model → 0|1','C10.02',0.2),
    ]),
    ('C11.死亡清理',[
        ('L3','C11.01','P4','清理死亡单位','for(p of players){ p.units=p.units.filter(u=>{ if(u.state==="dead"){ if(kamikaze) CombatSystem.kamikaze(u,model); return false } return true }) }','BattleLogic.removeDead(model)','js/game/BattleLogic.js','model → void','C4.05',0.3),
        ('L3','C11.02','P4','清理被毁阵法','for(p of players){ p.formations=p.formations.filter(f=>{ if(f.hp<=0){ cooldowns[key]=elapsedTime+8; return false } return true }) }','BattleLogic.removeDeadFormations(model)','js/game/BattleLogic.js','model → void','C5.03',0.3),
    ]),
]:
    D.append(('L1','','',SC,SN,'','','','','','','','','',''))
    ph = 'P3' if mod.startswith('C1') or mod.startswith('C2') or mod.startswith('C3') or mod.startswith('C4') else 'P4'
    D.append(('L2',mod.split('.')[0],ph,SC,'',mod,'','','','','','','','V1',''))
    for it in items:
        D.append((it[0],it[1],it[2],SC,'',mod,it[3],it[4],it[5],it[6],it[7],it[8],'V1','待办'))

# D.AI系统
SC,SN='D','D.AI系统'
D.append(('L1','','',SC,SN,'','','','','','','','','',''))
D.append(('L2','D1','P5',SC,'','D1.AI决策','','','','','','','','V1',''))
for it in [
    ('L3','D1.01','P5','AI主循环','this.thinkTimer+=dt; if(thinkTimer>=thinkInterval){ thinkTimer=0; think(model) }','AI.update(dt,model)','js/game/AI.js','dt+model → void','C1.01',0.3),
    ('L3','D1.02','P5','★决策核心★','let player=players[playerId]; let ratio=decideAttackRatio(); let cardIdx=pickCard(energy,ratio,hand); if(null) return; let card=Cards.get(cardId); if(unit) spawnUnit(); else if(formation) placeFormation(); else castSpell()','AI.think(model)','js/game/AI.js','model → void','D1.03,D1.04,D1.05,C9.01,C9.02,C5.01',1.5),
    ('L3','D1.03','P5','攻守比计算','let pct=hallHp/HALL_HP; if(pct>0.6) return 0.7; if(pct<0.3) return 0.2; return 0.4','AI.decideAttackRatio(model)','js/game/AI.js','model → number','B1.01',0.3),
    ('L3','D1.04','P5','选牌逻辑','let playable=hand.map(...).filter(cost<=energy); if(empty) return null; if(ratio>0.5) sort by unit优先; else sort by formation优先; return [0].idx','AI.pickCard(energy,ratio,hand)','js/game/AI.js','energy+ratio+hand → int|null','B2.02',0.5),
    ('L3','D1.05','P5','布阵位置','easy: random pos; normal: 靠近己方大殿; hard: 找最快敌方单位前方布','AI.pickFormationPos(model)','js/game/AI.js','model → {gx,gy}','B1.01',0.5),
]:
    D.append((it[0],it[1],it[2],SC,'','D1.AI决策',it[3],it[4],it[5],it[6],it[7],it[8],'V1','待办'))

# E.渲染与UI
SC,SN='E','E.渲染与UI'
for mod,items in [
    ('E1.战斗场景',[
        ('L3','E1.01','P6','场景进入','this.model=GameModel.init(deck0,deck1); this.battleLogic=new BattleLogic(model); registerInput();','BattleScene.onEnter()','js/scenes/BattleScene.js','void → void','A4.02,C1.01',0.5),
        ('L3','E1.02','P6','场景更新','this.battleLogic.update(dt)','BattleScene.onUpdate(dt)','js/scenes/BattleScene.js','dt → void','C1.01',0.2),
        ('L3','E1.03','P6','场景渲染(分层)','renderBackground(); renderHalls(); renderFormations(); renderUnits(); particleSystem.render(); renderUI()','BattleScene.onRender(ctx)','js/scenes/BattleScene.js','ctx → void','E2.01,E3.01,E4.01,E5.01,E6.01',0.5),
        ('L3','E1.04','P6','场景退出','input.clear(); model=null; battleLogic=null','BattleScene.onExit()','js/scenes/BattleScene.js','void → void','A2.04',0.2),
    ]),
    ('E2.背景渲染',[
        ('L3','E2.01','P6','山道背景','let grad=ctx.createLinearGradient(0,0,0,H); grad.addColorStop(0,"#1a3a2a"); grad.addColorStop(1,"#2d5a3d"); fillRect','BattleScene.renderBackground(ctx)','js/scenes/BattleScene.js','ctx → void','A5.01',0.3),
        ('L3','E2.02','P6','棋盘格子线','for(i=0;i<=BOARD_LENGTH;i++){ y=i*cellH; strokeStyle="rgba(255,255,255,0.1)"; drawLine }','BattleScene.renderGrid(ctx)','js/scenes/BattleScene.js','ctx → void','B1.01',0.3),
    ]),
    ('E3.大殿渲染',[
        ('L3','E3.01','P6','大殿色块','Renderer.rect(ctx,0,0,W,60,"#F39C12"); Renderer.rect(ctx,0,H-60,W,60,"#F39C12"); Renderer.text("敌方宗门"/"我方宗门")','BattleScene.renderHalls(ctx)','js/scenes/BattleScene.js','ctx → void','A5.01,A5.02',0.3),
        ('L3','E3.02','P6','大殿血条','Renderer.healthBar(ctx,0,H-70,W,8,p0.hp/p0.maxHp); Renderer.healthBar(ctx,0,62,W,8,p1.hp/p1.maxHp)','BattleScene.renderHallHp(ctx)','js/scenes/BattleScene.js','ctx → void','A5.03',0.3),
        ('L3','E3.03','P6','受击特效','if(shakeTimer>0){ shakeTimer-=dt; ctx.translate(random*6,random*6) }','BattleScene.renderHallHit(ctx,dt)','js/scenes/BattleScene.js','ctx+dt → void','—',0.3),
    ]),
    ('E4.单位渲染',[
        ('L3','E4.01','P6','单位色块+y排序','let all=getAllUnits(); all.sort((a,b)=>a.y-b.y); for(u of all){ color=owner===0?"#E74C3C":"#3498DB"; Renderer.rect(); Renderer.text(name[0]) }','BattleScene.renderUnits(ctx)','js/scenes/BattleScene.js','ctx → void','A5.01,A5.02',0.5),
        ('L3','E4.02','P6','单位血条','for(u of all){ Renderer.healthBar(ctx,u.px-u.w/2,u.py-u.h/2-6,u.w,4,u.hp/u.maxHp) }','BattleScene.renderUnitHp(ctx)','js/scenes/BattleScene.js','ctx → void','A5.03',0.3),
    ]),
    ('E5.阵法渲染',[
        ('L3','E5.01','P6','阵法光阵','for(p of players) for(f of formations){ color=isActive?"rgba(26,188,156,0.4)":"rgba(100,100,100,0.3)"; Renderer.rect(); Renderer.text(name) }','BattleScene.renderFormations(ctx)','js/scenes/BattleScene.js','ctx → void','A5.01,A5.02',0.5),
    ]),
    ('E6.UI-手牌栏',[
        ('L3','E6.01','P6','手牌显示','let hand=player.deck.hand; let cardW=(W-40)/4; for(i=0;i<hand.length;i++){ card=Cards.get(); canPlay=energy>=cost; Renderer.rect(); Renderer.text(name,cost) }','HandBar.render(ctx)','js/ui/HandBar.js','ctx → void','A5.01,A5.02,B2.02',0.8),
        ('L3','E6.02','P6','点击选中','this.selectedCard=(selectedCard===index)?-1:index','HandBar.onTap(index)','js/ui/HandBar.js','index:int → void','A2.02',0.2),
    ]),
    ('E7.UI-灵力条',[('L3','E7.01','P6','灵力显示','Renderer.rect(背景); Renderer.rect(前景按比例); Renderer.text(energy/max)','EnergyBar.render(ctx)','js/ui/EnergyBar.js','ctx → void','A5.01,A5.02',0.3)]),
    ('E8.UI-顶部HUD',[
        ('L3','E8.01','P6','双方血条','Renderer.rect(左半红); Renderer.rect(左半绿按比例); Renderer.rect(右半红); Renderer.rect(右半绿按比例)','HUD.renderHallBars(ctx)','js/ui/HUD.js','ctx → void','A5.01',0.3),
        ('L3','E8.02','P6','计时器','let t=ceil(model.time); let color=t<=30?"#E74C3C":"#FFF"; Renderer.text(t+"s")','HUD.renderTimer(ctx)','js/ui/HUD.js','ctx → void','A5.02',0.2),
    ]),
    ('E9.出牌交互',[
        ('L3','E9.01','P6','手牌→选目标','if(selectedCard===index){ selectedCard=-1; return } card=Cards.get(); if(energy<cost) return; selectedCard=index','BattleScene.onCardTap(index)','js/scenes/BattleScene.js','index → void','B2.02,B3.05',0.3),
        ('L3','E9.02','P6','格子点击执行','if(selectedCard===-1) return; card=Cards.get(); if(unit) spawnUnit(); else if(formation) placeFormation(); else castSpell(); selectedCard=-1','BattleScene.onGridTap(gx,gy)','js/scenes/BattleScene.js','gx+gy → void','C9.01,C5.01,C9.02',0.5),
    ]),
    ('E10.结算场景',[
        ('L3','E10.01','P6','胜负展示','let text=winner===0?"胜利":"失败"; let color=winner===0?"#27AE60":"#E74C3C"; Renderer.text(text,48px)','ResultScene.renderResult(ctx)','js/scenes/ResultScene.js','ctx → void','A5.02',0.3),
        ('L3','E10.02','P6','摧毁度','let r0=round((1-p1.hp/30)*100); let r1=round((1-p0.hp/30)*100); Renderer.text("我方"+r0+"% vs 敌方"+r1+"%")','ResultScene.renderDestroyRate(ctx)','js/scenes/ResultScene.js','ctx → void','A5.02',0.3),
        ('L3','E10.03','P6','再来一局','SceneManager.switch("battle")','ResultScene.onReplay()','js/scenes/ResultScene.js','void → void','A4.02',0.2),
    ]),
    ('E11.新手引导',[
        ('L3','E11.01','P6','引导1-出牌','高亮第一张手牌+向下箭头+"点击出兵"→监听到点击后step=2','TutorialGuide.step1(ctx)','js/ui/TutorialGuide.js','ctx → void','A5.01,A5.02',0.5),
        ('L3','E11.02','P6','引导2-灵力','高亮灵力条+"灵力不够时等待回复"→等3秒后step=3','TutorialGuide.step2(ctx)','js/ui/TutorialGuide.js','ctx → void','A5.01',0.3),
        ('L3','E11.03','P6','引导3-布阵','高亮阵法区格子+"点击布阵拦截"→监听到布阵后step=4','TutorialGuide.step3(ctx)','js/ui/TutorialGuide.js','ctx → void','A5.01',0.3),
        ('L3','E11.04','P6','引导4-目标','画指向敌方大殿箭头+"摧毁大殿获胜"→2秒后引导结束','TutorialGuide.step4(ctx)','js/ui/TutorialGuide.js','ctx → void','A5.01,A5.02',0.3),
    ]),
]:
    D.append(('L1','','',SC,SN,'','','','','','','','','',''))
    D.append(('L2',mod.split('.')[0],'P6',SC,'',mod,'','','','','','','','V1',''))
    for it in items:
        D.append((it[0],it[1],it[2],SC,'',mod,it[3],it[4],it[5],it[6],it[7],it[8],'V1','待办'))

# F/G/H
SC,SN='F','F.社交系统(V1.5)'
D.append(('L1','','',SC,SN,'','','','','','','','','',''))
for it in [
    ('L3','F1.01','P9','微信登录','wx.login→code→wx.cloud.callFunction("login",{code})→openid→setStorageSync','CloudManager.login()','js/social/CloudManager.js','void → openid','—',1.0),
    ('L3','F1.02','P10','云存档','wx.cloud.database().collection("users").doc(openid).set({data})','CloudManager.save(data)','js/social/CloudManager.js','data → void','F1.01',0.8),
    ('L3','F2.01','P11','布阵上传','wx.cloud.database().collection("challenges").add({data:{layout,openid}})→_id','AsyncPvP.uploadLayout(layout)','js/social/AsyncPvP.js','layout → challengeId','F1.01',1.0),
    ('L3','F2.02','P11','加载对手','wx.cloud.database().collection("challenges").doc(id).get()→layout→本地AI模拟','AsyncPvP.loadOpponent(id)','js/social/AsyncPvP.js','id → layout','F2.01',1.0),
    ('L3','F3.01','P12','好友排行','wx.getOpenDataContext().postMessage({action:"rank"})→开放数据域绘制','RankManager.renderFriendRank()','js/social/RankManager.js','void → void','F1.01',1.0),
    ('L3','F3.02','P12','分享','wx.shareAppMessage({title,imageUrl})','ShareManager.share(title,img)','js/social/ShareManager.js','title+img → void','—',0.5),
]:
    mod='F1.登录' if 'F1' in it[1] else ('F2.异步PvP' if 'F2' in it[1] else 'F3.排行分享')
    D.append((it[0],it[1],it[2],SC,'',mod,it[3],it[4],it[5],it[6],it[7],it[8],'V1.5','待办'))
SC,SN='G','G.养成变现(V1.5)'
D.append(('L1','','',SC,SN,'','','','','','','','','',''))
for it in [
    ('L3','G1.01','P13','弟子升级','unit.level++; unit.hp=baseHp*(1+level*0.1); unit.atk=baseAtk*(1+level*0.08)','ProgressionSystem.levelUp(unitId)','js/game/ProgressionSystem.js','unitId → void','—',1.0),
    ('L3','G1.02','P13','弟子升星','unit.star++; //练气→筑基→金丹→元婴','ProgressionSystem.starUp(unitId)','js/game/ProgressionSystem.js','unitId → void','G1.01',1.0),
    ('L3','G2.01','P14','激励广告','let ad=wx.createRewardedVideoAd({adUnitId}); ad.show(); ad.onClose(res=>{ if(res.isEnded) cb() })','AdManager.showReward(cb)','js/monetize/AdManager.js','cb → void','—',1.0),
    ('L3','G2.02','P14','内购','wx.requestPayment({timeStamp,nonceStr,package,signType,paySign,success:cb})','IAPManager.buy(productId,cb)','js/monetize/IAPManager.js','productId+cb → void','—',1.0),
]:
    mod='G1.养成' if 'G1' in it[1] else 'G2.变现'
    D.append((it[0],it[1],it[2],SC,'',mod,it[3],it[4],it[5],it[6],it[7],it[8],'V1.5','待办'))
SC,SN='H','H.实时PvP(V2)'
D.append(('L1','','',SC,SN,'','','','','','','','','',''))
for it in [
    ('L3','H1.01','P16','WebSocket','this.ws=wx.connectSocket({url}); onOpen/onMessage/onClose; setInterval(ping,10000)','NetworkClient.connect(url)','js/net/NetworkClient.js','url → void','—',2.0),
    ('L3','H1.02','P16','断线重连','onClose→保存model→重连→send({type:"reconnect",roomId})→服务端恢复','NetworkClient.reconnect()','js/net/NetworkClient.js','void → void','H1.01',1.0),
    ('L3','H2.01','P17','操作上报','ws.send(JSON.stringify({type:"action",action:{cardId,target,timestamp}}))','SyncManager.sendAction(action)','js/net/SyncManager.js','action → void','H1.01',1.0),
    ('L3','H2.02','P17','状态接收','onMessage→JSON.parse→if(type==="state") model.applyState(state)','SyncManager.onState(state)','js/net/SyncManager.js','state → void','H2.01',1.0),
    ('L3','H3.01','P18','匹配系统','wx.cloud.callFunction("match",{elo})→{roomId,opponentElo}','MatchMaker.match(elo)','js/net/MatchMaker.js','elo → roomId','F1.01',1.5),
    ('L3','H3.02','P18','段位赛','if(win) elo+=20; else elo-=15; rank=eloToRank(elo)','RankSystem.updateRank(win)','js/net/RankSystem.js','win:bool → void','—',1.0),
]:
    mod='H1.网络' if 'H1' in it[1] else ('H2.同步' if 'H2' in it[1] else 'H3.匹配')
    D.append((it[0],it[1],it[2],SC,'',mod,it[3],it[4],it[5],it[6],it[7],it[8],'V2','待办'))

print(f'数据准备完成: {len(D)}行, L3={len([d for d in D if d[0]=="L3"])}个功能点')
# 保存数据供第二部分使用
import json
with open('/workspace/_data.json','w') as f:
    json.dump(D,f)
print('数据已保存到_data.json')
