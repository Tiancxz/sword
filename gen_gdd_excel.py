#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成《宗门论道》游戏设计方案 Excel 表格"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== 样式定义 =====
title_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
title_fill = PatternFill('solid', fgColor='2F5496')
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='4472C4')
section_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
section_fill = PatternFill('solid', fgColor='5B9BD5')
cell_font = Font(name='微软雅黑', size=10)
cat_fill_attacker = PatternFill('solid', fgColor='FCE4D6')  # 攻方浅橙
cat_fill_defender = PatternFill('solid', fgColor='D6E4F0')  # 守方浅蓝
cat_fill_spell = PatternFill('solid', fgColor='E2EFDA')    # 法术浅绿
cat_fill_array = PatternFill('solid', fgColor='FFF2CC')     # 阵法浅黄
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='B4C7E7'),
    right=Side(style='thin', color='B4C7E7'),
    top=Side(style='thin', color='B4C7E7'),
    bottom=Side(style='thin', color='B4C7E7')
)

def style_header_row(ws, row, cols, fill=header_fill, font=header_font):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = center
        cell.border = thin_border

def style_data_rows(ws, start_row, end_row, cols, fill=None):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = cell_font
            cell.alignment = center if c > 1 else left
            cell.border = thin_border
            if fill:
                cell.fill = fill

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_title(ws, text, cols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = center
    ws.row_dimensions[row].height = 30

def add_section(ws, text, cols, row):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = left
    ws.row_dimensions[row].height = 24

# ================================================================
# Sheet 1: 游戏概述
# ================================================================
ws1 = wb.active
ws1.title = '游戏概述'
set_col_widths(ws1, [20, 50])
add_title(ws1, '《宗门论道》游戏概述', 2)
overview = [
    ('项目', '内容'),
    ('名称（暂定）', '宗门论道'),
    ('备选名', '御剑对决 / 仙门攻防 / 阵破苍穹 [?]'),
    ('类型', '实时策略卡牌（RTS-lite + 卡牌费用）'),
    ('题材', '东方修仙 / 宗门大战'),
    ('平台', '微信小游戏'),
    ('渲染', '2D，原生 Canvas 2D 起步（后期可迁 WebGL）'),
    ('美术', 'Spine 骨骼动画 + 序列帧 + 粒子特效'),
    ('后端', '微信云开发（V1 纯本地 MVP，V1.5 上云）'),
    ('目标用户', '喜欢轻度策略、修仙题材、3~5 分钟碎片对战'),
    ('单局时长', '3~5 分钟（180秒+60秒加时）'),
    ('核心一句话', '皇室战争式的实时卡牌攻防，修仙皮——你与对手各据山门一座，实时派弟子出征、布阵拦截、长老施法，摧毁对方宗门大殿者胜。'),
]
style_header_row(ws1, 3, 2)
for i, (k, v) in enumerate(overview):
    r = 3 + i
    ws1.cell(row=r, column=1, value=k)
    ws1.cell(row=r, column=2, value=v)
style_data_rows(ws1, 3, 3 + len(overview) - 1, 2)
ws1.cell(row=3, column=1).font = header_font
ws1.cell(row=3, column=2).font = header_font
ws1.cell(row=3, column=1).fill = header_fill
ws1.cell(row=3, column=2).fill = header_fill
for r in range(4, 3 + len(overview)):
    ws1.cell(row=r, column=2).alignment = left

# ================================================================
# Sheet 2: 攻方人物卡
# ================================================================
ws2 = wb.create_sheet('攻方人物卡')
cols2 = 9
set_col_widths(ws2, [16, 10, 8, 8, 8, 10, 12, 30, 6])
add_title(ws2, '攻方·人物卡（出场即向敌方大殿推进）', cols2)
headers2 = ['名称', '类型', '费用', '血量', '攻击', '移速', '攻击间隔', '特性', '稀有度']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, cols2)
data2 = [
    ('宗门体修弟子', '普通弟子', 2, 4, 2, 1.0, '1.0s', '普通近战士兵，基础兵', '凡品'),
    ('宗门剑修弟子', '普通弟子', 3, 3, 3, 0.9, '1.3s', '远程(射程3)，用飞剑，不掉血', '凡品'),
    ('宗门御兽弟子', '普通弟子', 3, 6, 2, 0.8, '1.2s', '控兽当肉盾，高血低速', '凡品'),
    ('金丹期长老', '精英长老', 6, 10, 3, 0.8, '1.5s', '定期随机释放分支技能：飞剑(范围伤)/丹药(回血)/符箓(清兵)/御兽(召唤)', '宝品'),
]
for i, row in enumerate(data2):
    r = 4 + i
    for j, val in enumerate(row, 1):
        ws2.cell(row=r, column=j, value=val)
style_data_rows(ws2, 4, 3 + len(data2), cols2, cat_fill_attacker)

# ================================================================
# Sheet 3: 守方人物卡
# ================================================================
ws3 = wb.create_sheet('守方人物卡')
cols3 = 8
set_col_widths(ws3, [16, 10, 8, 8, 8, 8, 30, 6])
add_title(ws3, '守方·人物卡（布在阵法区/大殿前，不推进或半动）', cols3)
headers3 = ['名称', '类型', '费用', '血量', '攻击', '射程', '特性', '稀有度']
for i, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=i, value=h)
style_header_row(ws3, 3, cols3)
data3 = [
    ('护山傀儡', '防守单位', 3, 6, 2, 1, '缓慢移动的肉盾守卫，拦截', '凡品'),
    ('护山灵兽', '防守单位', 4, 8, 2, 1, '高血拦截，死亡时自爆1伤', '灵品'),
    ('护法长老', '精英长老', 6, 10, 3, 1, '镇守大殿前，定期随机释放：阵法(补阵)/符箓(清兵)/丹药(回血)/御兽(召唤)', '宝品'),
]
for i, row in enumerate(data3):
    r = 4 + i
    for j, val in enumerate(row, 1):
        ws3.cell(row=r, column=j, value=val)
style_data_rows(ws3, 4, 3 + len(data3), cols3, cat_fill_defender)

# ================================================================
# Sheet 4: 阵法卡（双方通用）
# ================================================================
ws4 = wb.create_sheet('阵法卡(通用)')
cols4 = 7
set_col_widths(ws4, [14, 8, 8, 8, 8, 30, 6])
add_title(ws4, '技能卡·阵法（双方通用，布在阵法区格子，不动，有血量）', cols4)
headers4 = ['名称', '费用', '血量', '攻击', '射程', '特性', '稀有度']
for i, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=i, value=h)
style_header_row(ws4, 3, cols4)
data4 = [
    ('截脉阵', 2, 4, 2, 1, '基础拦截，便宜', '凡品'),
    ('寒霜阵', 3, 3, 1, 1, '命中后敌人移速-0.5(2s)', '凡品'),
    ('万刃阵', 4, 5, 3, 1, '高输出拦截', '灵品'),
    ('反震阵', 3, 3, 0, '-', '反伤50%', '灵品'),
    ('天罗阵', 5, 6, 2, 1, '范围(打相邻所有敌人)', '宝品'),
]
for i, row in enumerate(data4):
    r = 4 + i
    for j, val in enumerate(row, 1):
        ws4.cell(row=r, column=j, value=val)
style_data_rows(ws4, 4, 3 + len(data4), cols4, cat_fill_array)

# ================================================================
# Sheet 5: 法术卡
# ================================================================
ws5 = wb.create_sheet('法术卡')
cols5 = 5
set_col_widths(ws5, [14, 8, 40, 12, 6])
add_title(ws5, '技能卡·法术（即时效果，需选目标）', cols5)
headers5 = ['名称', '费用', '效果', '偏向', '稀有度']
for i, h in enumerate(headers5, 1):
    ws5.cell(row=3, column=i, value=h)
style_header_row(ws5, 3, cols5)
data5 = [
    ('万剑归宗', 5, '全己方单位+1攻、移速+0.3(5s)', '攻方', '宝品'),
    ('五雷正法', 4, '区域3格内敌方受4伤', '攻方/通用', '灵品'),
    ('御风诀', 2, '指定己方单位移速+0.5(5s)', '通用', '凡品'),
    ('镇魂符', 3, '指定敌方阵法失效3秒', '通用', '凡品'),
    ('金钟罩', 3, '己方大殿免疫伤害3秒', '守方', '灵品'),
    ('移山倒海', 4, '区域敌人推后2格+1伤', '守方', '灵品'),
    ('困仙索', 2, '指定敌人定身2秒', '守方', '凡品'),
    ('天雷诀', 4, '范围3格内敌方单位受4伤(清兵)', '守方/通用', '灵品'),
]
for i, row in enumerate(data5):
    r = 4 + i
    for j, val in enumerate(row, 1):
        ws5.cell(row=r, column=j, value=val)
style_data_rows(ws5, 4, 3 + len(data5), cols5, cat_fill_spell)

# ================================================================
# Sheet 6: 长老技能分支
# ================================================================
ws6 = wb.create_sheet('长老技能分支')
cols6 = 5
set_col_widths(ws6, [14, 14, 14, 30, 10])
add_title(ws6, '金丹长老 / 护法长老 随机分支技能（每5秒随机释放1个）', cols6)
headers6 = ['长老', '分支', '对应流派', '效果', '触发方式']
for i, h in enumerate(headers6, 1):
    ws6.cell(row=3, column=i, value=h)
style_header_row(ws6, 3, cols6)
data6 = [
    ('金丹期长老', '飞剑分支', '剑修', '万剑齐发(范围伤害)', '随机(每5s)'),
    ('金丹期长老', '丹药分支', '丹修', '自身+周围己方回血', '随机(每5s)'),
    ('金丹期长老', '符箓分支', '符修', '天雷诀(范围清兵)', '随机(每5s)'),
    ('金丹期长老', '御兽分支', '御兽', '召唤灵兽助战', '随机(每5s)'),
    ('护法长老', '阵法分支', '阵修', '自动补截脉阵', '随机(每5s)'),
    ('护法长老', '符箓分支', '符修', '天雷诀(清兵)', '随机(每5s)'),
    ('护法长老', '丹药分支', '丹修', '阵法/大殿回血', '随机(每5s)'),
    ('护法长老', '御兽分支', '御兽', '召唤护山灵兽', '随机(每5s)'),
]
for i, row in enumerate(data6):
    r = 4 + i
    for j, val in enumerate(row, 1):
        ws6.cell(row=r, column=j, value=val)
style_data_rows(ws6, 4, 3 + len(data6), cols6)
# 攻方长老橙色，守方长老蓝色
for r in range(4, 8):
    for c in range(1, cols6 + 1):
        ws6.cell(row=r, column=c).fill = cat_fill_attacker
for r in range(8, 12):
    for c in range(1, cols6 + 1):
        ws6.cell(row=r, column=c).fill = cat_fill_defender

# ================================================================
# Sheet 7: 战斗机制
# ================================================================
ws7 = wb.create_sheet('战斗机制')
cols7 = 3
set_col_widths(ws7, [16, 16, 40])
add_title(ws7, '战斗机制总览', cols7)
headers7 = ['类别', '机制', '说明']
for i, h in enumerate(headers7, 1):
    ws7.cell(row=3, column=i, value=h)
style_header_row(ws7, 3, cols7)
data7 = [
    ('单位行为', '寻路', '沿主路向敌方大殿移动'),
    ('单位行为', '交战', '进入攻击范围→停下→按攻击间隔造成伤害'),
    ('单位行为', '击杀后继续推进', '击杀目标后立即继续推进（不消失，核心修复）'),
    ('单位行为', '到达大殿', '对大殿造成1次伤害后消失'),
    ('伤害结算', '近战', '互相扣血（攻击力互减）'),
    ('伤害结算', '远程', '只打敌人，自己不掉血，射程3格'),
    ('伤害结算', '范围', '打相邻格所有敌人'),
    ('伤害结算', '自爆', '护山灵兽死亡时对相邻格造成1伤'),
    ('特殊机制', '减速', '寒霜阵命中后敌人移速-0.5，持续2秒'),
    ('特殊机制', '反震', '反震阵被攻击时反弹50%伤害给攻击者'),
    ('特殊机制', '加速', '御风诀/万剑归宗移速提升，持续X秒'),
    ('特殊机制', '禁阵', '镇魂符使阵法失效3秒（不攻击不拦截）'),
    ('特殊机制', '护盾', '金钟罩使大殿免疫3秒'),
    ('特殊机制', '定身', '困仙索使敌人无法移动2秒'),
    ('特殊机制', '推后', '移山倒海使区域敌人强制后退2格+1伤'),
    ('阵法机制', '固定建筑', '布在阵法区格子，不动，有血量'),
    ('阵法机制', '冷却', '被击毁的格子8秒内不能再布阵（防无限堵路）'),
    ('阵法机制', '主动攻击', '攻击范围1格，会主动攻击经过的敌方单位'),
    ('阵法机制', '双方通用', '攻方可铺阵掩护，守方可布阵拦截'),
]
for i, row in enumerate(data7):
    r = 4 + i
    for j, val in enumerate(row, 1):
        ws7.cell(row=r, column=j, value=val)
style_data_rows(ws7, 4, 3 + len(data7), cols7)

# ================================================================
# Sheet 8: 资源与手牌
# ================================================================
ws8 = wb.create_sheet('资源与手牌')
cols8 = 3
set_col_widths(ws8, [16, 20, 30])
add_title(ws8, '灵力系统与手牌系统', cols8)
headers8 = ['系统', '参数', '数值/说明']
for i, h in enumerate(headers8, 1):
    ws8.cell(row=3, column=i, value=h)
style_header_row(ws8, 3, cols8)
data8 = [
    ('灵力系统', '实时回复', '每 2.8 秒 +1 [初值]'),
    ('灵力系统', '上限', '开局 5，每 30 秒 +1，封顶 10 [初值]'),
    ('灵力系统', '加时赛', '回复速度 ×1.5 [?]'),
    ('灵力系统', '卡牌费用', '1~6（小单位1~2，中单位3~4，大单位/法术5~6）'),
    ('手牌系统', '卡组', '8 张 [初值]，可重复携带'),
    ('手牌系统', '手牌', '4 张 [初值]，打出1张后从牌库抽1张补位'),
    ('手牌系统', '抽牌延迟', '打出后 2 秒补新牌 [初值]'),
    ('手牌系统', '稀有度', '凡品/灵/宝/仙（影响养成，不影响单局公平）'),
]
for i, row in enumerate(data8):
    r = 4 + i
    for j, val in enumerate(row, 1):
        ws8.cell(row=r, column=j, value=val)
style_data_rows(ws8, 4, 3 + len(data8), cols8)

# ================================================================
# Sheet 9: 平衡设计
# ================================================================
ws9 = wb.create_sheet('平衡设计')
cols9 = 3
set_col_widths(ws9, [20, 16, 16])
add_title(ws9, '平衡杠杆（可调参数）', cols9)
headers9 = ['杠杆', '调高→偏向', '调低→偏向']
for i, h in enumerate(headers9, 1):
    ws9.cell(row=3, column=i, value=h)
style_header_row(ws9, 3, cols9)
data9 = [
    ('阵法冷却时间', '攻方', '守方'),
    ('阵法血量', '守方', '攻方'),
    ('阵法费用', '攻方', '守方'),
    ('单位移速', '攻方', '守方'),
    ('大殿血量', '守方', '攻方'),
    ('灵力回复速度', '攻方(出兵多)', '守方(布阵少)'),
]
for i, row in enumerate(data9):
    r = 4 + i
    for j, val in enumerate(row, 1):
        ws9.cell(row=r, column=j, value=val)
style_data_rows(ws9, 4, 3 + len(data9), cols9)
# 验证标准
vr = 4 + len(data9) + 1
add_section(ws9, '验证标准', cols9, vr)
verify = [
    '模拟10局：双方摧毁度都在 20%~70% 区间',
    '胜率：攻守方各 ~50%',
    '单局有伤害：不再出现0%摧毁度',
    '有来有回：双方大殿都被打到',
]
for i, v in enumerate(verify):
    r = vr + 1 + i
    ws9.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols9)
    ws9.cell(row=r, column=1, value='• ' + v)
    ws9.cell(row=r, column=1).font = cell_font
    ws9.cell(row=r, column=1).alignment = left
    ws9.cell(row=r, column=1).border = thin_border

# ================================================================
# Sheet 10: 版本规划
# ================================================================
ws10 = wb.create_sheet('版本规划')
cols10 = 3
set_col_widths(ws10, [12, 30, 40])
add_title(ws10, '版本规划', cols10)
headers10 = ['版本', '主题', '内容']
for i, h in enumerate(headers10, 1):
    ws10.cell(row=3, column=i, value=h)
style_header_row(ws10, 3, cols10)
data10 = [
    ('V1.0 MVP', '实时制 PvE', '实时核心循环（单位自动走/打/突破后继续推进）\n折中视角（上下直推主路+两侧阵法区+山道视觉）\n弟子职业体系（体修/剑修/御兽+金丹长老）\n阵法+法术技能卡（双方通用）\n灵力实时回复+手牌系统\n简单 AI 对手\n占位美术（色块+简单动画）\n胜负结算\n验证标准：模拟能打出30%+摧毁度'),
    ('V1.5', '社交+养成+美术', 'Spine 美术接入\n微信登录+云开发存档\n异步 PvP（挑战好友布阵+回放）\n排行榜+分享\n基础养成（单位升级）\n混合变现（广告+通行证）'),
    ('V2.0', '实时PvP', '实时对战（WebSocket+状态同步）\n匹配+段位赛\n赛季+锦标赛'),
]
for i, row in enumerate(data10):
    r = 4 + i
    for j, val in enumerate(row, 1):
        ws10.cell(row=r, column=j, value=val)
style_data_rows(ws10, 4, 3 + len(data10), cols10)
for r in range(4, 4 + len(data10)):
    ws10.row_dimensions[r].height = 120
    ws10.cell(row=r, column=3).alignment = left

# ================================================================
# Sheet 11: 待定问题
# ================================================================
ws11 = wb.create_sheet('待定问题')
cols11 = 3
set_col_widths(ws11, [6, 30, 20])
add_title(ws11, '待定/开放问题（需决策）', cols11)
headers11 = ['序号', '问题', '当前初值/状态']
for i, h in enumerate(headers11, 1):
    ws11.cell(row=3, column=i, value=h)
style_header_row(ws11, 3, cols11)
data11 = [
    (1, '最终名称', '宗门论道/御剑对决/其他'),
    (2, '单局时长', '180秒+60秒加时'),
    (3, '灵力回复速度', '2.8秒/+1'),
    (4, '手牌数量', '4张'),
    (5, '大殿血量', '30'),
    (6, '阵法冷却时间', '8秒'),
    (7, '金丹长老技能触发间隔', '5秒'),
    (8, '金丹长老是否随境界升级解锁更强分支', '待定'),
    (9, '美术来源', '自画/AI生成占位/买素材/找美术'),
    (10, '是否上云开发存档', 'V1本地 vs V1.5上云'),
    (11, '加时赛机制', '加速灵力/突然死亡'),
    (12, '是否引入"境界/段位"成长线', '待定'),
]
for i, row in enumerate(data11):
    r = 4 + i
    for j, val in enumerate(row, 1):
        ws11.cell(row=r, column=j, value=val)
style_data_rows(ws11, 4, 3 + len(data11), cols11)

# ===== 保存 =====
output_path = '/workspace/宗门论道_游戏设计方案.xlsx'
wb.save(output_path)
print(f'Excel 已生成: {output_path}')
print(f'共 {len(wb.sheetnames)} 个工作表: {wb.sheetnames}')
