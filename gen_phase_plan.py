#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
在现有Excel最前面插入「0.开发阶段规划」Sheet
按依赖关系排出8个开发阶段，每阶段说明：做什么/目的/能实现什么进展/验证标准/对应主表编号
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = '/workspace/宗门论道_开发管理工具包.xlsx'
wb = openpyxl.load_workbook(PATH)

# 配色
C_INK   = '2C3E50'
C_GOLD  = 'B7950B'
C_BLUE  = '2874A6'
C_TEAL  = '117A65'
C_RED   = 'C0392B'
C_PURPLE= '8E44AD'
C_GREEN = '27AE60'

# 8个阶段配色（渐变）
PHASE_COLORS = [
    '1A5276',  # P1 深蓝
    '2874A6',  # P2 蓝
    '117A65',  # P3 青
    'B7950B',  # P4 金
    '8E44AD',  # P5 紫
    'C0392B',  # P6 红
    'D35400',  # P7 橙
    '27AE60',  # P8 绿
]

F_BIG    = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
F_TITLE  = Font(name='微软雅黑', size=13, bold=True, color='FFFFFF')
F_HEADER = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
F_PHASE  = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
F_CELL   = Font(name='微软雅黑', size=10, color='2C3E50')
F_BOLD   = Font(name='微软雅黑', size=10, bold=True, color='2C3E50')
F_CODE   = Font(name='Consolas', size=10, color='1A5276')
F_SMALL  = Font(name='微软雅黑', size=9, color='7F8C8D')
F_GOAL   = Font(name='微软雅黑', size=10, bold=True, color='B7950B')
F_RESULT = Font(name='微软雅黑', size=10, bold=True, color='117A65')

A_C  = Alignment(horizontal='center', vertical='center', wrap_text=True)
A_L  = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
A_LT = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=1)

BD = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC')
)

# 创建Sheet并放到最前面
if '0.开发阶段规划' in wb.sheetnames:
    del wb['0.开发阶段规划']
ws = wb.create_sheet('0.开发阶段规划', 0)
ws.sheet_view.showGridLines = False

# 列宽
widths = [4, 6, 16, 26, 30, 36, 36, 22, 8]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 大标题
ws.merge_cells('A1:I1')
c = ws.cell(row=1, column=1, value='开发阶段规划 — 先做什么后做什么，每步目的是什么、能实现什么')
c.font = F_BIG
c.fill = PatternFill('solid', fgColor=C_INK)
c.alignment = A_C
ws.row_dimensions[1].height = 46

# 说明
ws.merge_cells('A2:I2')
c = ws.cell(row=2, column=1, value='按依赖关系排出8个阶段，必须从P1到P8顺序做（后面的依赖前面的）。每个阶段做完都有"能跑起来看到的东西"，不是做完一堆代码什么反应都没有。')
c.font = F_SMALL
c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
ws.row_dimensions[2].height = 24

# 表头
headers = ['', '阶段', '阶段名称', '做什么（内容）', '目的（为什么先做这个）', '能实现的进展（做完能看到什么）', '验证标准（怎么算做完）', '对应主表编号', '周期']
for i, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=i, value=h)
    c.font = F_HEADER
    c.fill = PatternFill('solid', fgColor=C_INK)
    c.alignment = A_C
    c.border = BD
ws.row_dimensions[3].height = 32

# 8个阶段数据
phases = [
    {
        'num': 'P1',
        'name': '引擎骨架',
        'content': '渲染循环(Director)\n输入系统(Input)\n事件总线(EventSystem)\n场景管理(SceneManager)\n渲染器(Renderer)',
        'purpose': '这是整个游戏的地基。没有渲染循环什么也跑不起来，没有输入系统什么也点不了。必须先搭好引擎框架，后面所有功能才能挂上去。',
        'progress': '能看到一个黑色画布在60fps跑\n能在画布上画矩形/文字\n点屏幕能触发回调\n能切换空场景A和空场景B',
        'verify': '①Director.loop 60fps跑通\n②点击屏幕console打印坐标\n③画矩形和文字能显示\n④场景A能切换到场景B',
        'ref': 'A1.01~A1.04\nA2.01~A2.04\nA3.01~A3.03\nA4.01~A4.02\nA5.01~A5.03',
        'period': '1周',
    },
    {
        'num': 'P2',
        'name': '数据层',
        'content': '常量配置(Constants)\n卡牌数据表(Cards)\n卡组手牌(Deck)\n预设卡组(DeckPresets)',
        'purpose': '引擎能跑了，但游戏需要数据。卡牌属性、灵力参数、棋盘尺寸等所有数值必须先配置好，后面的战斗逻辑才能读数据创建单位。配置驱动是核心原则——改数值只改配置不改代码。',
        'progress': '19张卡牌数据完整可查\n8张卡组能洗牌+抽牌\n能console打印手牌和卡牌属性\n改Cards.js数值能立即生效',
        'verify': '①Cards.get(id)返回完整属性\n②Deck.init后hand有3张牌\n③Deck.draw能抽牌\n④canPlay能正确判断灵力',
        'ref': 'B1.01\nB2.01~B2.02\nB3.01~B3.06\nB4.01',
        'period': '1周',
    },
    {
        'num': 'P3',
        'name': '核心战斗逻辑',
        'content': '战斗主循环(BattleLogic)\n单位实体(Unit)\n移动系统(Movement)\n战斗系统(Combat)\n★击杀后继续推进★',
        'purpose': '这是整个游戏好不好玩的命门。单位能不能走、能不能打、打完能不能继续冲——这3件事决定了游戏核心循环是否成立。必须在这个阶段验证"攻方打得动"。',
        'progress': '★最重要的里程碑★\nAI vs AI模拟能跑通\n能看到单位在棋盘上移动\n能看到单位互相打\n能看到单位杀掉敌人后继续冲\n能看到单位到大殿造成伤害',
        'verify': '①单位y坐标每帧变化(在走)\n②两个单位接触后互扣血\n③单位击杀后state=walking继续走\n④单位到大殿后大殿血量减少',
        'ref': 'C1.01\nC2.01~C2.04\nC3.01~C3.04\nC4.01~C4.06',
        'period': '2周',
    },
    {
        'num': 'P4',
        'name': '扩展战斗机制',
        'content': '阵法系统(布阵/攻击/冷却)\n法术系统(8种法术效果)\n长老技能(4分支随机)\n灵力系统(实时回复)\n出牌执行\n胜负判定\n死亡清理',
        'purpose': '核心循环跑通了，但只有"走和打"太单调。阵法让守方能防，法术增加策略深度，长老增加变化性，灵力系统驱动出牌节奏。这些让游戏从"能跑"变成"好玩"。',
        'progress': '完整单局能跑通\n能布阵拦截(阵法有血量会攻击)\n能放法术(8种效果各不相同)\n长老每5秒随机放技能\n灵力实时回复能出牌\n能判定胜负(大殿破/超时)',
        'verify': '①布阵后阵法能攻击经过敌人\n②阵法被毁后格子8秒冷却\n③8种法术效果各自正确生效\n④长老每5秒释放1个随机分支\n⑤灵力每2.8秒+1\n⑥大殿血量归零能判定胜负',
        'ref': 'C5.01~C5.04\nC6.01~C6.09\nC7.01~C7.06\nC8.01~C8.02\nC9.01~C9.02\nC10.01~C10.03\nC11.01~C11.02',
        'period': '2周',
    },
    {
        'num': 'P5',
        'name': 'AI对手',
        'content': 'AI决策(灵力管理/出牌/布阵)\n3档难度(简单/普通/困难)',
        'purpose': '战斗逻辑完整了，但需要对手。没有AI玩家无法测试游戏好不好玩。AI不需要完美，但要能模拟真人节奏出牌布阵，让你能完整体验一局对战。',
        'progress': '能和AI完整打一局(纯逻辑)\nAI会出兵也会布阵\nAI会根据血量调整攻守\n简单AI弱/困难AI强有明显差异',
        'verify': '①AI每2.5秒(普通)决策一次\n②AI会根据大殿血量调攻守比\n③简单/普通/困难行为有差异\n④AI vs AI能完整跑完一局不卡死',
        'ref': 'D1.01~D1.05\nD2.01~D2.03',
        'period': '1周',
    },
    {
        'num': 'P6',
        'name': '渲染与交互',
        'content': '战斗场景(分层渲染)\n背景/大殿/单位/阵法渲染\n手牌栏/灵力条/HUD\n出牌交互(选牌→选目标)\n结算场景\n新手引导\n粒子特效\n音频',
        'purpose': '到这一步为止游戏都是"纯逻辑"——console里跑数字。P6把所有逻辑变成"能看到能操作的画面"。这是从"程序员眼里的游戏"变成"玩家眼里的游戏"的关键。',
        'progress': '★玩家可见里程碑★\n能在微信开发者工具里看到画面\n能看到棋盘/大殿/单位/阵法\n能点手牌出牌、点格子布阵\n能看到灵力条/血条/计时器\n战斗结束有结算页\n新手引导能走通4步',
        'verify': '①真机/开发者工具能显示画面\n②点手牌→选目标→出牌流程通\n③单位移动/交战有视觉反馈\n④大殿受击有震动+闪烁\n⑤结算页显示胜负+摧毁度\n⑥新手引导4步走通',
        'ref': 'E1.01~E1.04\nE2.01~E2.02\nE3.01~E3.03\nE4.01~E4.03\nE5.01~E5.02\nE6.01~E6.03\nE7.01\nE8.01~E8.02\nE9.01~E9.02\nE10.01~E10.03\nE11.01~E11.04',
        'period': '2周',
    },
    {
        'num': 'P7',
        'name': '测试与平衡调优',
        'content': '纯逻辑模拟器(AI vs AI 10局)\n平衡性调优(参数调整)\n性能测试(60fps/内存)\n真机测试\nBug修复',
        'purpose': '游戏能玩了，但好不好玩是另一回事。P7用模拟器跑大量对局验证平衡——攻方能不能打进去、守方能不能防下来、有没有0%死局。调到"双方都有来有回"才算过关。',
        'progress': '★平衡达标里程碑★\n模拟10局输出摧毁度/胜率数据\n参数调到双方摧毁度20~70%\n真机60fps不卡\n无崩溃无死锁',
        'verify': '①模拟10局: 摧毁度20~70%\n②胜率40~60%\n③0%摧毁度对局<10%\n④90%对局150~210秒结束\n⑤真机60fps同屏20单位不卡',
        'ref': '全部联调\n(卡牌数据/数值配置)',
        'period': '1.5周',
    },
    {
        'num': 'P8',
        'name': '提交审核上线',
        'content': '最终Bug修复\n微信小游戏审核提交\n审核反馈处理\n正式上线',
        'purpose': '游戏做完要上线才能让玩家玩到。微信小游戏需要提交审核，修仙题材过审风险低但仍需确保无违规内容。这是V1的终点。',
        'progress': '★V1上线里程碑★\n微信审核通过\n玩家能搜索到并玩到游戏\nV1 MVP完成',
        'verify': '①微信审核通过\n②玩家能正常游玩\n③无严重Bug\n④新手引导完成率>80%',
        'ref': '—',
        'period': '1周',
    },
]

# 写入阶段数据
r = 4
for i, p in enumerate(phases):
    color = PHASE_COLORS[i]
    row_height = 120

    # 阶段编号列(大色块)
    c = ws.cell(row=r, column=2, value=p['num'])
    c.font = F_PHASE
    c.fill = PatternFill('solid', fgColor=color)
    c.alignment = A_C
    c.border = BD

    # 阶段名称
    c = ws.cell(row=r, column=3, value=p['name'])
    c.font = F_PHASE
    c.fill = PatternFill('solid', fgColor=color)
    c.alignment = A_C
    c.border = BD

    # 做什么
    c = ws.cell(row=r, column=4, value=p['content'])
    c.font = F_CELL
    c.alignment = A_LT
    c.border = BD

    # 目的
    c = ws.cell(row=r, column=5, value=p['purpose'])
    c.font = F_CELL
    c.alignment = A_LT
    c.border = BD

    # 能实现的进展
    c = ws.cell(row=r, column=6, value=p['progress'])
    c.font = F_RESULT
    c.alignment = A_LT
    c.border = BD
    c.fill = PatternFill('solid', fgColor='E8F8F5')

    # 验证标准
    c = ws.cell(row=r, column=7, value=p['verify'])
    c.font = F_CODE
    c.alignment = A_LT
    c.border = BD

    # 对应主表编号
    c = ws.cell(row=r, column=8, value=p['ref'])
    c.font = F_CODE
    c.alignment = A_LT
    c.border = BD
    c.fill = PatternFill('solid', fgColor='FDF2E9')

    # 周期
    c = ws.cell(row=r, column=9, value=p['period'])
    c.font = F_BOLD
    c.alignment = A_C
    c.border = BD

    ws.row_dimensions[r].height = row_height
    r += 1

# 箭头连接说明
r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
c = ws.cell(row=r, column=2, value='P1引擎骨架 → P2数据层 → P3核心战斗 → P4扩展战斗 → P5 AI对手 → P6渲染交互 → P7测试平衡 → P8上线   （必须按顺序，后面依赖前面）')
c.font = Font(name='微软雅黑', size=11, bold=True, color=C_GOLD)
c.alignment = A_C
c.fill = PatternFill('solid', fgColor='FEF9E7')
c.border = BD
ws.row_dimensions[r].height = 30
r += 2

# 关键里程碑总结
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
c = ws.cell(row=r, column=2, value='关键里程碑（做完能看到的东西）')
c.font = F_TITLE
c.fill = PatternFill('solid', fgColor=C_GOLD)
c.alignment = A_L
c.border = BD
ws.row_dimensions[r].height = 28
r += 1

milestones = [
    ('P1做完', '黑色画布在跑，能画能点', '代码能跑'),
    ('P3做完', '★最重要★ AI vs AI模拟能看到单位走、打、突破', '核心循环成立'),
    ('P5做完', '纯逻辑能和AI打一局完整对战', '逻辑层完成'),
    ('P6做完', '★玩家可见★ 微信工具里能看到画面能操作', '能玩了'),
    ('P7做完', '★平衡达标★ 模拟10局双方都能打出伤害', '好玩了'),
    ('P8做完', '★V1上线★ 微信审核通过玩家能玩', 'V1完成'),
]
for phase, desc, tag in milestones:
    c = ws.cell(row=r, column=2, value=phase)
    c.font = F_BOLD; c.alignment = A_C; c.border = BD
    c.fill = PatternFill('solid', fgColor='D6EAF8')
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    c = ws.cell(row=r, column=3, value=desc)
    c.font = F_CELL; c.alignment = A_L; c.border = BD
    c = ws.cell(row=r, column=8, value=tag)
    c.font = F_GOAL; c.alignment = A_C; c.border = BD
    c.fill = PatternFill('solid', fgColor='FEF9E7')
    ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
    ws.row_dimensions[r].height = 26
    r += 1

r += 2
# 每日开发指引
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
c = ws.cell(row=r, column=2, value='每日开发指引（照着做就行）')
c.font = F_TITLE
c.fill = PatternFill('solid', fgColor=C_TEAL)
c.alignment = A_L
c.border = BD
ws.row_dimensions[r].height = 28
r += 1

daily_steps = [
    '第1步：看本表，确认当前在哪个阶段（比如P3）',
    '第2步：打开「2.开发主表」，找到当前阶段对应的编号（如C3.01）',
    '第3步：看那一行的"L4实现原理"和"L5函数"——那就是要做的',
    '第4步：用「AI提问指南」的模板，把原理+函数名发给AI让它实现',
    '第5步：AI做完后，按本阶段的"验证标准"自己验证',
    '第6步：验证通过→主表状态改"已完成"→做下一个',
    '第7步：当前阶段所有功能都做完→进入下一阶段',
    '原则：一个阶段做完再进下一个，不要跳着做。每阶段都有里程碑，做完能看到东西。',
]
for step in daily_steps:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    c = ws.cell(row=r, column=2, value='  ' + step)
    c.font = F_CELL; c.alignment = A_L; c.border = BD
    ws.row_dimensions[r].height = 24
    r += 1

# 保存
wb.save(PATH)
print(f'Excel已更新: {PATH}')
print(f'Sheet列表: {wb.sheetnames}')
print(f'开发阶段: 8个阶段, 总周期约10.5周')
