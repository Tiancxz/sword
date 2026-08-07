#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
《宗门论道》开发管理工具包 v4 — 修复全部7个缺陷
1.主表加"阶段"列+"文件路径"列
2.进度统计改成Excel公式自动计算
3.加回卡牌数据表+数值配置表
4.加回AI提问指南表
5.关键函数原理补充字段名和边界条件
6.(缺陷6=文件路径已在1中解决)
7.阶段内部依赖用编号顺序体现(编号本身就是顺序)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

PATH = '/workspace/宗门论道_开发管理工具包.xlsx'
wb = openpyxl.Workbook()

# ===== 配色 =====
C_INK   = '2C3E50'
C_GOLD  = 'B7950B'
C_BLUE  = '2874A6'
C_TEAL  = '117A65'
C_RED   = 'C0392B'
C_PURPLE= '8E44AD'
C_GREEN = '27AE60'

L1_FILLS = {
    'A': ('1A5276', 'FFFFFF'),
    'B': ('2874A6', 'FFFFFF'),
    'C': ('117A65', 'FFFFFF'),
    'D': ('B7950B', 'FFFFFF'),
    'E': ('8E44AD', 'FFFFFF'),
    'F': ('C0392B', 'FFFFFF'),
    'G': ('D35400', 'FFFFFF'),
    'H': ('34495E', 'FFFFFF'),
}
L2_FILL = 'D6EAF8'
L3_FILL = 'FDF2E9'
L4_FILL = 'F4ECF7'
L5_FILL = 'E8F8F5'
ST_TODO  = 'FADBD8'
ST_DOING = 'FEF9E7'
ST_DONE  = 'D5F5E3'

# 阶段配色
PHASE_COLORS = {
    'P1': '1A5276', 'P2': '2874A6', 'P3': '117A65', 'P4': 'B7950B',
    'P5': '8E44AD', 'P6': 'C0392B', 'P7': 'D35400', 'P8': '27AE60',
}

F_BIG    = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
F_TITLE  = Font(name='微软雅黑', size=13, bold=True, color='FFFFFF')
F_HEADER = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
F_L1     = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
F_L2     = Font(name='微软雅黑', size=11, bold=True, color=C_INK)
F_CELL   = Font(name='微软雅黑', size=10, color='2C3E50')
F_BOLD   = Font(name='微软雅黑', size=10, bold=True, color='2C3E50')
F_CODE   = Font(name='Consolas', size=10, color='1A5276')
F_SMALL  = Font(name='微软雅黑', size=9, color='7F8C8D')
F_TAG    = Font(name='微软雅黑', size=8, bold=True, color='FFFFFF')
F_GOLD   = Font(name='微软雅黑', size=11, bold=True, color=C_GOLD)
F_RESULT = Font(name='微软雅黑', size=10, bold=True, color=C_TEAL)
F_PHASE  = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')

A_C  = Alignment(horizontal='center', vertical='center', wrap_text=True)
A_L  = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
A_LT = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=1)
A_LM = Alignment(horizontal='left', vertical='center', wrap_text=True)

BD = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC')
)

def set_w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def big_title(ws, text, cols, row=1, color=C_INK):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_BIG; c.fill = PatternFill('solid', fgColor=color); c.alignment = A_C
    ws.row_dimensions[row].height = 44

def sub_title(ws, text, cols, row, color=C_GOLD):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_TITLE; c.fill = PatternFill('solid', fgColor=color); c.alignment = A_L
    ws.row_dimensions[row].height = 30

def header_row(ws, headers, row, color=C_INK):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = F_HEADER; c.fill = PatternFill('solid', fgColor=color)
        c.alignment = A_C; c.border = BD
    ws.row_dimensions[row].height = 30

# ================================================================
# 开发主表数据（含阶段+文件路径+细化原理）
# ================================================================
# 格式: (层级, 编号, 阶段, 系统, 系统名, 模块, 子功能, 实现原理(细化), 函数, 文件路径, 输入→输出, 版本, 状态)
# L3行才有完整数据，L1/L2行只有层级标记

D = []

# === A.引擎框架 ===
SC, SN = 'A', 'A.引擎框架'
M = 'A1.渲染循环'
D += [
    ('L1', '',   '',    SC, SN, '', '', '', '', '', '', '', '', ''),
    ('L2', 'A1', 'P1', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'A1.01', 'P1', SC, '', M, '主循环驱动',
     'requestAnimationFrame((ts)=>{ dt=(ts-lastTime)/1000; dt=Math.min(dt,0.033); this.update(dt); this.render(); this.lastTime=ts; requestAnimationFrame下一帧 })。lastTime初始=0，第一帧dt可能异常大，用Math.min钳制。running=false时停止循环。',
     'Director.loop(ts)', 'js/core/Director.js', 'ts:number → void', 'V1', '待办'),
    ('L3', 'A1.02', 'P1', SC, '', M, '逻辑更新',
     'let scene=this.sceneManager.current; if(scene) scene.onUpdate(dt)。若scene为null则跳过。',
     'Director.update(dt)', 'js/core/Director.js', 'dt:number → void', 'V1', '待办'),
    ('L3', 'A1.03', 'P1', SC, '', M, '画面渲染',
     'let ctx=this.canvas.getContext("2d"); ctx.clearRect(0,0,W,H); let scene=...; if(scene) scene.onRender(ctx)。清屏后调用场景渲染。',
     'Director.render()', 'js/core/Director.js', 'void → void', 'V1', '待办'),
    ('L3', 'A1.04', 'P1', SC, '', M, '引擎初始化',
     'this.canvas=wx.createCanvas(); this.ctx=canvas.getContext("2d"); 获取屏幕尺寸wx.getSystemInfoSync().screenWidth/Height; this.sceneManager=new SceneManager(); this.input=new Input(); this.input.init(); this.start()。',
     'Director.init()', 'js/core/Director.js', 'void → void', 'V1', '待办'),
]
M = 'A2.输入系统'
D += [
    ('L2', 'A2', 'P1', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'A2.01', 'P1', SC, '', M, '触摸监听',
     'wx.onTouchStart((e)=>this.onTouchStart(e.touches[0].clientX, e.touches[0].clientY)); 同理onTouchEnd。坐标从e.touches[0]取。',
     'Input.init()', 'js/core/Input.js', 'void → void', 'V1', '待办'),
    ('L3', 'A2.02', 'P1', SC, '', M, '点击区域注册',
     'this.regions.push({id,x,y,w,h,cb})。id用于取消注册。场景onEnter时注册，onExit时clear。',
     'Input.register(id,x,y,w,h,cb)', 'js/core/Input.js', 'id+坐标+回调 → void', 'V1', '待办'),
    ('L3', 'A2.03', 'P1', SC, '', M, '命中检测',
     'for(let r of this.regions){ if(tx>=r.x && tx<=r.x+r.w && ty>=r.y && ty<=r.y+r.h){ r.cb(tx,ty); break } }。命中第一个就break。',
     'Input.onTouchStart(tx,ty)', 'js/core/Input.js', '坐标 → 触发cb', 'V1', '待办'),
    ('L3', 'A2.04', 'P1', SC, '', M, '区域清理',
     'this.regions=[]。场景切换onExit时必须调用，否则旧场景按钮还会响应。',
     'Input.clear()', 'js/core/Input.js', 'void → void', 'V1', '待办'),
]
M = 'A3.事件总线'
D += [
    ('L2', 'A3', 'P1', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'A3.01', 'P1', SC, '', M, '事件注册',
     'if(!this.listeners.has(event)) this.listeners.set(event,[]); this.listeners.get(event).push(cb)。',
     'EventSystem.on(event,cb)', 'js/core/EventSystem.js', '事件名+回调 → void', 'V1', '待办'),
    ('L3', 'A3.02', 'P1', SC, '', M, '事件触发',
     'let cbs=this.listeners.get(event); if(cbs) for(let cb of cbs) cb(data)。',
     'EventSystem.emit(event,data)', 'js/core/EventSystem.js', '事件名+数据 → void', 'V1', '待办'),
    ('L3', 'A3.03', 'P1', SC, '', M, '事件移除',
     'let cbs=this.listeners.get(event); if(cbs){ let i=cbs.indexOf(cb); if(i>=0) cbs.splice(i,1) }。',
     'EventSystem.off(event,cb)', 'js/core/EventSystem.js', '事件名+回调 → void', 'V1', '待办'),
]
M = 'A4.场景管理'
D += [
    ('L2', 'A4', 'P1', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'A4.01', 'P1', SC, '', M, '场景注册',
     'this.scenes.set(name,scene)。',
     'SceneManager.register(name,scene)', 'js/core/SceneManager.js', '名称+场景 → void', 'V1', '待办'),
    ('L3', 'A4.02', 'P1', SC, '', M, '场景切换',
     'if(this.current) this.current.onExit(); this.current=this.scenes.get(name); if(this.current) this.current.onEnter()。',
     'SceneManager.switch(name)', 'js/core/SceneManager.js', '场景名 → void', 'V1', '待办'),
]
M = 'A5.渲染器'
D += [
    ('L2', 'A5', 'P1', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'A5.01', 'P1', SC, '', M, '矩形绘制',
     'ctx.fillStyle=color; ctx.fillRect(x,y,w,h)。',
     'Renderer.rect(ctx,x,y,w,h,color)', 'js/core/Renderer.js', 'ctx+坐标+颜色 → void', 'V1', '待办'),
    ('L3', 'A5.02', 'P1', SC, '', M, '文字绘制',
     'ctx.fillStyle=color; ctx.font=size+"px 微软雅黑"; ctx.fillText(text,x,y)。',
     'Renderer.text(ctx,text,x,y,color,size)', 'js/core/Renderer.js', 'ctx+文本+坐标 → void', 'V1', '待办'),
    ('L3', 'A5.03', 'P1', SC, '', M, '血条绘制',
     'ctx.fillStyle="#E74C3C"; ctx.fillRect(x,y,w,h); //背景红 ctx.fillStyle="#27AE60"; ctx.fillRect(x,y,w*ratio,h); //前景绿',
     'Renderer.healthBar(ctx,x,y,w,h,ratio)', 'js/core/Renderer.js', 'ctx+坐标+比例 → void', 'V1', '待办'),
]
M = 'A6.粒子系统'
D += [
    ('L2', 'A6', 'P6', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'A6.01', 'P6', SC, '', M, '粒子发射',
     'for(let i=0;i<count;i++){ let angle=Math.random()*Math.PI*2; let speed=20+Math.random()*40; this.particles.push({x,y,vx:Math.cos(angle)*speed,vy:Math.sin(angle)*speed,life:0.3+Math.random()*0.2,maxLife:0.5,color,size:3+Math.random()*3}) }',
     'ParticleSystem.emit(x,y,count,color)', 'js/core/ParticleSystem.js', '坐标+数量+颜色 → void', 'V1', '待办'),
    ('L3', 'A6.02', 'P6', SC, '', M, '粒子更新',
     'for(let p of this.particles){ p.x+=p.vx*dt; p.y+=p.vy*dt; p.life-=dt; if(p.life<=0)标记移除 } 过滤掉life<=0的。',
     'ParticleSystem.update(dt)', 'js/core/ParticleSystem.js', 'dt → void', 'V1', '待办'),
    ('L3', 'A6.03', 'P6', SC, '', M, '粒子渲染',
     'for(let p of this.particles){ ctx.globalAlpha=p.life/p.maxLife; Renderer.circle(ctx,p.x,p.y,p.size,p.color) } ctx.globalAlpha=1。',
     'ParticleSystem.render(ctx)', 'js/core/ParticleSystem.js', 'ctx → void', 'V1', '待办'),
]
M = 'A7.音频管理'
D += [
    ('L2', 'A7', 'P6', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'A7.01', 'P6', SC, '', M, 'BGM播放',
     'this.bgm=wx.getBackgroundAudioManager(); this.bgm.src=url; this.bgm.loop=true; this.bgm.play()。',
     'AudioManager.playBGM(url)', 'js/core/AudioManager.js', 'URL → void', 'V1', '待办'),
    ('L3', 'A7.02', 'P6', SC, '', M, 'SFX播放',
     'let audio=wx.createInnerAudioContext(); audio.src=url; audio.play(); audio.onEnded(()=>audio.destroy())。',
     'AudioManager.playSFX(url)', 'js/core/AudioManager.js', 'URL → void', 'V1', '待办'),
]

# === B.游戏数据 ===
SC, SN = 'B', 'B.游戏数据'
M = 'B1.常量配置'
D += [
    ('L1', '',   '',    SC, SN, '', '', '', '', '', '', '', '', ''),
    ('L2', 'B1', 'P2', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'B1.01', 'P2', SC, '', M, '全局常量',
     'module.exports={ BOARD_LENGTH:9, HALL_HP:30, ENERGY_START:5, ENERGY_REGEN:2.8, HAND_SIZE:4, ... } 所有硬编码值集中，代码中只引用Constants.xxx。',
     'Constants.js', 'js/config/Constants.js', '文件 → 引用', 'V1', '待办'),
]
M = 'B2.卡牌数据表'
D += [
    ('L2', 'B2', 'P2', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'B2.01', 'P2', SC, '', M, '卡牌配置',
     'module.exports={ "body_disciple":{name:"体修弟子",type:"unit",faction:"attack",cost:2,hp:4,atk:2,speed:1.0,range:1,interval:1.0,traits:[]}, ...共19张 }',
     'Cards.js', 'js/config/Cards.js', '文件 → 引用', 'V1', '待办'),
    ('L3', 'B2.02', 'P2', SC, '', M, '卡牌查表',
     'let card=this.data[cardId]; if(!card) console.error("未知卡牌:"+cardId); return card。',
     'Cards.get(cardId)', 'js/config/Cards.js', 'cardId:string → Card Object', 'V1', '待办'),
]
M = 'B3.卡组手牌'
D += [
    ('L2', 'B3', 'P2', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'B3.01', 'P2', SC, '', M, '卡组初始化',
     'this.drawPile=[...cardIds]; this.shuffle(); this.hand=[]; for(let i=0;i<3;i++) this.draw()。初始抽3张。',
     'Deck.init(cardIds)', 'js/game/Deck.js', 'cardIds:string[] → void', 'V1', '待办'),
    ('L3', 'B3.02', 'P2', SC, '', M, '洗牌算法',
     'for(let i=this.drawPile.length-1;i>0;i--){ let j=Math.floor(Math.random()*(i+1)); [this.drawPile[i],this.drawPile[j]]=[this.drawPile[j],this.drawPile[i]] }',
     'Deck.shuffle()', 'js/game/Deck.js', 'void → void', 'V1', '待办'),
    ('L3', 'B3.03', 'P2', SC, '', M, '抽牌',
     'if(this.hand.length>=4) return null; if(this.drawPile.length==0) return null; let card=this.drawPile.pop(); this.hand.push(card); return card',
     'Deck.draw()', 'js/game/Deck.js', 'void → cardId:string|null', 'V1', '待办'),
    ('L3', 'B3.04', 'P2', SC, '', M, '打出后补牌',
     'if(this.drawTimer>0){ this.drawTimer-=dt; if(this.drawTimer<=0){ this.draw(); this.drawTimer=0 } }',
     'Deck.update(dt)', 'js/game/Deck.js', 'dt:number → void', 'V1', '待办'),
    ('L3', 'B3.05', 'P2', SC, '', M, '出牌检查',
     'let card=Cards.get(this.hand[index]); return energy>=card.cost。',
     'Deck.canPlay(index,energy)', 'js/game/Deck.js', 'index:int+energy:int → bool', 'V1', '待办'),
    ('L3', 'B3.06', 'P2', SC, '', M, '打出牌',
     'let cardId=this.hand.splice(index,1)[0]; this.drawTimer=2.0; return cardId。',
     'Deck.playCard(index)', 'js/game/Deck.js', 'index:int → cardId:string', 'V1', '待办'),
]
M = 'B4.预设卡组'
D += [
    ('L2', 'B4', 'P2', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'B4.01', 'P2', SC, '', M, '3套预设卡组',
     'module.exports={ rush:["body_disciple","body_disciple",...共8张], control:[...], tank:[...] }',
     'DeckPresets.js', 'js/config/DeckPresets.js', '文件 → Deck.init使用', 'V1', '待办'),
]

# === C.战斗逻辑 ===
SC, SN = 'C', 'C.战斗逻辑'
M = 'C1.战斗主循环'
D += [
    ('L1', '',   '',    SC, SN, '', '', '', '', '', '', '', '', ''),
    ('L2', 'C1', 'P3', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'C1.01', 'P3', SC, '', M, '每帧总入口',
     'this.updateEnergy(dt); this.updateDraw(dt); this.updateUnits(dt); this.updateFormations(dt); this.updateElders(dt); this.updateAI(dt); this.removeDead(); this.checkEnd(); model.time-=dt; model.elapsedTime+=dt',
     'BattleLogic.update(dt)', 'js/game/BattleLogic.js', 'dt:number → void', 'V1', '待办'),
]
M = 'C2.单位实体'
D += [
    ('L2', 'C2', 'P3', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'C2.01', 'P3', SC, '', M, '从卡牌创建单位',
     'let c=Cards.get(cardId); return { id:uid++, cardId, owner, x, y, hp:c.hp, maxHp:c.hp, atk:c.atk, speed:c.speed, baseSpeed:c.speed, attackRange:c.range||1, attackInterval:c.interval, lastAttackTime:0, target:null, state:"walking", traits:c.traits||[], buffs:[], facing:owner===0?1:-1, isElder:c.type==="elder", elderTimer:0 }',
     'Unit.fromCard(cardId,owner,x,y)', 'js/game/Unit.js', 'cardId+owner:int+x+y → Unit Object', 'V1', '待办'),
    ('L3', 'C2.02', 'P3', SC, '', M, 'buff系统',
     'this.buffs.push({type,value,duration,elapsed:0})。类型: "slow"(-0.5速度)/"speed"(+0.5)/"stun"(速度=0)/"shield"(免疫)/"atkBoost"(+攻击)。',
     'Unit.addBuff(type,value,duration)', 'js/game/Unit.js', 'type+value+duration → void', 'V1', '待办'),
    ('L3', 'C2.03', 'P3', SC, '', M, '实际移速计算',
     'let spd=this.baseSpeed; for(let b of this.buffs){ if(b.type==="slow") spd-=0.5; if(b.type==="speed") spd+=0.5; if(b.type==="stun") return 0 } return Math.max(0,spd)',
     'Unit.getEffectiveSpeed()', 'js/game/Unit.js', 'void → number', 'V1', '待办'),
    ('L3', 'C2.04', 'P3', SC, '', M, '受到伤害',
     'if(this.hp<=0) return 0; //已死不处理 this.hp-=amount; if(this.hp<=0){ this.hp=0; this.state="dead"; //触发死亡回调 } return amount',
     'Unit.takeDamage(amount,attacker)', 'js/game/Unit.js', 'amount:int+attacker:Unit → int(实际伤害)', 'V1', '待办'),
]
M = 'C3.移动系统'
D += [
    ('L2', 'C3', 'P3', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'C3.01', 'P3', SC, '', M, '★单位移动★',
     'if(unit.state!=="walking") return; let target=MovementSystem.findTarget(unit,model); if(target && MovementSystem.checkCollision(unit,target)){ unit.state="fighting"; unit.target=target } else { let spd=unit.getEffectiveSpeed(); unit.y+=spd*dt*unit.facing; if(MovementSystem.checkHallReach(unit,model)){ unit.state="dead"; //到大殿→造成伤害 CombatSystem.damageHall(model.players[1-unit.owner],unit.atk) } }',
     'MovementSystem.moveUnit(unit,dt,model)', 'js/game/MovementSystem.js', 'unit+dt+model → void', 'V1', '待办'),
    ('L3', 'C3.02', 'P3', SC, '', M, '★目标选择★',
     'let enemies=model.getEnemyUnits(unit.owner); let formations=model.getEnemyFormations(unit.owner); let candidates=[...enemies,...formations]; //过滤同列(x相同) candidates=candidates.filter(e=>e.x===unit.x); if(candidates.length===0) return null; //按y距离排序 candidates.sort((a,b)=>Math.abs(a.y-unit.y)-Math.abs(b.y-unit.y)); let nearest=candidates[0]; let dist=Math.abs(nearest.y-unit.y); if(dist<=unit.attackRange) return nearest; return null //不在范围内继续走',
     'MovementSystem.findTarget(unit,model)', 'js/game/MovementSystem.js', 'unit+model → Unit|Formation|null', 'V1', '待办'),
    ('L3', 'C3.03', 'P3', SC, '', M, '碰撞检测',
     'let dist=Math.abs(target.y-unit.y); return dist<=unit.attackRange;',
     'MovementSystem.checkCollision(unit,target)', 'js/game/MovementSystem.js', 'unit+target → bool', 'V1', '待办'),
    ('L3', 'C3.04', 'P3', SC, '', M, '到达大殿检测',
     'if(unit.owner===0 && unit.y>=Constants.BOARD_LENGTH-1) return true; if(unit.owner===1 && unit.y<=0) return true; return false',
     'MovementSystem.checkHallReach(unit,model)', 'js/game/MovementSystem.js', 'unit+model → bool', 'V1', '待办'),
]
M = 'C4.战斗系统'
D += [
    ('L2', 'C4', 'P3', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'C4.01', 'P3', SC, '', M, '攻击主逻辑',
     'if(unit.state!=="fighting"||!unit.target) return; let now=model.elapsedTime; if(now-unit.lastAttackTime>=unit.attackInterval){ if(unit.traits.includes("ranged")) CombatSystem.rangedAttack(unit,unit.target); else CombatSystem.meleeAttack(unit,unit.target); unit.lastAttackTime=now; //检查击杀 if(unit.target.hp<=0) CombatSystem.handleKill(unit,unit.target) }',
     'CombatSystem.attack(unit,dt,model)', 'js/game/CombatSystem.js', 'unit+dt+model → void', 'V1', '待办'),
    ('L3', 'C4.02', 'P3', SC, '', M, '近战互扣',
     'target.takeDamage(attacker.atk,attacker); attacker.takeDamage(target.atk,target); //双方同时掉血',
     'CombatSystem.meleeAttack(attacker,target)', 'js/game/CombatSystem.js', 'attacker+target → void', 'V1', '待办'),
    ('L3', 'C4.03', 'P3', SC, '', M, '远程单方',
     'target.takeDamage(attacker.atk); //攻击者不掉血',
     'CombatSystem.rangedAttack(attacker,target)', 'js/game/CombatSystem.js', 'attacker+target → void', 'V1', '待办'),
    ('L3', 'C4.04', 'P4', SC, '', M, '范围伤害',
     'for(let t of targets) t.takeDamage(damage)',
     'CombatSystem.aoeAttack(source,targets,damage)', 'js/game/CombatSystem.js', 'source+targets[]+damage → void', 'V1', '待办'),
    ('L3', 'C4.05', 'P3', SC, '', M, '★击杀后继续推进★',
     'killer.state="walking"; killer.target=null; //不消失，继续走！这是核心修复',
     'CombatSystem.handleKill(killer,victim)', 'js/game/CombatSystem.js', 'killer+victim → void', 'V1', '待办'),
    ('L3', 'C4.06', 'P3', SC, '', M, '大殿受伤',
     'if(player.hallShield>0) return; //金钟罩免疫 player.hallHp-=amount; if(player.hallHp<=0){ player.hallHp=0; model.state="ended"; model.winner=1-player.id }',
     'CombatSystem.damageHall(player,amount,model)', 'js/game/CombatSystem.js', 'player+amount+model → void', 'V1', '待办'),
]
M = 'C5.阵法系统'
D += [
    ('L2', 'C5', 'P4', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'C5.01', 'P4', SC, '', M, '布阵',
     'let key=gx+","+gy; if(player.formationCooldowns[key] && model.elapsedTime<player.formationCooldowns[key]) return false; //冷却中 let f=Formation.fromCard(cardId,player.id,gx,gy); player.formations.push(f); player.spendEnergy(card.cost); deck.playCard(handIndex); return true',
     'BattleLogic.placeFormation(player,cardId,gx,gy,model)', 'js/game/BattleLogic.js', 'player+cardId+gx+gy+model → bool', 'V1', '待办'),
    ('L3', 'C5.02', 'P4', SC, '', M, '阵法攻击',
     'if(!f.isActive) return; //被禁 let enemies=model.getEnemyUnits(f.owner); for(let e of enemies){ if(e.x===f.gridX && Math.abs(e.y-f.gridY)<=f.range){ let now=model.elapsedTime; if(now-f.lastAttackTime>=f.attackInterval){ CombatSystem.rangedAttack(f,e); f.lastAttackTime=now; break } } }',
     'Formation.update(f,dt,model)', 'js/game/Formation.js', 'f+dt+model → void', 'V1', '待办'),
    ('L3', 'C5.03', 'P4', SC, '', M, '阵法冷却',
     'let key=gx+","+gy; return !player.formationCooldowns[key] || model.elapsedTime>=player.formationCooldowns[key]; //true=可布阵',
     'Player.checkCooldown(gx,gy,model)', 'js/game/Player.js', 'gx+gy+model → bool', 'V1', '待办'),
    ('L3', 'C5.04', 'P4', SC, '', M, '阵法被禁',
     'f.isActive=false; f.silenceTimer=duration; //在update中倒数，到0恢复isActive=true',
     'Formation.setSilence(duration)', 'js/game/Formation.js', 'duration:number → void', 'V1', '待办'),
]
M = 'C6.法术系统'
D += [
    ('L2', 'C6', 'P4', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'C6.01', 'P4', SC, '', M, '法术入口',
     'switch(cardId){ case "wan_jian": this.castWanJian(caster,model); break; case "wu_lei": this.castWuLei(target,model); break; ... }',
     'SpellSystem.cast(cardId,caster,target,model)', 'js/game/SpellSystem.js', 'cardId+caster+target+model → void', 'V1', '待办'),
    ('L3', 'C6.02', 'P4', SC, '', M, '万剑归宗',
     'let units=model.players[caster].units; for(let u of units){ u.atk+=1; u.addBuff("speed",0.3,5) }',
     'SpellSystem.castWanJian(casterId,model)', 'js/game/SpellSystem.js', 'casterId:int+model → void', 'V1', '待办'),
    ('L3', 'C6.03', 'P4', SC, '', M, '五雷正法',
     'let enemies=model.getEnemyUnits(caster).concat(model.getEnemyFormations(caster)); let inRange=enemies.filter(e=>Math.abs(e.y-target.y)<=3 && Math.abs(e.x-target.x)<=1); for(let e of inRange) e.takeDamage(4)',
     'SpellSystem.castWuLei(target,casterId,model)', 'js/game/SpellSystem.js', 'target{x,y}+casterId+model → void', 'V1', '待办'),
    ('L3', 'C6.04', 'P4', SC, '', M, '御风诀',
     'target.addBuff("speed",0.5,5)',
     'SpellSystem.castYuFeng(target)', 'js/game/SpellSystem.js', 'target:Unit → void', 'V1', '待办'),
    ('L3', 'C6.05', 'P4', SC, '', M, '镇魂符',
     'target.setSilence(3)',
     'SpellSystem.castZhenHun(target)', 'js/game/SpellSystem.js', 'target:Formation → void', 'V1', '待办'),
    ('L3', 'C6.06', 'P4', SC, '', M, '金钟罩',
     'model.players[casterId].hallShield=3',
     'SpellSystem.castJinZhong(casterId,model)', 'js/game/SpellSystem.js', 'casterId+model → void', 'V1', '待办'),
    ('L3', 'C6.07', 'P4', SC, '', M, '移山倒海',
     'for(let t of targets){ t.y-=2*t.facing; t.takeDamage(1) }',
     'SpellSystem.castYiShan(targets)', 'js/game/SpellSystem.js', 'targets:Unit[] → void', 'V1', '待办'),
    ('L3', 'C6.08', 'P4', SC, '', M, '困仙索',
     'target.addBuff("stun",0,2)',
     'SpellSystem.castKunXian(target)', 'js/game/SpellSystem.js', 'target:Unit → void', 'V1', '待办'),
    ('L3', 'C6.09', 'P4', SC, '', M, '天雷诀',
     'for(let t of targets) t.takeDamage(4)',
     'SpellSystem.castTianLei(targets)', 'js/game/SpellSystem.js', 'targets:Unit[] → void', 'V1', '待办'),
]
M = 'C7.长老技能'
D += [
    ('L2', 'C7', 'P4', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'C7.01', 'P4', SC, '', M, '技能计时',
     'if(!unit.isElder) return; unit.elderTimer+=dt; if(unit.elderTimer>=5.0){ unit.elderTimer=0; ElderSkillSystem.triggerRandom(unit,model) }',
     'ElderSkillSystem.update(unit,dt,model)', 'js/game/ElderSkillSystem.js', 'unit+dt+model → void', 'V1', '待办'),
    ('L3', 'C7.02', 'P4', SC, '', M, '★随机分支★',
     'let branches=["flyingSword","pill","talisman","beast"]; let idx=Math.floor(Math.random()*4); this[branches[idx]](elder,model);',
     'ElderSkillSystem.triggerRandom(elder,model)', 'js/game/ElderSkillSystem.js', 'elder+model → void', 'V1', '待办'),
    ('L3', 'C7.03', 'P4', SC, '', M, '飞剑分支',
     'let enemies=model.getEnemyUnits(elder.owner); let nearby=enemies.filter(e=>Math.abs(e.y-elder.y)<=3); CombatSystem.aoeAttack(elder,nearby,3)',
     'ElderSkillSystem.flyingSword(elder,model)', 'js/game/ElderSkillSystem.js', 'elder+model → void', 'V1', '待办'),
    ('L3', 'C7.04', 'P4', SC, '', M, '丹药分支',
     'elder.hp=Math.min(elder.maxHp,elder.hp+3); let allies=model.players[elder.owner].units; for(let a of allies){ if(a!==elder && Math.abs(a.y-elder.y)<=2) a.hp=Math.min(a.maxHp,a.hp+2) }',
     'ElderSkillSystem.pill(elder,model)', 'js/game/ElderSkillSystem.js', 'elder+model → void', 'V1', '待办'),
    ('L3', 'C7.05', 'P4', SC, '', M, '符箓分支',
     'let enemies=model.getEnemyUnits(elder.owner); let front=enemies.filter(e=>Math.abs(e.y-elder.y)<=3); SpellSystem.castTianLei(front)',
     'ElderSkillSystem.talisman(elder,model)', 'js/game/ElderSkillSystem.js', 'elder+model → void', 'V1', '待办'),
    ('L3', 'C7.06', 'P4', SC, '', M, '御兽分支',
     'let beast=Unit.fromCard("spirit_beast",elder.owner,elder.x,elder.y); model.players[elder.owner].units.push(beast)',
     'ElderSkillSystem.beast(elder,model)', 'js/game/ElderSkillSystem.js', 'elder+model → void', 'V1', '待办'),
]
M = 'C8.灵力系统'
D += [
    ('L2', 'C8', 'P4', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'C8.01', 'P4', SC, '', M, '实时回复',
     'let regenRate=model.state==="overtime"?Constants.ENERGY_REGEN/1.5:Constants.ENERGY_REGEN; player.energyTimer+=dt; if(player.energyTimer>=regenRate){ player.energyTimer-=regenRate; player.energy=Math.min(player.energyMax,player.energy+1) }',
     'Player.updateEnergy(dt,model)', 'js/game/Player.js', 'dt+model → void', 'V1', '待办'),
    ('L3', 'C8.02', 'P4', SC, '', M, '上限增长',
     'let newMax=Math.min(Constants.ENERGY_MAX_CAP, Constants.ENERGY_START+Math.floor(model.elapsedTime/30)); player.energyMax=newMax',
     'Player.updateEnergyMax(model)', 'js/game/Player.js', 'model → void', 'V1', '待办'),
]
M = 'C9.出牌执行'
D += [
    ('L2', 'C9', 'P4', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'C9.01', 'P4', SC, '', M, '出兵',
     'let card=Cards.get(cardId); if(!player.spendEnergy(card.cost)) return null; let spawnY=player.id===0?0:Constants.BOARD_LENGTH-1; let unit=Unit.fromCard(cardId,player.id,x,spawnY); player.units.push(unit); player.deck.playCard(handIndex); return unit',
     'BattleLogic.spawnUnit(player,cardId,x,model)', 'js/game/BattleLogic.js', 'player+cardId+x+model → Unit', 'V1', '待办'),
    ('L3', 'C9.02', 'P4', SC, '', M, '施法',
     'let card=Cards.get(cardId); if(!player.spendEnergy(card.cost)) return; SpellSystem.cast(cardId,player.id,target,model); player.deck.playCard(handIndex)',
     'BattleLogic.castSpell(player,cardId,target,model)', 'js/game/BattleLogic.js', 'player+cardId+target+model → void', 'V1', '待办'),
]
M = 'C10.胜负判定'
D += [
    ('L2', 'C10', 'P4', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'C10.01', 'P4', SC, '', M, '大殿摧毁',
     'if(model.players[0].hallHp<=0) return 1; if(model.players[1].hallHp<=0) return 0; return null',
     'BattleChecker.checkHall(model)', 'js/game/BattleChecker.js', 'model → 0|1|null', 'V1', '待办'),
    ('L3', 'C10.02', 'P4', SC, '', M, '时限检测',
     'if(model.time>0) return null; if(model.players[0].hallHp!==model.players[1].hallHp) return model.players[0].hallHp>model.players[1].hallHp?0:1; model.state="overtime"; model.time=60; return null',
     'BattleChecker.checkTime(model)', 'js/game/BattleChecker.js', 'model → 0|1|null', 'V1', '待办'),
    ('L3', 'C10.03', 'P4', SC, '', M, '加时结算',
     'if(model.time>0) return null; return model.players[0].energy>model.players[1].energy?0:1',
     'BattleChecker.checkOvertime(model)', 'js/game/BattleChecker.js', 'model → 0|1', 'V1', '待办'),
]
M = 'C11.死亡清理'
D += [
    ('L2', 'C11', 'P4', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'C11.01', 'P4', SC, '', M, '清理死亡单位',
     'for(let p of model.players){ p.units=p.units.filter(u=>{ if(u.state==="dead"){ if(u.traits.includes("kamikaze")) CombatSystem.kamikaze(u,model); return false } return true }) }',
     'BattleLogic.removeDead(model)', 'js/game/BattleLogic.js', 'model → void', 'V1', '待办'),
    ('L3', 'C11.02', 'P4', SC, '', M, '清理被毁阵法',
     'for(let p of model.players){ p.formations=p.formations.filter(f=>{ if(f.hp<=0){ let key=f.gridX+","+f.gridY; p.formationCooldowns[key]=model.elapsedTime+8; return false } return true }) }',
     'BattleLogic.removeDeadFormations(model)', 'js/game/BattleLogic.js', 'model → void', 'V1', '待办'),
]

# === D.AI系统 ===
SC, SN = 'D', 'D.AI系统'
M = 'D1.AI决策'
D += [
    ('L1', '',   '',    SC, SN, '', '', '', '', '', '', '', '', ''),
    ('L2', 'D1', 'P5', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'D1.01', 'P5', SC, '', M, 'AI主循环',
     'this.thinkTimer+=dt; if(this.thinkTimer>=this.thinkInterval){ this.thinkTimer=0; this.think(model) }',
     'AI.update(dt,model)', 'js/game/AI.js', 'dt+model → void', 'V1', '待办'),
    ('L3', 'D1.02', 'P5', SC, '', M, '★决策核心★',
     'let player=model.players[this.playerId]; let ratio=this.decideAttackRatio(model); let cardIdx=this.pickCard(player.energy,ratio,player.deck.hand); if(cardIdx===null) return; let cardId=player.deck.hand[cardIdx]; let card=Cards.get(cardId); if(card.type==="unit") BattleLogic.spawnUnit(player,cardId,0,model); else if(card.type==="formation"){ let pos=this.pickFormationPos(model); BattleLogic.placeFormation(player,cardId,pos.gx,pos.gy,model) } else BattleLogic.castSpell(player,cardId,null,model)',
     'AI.think(model)', 'js/game/AI.js', 'model → void', 'V1', '待办'),
    ('L3', 'D1.03', 'P5', SC, '', M, '攻守比计算',
     'let hp=model.players[this.playerId].hallHp; let maxHp=Constants.HALL_HP; let pct=hp/maxHp; if(pct>0.6) return 0.7; if(pct<0.3) return 0.2; return 0.4',
     'AI.decideAttackRatio(model)', 'js/game/AI.js', 'model → number(0~1)', 'V1', '待办'),
    ('L3', 'D1.04', 'P5', SC, '', M, '选牌逻辑',
     'let playable=hand.map((id,idx)=>({id,idx,card:Cards.get(id)})).filter(h=>h.card.cost<=energy); if(playable.length===0) return null; if(ratio>0.5) playable.sort((a,b)=>(b.card.type==="unit")-(a.card.type==="unit")); else playable.sort((a,b)=>(b.card.type==="formation")-(a.card.type==="formation")); return playable[0].idx',
     'AI.pickCard(energy,ratio,hand)', 'js/game/AI.js', 'energy+ratio+hand → int|null', 'V1', '待办'),
    ('L3', 'D1.05', 'P5', SC, '', M, '布阵位置',
     'if(this.difficulty==="easy") return {gx:Math.random()<0.5?-1:1, gy:1+Math.floor(Math.random()*7)}; if(this.difficulty==="normal") return {gx:Math.random()<0.5?-1:1, gy:Constants.BOARD_LENGTH-3+Math.floor(Math.random()*2)}; //hard: 找最快敌方单位前方布 let fastest=model.players[0].units.sort((a,b)=>b.y-a.y)[0]; return {gx:Math.random()<0.5?-1:1, gy:Math.max(1,fastest.y+1)}',
     'AI.pickFormationPos(model)', 'js/game/AI.js', 'model → {gx,gy}', 'V1', '待办'),
]

# === E.渲染与UI ===
SC, SN = 'E', 'E.渲染与UI'
M = 'E1.战斗场景'
D += [
    ('L1', '',   '',    SC, SN, '', '', '', '', '', '', '', '', ''),
    ('L2', 'E1', 'P6', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'E1.01', 'P6', SC, '', M, '场景进入',
     'this.model=GameModel.init(deck0,deck1); this.battleLogic=new BattleLogic(this.model); this.registerInput(); ResourceLoader.load(...)',
     'BattleScene.onEnter()', 'js/scenes/BattleScene.js', 'void → void', 'V1', '待办'),
    ('L3', 'E1.02', 'P6', SC, '', M, '场景更新',
     'this.battleLogic.update(dt)',
     'BattleScene.onUpdate(dt)', 'js/scenes/BattleScene.js', 'dt → void', 'V1', '待办'),
    ('L3', 'E1.03', 'P6', SC, '', M, '场景渲染(分层)',
     'this.renderBackground(ctx); this.renderHalls(ctx); this.renderFormations(ctx); this.renderUnits(ctx); this.particleSystem.render(ctx); this.renderUI(ctx)',
     'BattleScene.onRender(ctx)', 'js/scenes/BattleScene.js', 'ctx → void', 'V1', '待办'),
    ('L3', 'E1.04', 'P6', SC, '', M, '场景退出',
     'input.clear(); this.model=null; this.battleLogic=null',
     'BattleScene.onExit()', 'js/scenes/BattleScene.js', 'void → void', 'V1', '待办'),
]
M = 'E2.背景渲染'
D += [
    ('L2', 'E2', 'P6', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'E2.01', 'P6', SC, '', M, '山道背景',
     'V1: let grad=ctx.createLinearGradient(0,0,0,H); grad.addColorStop(0,"#1a3a2a"); grad.addColorStop(1,"#2d5a3d"); ctx.fillStyle=grad; ctx.fillRect(0,0,W,H)',
     'BattleScene.renderBackground(ctx)', 'js/scenes/BattleScene.js', 'ctx → void', 'V1', '待办'),
    ('L3', 'E2.02', 'P6', SC, '', M, '棋盘格子线',
     'let cellW=W/3; for(let i=0;i<=BOARD_LENGTH;i++){ let y=i*cellH; ctx.strokeStyle="rgba(255,255,255,0.1)"; ctx.beginPath(); ctx.moveTo(0,y); ctx.lineTo(W,y); ctx.stroke() }',
     'BattleScene.renderGrid(ctx)', 'js/scenes/BattleScene.js', 'ctx → void', 'V1', '待办'),
]
M = 'E3.大殿渲染'
D += [
    ('L2', 'E3', 'P6', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'E3.01', 'P6', SC, '', M, '大殿色块',
     'Renderer.rect(ctx,0,0,W,60,"#F39C12"); //顶部 Renderer.rect(ctx,0,H-60,W,60,"#F39C12"); //底部 Renderer.text(ctx,"敌方宗门",W/2,30,"#FFF",16); Renderer.text(ctx,"我方宗门",W/2,H-30,"#FFF",16)',
     'BattleScene.renderHalls(ctx)', 'js/scenes/BattleScene.js', 'ctx → void', 'V1', '待办'),
    ('L3', 'E3.02', 'P6', SC, '', M, '大殿血条',
     'let p0=model.players[0]; Renderer.healthBar(ctx,0,H-70,W,8,p0.hallHp/p0.hallMaxHp); let p1=model.players[1]; Renderer.healthBar(ctx,0,62,W,8,p1.hallHp/p1.hallMaxHp)',
     'BattleScene.renderHallHp(ctx)', 'js/scenes/BattleScene.js', 'ctx → void', 'V1', '待办'),
    ('L3', 'E3.03', 'P6', SC, '', M, '受击特效',
     'if(this.shakeTimer>0){ this.shakeTimer-=dt; ctx.translate((Math.random()-0.5)*6,(Math.random()-0.5)*6) }',
     'BattleScene.renderHallHit(ctx,dt)', 'js/scenes/BattleScene.js', 'ctx+dt → void', 'V1', '待办'),
]
M = 'E4.单位渲染'
D += [
    ('L2', 'E4', 'P6', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'E4.01', 'P6', SC, '', M, '单位色块+y排序',
     'let all=model.getAllUnits(); all.sort((a,b)=>a.y-b.y); for(let u of all){ let color=u.owner===0?"#E74C3C":"#3498DB"; Renderer.rect(ctx,u.px-u.w/2,u.py-u.h/2,u.w,u.h,color); Renderer.text(ctx,Cards.get(u.cardId).name[0],u.px,u.py,"#FFF",10) }',
     'BattleScene.renderUnits(ctx)', 'js/scenes/BattleScene.js', 'ctx → void', 'V1', '待办'),
    ('L3', 'E4.02', 'P6', SC, '', M, '单位血条',
     'for(let u of all){ Renderer.healthBar(ctx,u.px-u.w/2,u.py-u.h/2-6,u.w,4,u.hp/u.maxHp) }',
     'BattleScene.renderUnitHp(ctx)', 'js/scenes/BattleScene.js', 'ctx → void', 'V1', '待办'),
]
M = 'E5.阵法渲染'
D += [
    ('L2', 'E5', 'P6', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'E5.01', 'P6', SC, '', M, '阵法光阵',
     'for(let p of model.players) for(let f of p.formations){ let color=f.isActive?"rgba(26,188,156,0.4)":"rgba(100,100,100,0.3)"; Renderer.rect(ctx,f.px,f.py,cellW,cellH,color); Renderer.text(ctx,Cards.get(f.cardId).name,f.px+cellW/2,f.py+cellH/2,"#FFF",10) }',
     'BattleScene.renderFormations(ctx)', 'js/scenes/BattleScene.js', 'ctx → void', 'V1', '待办'),
]
M = 'E6.UI-手牌栏'
D += [
    ('L2', 'E6', 'P6', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'E6.01', 'P6', SC, '', M, '手牌显示',
     'let hand=player.deck.hand; let cardW=(W-40)/4; for(let i=0;i<hand.length;i++){ let card=Cards.get(hand[i]); let x=20+i*cardW; let y=H-80; if(i===this.selectedCard) y-=20; let canPlay=player.energy>=card.cost; Renderer.rect(ctx,x,y,cardW-10,70,canPlay?"#FFF":"#999"); Renderer.text(ctx,card.name,x+cardW/2,y+20,"#333",12); Renderer.text(ctx,"费"+card.cost,x+cardW/2,y+50,"#B7950B",14) }',
     'HandBar.render(ctx)', 'js/ui/HandBar.js', 'ctx → void', 'V1', '待办'),
    ('L3', 'E6.02', 'P6', SC, '', M, '点击选中',
     'this.selectedCard = (this.selectedCard===index) ? -1 : index',
     'HandBar.onTap(index)', 'js/ui/HandBar.js', 'index:int → void', 'V1', '待办'),
]
M = 'E7.UI-灵力条'
D += [
    ('L2', 'E7', 'P6', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'E7.01', 'P6', SC, '', M, '灵力显示',
     'let p=model.players[0]; Renderer.rect(ctx,20,H-90,W-40,8,"#333"); Renderer.rect(ctx,20,H-90,(W-40)*(p.energy/p.energyMax),8,"#9B59B6"); Renderer.text(ctx,p.energy+"/"+p.energyMax,W/2,H-95,"#FFF",12)',
     'EnergyBar.render(ctx)', 'js/ui/EnergyBar.js', 'ctx → void', 'V1', '待办'),
]
M = 'E8.UI-顶部HUD'
D += [
    ('L2', 'E8', 'P6', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'E8.01', 'P6', SC, '', M, '双方血条',
     'let p0=model.players[0]; let p1=model.players[1]; Renderer.rect(ctx,0,0,W/2,6,"#E74C3C"); Renderer.rect(ctx,0,0,(W/2)*(p0.hallHp/p0.hallMaxHp),6,"#27AE60"); Renderer.rect(ctx,W/2,0,W/2,6,"#E74C3C"); Renderer.rect(ctx,W/2+(W/2)*(1-p1.hallHp/p1.hallMaxHp),0,(W/2)*(p1.hallHp/p1.hallMaxHp),6,"#27AE60")',
     'HUD.renderHallBars(ctx)', 'js/ui/HUD.js', 'ctx → void', 'V1', '待办'),
    ('L3', 'E8.02', 'P6', SC, '', M, '计时器',
     'let t=Math.ceil(model.time); let color=t<=30?"#E74C3C":"#FFF"; Renderer.text(ctx,t+"s",W/2,20,color,20)',
     'HUD.renderTimer(ctx)', 'js/ui/HUD.js', 'ctx → void', 'V1', '待办'),
]
M = 'E9.出牌交互'
D += [
    ('L2', 'E9', 'P6', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'E9.01', 'P6', SC, '', M, '手牌→选目标',
     'if(this.selectedCard===index){ this.selectedCard=-1; return } let card=Cards.get(player.deck.hand[index]); if(player.energy<card.cost){ //提示灵力不足 return } this.selectedCard=index',
     'BattleScene.onCardTap(index)', 'js/scenes/BattleScene.js', 'index → void', 'V1', '待办'),
    ('L3', 'E9.02', 'P6', SC, '', M, '格子点击执行',
     'if(this.selectedCard===-1) return; let card=Cards.get(player.deck.hand[this.selectedCard]); if(card.type==="unit") this.battleLogic.spawnUnit(player,card.cardId,0,model); else if(card.type==="formation") this.battleLogic.placeFormation(player,card.cardId,gx,gy,model); else this.battleLogic.castSpell(player,card.cardId,{x:gx,y:gy},model); this.selectedCard=-1',
     'BattleScene.onGridTap(gx,gy)', 'js/scenes/BattleScene.js', 'gx+gy → void', 'V1', '待办'),
]
M = 'E10.结算场景'
D += [
    ('L2', 'E10', 'P6', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'E10.01', 'P6', SC, '', M, '胜负展示',
     'let text=model.winner===0?"胜利":"失败"; let color=model.winner===0?"#27AE60":"#E74C3C"; Renderer.text(ctx,text,W/2,H/2,color,48)',
     'ResultScene.renderResult(ctx)', 'js/scenes/ResultScene.js', 'ctx → void', 'V1', '待办'),
    ('L3', 'E10.02', 'P6', SC, '', M, '摧毁度',
     'let r0=Math.round((1-model.players[1].hallHp/30)*100); let r1=Math.round((1-model.players[0].hallHp/30)*100); Renderer.text(ctx,"我方"+r0+"% vs 敌方"+r1+"%",W/2,H/2+60,"#FFF",16)',
     'ResultScene.renderDestroyRate(ctx)', 'js/scenes/ResultScene.js', 'ctx → void', 'V1', '待办'),
    ('L3', 'E10.03', 'P6', SC, '', M, '再来一局',
     'SceneManager.switch("battle")',
     'ResultScene.onReplay()', 'js/scenes/ResultScene.js', 'void → void', 'V1', '待办'),
]
M = 'E11.新手引导'
D += [
    ('L2', 'E11', 'P6', SC, '', M, '', '', '', '', '', 'V1', ''),
    ('L3', 'E11.01', 'P6', SC, '', M, '引导1-出牌',
     '高亮第一张手牌(画黄色边框)+画向下箭头+显示文字"点击出兵"→监听到手牌点击后step=2',
     'TutorialGuide.step1(ctx)', 'js/ui/TutorialGuide.js', 'ctx → void', 'V1', '待办'),
    ('L3', 'E11.02', 'P6', SC, '', M, '引导2-灵力',
     '高亮灵力条+文字"灵力不够时等待回复"→等3秒后step=3',
     'TutorialGuide.step2(ctx)', 'js/ui/TutorialGuide.js', 'ctx → void', 'V1', '待办'),
    ('L3', 'E11.03', 'P6', SC, '', M, '引导3-布阵',
     '高亮阵法区格子+文字"点击布阵拦截"→监听到布阵后step=4',
     'TutorialGuide.step3(ctx)', 'js/ui/TutorialGuide.js', 'ctx → void', 'V1', '待办'),
    ('L3', 'E11.04', 'P6', SC, '', M, '引导4-目标',
     '画指向敌方大殿的箭头+文字"摧毁大殿获胜"→2秒后step=0(引导结束)',
     'TutorialGuide.step4(ctx)', 'js/ui/TutorialGuide.js', 'ctx → void', 'V1', '待办'),
]

# === F/G/H (V1.5/V2) ===
SC, SN = 'F', 'F.社交系统(V1.5)'
D += [
    ('L1', '',   '',    SC, SN, '', '', '', '', '', '', '', '', ''),
    ('L3', 'F1.01', 'P9', SC, '', 'F1.登录', '微信登录', 'wx.login→code→wx.cloud.callFunction("login",{code})→返回openid→存wx.setStorageSync("openid",openid)', 'CloudManager.login()', 'js/social/CloudManager.js', 'void → openid', 'V1.5', '待办'),
    ('L3', 'F1.02', 'P10', SC, '', 'F1.存档', '云存档', 'wx.cloud.database().collection("users").doc(openid).set({data})', 'CloudManager.save(data)', 'js/social/CloudManager.js', 'data → void', 'V1.5', '待办'),
    ('L3', 'F2.01', 'P11', SC, '', 'F2.异步PvP', '布阵上传', 'wx.cloud.database().collection("challenges").add({data:{layout,openid}})→返回_id', 'AsyncPvP.uploadLayout(layout)', 'js/social/AsyncPvP.js', 'layout → challengeId', 'V1.5', '待办'),
    ('L3', 'F2.02', 'P11', SC, '', 'F2.异步PvP', '加载对手', 'wx.cloud.database().collection("challenges").doc(id).get()→返回layout→本地AI模拟攻打', 'AsyncPvP.loadOpponent(id)', 'js/social/AsyncPvP.js', 'id → layout', 'V1.5', '待办'),
    ('L3', 'F3.01', 'P12', SC, '', 'F3.排行分享', '好友排行', 'wx.getOpenDataContext().postMessage({action:"rank"})→开放数据域绘制排行', 'RankManager.renderFriendRank()', 'js/social/RankManager.js', 'void → void', 'V1.5', '待办'),
    ('L3', 'F3.02', 'P12', SC, '', 'F3.排行分享', '分享', 'wx.shareAppMessage({title,imageUrl})', 'ShareManager.share(title,img)', 'js/social/ShareManager.js', 'title+img → void', 'V1.5', '待办'),
]
SC, SN = 'G', 'G.养成变现(V1.5)'
D += [
    ('L1', '',   '',    SC, SN, '', '', '', '', '', '', '', '', ''),
    ('L3', 'G1.01', 'P13', SC, '', 'G1.养成', '弟子升级', 'unit.level++; unit.hp=baseHp*(1+level*0.1); unit.atk=baseAtk*(1+level*0.08)', 'ProgressionSystem.levelUp(unitId)', 'js/game/ProgressionSystem.js', 'unitId → void', 'V1.5', '待办'),
    ('L3', 'G1.02', 'P13', SC, '', 'G1.养成', '弟子升星', 'unit.star++; //练气→筑基→金丹→元婴 解锁更高上限+新技能', 'ProgressionSystem.starUp(unitId)', 'js/game/ProgressionSystem.js', 'unitId → void', 'V1.5', '待办'),
    ('L3', 'G2.01', 'P14', SC, '', 'G2.变现', '激励广告', 'let ad=wx.createRewardedVideoAd({adUnitId}); ad.show(); ad.onClose(res=>{ if(res.isEnded) cb() })', 'AdManager.showReward(cb)', 'js/monetize/AdManager.js', 'cb → void', 'V1.5', '待办'),
    ('L3', 'G2.02', 'P14', SC, '', 'G2.变现', '内购', 'wx.requestPayment({timeStamp,nonceStr,package,signType,paySign,success:cb})', 'IAPManager.buy(productId,cb)', 'js/monetize/IAPManager.js', 'productId+cb → void', 'V1.5', '待办'),
]
SC, SN = 'H', 'H.实时PvP(V2)'
D += [
    ('L1', '',   '',    SC, SN, '', '', '', '', '', '', '', '', ''),
    ('L3', 'H1.01', 'P16', SC, '', 'H1.网络', 'WebSocket', 'this.ws=wx.connectSocket({url}); this.ws.onOpen(cb); this.ws.onMessage(cb); this.ws.onClose(cb); setInterval(()=>this.ws.send({type:"ping"}),10000)', 'NetworkClient.connect(url)', 'js/net/NetworkClient.js', 'url → void', 'V2', '待办'),
    ('L3', 'H1.02', 'P16', SC, '', 'H1.网络', '断线重连', 'onClose时保存model状态→wx.connectSocket重连→send({type:"reconnect",roomId})→服务端恢复', 'NetworkClient.reconnect()', 'js/net/NetworkClient.js', 'void → void', 'V2', '待办'),
    ('L3', 'H2.01', 'P17', SC, '', 'H2.同步', '操作上报', 'this.ws.send(JSON.stringify({type:"action",action:{cardId,target,timestamp}}))', 'SyncManager.sendAction(action)', 'js/net/SyncManager.js', 'action → void', 'V2', '待办'),
    ('L3', 'H2.02', 'P17', SC, '', 'H2.同步', '状态接收', 'onMessage→JSON.parse→if(type==="state") model.applyState(state)', 'SyncManager.onState(state)', 'js/net/SyncManager.js', 'state → void', 'V2', '待办'),
    ('L3', 'H3.01', 'P18', SC, '', 'H3.匹配', '匹配系统', 'wx.cloud.callFunction("match",{elo})→返回{roomId,opponentElo}', 'MatchMaker.match(elo)', 'js/net/MatchMaker.js', 'elo → roomId', 'V2', '待办'),
    ('L3', 'H3.02', 'P18', SC, '', 'H3.匹配', '段位赛', 'if(win) elo+=20; else elo-=15; rank=eloToRank(elo); //青铜→宗师', 'RankSystem.updateRank(win)', 'js/net/RankSystem.js', 'win:bool → void', 'V2', '待办'),
]

# ================================================================
# Sheet 0: 开发阶段规划
# ================================================================
ws = wb.active
ws.title = '0.开发阶段规划'
ws.sheet_view.showGridLines = False
widths = [4, 6, 16, 26, 30, 36, 36, 22, 8]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

big_title(ws, '开发阶段规划 — 先做什么后做什么，每步目的是什么、能实现什么', 9)
ws.merge_cells('A2:I2')
c = ws.cell(row=2, column=1, value='按依赖关系排出8个阶段，必须从P1到P8顺序做。每个阶段做完都有"能跑起来看到的东西"。')
c.font = F_SMALL; c.alignment = A_L; ws.row_dimensions[2].height = 24

headers = ['', '阶段', '阶段名称', '做什么（内容）', '目的（为什么先做这个）', '能实现的进展（做完能看到什么）', '验证标准（怎么算做完）', '对应主表编号', '周期']
header_row(ws, headers, 3)

phases = [
    {'num':'P1','name':'引擎骨架','content':'渲染循环(Director)\n输入系统(Input)\n事件总线(EventSystem)\n场景管理(SceneManager)\n渲染器(Renderer)','purpose':'这是整个游戏的地基。没有渲染循环什么也跑不起来，没有输入系统什么也点不了。必须先搭好引擎框架，后面所有功能才能挂上去。','progress':'能看到一个黑色画布在60fps跑\n能在画布上画矩形/文字\n点屏幕能触发回调\n能切换空场景A和空场景B','verify':'①Director.loop 60fps跑通\n②点击屏幕console打印坐标\n③画矩形和文字能显示\n④场景A能切换到场景B','ref':'A1~A5','period':'1周'},
    {'num':'P2','name':'数据层','content':'常量配置(Constants)\n卡牌数据表(Cards)\n卡组手牌(Deck)\n预设卡组(DeckPresets)','purpose':'引擎能跑了，但游戏需要数据。卡牌属性、灵力参数、棋盘尺寸等所有数值必须先配置好，后面的战斗逻辑才能读数据创建单位。','progress':'19张卡牌数据完整可查\n8张卡组能洗牌+抽牌\n能console打印手牌和卡牌属性\n改Cards.js数值能立即生效','verify':'①Cards.get(id)返回完整属性\n②Deck.init后hand有3张牌\n③Deck.draw能抽牌\n④canPlay能正确判断灵力','ref':'B1~B4','period':'1周'},
    {'num':'P3','name':'核心战斗逻辑','content':'战斗主循环(BattleLogic)\n单位实体(Unit)\n移动系统(Movement)\n战斗系统(Combat)\n★击杀后继续推进★','purpose':'这是整个游戏好不好玩的命门。单位能不能走、能不能打、打完能不能继续冲——这3件事决定了游戏核心循环是否成立。必须验证"攻方打得动"。','progress':'★最重要的里程碑★\nAI vs AI模拟能跑通\n能看到单位在棋盘上移动\n能看到单位互相打\n能看到单位杀掉敌人后继续冲\n能看到单位到大殿造成伤害','verify':'①单位y坐标每帧变化(在走)\n②两个单位接触后互扣血\n③单位击杀后state=walking继续走\n④单位到大殿后大殿血量减少','ref':'C1~C4','period':'2周'},
    {'num':'P4','name':'扩展战斗机制','content':'阵法系统(布阵/攻击/冷却)\n法术系统(8种法术效果)\n长老技能(4分支随机)\n灵力系统(实时回复)\n出牌执行\n胜负判定\n死亡清理','purpose':'核心循环跑通了，但只有"走和打"太单调。阵法让守方能防，法术增加策略深度，长老增加变化性，灵力系统驱动出牌节奏。这些让游戏从"能跑"变成"好玩"。','progress':'完整单局能跑通\n能布阵拦截(阵法有血量会攻击)\n能放法术(8种效果各不相同)\n长老每5秒随机放技能\n灵力实时回复能出牌\n能判定胜负(大殿破/超时)','verify':'①布阵后阵法能攻击经过敌人\n②阵法被毁后格子8秒冷却\n③8种法术效果各自正确生效\n④长老每5秒释放1个随机分支\n⑤灵力每2.8秒+1\n⑥大殿血量归零能判定胜负','ref':'C5~C11','period':'2周'},
    {'num':'P5','name':'AI对手','content':'AI决策(灵力管理/出牌/布阵)\n3档难度(简单/普通/困难)','purpose':'战斗逻辑完整了，但需要对手。没有AI玩家无法测试游戏好不好玩。AI不需要完美，但要能模拟真人节奏出牌布阵。','progress':'能和AI完整打一局(纯逻辑)\nAI会出兵也会布阵\nAI会根据血量调整攻守\n简单AI弱/困难AI强有明显差异','verify':'①AI每2.5秒(普通)决策一次\n②AI会根据大殿血量调攻守比\n③简单/普通/困难行为有差异\n④AI vs AI能完整跑完一局不卡死','ref':'D1','period':'1周'},
    {'num':'P6','name':'渲染与交互','content':'战斗场景(分层渲染)\n背景/大殿/单位/阵法渲染\n手牌栏/灵力条/HUD\n出牌交互(选牌→选目标)\n结算场景\n新手引导\n粒子特效\n音频','purpose':'到这一步为止游戏都是"纯逻辑"。P6把所有逻辑变成"能看到能操作的画面"。这是从"程序员眼里的游戏"变成"玩家眼里的游戏"的关键。','progress':'★玩家可见里程碑★\n能在微信开发者工具里看到画面\n能看到棋盘/大殿/单位/阵法\n能点手牌出牌、点格子布阵\n能看到灵力条/血条/计时器\n战斗结束有结算页\n新手引导能走通4步','verify':'①真机/开发者工具能显示画面\n②点手牌→选目标→出牌流程通\n③单位移动/交战有视觉反馈\n④大殿受击有震动+闪烁\n⑤结算页显示胜负+摧毁度\n⑥新手引导4步走通','ref':'E1~E11, A6~A7','period':'2周'},
    {'num':'P7','name':'测试与平衡调优','content':'纯逻辑模拟器(AI vs AI 10局)\n平衡性调优(参数调整)\n性能测试(60fps/内存)\n真机测试\nBug修复','purpose':'游戏能玩了，但好不好玩是另一回事。P7用模拟器跑大量对局验证平衡——攻方能不能打进去、守方能不能防下来、有没有0%死局。','progress':'★平衡达标里程碑★\n模拟10局输出摧毁度/胜率数据\n参数调到双方摧毁度20~70%\n真机60fps不卡\n无崩溃无死锁','verify':'①模拟10局: 摧毁度20~70%\n②胜率40~60%\n③0%摧毁度对局<10%\n④90%对局150~210秒结束\n⑤真机60fps同屏20单位不卡','ref':'全部联调','period':'1.5周'},
    {'num':'P8','name':'提交审核上线','content':'最终Bug修复\n微信小游戏审核提交\n审核反馈处理\n正式上线','purpose':'游戏做完要上线才能让玩家玩到。修仙题材过审风险低但仍需确保无违规内容。','progress':'★V1上线里程碑★\n微信审核通过\n玩家能搜索到并玩到游戏\nV1 MVP完成','verify':'①微信审核通过\n②玩家能正常游玩\n③无严重Bug\n④新手引导完成率>80%','ref':'—','period':'1周'},
]

r = 4
for i, p in enumerate(phases):
    color = PHASE_COLORS[p['num']]
    c = ws.cell(row=r, column=2, value=p['num']); c.font = F_PHASE; c.fill = PatternFill('solid', fgColor=color); c.alignment = A_C; c.border = BD
    c = ws.cell(row=r, column=3, value=p['name']); c.font = F_PHASE; c.fill = PatternFill('solid', fgColor=color); c.alignment = A_C; c.border = BD
    c = ws.cell(row=r, column=4, value=p['content']); c.font = F_CELL; c.alignment = A_LT; c.border = BD
    c = ws.cell(row=r, column=5, value=p['purpose']); c.font = F_CELL; c.alignment = A_LT; c.border = BD
    c = ws.cell(row=r, column=6, value=p['progress']); c.font = F_RESULT; c.alignment = A_LT; c.border = BD; c.fill = PatternFill('solid', fgColor='E8F8F5')
    c = ws.cell(row=r, column=7, value=p['verify']); c.font = F_CODE; c.alignment = A_LT; c.border = BD
    c = ws.cell(row=r, column=8, value=p['ref']); c.font = F_CODE; c.alignment = A_LT; c.border = BD; c.fill = PatternFill('solid', fgColor='FDF2E9')
    c = ws.cell(row=r, column=9, value=p['period']); c.font = F_BOLD; c.alignment = A_C; c.border = BD
    ws.row_dimensions[r].height = 120; r += 1

r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
c = ws.cell(row=2, column=1)
c = ws.cell(row=r, column=2, value='P1→P2→P3→P4→P5→P6→P7→P8 （必须按顺序，后面依赖前面）')
c.font = Font(name='微软雅黑', size=11, bold=True, color=C_GOLD); c.alignment = A_C; c.fill = PatternFill('solid', fgColor='FEF9E7'); c.border = BD
ws.row_dimensions[r].height = 30

# ================================================================
# Sheet 1: 设计思路与玩法
# ================================================================
ws = wb.create_sheet('1.设计思路与玩法')
ws.sheet_view.showGridLines = False
set_w(ws, [4, 20, 20, 20, 20, 20, 4])
big_title(ws, '《宗门论道》 — 游戏设计思路与玩法', 7)
r = 3
sub_title(ws, '一、游戏概述', 7, r, C_GOLD); r += 1
overview = [
    ('游戏名称','宗门论道（暂定）'),('游戏类型','实时策略卡牌（RTS-lite + 卡牌费用制）'),
    ('游戏题材','东方修仙 / 宗门大战'),('运行平台','微信小游戏（Canvas 2D 渲染）'),
    ('单局时长','3~5分钟（180秒正赛+60秒加时）'),('目标用户','18~35岁，喜欢轻度策略、修仙题材、碎片化对战的玩家'),
    ('一句话简介','皇室战争式的实时卡牌攻防，修仙皮——你与对手各据山门一座，实时派弟子出征、布阵拦截、长老施法，摧毁对方宗门大殿者胜。'),
]
for k, v in overview:
    ws.cell(row=r, column=2, value=k).font = F_L2; ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=L2_FILL); ws.cell(row=r, column=2).alignment = A_C; ws.cell(row=r, column=2).border = BD
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.cell(row=r, column=3, value=v).font = F_CELL; ws.cell(row=r, column=3).alignment = A_LM; ws.cell(row=r, column=3).border = BD
    ws.row_dimensions[r].height = 28; r += 1
r += 1
sub_title(ws, '二、核心玩法 — 实时双线互推', 7, r, C_BLUE); r += 1
play_lines = [
    '你和对手各据一座宗门大殿（30血），实时同时出兵互攻，谁先拆掉对方大殿谁赢，3分钟时限。','',
    '【你做什么】','  1. 出兵：花灵力出弟子，单位自动沿山道推进，遇敌自动交战，杀完继续冲',
    '  2. 布阵：在主路两侧阵法区布阵法（不动，拦截+攻击经过的敌人），剑阵双方通用',
    '  3. 施法：放法术（五雷正法范围炸、万剑归宗全队加速、金钟罩大殿免疫等）',
    '  4. 出长老：花6费出金丹长老，每5秒随机释放飞剑/丹药/符箓/御兽分支技能','',
    '【灵力系统】','  实时回复：每2.8秒+1，开局上限5随时间涨到10。灵力不够出不了牌。','',
    '【地图视角】折中方案：逻辑上下直推 + 视觉山道攻山 + 两侧阵法区',
]
for line in play_lines:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(row=r, column=2, value=line)
    c.font = F_GOLD if line.startswith('【') else F_CELL; c.alignment = A_LM; c.border = BD
    ws.row_dimensions[r].height = 24 if line else 8; r += 1

r += 1
sub_title(ws, '三、设计思路 — 为什么这么设计', 7, r, C_PURPLE); r += 1
thoughts = [
    ('为什么做实时不做回合制？','回合制等对方出招很无聊。实时制双方同时操作互不等待，每秒都有决策压力，紧张感是核心体验。CR验证了这一点。'),
    ('为什么"击杀后继续推进"？','★最核心的设计修复★。旧版"遇阵即停打完即没"导致攻方永远过不去，0%摧毁度。改成击杀后不消失继续冲，攻方才有突破感。'),
    ('为什么阵法有8秒冷却？','防止守方无限补阵堵路。阵被破后该格8秒不能再布，攻方有突破窗口。创造"攻守博弈"节奏。'),
    ('为什么用折中视角？','纯斜45度等距视角美术成本3~4倍，碰撞复杂。折中方案：逻辑上下直推（实现简单），视觉做成山道攻山（保留修仙味）。'),
    ('为什么金丹长老是随机技能？','固定技能会变成"最优解"。随机性增加不可预测性，每次出长老都有惊喜/惊吓，增加对局变化和观赏性。'),
    ('为什么剑阵双方通用？','阵法只能防守时攻方缺乏掩护。双方通用后攻方可铺阵掩护推进，守方布阵拦截，策略空间更大。'),
    ('为什么选修仙题材？','①国内过审零风险②修仙认知度高③"攻山"主题与推塔玩法天然契合④市场有验证。现代战争过审风险极高。'),
    ('为什么V1先做3个流派？','MVP原则：用最少的验证核心玩法。剑修+符修+阵修覆盖攻防基本循环，验证好玩后再加更多。'),
]
for q, a in thoughts:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
    ws.cell(row=r, column=2, value='问').font = F_TAG; ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=C_PURPLE); ws.cell(row=r, column=2).alignment = A_C; ws.cell(row=r, column=2).border = BD
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.cell(row=r, column=3, value=q).font = F_L2; ws.cell(row=r, column=3).alignment = A_LM; ws.cell(row=r, column=3).border = BD; ws.cell(row=r, column=3).fill = PatternFill('solid', fgColor=L4_FILL)
    ws.row_dimensions[r].height = 26; r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
    ws.cell(row=r, column=2, value='答').font = F_TAG; ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=C_TEAL); ws.cell(row=r, column=2).alignment = A_C; ws.cell(row=r, column=2).border = BD
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.cell(row=r, column=3, value=a).font = F_CELL; ws.cell(row=r, column=3).alignment = A_LT; ws.cell(row=r, column=3).border = BD
    ws.row_dimensions[r].height = 50; r += 1

# ================================================================
# Sheet 2: 开发主表（带阶段+文件路径）
# ================================================================
ws = wb.create_sheet('2.开发主表')
ws.sheet_view.showGridLines = False
# 列: 层级 | 编号 | 阶段 | L1系统 | L2模块 | L3子功能 | L4实现原理 | L5函数 | 文件路径 | 输入→输出 | 版本 | 状态
set_w(ws, [5, 8, 6, 12, 14, 18, 44, 24, 22, 20, 6, 8])
big_title(ws, '开发主表 — 从系统到函数，逐层展开（含阶段+文件路径）', 12)
header_row(ws, ['层级','编号','阶段','L1 系统','L2 模块','L3 子功能','L4 实现原理','L5 函数','文件路径','输入 → 输出','版本','状态'], 2)
ws.freeze_panes = 'A3'

r = 3
for row_data in D:
    row_list = list(row_data)
    while len(row_list) < 13:
        row_list.append('')
    level = row_list[0]
    num = row_list[1]
    phase = row_list[2]
    sys_code = row_list[3]
    sys_name = row_list[4]
    mod = row_list[5]
    func = row_list[6]
    principle = row_list[7]
    fn = row_list[8]
    filepath = row_list[9]
    io = row_list[10]
    ver = row_list[11]
    status = row_list[12]

    vals = [level, num, phase, sys_name if level=='L1' else '', mod, func, principle, fn, filepath, io, ver, status]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.border = BD
        c.alignment = A_LT

    # 层级标签
    lc = ws.cell(row=r, column=1)
    lc.font = F_TAG; lc.alignment = A_C
    if level == 'L1':
        bg, fg = L1_FILLS.get(sys_code, ('333333','FFFFFF'))
        lc.fill = PatternFill('solid', fgColor=bg); lc.value = 'L1\n系统'; ws.row_dimensions[r].height = 32
    elif level == 'L2':
        lc.fill = PatternFill('solid', fgColor=L2_FILL); lc.value = 'L2\n模块'; ws.row_dimensions[r].height = 24
    elif level == 'L3':
        lc.fill = PatternFill('solid', fgColor=L3_FILL); lc.value = 'L3\n功能'; ws.row_dimensions[r].height = 56

    # 编号
    ws.cell(row=r, column=2).font = F_CODE; ws.cell(row=r, column=2).alignment = A_C

    # 阶段
    pc = ws.cell(row=r, column=3)
    pc.font = F_BOLD; pc.alignment = A_C
    if phase and phase in PHASE_COLORS:
        pc.fill = PatternFill('solid', fgColor=PHASE_COLORS[phase])
        pc.font = Font(name='微软雅黑', size=9, bold=True, color='FFFFFF')

    # L1系统
    if level == 'L1':
        bg, fg = L1_FILLS.get(sys_code, ('333333','FFFFFF'))
        c = ws.cell(row=r, column=4); c.fill = PatternFill('solid', fgColor=bg); c.font = F_L1; c.alignment = A_C

    # L2模块
    if level == 'L2':
        c = ws.cell(row=r, column=5); c.fill = PatternFill('solid', fgColor=L2_FILL); c.font = F_L2; c.alignment = A_L

    # L3子功能
    if level == 'L3':
        c = ws.cell(row=r, column=6); c.fill = PatternFill('solid', fgColor=L3_FILL); c.font = F_CELL
        if '★' in (func or ''): c.font = Font(name='微软雅黑', size=10, bold=True, color=C_RED)
        # 原理
        ws.cell(row=r, column=7).fill = PatternFill('solid', fgColor=L4_FILL); ws.cell(row=r, column=7).font = F_CODE
        # 函数
        ws.cell(row=r, column=8).fill = PatternFill('solid', fgColor=L5_FILL); ws.cell(row=r, column=8).font = F_CODE
        # 文件路径
        ws.cell(row=r, column=9).font = F_CODE
        # 输入输出
        ws.cell(row=r, column=10).font = F_CODE
        # 版本
        vc = ws.cell(row=r, column=11); vc.alignment = A_C
        if ver == 'V1': vc.fill = PatternFill('solid', fgColor='D6EAF8')
        elif ver == 'V1.5': vc.fill = PatternFill('solid', fgColor='FADBD8')
        elif ver == 'V2': vc.fill = PatternFill('solid', fgColor='D5D8DC')
        # 状态
        sc = ws.cell(row=r, column=12); sc.font = F_BOLD; sc.alignment = A_C
        if status == '待办': sc.fill = PatternFill('solid', fgColor=ST_TODO)
        elif status == '进行中': sc.fill = PatternFill('solid', fgColor=ST_DOING)
        elif status == '已完成': sc.fill = PatternFill('solid', fgColor=ST_DONE)

    r += 1

# 条件格式
ws.conditional_formatting.add(f'L3:L{r-1}', FormulaRule(formula=['$L3="已完成"'], fill=PatternFill('solid', fgColor=ST_DONE)))
ws.conditional_formatting.add(f'L3:L{r-1}', FormulaRule(formula=['$L3="进行中"'], fill=PatternFill('solid', fgColor=ST_DOING)))
ws.conditional_formatting.add(f'L3:L{r-1}', FormulaRule(formula=['$L3="待办"'], fill=PatternFill('solid', fgColor=ST_TODO)))

last_row = r - 1
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
c = ws.cell(row=r, column=1, value='层级：L1系统(色块)→L2模块(浅蓝)→L3功能(浅橙)→L4原理(浅紫)→L5函数(浅青) | 阶段列标P1~P8 | 文件路径列告诉你代码放哪 | 改状态→进度统计自动更新')
c.font = F_SMALL; c.alignment = A_L

# ================================================================
# Sheet 3: 进度统计（Excel公式自动计算）
# ================================================================
ws = wb.create_sheet('3.进度统计')
ws.sheet_view.showGridLines = False
set_w(ws, [4, 22, 8, 10, 10, 10, 10, 16, 4])
big_title(ws, '进度统计 — 自动计算（改主表状态后刷新即更新）', 9)
header_row(ws, ['','系统大类','阶段','总点数','已完成','进行中','待办','完成率','进度条'], 3)

# 用COUNTIF公式自动统计主表数据
MAIN = "'2.开发主表'"
# 主表中L3行才有状态，用层级列=L3来过滤
# 系统大类在D列(L1系统列)，但L3行该列为空。用编号首字母区分。
sys_map = [('A','A.引擎框架','P1/P6'),('B','B.游戏数据','P2'),('C','C.战斗逻辑','P3/P4'),('D','D.AI系统','P5'),('E','E.渲染与UI','P6'),('F','F.社交系统','P7+'),('G','G.养成变现','P7+'),('H','H.实时PvP','V2')]

r = 4
for code, name, phase_tag in sys_map:
    bg, fg = L1_FILLS.get(code, ('333333','FFFFFF'))
    # COUNTIFS: 编号列(B)以code开头 且 状态列(L)有值
    total_f = f'=COUNTIFS({MAIN}!B3:B{last_row},"{code}*",{MAIN}!L3:L{last_row},"待办")+COUNTIFS({MAIN}!B3:B{last_row},"{code}*",{MAIN}!L3:L{last_row},"进行中")+COUNTIFS({MAIN}!B3:B{last_row},"{code}*",{MAIN}!L3:L{last_row},"已完成")'
    done_f = f'=COUNTIFS({MAIN}!B3:B{last_row},"{code}*",{MAIN}!L3:L{last_row},"已完成")'
    doing_f = f'=COUNTIFS({MAIN}!B3:B{last_row},"{code}*",{MAIN}!L3:L{last_row},"进行中")'
    todo_f = f'=COUNTIFS({MAIN}!B3:B{last_row},"{code}*",{MAIN}!L3:L{last_row},"待办")'
    rate_f = f'=IF(D{r}=0,"0%",TEXT(E{r}/D{r},"0%"))'
    bar_f = f'=REPT("█",ROUND(E{r}/D{r}*20,0))&REPT("░",20-ROUND(E{r}/D{r}*20,0))'

    vals = ['', name, phase_tag, total_f, done_f, doing_f, todo_f, rate_f, bar_f]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.border = BD; c.alignment = A_C if i > 2 else A_L; c.font = F_CELL
    ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=bg); ws.cell(row=r, column=2).font = F_L1; ws.cell(row=r, column=2).alignment = A_C
    ws.cell(row=r, column=9).font = Font(name='Consolas', size=11, color=C_GREEN)
    ws.row_dimensions[r].height = 30; r += 1

# 总计行
total_f = f'=SUM(D4:D{r-1})'
done_f = f'=SUM(E4:E{r-1})'
doing_f = f'=SUM(F4:F{r-1})'
todo_f = f'=SUM(G4:G{r-1})'
rate_f = f'=IF(D{r}=0,"0%",TEXT(E{r}/D{r},"0%"))'
bar_f = f'=REPT("█",ROUND(E{r}/D{r}*20,0))&REPT("░",20-ROUND(E{r}/D{r}*20,0))'
vals = ['','总计','',total_f,done_f,doing_f,todo_f,rate_f,bar_f]
for i, v in enumerate(vals, 1):
    c = ws.cell(row=r, column=i, value=v)
    c.font = F_HEADER; c.fill = PatternFill('solid', fgColor=C_INK); c.alignment = A_C if i > 2 else A_L; c.border = BD
ws.row_dimensions[r].height = 34

# 按阶段统计
r += 2
sub_title(ws, '按阶段统计', 9, r, C_GOLD); r += 1
header_row(ws, ['','阶段','阶段名称','总点数','已完成','进行中','待办','完成率','进度条'], r)
r += 1
phase_map = [('P1','引擎骨架'),('P2','数据层'),('P3','核心战斗'),('P4','扩展战斗'),('P5','AI对手'),('P6','渲染交互'),('P7','测试平衡'),('P8','上线')]
for code, name in phase_map:
    total_f = f'=COUNTIF({MAIN}!C3:C{last_row},"{code}")'
    done_f = f'=COUNTIFS({MAIN}!C3:C{last_row},"{code}",{MAIN}!L3:L{last_row},"已完成")'
    doing_f = f'=COUNTIFS({MAIN}!C3:C{last_row},"{code}",{MAIN}!L3:L{last_row},"进行中")'
    todo_f = f'=COUNTIFS({MAIN}!C3:C{last_row},"{code}",{MAIN}!L3:L{last_row},"待办")'
    rate_f = f'=IF(D{r}=0,"0%",TEXT(E{r}/D{r},"0%"))'
    bar_f = f'=REPT("█",ROUND(E{r}/D{r}*20,0))&REPT("░",20-ROUND(E{r}/D{r}*20,0))'
    vals = ['', code, name, total_f, done_f, doing_f, todo_f, rate_f, bar_f]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.border = BD; c.alignment = A_C if i > 2 else A_L; c.font = F_CELL
    ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=PHASE_COLORS[code]); ws.cell(row=r, column=2).font = F_L1; ws.cell(row=r, column=2).alignment = A_C
    ws.cell(row=r, column=9).font = Font(name='Consolas', size=11, color=C_GREEN)
    ws.row_dimensions[r].height = 28; r += 1

r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
c = ws.cell(row=r, column=2, value='说明：改「2.开发主表」的"状态"列后，回到本表按Ctrl+Z再Ctrl+Y（或重新打开文件）即可刷新数字。')
c.font = F_SMALL; c.alignment = A_L

# ================================================================
# Sheet 4: 卡牌数据表
# ================================================================
ws = wb.create_sheet('4.卡牌数据')
ws.sheet_view.showGridLines = False
set_w(ws, [16, 10, 8, 8, 8, 8, 8, 30, 8])
big_title(ws, '卡牌数据表 — 改数值直接在这改', 9)
r = 3
sub_title(ws, '攻方·人物卡', 9, r, C_RED); r += 1
header_row(ws, ['名称','类型','费用','血量','攻击','移速','间隔','特性','稀有度'], r); r += 1
attack_cards = [
    ('宗门体修弟子','普通弟子',2,4,2,1.0,1.0,'普通近战士兵，基础兵','凡品'),
    ('宗门剑修弟子','普通弟子',3,3,3,0.9,1.3,'远程(射程3)，用飞剑，不掉血','凡品'),
    ('宗门御兽弟子','普通弟子',3,6,2,0.8,1.2,'控兽当肉盾，高血低速','凡品'),
    ('金丹期长老','精英长老',6,10,3,0.8,1.5,'每5s随机释放4分支技能','宝品'),
]
for c in attack_cards:
    for i, v in enumerate(c, 1):
        cell = ws.cell(row=r, column=i, value=v)
        cell.font = F_CELL; cell.alignment = A_C; cell.border = BD
    ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor='FADBD8')
    ws.row_dimensions[r].height = 24; r += 1

r += 1
sub_title(ws, '守方·人物卡', 9, r, C_BLUE); r += 1
header_row(ws, ['名称','类型','费用','血量','攻击','射程','间隔','特性','稀有度'], r); r += 1
defend_cards = [
    ('护山傀儡','防守单位',3,6,2,1,1.2,'缓慢移动肉盾守卫','凡品'),
    ('护山灵兽','防守单位',4,8,2,1,1.0,'高血拦截，死亡自爆1伤','灵品'),
    ('护法长老','精英长老',6,10,3,1,1.5,'镇守大殿前，每5s随机释放4分支','宝品'),
]
for c in defend_cards:
    for i, v in enumerate(c, 1):
        cell = ws.cell(row=r, column=i, value=v)
        cell.font = F_CELL; cell.alignment = A_C; cell.border = BD
    ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor='D6EAF8')
    ws.row_dimensions[r].height = 24; r += 1

r += 1
sub_title(ws, '阵法（双方通用）', 9, r, C_TEAL); r += 1
header_row(ws, ['名称','类型','费用','血量','攻击','射程','间隔','特性','稀有度'], r); r += 1
array_cards = [
    ('截脉阵','阵法',2,4,2,1,1.0,'基础拦截，便宜','凡品'),
    ('寒霜阵','阵法',3,3,1,1,1.0,'命中后敌人移速-0.5(2s)','凡品'),
    ('万刃阵','阵法',4,5,3,1,1.0,'高输出拦截','灵品'),
    ('反震阵','阵法',3,3,0,'-',1.0,'反伤50%','灵品'),
    ('天罗阵','阵法',5,6,2,1,1.0,'范围(打相邻所有敌人)','宝品'),
]
for c in array_cards:
    for i, v in enumerate(c, 1):
        cell = ws.cell(row=r, column=i, value=v)
        cell.font = F_CELL; cell.alignment = A_C; cell.border = BD
    ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor='E8F8F5')
    ws.row_dimensions[r].height = 24; r += 1

r += 1
sub_title(ws, '法术', 9, r, C_PURPLE); r += 1
header_row(ws, ['名称','类型','费用','效果','','','','偏向','稀有度'], r); r += 1
spell_cards = [
    ('万剑归宗','法术',5,'全己方单位+1攻、移速+0.3(5s)','','','','攻方','宝品'),
    ('五雷正法','法术',4,'区域3格内敌方受4伤','','','','通用','灵品'),
    ('御风诀','法术',2,'指定己方单位移速+0.5(5s)','','','','通用','凡品'),
    ('镇魂符','法术',3,'指定敌方阵法失效3秒','','','','通用','凡品'),
    ('金钟罩','法术',3,'己方大殿免疫伤害3秒','','','','守方','灵品'),
    ('移山倒海','法术',4,'区域敌人推后2格+1伤','','','','守方','灵品'),
    ('困仙索','法术',2,'指定敌人定身2秒','','','','守方','凡品'),
    ('天雷诀','法术',4,'范围3格内敌方受4伤(清兵)','','','','通用','灵品'),
]
for c in spell_cards:
    for i, v in enumerate(c, 1):
        cell = ws.cell(row=r, column=i, value=v)
        cell.font = F_CELL; cell.alignment = A_C; cell.border = BD
    ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor='F4ECF7')
    ws.row_dimensions[r].height = 24; r += 1

# ================================================================
# Sheet 5: 数值配置表
# ================================================================
ws = wb.create_sheet('5.数值配置')
ws.sheet_view.showGridLines = False
set_w(ws, [20, 12, 32, 14, 14])
big_title(ws, '数值配置 — 所有可调参数集中管理', 5)
header_row(ws, ['参数名','当前值','说明','调高偏向','调低偏向'], 3)
params = [
    ('灵力回复间隔','2.8秒','每多少秒回复1点灵力','攻方(出兵多)','守方(布阵少)'),
    ('灵力初始上限','5','开局灵力上限','攻方','守方'),
    ('灵力上限增长间隔','30秒','每多少秒上限+1','攻方','守方'),
    ('灵力上限封顶','10','灵力上限最大值','攻方','守方'),
    ('大殿血量','30','大殿初始血量','守方','攻方'),
    ('单局时长','180秒','正常对战时间','守方','攻方'),
    ('加时时长','60秒','平局后加时时间','中立','中立'),
    ('加时灵力倍率','1.5','加时赛灵力回复倍率','攻方','守方'),
    ('手牌数量','4','同时持有手牌数','攻方(选择多)','守方'),
    ('卡组数量','8','卡组总卡牌数','中立','中立'),
    ('抽牌延迟','2秒','打出后多久抽新牌','守方','攻方'),
    ('阵法冷却时间','8秒','阵法被毁后格子冷却','攻方','守方'),
    ('长老技能间隔','5秒','长老多久释放一次技能','守方','攻方'),
    ('棋盘长度','9格','大殿到大殿格子数(含大殿)','守方','攻方'),
    ('同屏单位上限','30','最多同时存在单位数','中立(性能)','中立(性能)'),
    ('远程单位射程','3格','剑修弟子攻击射程','攻方','守方'),
    ('减速幅度','0.5','寒霜阵减速值','守方','攻方'),
    ('减速持续','2秒','减速持续时间','守方','攻方'),
    ('反震比例','50%','反震阵反弹伤害比例','守方','攻方'),
    ('护盾持续','3秒','金钟罩大殿免疫时间','守方','攻方'),
    ('定身持续','2秒','困仙索定身时间','守方','攻方'),
    ('推后格数','2格','移山倒海推后距离','守方','攻方'),
    ('加速持续','5秒','御风诀/万剑归宗加速时间','攻方','守方'),
    ('禁阵持续','3秒','镇魂符禁阵时间','攻方','守方'),
    ('AI简单思考间隔','3.5秒','简单AI决策间隔','AI强','AI弱'),
    ('AI普通思考间隔','2.5秒','普通AI决策间隔','AI强','AI弱'),
    ('AI困难思考间隔','1.5秒','困难AI决策间隔','AI强','AI弱'),
]
r = 4
for p in params:
    for i, v in enumerate(p, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = F_CELL; c.alignment = A_C if i > 1 else A_L; c.border = BD
    ws.row_dimensions[r].height = 22; r += 1

r += 1
sub_title(ws, '平衡验证标准（模拟10局AI vs AI）', 5, r, C_GOLD); r += 1
verify = ['双方摧毁度都在 20%~70% 区间','胜率：攻守方各 ~50%（40%~60%）','0% 摧毁度对局 < 10%','90% 对局在 150~210 秒内结束','双方大殿都被打到（有来有回）']
for v in verify:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value='• ' + v)
    c.font = F_CELL; c.alignment = A_L; c.border = BD; r += 1

# ================================================================
# Sheet 6: AI提问指南
# ================================================================
ws = wb.create_sheet('6.AI提问指南')
ws.sheet_view.showGridLines = False
set_w(ws, [18, 75])
big_title(ws, 'AI提问指南 — 如何让AI替你干活', 2)
r = 3
sub_title(ws, '一、黄金公式', 2, r, C_GOLD); r += 1
ws.cell(row=r, column=1, value='公式').font = F_BOLD; ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor=L2_FILL); ws.cell(row=r, column=1).border = BD; ws.cell(row=r, column=1).alignment = A_C
ws.cell(row=r, column=2, value='【角色定位】+【具体任务】+【上下文/参考文件】+【完成标准】+【自主决策授权】').font = F_CELL; ws.cell(row=r, column=2).alignment = A_LT; ws.cell(row=r, column=2).border = BD
ws.row_dimensions[r].height = 30; r += 2

sub_title(ws, '二、提示词模板（直接复制使用）', 2, r, C_PURPLE); r += 1
templates = [
    ('场景1：实现功能','你是微信小游戏开发专家。\n请实现【功能名】：【把主表"实现原理"列的内容粘过来】\n函数签名：【把主表"L5函数"列粘过来】\n文件路径：【把主表"文件路径"列粘过来】\n完成标准：①代码能通过node --check语法检查 ②符合原理描述的逻辑\n遇到设计细节未明确的按合理默认自行决定，做完统一汇报，不要中途问我。'),
    ('场景2：修Bug','我遇到了一个Bug：\n【现象：什么情况发生了什么】\n【错误信息：贴上报错日志】\n相关文件：【文件路径】\n请分析原因并修复。修复后说明改了什么、为什么。'),
    ('场景3：平衡调优','请运行AI vs AI模拟10局，输出每局：双方摧毁度/胜方/时长。\n当前参数见「数值配置」表。\n如发现平衡问题（摧毁度0%或一边倒），调整参数重新模拟，直到：双方摧毁度20~70%，胜率~50%。\n每次调参记录：调了什么、从多少到多少、为什么。'),
    ('场景4：自主连续开发','以下是连续任务（按顺序）：\n【粘贴主表中连续3~5个L3行的"编号+原理+函数"】\n请按顺序逐个完成，每个完成后：①标记完成 ②记录做了什么 ③立即开始下一个\n全部做完统一汇报。遇到细节自行决定，不要中途问我。\n每个任务必须通过完成标准验证。'),
]
for name, tmpl in templates:
    ws.cell(row=r, column=1, value=name).font = F_BOLD; ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor=L2_FILL); ws.cell(row=r, column=1).border = BD; ws.cell(row=r, column=1).alignment = A_C
    ws.cell(row=r, column=2, value=tmpl).font = F_CELL; ws.cell(row=r, column=2).alignment = A_LT; ws.cell(row=r, column=2).border = BD
    ws.row_dimensions[r].height = 100; r += 1

r += 1
sub_title(ws, '三、常见错误与纠正', 2, r, C_RED); r += 1
mistakes = [
    ('❌太模糊','"做个战斗系统" → ✅"实现单位移动：沿y轴推进，speed×dt，遇敌停下"'),
    ('❌太大','"把游戏做完" → ✅一次一个任务（主表一行）'),
    ('❌没标准','"帮我写代码" → ✅给完成标准"能通过node --check + 模拟能跑"'),
    ('❌没授权','AI每步都停下问你 → ✅"遇到细节自行决定，做完汇报，不要中途问我"'),
    ('❌没验证','AI说做完了你就信 → ✅按完成标准自己验证（跑模拟/查语法）'),
    ('❌没上下文','AI不知道已有代码 → ✅告诉它参考文件路径和设计文档'),
]
for wrong, right in mistakes:
    ws.cell(row=r, column=1, value=wrong).font = F_CELL; ws.cell(row=r, column=1).border = BD; ws.cell(row=r, column=1).alignment = A_LT
    ws.cell(row=r, column=2, value=right).font = F_CELL; ws.cell(row=r, column=2).border = BD; ws.cell(row=r, column=2).alignment = A_LT
    ws.row_dimensions[r].height = 30; r += 1

r += 1
sub_title(ws, '四、每日开发流程', 2, r, C_TEAL); r += 1
daily = [
    '1. 看「0.开发阶段规划」确认当前阶段（如P3）',
    '2. 打开「2.开发主表」筛选阶段=P3，找第一个"待办"的L3行',
    '3. 复制该行的"实现原理"+"函数"+"文件路径"',
    '4. 用「场景1」模板粘贴发给AI',
    '5. AI完成后你验证（node --check / 跑模拟）',
    '6. 通过→主表状态改"已完成"→进度统计自动更新',
    '7. 不通过→用「场景2」修Bug模板发给AI',
    '8. 重复，每天2~4个任务',
]
for step in daily:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws.cell(row=r, column=1, value=step)
    c.font = F_CELL; c.alignment = A_L; c.border = BD
    ws.row_dimensions[r].height = 24; r += 1

# 保存
wb.save(PATH)
l3_count = len([d for d in D if d[0]=='L3'])
print(f'Excel已生成: {PATH}')
print(f'Sheet: {wb.sheetnames}')
print(f'L3功能点: {l3_count}')
print(f'卡牌: {len(attack_cards)+len(defend_cards)+len(array_cards)+len(spell_cards)}张')
print(f'参数: {len(params)}个')
print(f'AI模板: {len(templates)}个')
