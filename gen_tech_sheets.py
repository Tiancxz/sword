#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
扩充《宗门论道》开发管理工具包 — 新增7个技术Sheet
在现有Excel基础上追加：技术架构图/类设计/函数清单/数据结构/调用关系/技术选型/代码规范
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 打开现有Excel
PATH = '/workspace/宗门论道_开发管理工具包.xlsx'
wb = openpyxl.load_workbook(PATH)

# ===== 样式库 =====
F_TITLE   = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
F_HEADER  = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
F_SECTION = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')
F_CELL    = Font(name='微软雅黑', size=10)
F_BOLD    = Font(name='微软雅黑', size=10, bold=True)
F_CODE    = Font(name='Consolas', size=10, color='333333')
F_NOTE    = Font(name='微软雅黑', size=10, color='666666', italic=True)

BG_TITLE   = PatternFill('solid', fgColor='2F5496')
BG_HEADER  = PatternFill('solid', fgColor='4472C4')
BG_SECTION = PatternFill('solid', fgColor='D6E4F0')
BG_ENGINE  = PatternFill('solid', fgColor='E2EFDA')
BG_GAME    = PatternFill('solid', fgColor='FCE4D6')
BG_RENDER  = PatternFill('solid', fgColor='FFF2CC')
BG_DATA    = PatternFill('solid', fgColor='D6E4F0')
BG_V15     = PatternFill('solid', fgColor='F8CBAD')
BG_V2      = PatternFill('solid', fgColor='D9D9D9')

A_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
A_LEFT   = Alignment(horizontal='left', vertical='center', wrap_text=True)
A_LEFTT  = Alignment(horizontal='left', vertical='top', wrap_text=True)

BD = Border(
    left=Side(style='thin', color='B4C7E7'),
    right=Side(style='thin', color='B4C7E7'),
    top=Side(style='thin', color='B4C7E7'),
    bottom=Side(style='thin', color='B4C7E7')
)

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_row(ws, text, cols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_TITLE; c.fill = BG_TITLE; c.alignment = A_CENTER
    ws.row_dimensions[row].height = 32

def header_row(ws, headers, row):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = F_HEADER; c.fill = BG_HEADER; c.alignment = A_CENTER; c.border = BD
    ws.row_dimensions[row].height = 24

def data_row(ws, values, row, fill=None, code_cols=None):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        if code_cols and i in code_cols:
            c.font = F_CODE
        else:
            c.font = F_CELL
        c.alignment = A_LEFTT
        c.border = BD
        if fill: c.fill = fill

def section_row(ws, text, cols, row):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_SECTION; c.fill = BG_SECTION; c.alignment = A_LEFT
    ws.row_dimensions[row].height = 22

# ================================================================
# Sheet 11: 技术架构总览
# ================================================================
ws = wb.create_sheet('11.技术架构图')
set_widths(ws, [16, 22, 40, 14, 8])
title_row(ws, '技术架构总览 — 模块分层与职责', 5)
header_row(ws, ['层级', '模块名', '职责描述', '文件路径', '版本'], 3)

arch = [
    # 引擎层
    ('引擎层', 'Director', '渲染循环主驱动：requestAnimationFrame→update(dt)+render()，帧率控制，场景切换', 'js/core/Director.js', 'V1'),
    ('引擎层', 'Input', '触摸输入管理：wx.onTouchStart/Move/End→坐标转换→点击区域注册→命中检测→回调', 'js/core/Input.js', 'V1'),
    ('引擎层', 'EventSystem', '事件总线：on(event,cb)/emit(event,data)/off(event)，模块间解耦通信', 'js/core/EventSystem.js', 'V1'),
    ('引擎层', 'ResourceLoader', '资源预加载：图片/音频/Spine加载，加载进度回调，加载完成后通知', 'js/core/ResourceLoader.js', 'V1'),
    ('引擎层', 'SceneManager', '场景管理：注册/切换/生命周期(onEnter/onExit/onUpdate/onRender)', 'js/core/SceneManager.js', 'V1'),
    ('引擎层', 'ObjectPool', '对象池：单位/粒子复用，避免频繁GC，提升性能', 'js/core/ObjectPool.js', 'V1.5'),
    # 配置层
    ('配置层', 'Constants', '全局常量：棋盘尺寸/灵力参数/时间/颜色/坐标等所有硬编码值集中管理', 'js/config/Constants.js', 'V1'),
    ('配置层', 'Cards', '卡牌数据表：19张卡牌完整属性，配置驱动，改数值只改这里', 'js/config/Cards.js', 'V1'),
    ('配置层', 'DeckPresets', '预设卡组：3套预设(快攻/控制/坦克)，每套8张', 'js/config/DeckPresets.js', 'V1'),
    # 数据层
    ('数据层', 'GameModel', '游戏状态容器：时间/双方玩家状态(大殿/灵力/单位/阵法)/游戏阶段', 'js/game/GameModel.js', 'V1'),
    ('数据层', 'Unit', '单位实体：属性+状态机+buff列表，可从卡牌数据实例化', 'js/game/Unit.js', 'V1'),
    ('数据层', 'Formation', '阵法实体：属性+位置+冷却状态，固定建筑', 'js/game/Formation.js', 'V1'),
    ('数据层', 'Deck', '卡组手牌：洗牌/抽牌/补位/手牌管理', 'js/game/Deck.js', 'V1'),
    ('数据层', 'Player', '玩家状态：大殿血量/灵力/卡组/手牌/单位列表/阵法列表', 'js/game/Player.js', 'V1'),
    # 逻辑层
    ('逻辑层', 'BattleLogic', '★核心战斗逻辑★：移动/碰撞/交战/伤害/击杀推进/阵法/法术/长老技能', 'js/game/BattleLogic.js', 'V1'),
    ('逻辑层', 'MovementSystem', '移动系统：单位寻路+移动+碰撞检测+目标选择', 'js/game/MovementSystem.js', 'V1'),
    ('逻辑层', 'CombatSystem', '战斗系统：近战/远程/范围/自爆伤害结算+击杀判定', 'js/game/CombatSystem.js', 'V1'),
    ('逻辑层', 'SpellSystem', '法术系统：8种法术效果执行(减速/反震/定身/推后/护盾/加速/禁阵/范围)', 'js/game/SpellSystem.js', 'V1'),
    ('逻辑层', 'ElderSkillSystem', '长老技能系统：4分支随机释放逻辑(飞剑/丹药/符箓/御兽)', 'js/game/ElderSkillSystem.js', 'V1'),
    ('逻辑层', 'AI', 'AI决策：灵力管理/出牌策略/布阵位置/难度分级', 'js/game/AI.js', 'V1'),
    ('逻辑层', 'EnergySystem', '灵力系统：实时回复/上限增长/加时倍率', 'js/game/EnergySystem.js', 'V1'),
    ('逻辑层', 'BattleChecker', '胜负判定：大殿血量/时限/加时/结算', 'js/game/BattleChecker.js', 'V1'),
    # 渲染层
    ('渲染层', 'BattleScene', '战斗主场景：背景+大殿+单位+阵法+特效+UI整合渲染', 'js/scenes/BattleScene.js', 'V1'),
    ('渲染层', 'ResultScene', '结算场景：胜负展示+摧毁度+奖励+再来一局', 'js/scenes/ResultScene.js', 'V1'),
    ('渲染层', 'MenuScene', '主菜单：开始对战/卡组选择/设置(V1.5+)', 'js/scenes/MenuScene.js', 'V1.5'),
    ('渲染层', 'Renderer', '渲染器：Canvas 2D绘制封装(矩形/圆/文字/图片/粒子)', 'js/core/Renderer.js', 'V1'),
    ('渲染层', 'Camera', '摄像机：屏幕震动/缩放(V2考虑)', 'js/core/Camera.js', 'V1.5'),
    ('渲染层', 'ParticleSystem', '粒子系统：命中/死亡/法术/大殿受击特效', 'js/core/ParticleSystem.js', 'V1'),
    # UI层
    ('UI层', 'HandBar', '手牌栏：4张牌显示/选中/灰显/出牌动画', 'js/ui/HandBar.js', 'V1'),
    ('UI层', 'EnergyBar', '灵力条：数字+液条+回复动画', 'js/ui/EnergyBar.js', 'V1'),
    ('UI层', 'HUD', '顶部HUD：双方血条+计时器+加时提示', 'js/ui/HUD.js', 'V1'),
    ('UI层', 'TutorialGuide', '新手引导：4步教程+高亮+箭头指引', 'js/ui/TutorialGuide.js', 'V1'),
    ('UI层', 'CardSelectOverlay', '出牌选目标：高亮可放置区域+取消选择', 'js/ui/CardSelectOverlay.js', 'V1'),
    # 音频层
    ('音频层', 'AudioManager', '音频管理：BGM播放/SFX播放/音量控制/音频池', 'js/core/AudioManager.js', 'V1'),
    # V1.5+
    ('社交层(V1.5)', 'CloudManager', '云开发管理：登录/存档/云函数调用', 'js/social/CloudManager.js', 'V1.5'),
    ('社交层(V1.5)', 'ShareManager', '分享管理：对局分享/求助/炫耀', 'js/social/ShareManager.js', 'V1.5'),
    ('社交层(V1.5)', 'RankManager', '排行榜：开放数据域好友/群排行', 'js/social/RankManager.js', 'V1.5'),
    ('养成层(V1.5)', 'ProgressionSystem', '养成系统：弟子升级/升星/卡牌收集', 'js/game/ProgressionSystem.js', 'V1.5'),
    ('养成层(V1.5)', 'ShopSystem', '商店系统：宝箱/碎片/礼包购买', 'js/game/ShopSystem.js', 'V1.5'),
    ('变现层(V1.5)', 'AdManager', '广告管理：激励视频/插屏/Banner', 'js/monetize/AdManager.js', 'V1.5'),
    ('变现层(V1.5)', 'IAPManager', '内购管理：仙玉/月卡/通行证购买', 'js/monetize/IAPManager.js', 'V1.5'),
    # V2
    ('PvP层(V2)', 'NetworkClient', '网络客户端：WebSocket连接/心跳/断线重连', 'js/net/NetworkClient.js', 'V2'),
    ('PvP层(V2)', 'SyncManager', '同步管理：状态同步/帧同步/延迟补偿', 'js/net/SyncManager.js', 'V2'),
    ('PvP层(V2)', 'MatchMaker', '匹配系统：按段位/ELO匹配对手', 'js/net/MatchMaker.js', 'V2'),
]
r = 4
for a in arch:
    layer = a[0]
    if 'V1.5' in layer: fill = BG_V15
    elif 'V2' in layer: fill = BG_V2
    elif '引擎' in layer: fill = BG_ENGINE
    elif '渲染' in layer: fill = BG_RENDER
    elif '数据' in layer or '配置' in layer: fill = BG_DATA
    else: fill = BG_GAME
    data_row(ws, list(a), r, fill=fill)
    ws.row_dimensions[r].height = 30; r += 1

# 架构分层图
r += 1
section_row(ws, '架构分层关系（上层依赖下层）', 5, r); r += 1
layers_text = [
    '第5层 UI层      ← HandBar/EnergyBar/HUD/TutorialGuide',
    '第4层 渲染层     ← BattleScene/ResultScene/Renderer/ParticleSystem',
    '第3层 逻辑层     ← BattleLogic/MovementSystem/CombatSystem/SpellSystem/AI',
    '第2层 数据层     ← GameModel/Unit/Formation/Deck/Player',
    '第1层 配置层     ← Constants/Cards/DeckPresets',
    '第0层 引擎层     ← Director/Input/EventSystem/ResourceLoader/SceneManager',
]
for t in layers_text:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value=t)
    c.font = F_CODE; c.alignment = A_LEFT; c.border = BD
    r += 1

# ================================================================
# Sheet 12: 类设计（每个类的属性+方法）
# ================================================================
ws = wb.create_sheet('12.类设计')
set_widths(ws, [14, 24, 8, 50, 14])
title_row(ws, '类设计 — 每个类的属性与方法清单', 5)
header_row(ws, ['类名', '成员名', '类型', '说明', '所属文件'], 3)

classes = [
    # Director
    ('Director', 'instance', '静态属性', '单例实例', 'core/Director.js'),
    ('Director', 'canvas', '属性', 'Canvas 2D上下文(wx.createCanvas())', 'core/Director.js'),
    ('Director', 'sceneManager', '属性', 'SceneManager实例', 'core/Director.js'),
    ('Director', 'lastTime', '属性', '上一帧时间戳(用于计算dt)', 'core/Director.js'),
    ('Director', 'running', '属性', '是否正在运行循环', 'core/Director.js'),
    ('Director', 'init()', '方法', '初始化:获取canvas/创建SceneManager/Input/启动循环', 'core/Director.js'),
    ('Director', 'start()', '方法', '启动requestAnimationFrame循环', 'core/Director.js'),
    ('Director', 'loop(timestamp)', '方法', '主循环:计算dt→update(dt)→render()', 'core/Director.js'),
    ('Director', 'update(dt)', '方法', '调用当前场景的update(dt)', 'core/Director.js'),
    ('Director', 'render()', '方法', '调用当前场景的render()', 'core/Director.js'),
    ('Director', 'getInstance()', '静态方法', '获取单例', 'core/Director.js'),

    # Input
    ('Input', 'touchRegions', '属性', '注册的可点击区域数组[{x,y,w,h,cb,id}]', 'core/Input.js'),
    ('Input', 'touchingId', '属性', '当前正在触摸的区域id', 'core/Input.js'),
    ('Input', 'init()', '方法', '绑定wx.onTouchStart/Move/End事件', 'core/Input.js'),
    ('Input', 'register(id, x, y, w, h, cb)', '方法', '注册一个可点击区域', 'core/Input.js'),
    ('Input', 'unregister(id)', '方法', '取消注册一个区域', 'core/Input.js'),
    ('Input', 'clear()', '方法', '清空所有注册区域(场景切换时调用)', 'core/Input.js'),
    ('Input', 'onTouchStart(x, y)', '方法', '触摸开始:遍历区域命中检测→触发cb', 'core/Input.js'),
    ('Input', 'onTouchEnd(x, y)', '方法', '触摸结束', 'core/Input.js'),

    # EventSystem
    ('EventSystem', 'listeners', '属性', '事件监听器Map{event:[cb1,cb2,...]}', 'core/EventSystem.js'),
    ('EventSystem', 'on(event, cb)', '方法', '注册事件监听', 'core/EventSystem.js'),
    ('EventSystem', 'off(event, cb)', '方法', '移除事件监听', 'core/EventSystem.js'),
    ('EventSystem', 'emit(event, data)', '方法', '触发事件,传入data给所有监听器', 'core/EventSystem.js'),
    ('EventSystem', 'clear(event)', '方法', '清空指定事件的所有监听(或全部)', 'core/EventSystem.js'),

    # GameModel
    ('GameModel', 'time', '属性', '剩余时间(秒),初始180', 'game/GameModel.js'),
    ('GameModel', 'players', '属性', '[Player, Player] 双方玩家状态', 'game/GameModel.js'),
    ('GameModel', 'state', '属性', '游戏状态: playing/overtime/ended', 'game/GameModel.js'),
    ('GameModel', 'winner', '属性', '胜利方: 0/1/null', 'game/GameModel.js'),
    ('GameModel', 'elapsedTime', '属性', '已过时间(秒),用于灵力上限增长', 'game/GameModel.js'),
    ('GameModel', 'init(deck0, deck1)', '方法', '初始化游戏:创建双方Player/设置大殿/灵力/卡组', 'game/GameModel.js'),
    ('GameModel', 'getAllUnits()', '方法', '获取场上所有单位(双方)', 'game/GameModel.js'),
    ('GameModel', 'getAllFormations()', '方法', '获取场上所有阵法(双方)', 'game/GameModel.js'),
    ('GameModel', 'getEnemyUnits(owner)', '方法', '获取指定方的敌方单位', 'game/GameModel.js'),
    ('GameModel', 'getEnemyFormations(owner)', '方法', '获取指定方的敌方阵法', 'game/GameModel.js'),
    ('GameModel', 'toJSON()', '方法', '序列化为JSON(用于存档/回放)', 'game/GameModel.js'),

    # Unit
    ('Unit', 'id', '属性', '唯一实例ID', 'game/Unit.js'),
    ('Unit', 'cardId', '属性', '卡牌ID(对应Cards表)', 'game/Unit.js'),
    ('Unit', 'owner', '属性', '归属方: 0(我方)/1(敌方)', 'game/Unit.js'),
    ('Unit', 'x', '属性', 'x坐标(列: -1=左阵法区, 0=主路, 1=右阵法区)', 'game/Unit.js'),
    ('Unit', 'y', '属性', 'y坐标(行: 0=我方大殿, 8=敌方大殿)', 'game/Unit.js'),
    ('Unit', 'hp', '属性', '当前血量', 'game/Unit.js'),
    ('Unit', 'maxHp', '属性', '最大血量', 'game/Unit.js'),
    ('Unit', 'atk', '属性', '攻击力', 'game/Unit.js'),
    ('Unit', 'speed', '属性', '当前移速(可被buff修改)', 'game/Unit.js'),
    ('Unit', 'baseSpeed', '属性', '基础移速(卡牌原始值)', 'game/Unit.js'),
    ('Unit', 'attackRange', '属性', '攻击射程(近战=1, 远程=3)', 'game/Unit.js'),
    ('Unit', 'attackInterval', '属性', '攻击间隔(秒)', 'game/Unit.js'),
    ('Unit', 'lastAttackTime', '属性', '上次攻击时间戳', 'game/Unit.js'),
    ('Unit', 'target', '属性', '当前攻击目标(Unit/Formation/null)', 'game/Unit.js'),
    ('Unit', 'state', '属性', '状态: walking/fighting/dying/dead', 'game/Unit.js'),
    ('Unit', 'traits', '属性', '特性数组: ["ranged","kamikaze","reflect"等]', 'game/Unit.js'),
    ('Unit', 'buffs', '属性', 'buff列表[{type,value,duration}]', 'game/Unit.js'),
    ('Unit', 'facing', '属性', '朝向: 1(向上)/-1(向下)', 'game/Unit.js'),
    ('Unit', 'isElder', '属性', '是否长老(用于技能触发)', 'game/Unit.js'),
    ('Unit', 'elderTimer', '属性', '长老技能计时器', 'game/Unit.js'),
    ('Unit', 'static fromCard(cardId, owner, x, y)', '静态方法', '从卡牌数据创建Unit实例', 'game/Unit.js'),
    ('Unit', 'addBuff(type, value, duration)', '方法', '添加buff(减速/加速/定身/护盾等)', 'game/Unit.js'),
    ('Unit', 'updateBuffs(dt)', '方法', '更新buff计时,过期移除', 'game/Unit.js'),
    ('Unit', 'getEffectiveSpeed()', '方法', '获取考虑buff后的实际移速', 'game/Unit.js'),
    ('Unit', 'takeDamage(amount, attacker)', '方法', '受到伤害(处理反震/护盾)', 'game/Unit.js'),
    ('Unit', 'isDead()', '方法', '是否死亡(hp<=0)', 'game/Unit.js'),

    # Formation
    ('Formation', 'id', '属性', '唯一实例ID', 'game/Formation.js'),
    ('Formation', 'cardId', '属性', '卡牌ID', 'game/Formation.js'),
    ('Formation', 'owner', '属性', '归属方', 'game/Formation.js'),
    ('Formation', 'gridX', '属性', '格子x坐标(-1或1)', 'game/Formation.js'),
    ('Formation', 'gridY', '属性', '格子y坐标(0~8)', 'game/Formation.js'),
    ('Formation', 'hp', '属性', '当前血量', 'game/Formation.js'),
    ('Formation', 'maxHp', '属性', '最大血量', 'game/Formation.js'),
    ('Formation', 'atk', '属性', '攻击力', 'game/Formation.js'),
    ('Formation', 'range', '属性', '攻击范围(格)', 'game/Formation.js'),
    ('Formation', 'attackInterval', '属性', '攻击间隔', 'game/Formation.js'),
    ('Formation', 'lastAttackTime', '属性', '上次攻击时间', 'game/Formation.js'),
    ('Formation', 'isActive', '属性', '是否激活(被镇魂符禁时=false)', 'game/Formation.js'),
    ('Formation', 'traits', '属性', '特性: ["slow","reflect","aoe"]', 'game/Formation.js'),
    ('Formation', 'static fromCard(cardId, owner, gridX, gridY)', '静态方法', '从卡牌数据创建阵法', 'game/Formation.js'),
    ('Formation', 'takeDamage(amount)', '方法', '受到伤害', 'game/Formation.js'),
    ('Formation', 'isDead()', '方法', '是否被毁', 'game/Formation.js'),

    # Deck
    ('Deck', 'drawPile', '属性', '抽牌堆(剩余可抽的牌)', 'game/Deck.js'),
    ('Deck', 'hand', '属性', '手牌数组(当前持有的牌)', 'game/Deck.js'),
    ('Deck', 'maxHand', '属性', '手牌上限(4)', 'game/Deck.js'),
    ('Deck', 'drawTimer', '属性', '抽牌计时器(打出后2秒补牌)', 'game/Deck.js'),
    ('Deck', 'init(cardIds)', '方法', '用卡组ID数组初始化→洗牌→抽3张起手', 'game/Deck.js'),
    ('Deck', 'shuffle()', '方法', '洗牌(Fisher-Yates)', 'game/Deck.js'),
    ('Deck', 'draw()', '方法', '从牌堆抽1张到手牌', 'game/Deck.js'),
    ('Deck', 'playCard(handIndex)', '方法', '打出手牌:从hand移除→启动drawTimer', 'game/Deck.js'),
    ('Deck', 'update(dt)', '方法', '更新抽牌计时器→到时补牌', 'game/Deck.js'),
    ('Deck', 'canPlay(handIndex, energy)', '方法', '检查灵力是否足够出牌', 'game/Deck.js'),

    # Player
    ('Player', 'id', '属性', '玩家ID: 0/1', 'game/Player.js'),
    ('Player', 'hallHp', '属性', '大殿血量', 'game/Player.js'),
    ('Player', 'hallMaxHp', '属性', '大殿最大血量(30)', 'game/Player.js'),
    ('Player', 'hallShield', '属性', '大殿护盾剩余时间(金钟罩)', 'game/Player.js'),
    ('Player', 'energy', '属性', '当前灵力', 'game/Player.js'),
    ('Player', 'energyMax', '属性', '灵力上限', 'game/Player.js'),
    ('Player', 'energyTimer', '属性', '灵力回复计时器', 'game/Player.js'),
    ('Player', 'deck', '属性', 'Deck实例(卡组手牌)', 'game/Player.js'),
    ('Player', 'units', '属性', '场上己方单位列表[Unit]', 'game/Player.js'),
    ('Player', 'formations', '属性', '场上己方阵法列表[Formation]', 'game/Player.js'),
    ('Player', 'formationCooldowns', '属性', '格子冷却Map{gridKey:expireTime}', 'game/Player.js'),
    ('Player', 'addEnergy(amount)', '方法', '增加灵力(不超上限)', 'game/Player.js'),
    ('Player', 'spendEnergy(amount)', '方法', '消耗灵力(不足返回false)', 'game/Player.js'),
    ('Player', 'updateEnergy(dt)', '方法', '灵力实时回复+上限增长', 'game/Player.js'),
    ('Player', 'damageHall(amount)', '方法', '大殿受伤(处理护盾)', 'game/Player.js'),
    ('Player', 'isHallDestroyed()', '方法', '大殿是否被摧毁', 'game/Player.js'),

    # BattleLogic
    ('BattleLogic', 'model', '属性', 'GameModel实例引用', 'game/BattleLogic.js'),
    ('BattleLogic', 'init(model)', '方法', '初始化,绑定GameModel', 'game/BattleLogic.js'),
    ('BattleLogic', 'update(dt)', '方法', '★主更新入口★:更新所有单位/阵法/法术/AI/灵力/胜负', 'game/BattleLogic.js'),
    ('BattleLogic', 'updateUnits(dt)', '方法', '遍历所有单位→移动/找目标/交战/buff', 'game/BattleLogic.js'),
    ('BattleLogic', 'updateFormations(dt)', '方法', '遍历所有阵法→主动攻击经过的敌人', 'game/BattleLogic.js'),
    ('BattleLogic', 'updateElders(dt)', '方法', '长老技能计时→随机释放分支', 'game/BattleLogic.js'),
    ('BattleLogic', 'updateEnergy(dt)', '方法', '双方灵力实时回复', 'game/BattleLogic.js'),
    ('BattleLogic', 'updateAI(dt)', '方法', 'AI决策→出牌/布阵/施法', 'game/BattleLogic.js'),
    ('BattleLogic', 'updateDraw(dt)', '方法', '双方手牌抽牌计时', 'game/BattleLogic.js'),
    ('BattleLogic', 'checkBattleEnd()', '方法', '检查胜负条件', 'game/BattleLogic.js'),
    ('BattleLogic', 'spawnUnit(player, cardId, x)', '方法', '出兵:创建Unit→加入units', 'game/BattleLogic.js'),
    ('BattleLogic', 'placeFormation(player, cardId, gridX, gridY)', '方法', '布阵:创建Formation→检查冷却', 'game/BattleLogic.js'),
    ('BattleLogic', 'castSpell(player, cardId, target)', '方法', '施法:执行法术效果', 'game/BattleLogic.js'),
    ('BattleLogic', 'removeDeadUnits()', '方法', '清理死亡单位(触发死亡效果)', 'game/BattleLogic.js'),

    # MovementSystem
    ('MovementSystem', 'moveUnit(unit, dt)', '方法', '★单位移动★:y+=speed*dt*facing→检测前方→停或走', 'game/MovementSystem.js'),
    ('MovementSystem', 'findTarget(unit, model)', '方法', '★目标选择★:同列最近敌方单位/阵法', 'game/MovementSystem.js'),
    ('MovementSystem', 'checkCollision(unit, target)', '方法', '碰撞检测:y距离<attackRange', 'game/MovementSystem.js'),
    ('MovementSystem', 'checkHallReach(unit, model)', '方法', '检查是否到达大殿格', 'game/MovementSystem.js'),

    # CombatSystem
    ('CombatSystem', 'attack(attacker, target, dt)', '方法', '★攻击逻辑★:间隔检查→造成伤害', 'game/CombatSystem.js'),
    ('CombatSystem', 'meleeAttack(attacker, target)', '方法', '近战:双方互扣血', 'game/CombatSystem.js'),
    ('CombatSystem', 'rangedAttack(attacker, target)', '方法', '远程:只扣对方血,自己不掉', 'game/CombatSystem.js'),
    ('CombatSystem', 'aoeAttack(source, targets, damage)', '方法', '范围:对多目标造成伤害', 'game/CombatSystem.js'),
    ('CombatSystem', 'kamikaze(unit, targets)', '方法', '自爆:对相邻格范围伤害后消失', 'game/CombatSystem.js'),
    ('CombatSystem', 'handleKill(killer, victim)', '方法', '★击杀处理★:击杀者继续推进(核心修复)', 'game/CombatSystem.js'),
    ('CombatSystem', 'damageHall(player, amount)', '方法', '大殿受伤(处理护盾)', 'game/CombatSystem.js'),

    # SpellSystem
    ('SpellSystem', 'cast(cardId, caster, target, model)', '方法', '★法术入口★:根据cardId分支执行', 'game/SpellSystem.js'),
    ('SpellSystem', 'castWanJian(caster)', '方法', '万剑归宗:全己方单位+1攻+加速', 'game/SpellSystem.js'),
    ('SpellSystem', 'castWuLei(target, model)', '方法', '五雷正法:区域3格敌方受4伤', 'game/SpellSystem.js'),
    ('SpellSystem', 'castYuFeng(target)', '方法', '御风诀:目标移速+0.5(5s)', 'game/SpellSystem.js'),
    ('SpellSystem', 'castZhenHun(target)', '方法', '镇魂符:目标阵法失效3s', 'game/SpellSystem.js'),
    ('SpellSystem', 'castJinZhong(caster)', '方法', '金钟罩:大殿免疫3s', 'game/SpellSystem.js'),
    ('SpellSystem', 'castYiShan(targets)', '方法', '移山倒海:区域敌人推后2格+1伤', 'game/SpellSystem.js'),
    ('SpellSystem', 'castKunXian(target)', '方法', '困仙索:目标定身2s', 'game/SpellSystem.js'),
    ('SpellSystem', 'castTianLei(targets)', '方法', '天雷诀:范围敌方受4伤(清兵)', 'game/SpellSystem.js'),

    # ElderSkillSystem
    ('ElderSkillSystem', 'update(elder, dt, model)', '方法', '长老技能计时→到5s触发', 'game/ElderSkillSystem.js'),
    ('ElderSkillSystem', 'triggerRandom(elder, model)', '方法', '★随机选分支★:4选1释放', 'game/ElderSkillSystem.js'),
    ('ElderSkillSystem', 'castFlyingSword(elder, model)', '方法', '飞剑分支:万剑齐发范围伤', 'game/ElderSkillSystem.js'),
    ('ElderSkillSystem', 'castPill(elder, model)', '方法', '丹药分支:自身+周围己方回血', 'game/ElderSkillSystem.js'),
    ('ElderSkillSystem', 'castTalisman(elder, model)', '方法', '符箓分支:天雷诀清兵', 'game/ElderSkillSystem.js'),
    ('ElderSkillSystem', 'castBeast(elder, model)', '方法', '御兽分支:召唤灵兽', 'game/ElderSkillSystem.js'),

    # AI
    ('AI', 'playerId', '属性', 'AI控制的玩家ID(1)', 'game/AI.js'),
    ('AI', 'difficulty', '属性', '难度: easy/normal/hard', 'game/AI.js'),
    ('AI', 'thinkTimer', '属性', '思考计时器', 'game/AI.js'),
    ('AI', 'thinkInterval', '属性', '思考间隔: easy=3.5/normal=2.5/hard=1.5', 'game/AI.js'),
    ('AI', 'init(playerId, difficulty)', '方法', '初始化AI', 'game/AI.js'),
    ('AI', 'update(dt, model)', '方法', '★AI主循环★:计时→到点决策', 'game/AI.js'),
    ('AI', 'think(model)', '方法', '★决策核心★:灵力管理→攻守判断→选牌→执行', 'game/AI.js'),
    ('AI', 'decideAttackRatio(model)', '方法', '根据大殿血量决定攻守比(>60%偏攻/<30%偏守)', 'game/AI.js'),
    ('AI', 'pickCard(energy, ratio)', '方法', '根据灵力和攻守比选一张可出的牌', 'game/AI.js'),
    ('AI', 'pickFormationPos(model)', '方法', '选布阵位置: easy随机/normal大殿前/hard针对性', 'game/AI.js'),

    # BattleScene
    ('BattleScene', 'model', '属性', 'GameModel实例', 'scenes/BattleScene.js'),
    ('BattleScene', 'battleLogic', '属性', 'BattleLogic实例', 'scenes/BattleScene.js'),
    ('BattleScene', 'selectedCard', '属性', '当前选中的手牌index(-1=未选)', 'scenes/BattleScene.js'),
    ('BattleScene', 'onEnter()', '方法', '场景进入:初始化游戏/注册输入/加载资源', 'scenes/BattleScene.js'),
    ('BattleScene', 'onExit()', '方法', '场景退出:清理资源/注销输入', 'scenes/BattleScene.js'),
    ('BattleScene', 'onUpdate(dt)', '方法', '调用battleLogic.update(dt)', 'scenes/BattleScene.js'),
    ('BattleScene', 'onRender()', '方法', '渲染:背景→大殿→阵法→单位→特效→UI', 'scenes/BattleScene.js'),
    ('BattleScene', 'onCardTap(index)', '方法', '手牌点击回调:选中/取消/出牌', 'scenes/BattleScene.js'),
    ('BattleScene', 'onGridTap(gridX, gridY)', '方法', '格子点击回调:布阵/施法选目标', 'scenes/BattleScene.js'),
]
r = 4
for c in classes:
    layer = c[4].split('/')[0]
    if 'core' in layer: fill = BG_ENGINE
    elif 'scenes' in layer: fill = BG_RENDER
    elif 'config' in layer: fill = BG_DATA
    else: fill = BG_GAME
    data_row(ws, list(c), r, fill=fill, code_cols=[1,2])
    ws.row_dimensions[r].height = 24; r += 1

# ================================================================
# Sheet 13: 函数清单（每个函数的签名+输入输出+实现要点）
# ================================================================
ws = wb.create_sheet('13.函数清单')
set_widths(ws, [10, 20, 45, 18, 18, 30, 8])
title_row(ws, '函数清单 — 每个函数的签名/输入/输出/实现要点', 7)
header_row(ws, ['模块', '函数名', '函数签名', '输入参数', '返回值', '实现要点', '状态'], 3)

funcs = [
    # Director
    ('Director', 'init', 'init()', '无', 'void', '获取canvas 2D context→创建SceneManager→创建Input→调用start()', '待办'),
    ('Director', 'loop', 'loop(timestamp)', 'timestamp:帧时间戳', 'void', 'dt=(timestamp-lastTime)/1000→dt=Math.min(dt,0.033)→update(dt)→render()→requestAnimationFrame(loop)', '待办'),
    ('Director', 'update', 'update(dt)', 'dt:帧间隔(秒)', 'void', '获取当前场景→调用scene.onUpdate(dt)', '待办'),
    ('Director', 'render', 'render()', '无', 'void', '清屏→获取当前场景→调用scene.onRender()', '待办'),

    # Input
    ('Input', 'register', 'register(id,x,y,w,h,cb)', 'id:区域唯一标识, x/y/w/h:区域, cb:回调', 'void', 'push到touchRegions数组', '待办'),
    ('Input', 'onTouchStart', 'onTouchStart(touchX,touchY)', '触摸坐标', 'void', '遍历touchRegions→判断点是否在矩形内→命中则调用cb', '待办'),
    ('Input', 'clear', 'clear()', '无', 'void', 'touchRegions=[](场景切换时必须调用)', '待办'),

    # GameModel
    ('GameModel', 'init', 'init(deck0, deck1)', '双方卡组ID数组', 'void', '创建2个Player→设置大殿30血→灵力5→卡组→抽3张起手→time=180', '待办'),
    ('GameModel', 'getAllUnits', 'getAllUnits()', '无', '[Unit]', 'concat(players[0].units, players[1].units)', '待办'),
    ('GameModel', 'getEnemyUnits', 'getEnemyUnits(owner)', 'owner:0或1', '[Unit]', '返回对方玩家的units', '待办'),

    # Unit
    ('Unit', 'fromCard', 'Unit.fromCard(cardId,owner,x,y)', '卡牌ID/归属/坐标', 'Unit实例', '从Cards表读属性→创建Unit→设置facing(owner=0向上=1, owner=1向下=-1)', '待办'),
    ('Unit', 'takeDamage', 'takeDamage(amount,attacker)', '伤害量/攻击者', '实际伤害', 'hp-=amount→如果hp<=0则state=dead→触发死亡效果(自爆等)', '待办'),
    ('Unit', 'getEffectiveSpeed', 'getEffectiveSpeed()', '无', '实际移速', 'speed=baseSpeed→遍历buffs:减速则-0.5,加速则+0.5,定身则0→返回speed', '待办'),
    ('Unit', 'addBuff', 'addBuff(type,value,duration)', '类型/数值/持续秒', 'void', 'push到buffs数组→updateBuffs中倒计时', '待办'),
    ('Unit', 'updateBuffs', 'updateBuffs(dt)', '帧间隔', 'void', '遍历buffs→duration-=dt→<=0则移除→重算speed', '待办'),

    # MovementSystem
    ('MovementSystem', 'moveUnit', 'moveUnit(unit,dt)', '单位/帧间隔', 'void', '★核心★: if state==walking: target=findTarget()→if target且checkCollision: state=fighting→else: y+=speed*dt*facing→checkHallReach()', '待办'),
    ('MovementSystem', 'findTarget', 'findTarget(unit,model)', '单位/游戏状态', 'Unit/Formation/null', '获取敌方单位+阵法→过滤同列(x相同)→按y距离排序→返回最近的且在攻击范围内→不在范围内返回null(继续走)', '待办'),
    ('MovementSystem', 'checkHallReach', 'checkHallReach(unit,model)', '单位/游戏状态', 'bool', 'if owner==0且y>=8: return true→if owner==1且y<=0: return true', '待办'),

    # CombatSystem
    ('CombatSystem', 'attack', 'attack(attacker,target,dt)', '攻击者/目标/帧间隔', 'void', 'if now-lastAttackTime>=interval: 根据类型调用melee/ranged/aoe→重置lastAttackTime', '待办'),
    ('CombatSystem', 'meleeAttack', 'meleeAttack(attacker,target)', '攻击者/目标', 'void', 'target.takeDamage(attacker.atk)→attacker.takeDamage(target.atk)(互扣)→检查死亡', '待办'),
    ('CombatSystem', 'rangedAttack', 'rangedAttack(attacker,target)', '攻击者/目标', 'void', 'target.takeDamage(attacker.atk)→攻击者不掉血', '待办'),
    ('CombatSystem', 'handleKill', 'handleKill(killer,victim)', '击杀者/被杀者', 'void', '★核心修复★: killer.state=walking→killer.target=null→继续推进(不消失)→触发victim死亡效果', '待办'),
    ('CombatSystem', 'damageHall', 'damageHall(player,amount)', '玩家/伤害量', 'void', 'if player.hallShield>0: return(护盾免疫)→else: player.hallHp-=amount', '待办'),

    # SpellSystem
    ('SpellSystem', 'cast', 'cast(cardId,caster,target,model)', '卡牌ID/施法方/目标/状态', 'void', 'switch(cardId)→分支调用对应cast方法', '待办'),
    ('SpellSystem', 'castWanJian', 'castWanJian(caster)', '施法方', 'void', '遍历己方所有单位→atk+=1→addBuff("speed",0.3,5)', '待办'),
    ('SpellSystem', 'castWuLei', 'castWuLei(target,model)', '目标位置/状态', 'void', '获取目标3格内敌方单位/阵法→各受4伤', '待办'),
    ('SpellSystem', 'castJinZhong', 'castJinZhong(caster)', '施法方', 'void', 'player.hallShield=3(3秒免疫)', '待办'),
    ('SpellSystem', 'castYiShan', 'castYiShan(targets)', '目标区域单位数组', 'void', '遍历targets→y-=2*facing(推后)→takeDamage(1)', '待办'),
    ('SpellSystem', 'castKunXian', 'castKunXian(target)', '目标单位', 'void', 'target.addBuff("stun",0,2)(定身2秒,speed=0)', '待办'),

    # ElderSkillSystem
    ('ElderSkillSystem', 'triggerRandom', 'triggerRandom(elder,model)', '长老/状态', 'void', 'branches=[飞剑,丹药,符箓,御兽]→random=Math.floor(Math.random()*4)→调用对应方法', '待办'),
    ('ElderSkillSystem', 'castFlyingSword', 'castFlyingSword(elder,model)', '长老/状态', 'void', '获取长老周围3格内敌方→各受3伤(范围伤)', '待办'),
    ('ElderSkillSystem', 'castPill', 'castPill(elder,model)', '长老/状态', 'void', '长老自身hp+3→周围己方单位各hp+2(不超maxHp)', '待办'),
    ('ElderSkillSystem', 'castBeast', 'castBeast(elder,model)', '长老/状态', 'void', '在长老位置召唤1个灵兽Unit(owner同长老)', '待办'),

    # AI
    ('AI', 'update', 'update(dt,model)', '帧间隔/状态', 'void', 'thinkTimer+=dt→if thinkTimer>=thinkInterval: think(model)→thinkTimer=0', '待办'),
    ('AI', 'think', 'think(model)', '游戏状态', 'void', 'energy=model.players[1].energy→ratio=decideAttackRatio()→card=pickCard(energy,ratio)→if card: 执行出牌/布阵/施法', '待办'),
    ('AI', 'decideAttackRatio', 'decideAttackRatio(model)', '游戏状态', '0~1攻守比', 'hpPercent=hallHp/hallMaxHp→if>0.6: return 0.7→if<0.3: return 0.2→else: return 0.4', '待办'),
    ('AI', 'pickCard', 'pickCard(energy,ratio)', '灵力/攻守比', '卡牌index或null', '过滤hand中energy够的牌→if ratio>0.5: 优先攻方单位→else: 优先阵法→随机选一张', '待办'),

    # BattleLogic
    ('BattleLogic', 'update', 'update(dt)', '帧间隔', 'void', '★主入口★: updateEnergy→updateDraw→updateUnits→updateFormations→updateElders→updateAI→removeDead→checkBattleEnd', '待办'),
    ('BattleLogic', 'spawnUnit', 'spawnUnit(player,cardId,x)', '玩家/卡牌ID/列x', 'Unit', 'Unit.fromCard(cardId,player.id,x,出生y)→player.units.push(unit)→扣灵力→手牌打出', '待办'),
    ('BattleLogic', 'placeFormation', 'placeFormation(player,cardId,gridX,gridY)', '玩家/卡牌/格子坐标', 'bool', '检查格子冷却→if冷却中:return false→创建Formation→扣灵力→手牌打出→return true', '待办'),
    ('BattleLogic', 'castSpell', 'castSpell(player,cardId,target)', '玩家/卡牌/目标', 'void', '扣灵力→SpellSystem.cast(cardId,player,target,model)→手牌打出', '待办'),

    # Deck
    ('Deck', 'init', 'init(cardIds)', '卡组ID数组', 'void', 'drawPile=[...cardIds]→shuffle()→for i<3: draw()', '待办'),
    ('Deck', 'draw', 'draw()', '无', '卡牌ID或null', 'if hand.length>=maxHand: return null→if drawPile.length==0: return null→card=drawPile.pop()→hand.push(card)→return card', '待办'),
    ('Deck', 'canPlay', 'canPlay(handIndex,energy)', '手牌index/灵力', 'bool', 'card=hand[handIndex]→cost=Cards[card].cost→return energy>=cost', '待办'),

    # BattleScene
    ('BattleScene', 'onRender', 'onRender()', '无', 'void', 'renderBackground()→renderHalls()→renderFormations()→renderUnits()(y排序)→renderEffects()→renderUI()', '待办'),
    ('BattleScene', 'renderUnits', 'renderUnits()', '无', 'void', 'allUnits=getAllUnits()→sort by y(小的后画=远的先画)→遍历绘制色块+名字+血条', '待办'),
    ('BattleScene', 'onCardTap', 'onCardTap(index)', '手牌index', 'void', 'if selectedCard==index: 取消选择→else: if canPlay: selectedCard=index→else: 提示灵力不足', '待办'),
    ('BattleScene', 'onGridTap', 'onGridTap(gridX,gridY)', '格子坐标', 'void', 'if selectedCard==-1: return→card=hand[selectedCard]→if 阵法: placeFormation→if 法术: castSpell→if 单位: spawnUnit→selectedCard=-1', '待办'),

    # BattleChecker
    ('BattleChecker', 'checkEnd', 'checkEnd(model)', '游戏状态', '0/1/null', 'if 任一大殿hp<=0: return 胜者→if time<=0: if平局进入加时 else return 血量高者→加时结束return灵力多者', '待办'),
]
r = 4
for f in funcs:
    mod = f[0]
    if mod in ('Director','Input'): fill = BG_ENGINE
    elif mod in ('BattleScene',): fill = BG_RENDER
    elif mod in ('GameModel','Unit','Deck'): fill = BG_DATA
    else: fill = BG_GAME
    data_row(ws, list(f), r, fill=fill, code_cols=[2,3,4,5])
    ws.row_dimensions[r].height = 36; r += 1

# ================================================================
# Sheet 14: 数据结构（所有数据结构定义）
# ================================================================
ws = wb.create_sheet('14.数据结构')
set_widths(ws, [16, 16, 60, 14])
title_row(ws, '数据结构定义 — 所有核心数据格式', 4)
header_row(ws, ['结构名', '类型', '字段定义', '用途'], 3)

structs = [
    ('CardConfig', 'Object', '{id, name, type, faction, cost, hp, atk, speed, range, interval, traits[], rarity, desc}', '卡牌配置表(静态数据)'),
    ('Unit', 'Object', '{id, cardId, owner, x, y, hp, maxHp, atk, speed, baseSpeed, range, interval, lastAttackTime, target, state, traits[], buffs[], facing, isElder, elderTimer}', '单位实例(运行时)'),
    ('Formation', 'Object', '{id, cardId, owner, gridX, gridY, hp, maxHp, atk, range, interval, lastAttackTime, isActive, traits[]}', '阵法实例(运行时)'),
    ('Buff', 'Object', '{type: "slow"|"speed"|"stun"|"shield"|"atkBoost", value, duration, elapsed}', 'buff/debuff'),
    ('Player', 'Object', '{id, hallHp, hallMaxHp, hallShield, energy, energyMax, energyTimer, deck, units[], formations[], formationCooldowns:{}}', '玩家状态'),
    ('GameModel', 'Object', '{time, elapsedTime, state, winner, players:[Player,Player]}', '游戏总状态'),
    ('TouchRegion', 'Object', '{id, x, y, w, h, cb}', '可点击区域注册'),
    ('Particle', 'Object', '{x, y, vx, vy, life, maxLife, color, size}', '粒子(特效)'),
    ('AIState', 'Object', '{playerId, difficulty, thinkTimer, thinkInterval}', 'AI状态'),
    ('FormationCooldown', 'Map', '{"gridX,gridY": expireTimestamp}', '阵法格子冷却(防无限布阵)'),
    ('BattleResult', 'Object', '{winner, destroyRate0, destroyRate1, duration, rewards}', '战斗结果(结算用)'),
    ('DeckPreset', 'Object', '{name, desc, cardIds[]}', '预设卡组'),
    ('SaveData(V1.5)', 'Object', '{uid, nickname, realm, spirit, jade, cardsOwned{}, decks[], battleRecords[]}', '云存档数据'),
]
r = 4
for s in structs:
    data_row(ws, list(s), r, fill=BG_DATA, code_cols=[2,3])
    ws.row_dimensions[r].height = 36; r += 1

# 常量定义
r += 1
section_row(ws, 'Constants.js 全局常量定义', 4, r); r += 1
header_row(ws, ['常量名', '类型', '值', '说明'], r); r += 1
consts = [
    ('BOARD_LENGTH', 'int', '9', '棋盘长度(0~8, 含两端大殿)'),
    ('MAIN_LANE_X', 'int', '0', '主路x坐标'),
    ('LEFT_ARRAY_X', 'int', '-1', '左阵法区x坐标'),
    ('RIGHT_ARRAY_X', 'int', '1', '右阵法区x坐标'),
    ('HALL_HP', 'int', '30', '大殿血量'),
    ('ENERGY_START', 'int', '5', '灵力起始值'),
    ('ENERGY_MAX_START', 'int', '5', '灵力上限起始'),
    ('ENERGY_MAX_CAP', 'int', '10', '灵力上限封顶'),
    ('ENERGY_REGEN_INTERVAL', 'float', '2.8', '灵力回复间隔(秒)'),
    ('ENERGY_MAX_GROW_INTERVAL', 'int', '30', '灵力上限增长间隔(秒)'),
    ('BATTLE_TIME', 'int', '180', '单局时长(秒)'),
    ('OVERTIME', 'int', '60', '加时时长(秒)'),
    ('OVERTIME_ENERGY_MULT', 'float', '1.5', '加时灵力倍率'),
    ('HAND_SIZE', 'int', '4', '手牌数量'),
    ('DECK_SIZE', 'int', '8', '卡组数量'),
    ('DRAW_DELAY', 'float', '2.0', '打出后抽牌延迟(秒)'),
    ('FORMATION_COOLDOWN', 'float', '8.0', '阵法格子冷却(秒)'),
    ('ELDER_SKILL_INTERVAL', 'float', '5.0', '长老技能间隔(秒)'),
    ('MAX_UNITS', 'int', '30', '同屏单位上限'),
    ('REMOTE_RANGE', 'int', '3', '远程单位射程(格)'),
    ('SLOW_AMOUNT', 'float', '0.5', '减速幅度'),
    ('SLOW_DURATION', 'float', '2.0', '减速持续(秒)'),
    ('REFLECT_RATIO', 'float', '0.5', '反震比例'),
    ('SHIELD_DURATION', 'float', '3.0', '护盾持续(秒)'),
    ('STUN_DURATION', 'float', '2.0', '定身持续(秒)'),
    ('PUSH_BACK_GRIDS', 'int', '2', '推后格数'),
    ('SPEED_BUFF_DURATION', 'float', '5.0', '加速持续(秒)'),
    ('SILENCE_DURATION', 'float', '3.0', '禁阵持续(秒)'),
    ('AI_THINK_EASY', 'float', '3.5', '简单AI思考间隔'),
    ('AI_THINK_NORMAL', 'float', '2.5', '普通AI思考间隔'),
    ('AI_THINK_HARD', 'float', '1.5', '困难AI思考间隔'),
    ('TARGET_FPS', 'int', '60', '目标帧率'),
    ('MAX_DT', 'float', '0.033', '单帧最大时间步(秒)'),
    ('COLOR_ATTACKER', 'string', '#E74C3C', '攻方颜色(红)'),
    ('COLOR_DEFENDER', 'string', '#3498DB', '守方颜色(蓝)'),
    ('COLOR_ENERGY', 'string', '#9B59B6', '灵力条颜色(紫)'),
    ('COLOR_HALL', 'string', '#F39C12', '大殿颜色(金)'),
    ('COLOR_FORMATION', 'string', '#1ABC9C', '阵法颜色(青)'),
]
for c in consts:
    data_row(ws, list(c), r, fill=BG_ENGINE, code_cols=[1,3])
    ws.row_dimensions[r].height = 20; r += 1

# ================================================================
# Sheet 15: 调用关系（模块间依赖）
# ================================================================
ws = wb.create_sheet('15.调用关系')
set_widths(ws, [18, 14, 18, 40, 8])
title_row(ws, '模块调用关系 — 谁调用谁（依赖关系）', 5)
header_row(ws, ['调用方', '方向', '被调用方', '调用说明', '版本'], 3)

calls = [
    ('Director', '→', 'SceneManager', '主循环中获取当前场景,调用onUpdate/onRender', 'V1'),
    ('Director', '→', 'Input', 'init中创建Input,绑定触摸事件', 'V1'),
    ('SceneManager', '→', 'BattleScene/ResultScene', '管理场景切换,调用onEnter/onExit', 'V1'),
    ('BattleScene', '→', 'GameModel', '持有游戏状态,渲染时读取', 'V1'),
    ('BattleScene', '→', 'BattleLogic', '每帧调用battleLogic.update(dt)', 'V1'),
    ('BattleScene', '→', 'Input', 'onEnter时注册手牌/格子点击区域', 'V1'),
    ('BattleScene', '→', 'Renderer', '渲染时调用绘制方法', 'V1'),
    ('BattleScene', '→', 'ParticleSystem', '渲染特效', 'V1'),
    ('BattleScene', '→', 'HandBar/EnergyBar/HUD', '渲染UI组件', 'V1'),
    ('BattleLogic', '→', 'MovementSystem', 'updateUnits中调用moveUnit', 'V1'),
    ('BattleLogic', '→', 'CombatSystem', '交战时调用attack/handleKill', 'V1'),
    ('BattleLogic', '→', 'SpellSystem', 'castSpell时调用', 'V1'),
    ('BattleLogic', '→', 'ElderSkillSystem', 'updateElders中调用', 'V1'),
    ('BattleLogic', '→', 'AI', 'updateAI中调用ai.update', 'V1'),
    ('BattleLogic', '→', 'BattleChecker', 'checkBattleEnd中调用', 'V1'),
    ('BattleLogic', '→', 'GameModel', '读写游戏状态', 'V1'),
    ('BattleLogic', '→', 'Player', '读写玩家状态(灵力/单位/阵法)', 'V1'),
    ('MovementSystem', '→', 'GameModel', 'findTarget时获取敌方单位/阵法', 'V1'),
    ('MovementSystem', '→', 'Unit', '操作unit.state/y/target', 'V1'),
    ('CombatSystem', '→', 'Unit/Formation', '调用takeDamage', 'V1'),
    ('CombatSystem', '→', 'Player', 'damageHall时调用', 'V1'),
    ('SpellSystem', '→', 'Unit', 'addBuff/takeDamage', 'V1'),
    ('SpellSystem', '→', 'Formation', 'isActive=false(禁阵)', 'V1'),
    ('SpellSystem', '→', 'Player', 'hallShield(金钟罩)', 'V1'),
    ('ElderSkillSystem', '→', 'CombatSystem', '飞剑分支调用aoeAttack', 'V1'),
    ('ElderSkillSystem', '→', 'Unit', '丹药分支调用回血/御兽分支召唤', 'V1'),
    ('ElderSkillSystem', '→', 'BattleLogic', '召唤灵兽调用spawnUnit', 'V1'),
    ('AI', '→', 'GameModel', '读取状态决策', 'V1'),
    ('AI', '→', 'BattleLogic', '执行出牌spawnUnit/placeFormation/castSpell', 'V1'),
    ('AI', '→', 'Deck', '读取手牌决策', 'V1'),
    ('Deck', '→', 'Cards', '读卡牌属性(费用等)', 'V1'),
    ('Unit.fromCard', '→', 'Cards', '从卡牌表读属性创建实例', 'V1'),
    ('Formation.fromCard', '→', 'Cards', '从卡牌表读属性创建实例', 'V1'),
    ('HandBar', '→', 'Deck', '读取手牌数据渲染', 'V1'),
    ('HandBar', '→', 'Player', '读取灵力判断是否灰显', 'V1'),
    ('EnergyBar', '→', 'Player', '读取灵力/上限渲染', 'V1'),
    ('HUD', '→', 'GameModel', '读取时间/大殿血量渲染', 'V1'),
    ('ResultScene', '→', 'GameModel', '读取胜负/摧毁度渲染', 'V1'),
    ('CloudManager(V1.5)', '→', 'wx.cloud', '调用云函数/数据库', 'V1.5'),
    ('NetworkClient(V2)', '→', 'WebSocket', '实时通信', 'V2'),
    ('SyncManager(V2)', '→', 'BattleLogic', '同步状态到远端', 'V2'),
]
r = 4
for c in calls:
    v = c[4]
    if v == 'V1.5': fill = BG_V15
    elif v == 'V2': fill = BG_V2
    else: fill = None
    data_row(ws, list(c), r, fill=fill, code_cols=[1,3])
    ws.row_dimensions[r].height = 22; r += 1

# 每帧执行流程
r += 1
section_row(ws, '每帧执行流程（BattleLogic.update调用顺序）', 5, r); r += 1
flow = [
    '1. updateEnergy(dt)         → 双方灵力实时回复+上限增长',
    '2. updateDraw(dt)           → 双方手牌抽牌计时→到时补牌',
    '3. updateUnits(dt)          → 遍历所有单位:',
    '     3a. unit.updateBuffs(dt)      → 更新buff计时',
    '     3b. MovementSystem.moveUnit() → 移动/找目标/碰撞',
    '     3c. CombatSystem.attack()     → 交战伤害结算',
    '     3d. MovementSystem.checkHallReach() → 到大殿伤害',
    '4. updateFormations(dt)     → 遍历阵法→主动攻击经过敌人',
    '5. updateElders(dt)         → 长老技能计时→随机释放',
    '6. updateAI(dt)             → AI决策→出牌/布阵/施法',
    '7. removeDeadUnits()        → 清理死亡单位→触发死亡效果(自爆)',
    '8. checkBattleEnd()         → 检查胜负→设置state/winner',
    '9. time-=dt                 → 倒计时',
]
for line in flow:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value=line)
    c.font = F_CODE; c.alignment = A_LEFT; c.border = BD
    r += 1

# ================================================================
# Sheet 16: 技术选型与方案对比
# ================================================================
ws = wb.create_sheet('16.技术选型')
set_widths(ws, [16, 20, 22, 22, 30])
title_row(ws, '技术选型 — 每个技术决策的对比与理由', 5)
header_row(ws, ['决策点', '选定方案', '备选A', '备选B', '选择理由'], 3)

choices = [
    ('渲染方式', 'Canvas 2D', 'WebGL', 'DOM/CSS', 'V1单位少Canvas够用;WebGL复杂原生手写成本高;DOM不适合实时多单位'),
    ('框架', '自研轻量框架', 'Cocos Creator', 'Laya', '自研最轻量包体最小;Cocos包大但功能全(V1.5可考虑);Laya学习成本'),
    ('动画', '色块占位(V1)/Spine(V1.5)', '序列帧', '龙骨DragonBones', '色块验证玩法最快;Spine运行时换装方便生态好;序列帧包体大'),
    ('碰撞检测', 'AABB矩形', '圆形碰撞', '物理引擎', '棋盘格子制AABB最简单准确;圆形不够精确;物理引擎过度设计'),
    ('单位排序', 'y轴排序', 'z-index分层', '不排序', 'y排序模拟深度遮挡最自然;z-index手动管理麻烦;不排序会穿模'),
    ('AI决策', '状态机+规则', '行为树', '机器学习', '规则简单可控易调参;行为树过度设计;ML不切实际'),
    ('事件通信', '发布订阅(EventSystem)', '直接调用', '全局变量', '解耦最好;直接调用耦合高;全局变量不可维护'),
    ('存档(V1.5)', '微信云数据库', '本地wx.storage', '自建服务器', '云开发免运维与微信打通;本地不能跨设备;自建成本高'),
    ('PvP(V2)', 'WebSocket状态同步', '帧同步', 'HTTP轮询', '状态同步流量小断线易恢复;帧同步作弊风险大;轮询延迟高'),
    ('资源加载', '预加载+按需加载', '全预加载', '懒加载', '核心资源预加载体验好;全预加载启动慢;懒加载会卡'),
    ('包体管理', '主包<4MB+分包', '全主包', '全分包', '主包放核心代码快速启动;美术分包按需下载;全主包易超限'),
    ('音频', 'wx.createInnerAudioContext', 'Web Audio API', 'wx.getBackgroundAudioManager', 'InnerAudioContext适合SFX;WebAudio兼容性差;Background适合BGM'),
]
r = 4
for c in choices:
    data_row(ws, list(c), r, fill=BG_ENGINE)
    ws.row_dimensions[r].height = 36; r += 1

# ================================================================
# Sheet 17: 代码规范
# ================================================================
ws = wb.create_sheet('17.代码规范')
set_widths(ws, [16, 65])
title_row(ws, '代码规范 — 统一风格保证可维护性', 2)

r = 3
sections = [
    ('文件组织', [
        '每个类/模块一个文件,文件名=类名.js',
        '目录结构: js/core(引擎) / js/config(配置) / js/game(逻辑) / js/scenes(场景) / js/ui(界面)',
        '文件头注释: 文件用途+作者+日期',
    ]),
    ('命名规范', [
        '类名: PascalCase (Director, GameModel, BattleLogic)',
        '函数名/变量名: camelCase (moveUnit, lastAttackTime)',
        '常量: UPPER_SNAKE_CASE (BOARD_LENGTH, ENERGY_START)',
        '私有成员: 下划线前缀 (_privateMethod)',
        '事件名: namespace:action (unit:dead, spell:cast)',
    ]),
    ('代码结构', [
        '每个类: constructor→属性→公开方法→私有方法',
        '函数体不超过50行,超过则拆分',
        '每个函数有注释: 作用/参数/返回值',
        '复杂逻辑有行内注释解释"为什么"而非"做什么"',
    ]),
    ('数据驱动', [
        '所有卡牌属性在Cards.js配置,不在代码中硬编码',
        '所有可调参数在Constants.js,不在代码中硬编码',
        '数值平衡只改配置文件,不改逻辑代码',
        '新增卡牌只需在Cards.js加一条配置,不改逻辑',
    ]),
    ('性能规范', [
        '避免在update/render循环中创建对象(用对象池)',
        '避免在循环中频繁GC(复用数组/对象)',
        '同屏单位超过MAX_UNITS不再生成',
        '粒子用对象池复用',
        '渲染按y排序减少遮挡计算',
    ]),
    ('错误处理', [
        '外部输入(触摸/网络)需要容错',
        '数组访问前检查length',
        '单位/阵法操作前检查是否已死亡',
        '灵力扣除前检查是否足够',
    ]),
    ('微信小游戏适配', [
        '不使用DOM API(document/window等)',
        '使用wx.xxx API替代浏览器API',
        'canvas用wx.createCanvas()创建',
        '触摸用wx.onTouchStart/Move/End',
        '音频用wx.createInnerAudioContext()',
        '存储用wx.setStorageSync/getStorageSync',
    ]),
]
for sec_name, items in sections:
    section_row(ws, sec_name, 2, r); r += 1
    for item in items:
        ws.cell(row=r, column=1, value='•').font = F_CELL
        ws.cell(row=r, column=1).alignment = A_CENTER; ws.cell(row=r, column=1).border = BD
        ws.cell(row=r, column=2, value=item).font = F_CELL
        ws.cell(row=r, column=2).alignment = A_LEFTT; ws.cell(row=r, column=2).border = BD
        ws.row_dimensions[r].height = 24; r += 1
    r += 1

# 保存
wb.save(PATH)
print(f'Excel已扩充: {PATH}')
print(f'总Sheet数: {len(wb.sheetnames)}')
print(f'新增Sheet: {wb.sheetnames[10:]}')
print(f'技术架构: {len(arch)}个模块')
print(f'类设计: {len(classes)}个类成员')
print(f'函数清单: {len(funcs)}个函数')
print(f'数据结构: {len(structs)}个结构 + {len(consts)}个常量')
print(f'调用关系: {len(calls)}条依赖')
print(f'技术选型: {len(choices)}个决策')
