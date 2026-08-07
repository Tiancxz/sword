#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""《宗门论道》v5 Excel生成 — Sheet部分"""
import openpyxl, json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

PATH='/workspace/宗门论道_开发管理工具包.xlsx'
with open('/workspace/_data.json') as f:
    D=json.load(f)

wb=openpyxl.Workbook()
C_INK='2C3E50';C_GOLD='B7950B';C_BLUE='2874A6';C_TEAL='117A65';C_RED='C0392B';C_PURPLE='8E44AD';C_GREEN='27AE60';C_ORANGE='D35400'
L1_FILLS={'A':('1A5276','FFFFFF'),'B':('2874A6','FFFFFF'),'C':('117A65','FFFFFF'),'D':('B7950B','FFFFFF'),'E':('8E44AD','FFFFFF'),'F':('C0392B','FFFFFF'),'G':('D35400','FFFFFF'),'H':('34495E','FFFFFF')}
L2_FILL='D6EAF8';L3_FILL='FDF2E9';L4_FILL='F4ECF7';L5_FILL='E8F8F5'
ST_TODO='FADBD8';ST_DOING='FEF9E7';ST_DONE='D5F5E3'
PHASE_COLORS={'P1':'1A5276','P2':'2874A6','P3':'117A65','P4':'B7950B','P5':'8E44AD','P6':'C0392B','P7':'D35400','P8':'27AE60','P9':'2E86C1','P10':'85C1E9','P11':'F8C471','P12':'F0B27A','P13':'BB8FCE','P14':'F1948A','P16':'5D6D7E','P17':'85929E','P18':'AAB7B8'}
F_BIG=Font(name='微软雅黑',size=16,bold=True,color='FFFFFF');F_TITLE=Font(name='微软雅黑',size=13,bold=True,color='FFFFFF')
F_HEADER=Font(name='微软雅黑',size=11,bold=True,color='FFFFFF');F_L1=Font(name='微软雅黑',size=11,bold=True,color='FFFFFF')
F_L2=Font(name='微软雅黑',size=11,bold=True,color=C_INK);F_CELL=Font(name='微软雅黑',size=10,color='2C3E50')
F_BOLD=Font(name='微软雅黑',size=10,bold=True,color='2C3E50');F_CODE=Font(name='Consolas',size=10,color='1A5276')
F_SMALL=Font(name='微软雅黑',size=9,color='7F8C8D');F_TAG=Font(name='微软雅黑',size=8,bold=True,color='FFFFFF')
F_GOLD=Font(name='微软雅黑',size=11,bold=True,color=C_GOLD);F_RESULT=Font(name='微软雅黑',size=10,bold=True,color=C_TEAL)
F_PHASE=Font(name='微软雅黑',size=12,bold=True,color='FFFFFF')
A_C=Alignment(horizontal='center',vertical='center',wrap_text=True)
A_L=Alignment(horizontal='left',vertical='center',wrap_text=True,indent=1)
A_LT=Alignment(horizontal='left',vertical='top',wrap_text=True,indent=1)
A_LM=Alignment(horizontal='left',vertical='center',wrap_text=True)
BD=Border(left=Side(style='thin',color='D5D8DC'),right=Side(style='thin',color='D5D8DC'),top=Side(style='thin',color='D5D8DC'),bottom=Side(style='thin',color='D5D8DC'))

# ===== 完成标准字典（每个L3任务的验收标准） =====
CRITERIA={
    'A1.01':'①node --check通过 ②Director.loop 60fps跑通','A1.02':'①scene.onUpdate被调用 ②scene为null不报错','A1.03':'①clearRect后画面清空 ②scene.onRender执行','A1.04':'①wx.createCanvas成功 ②屏幕尺寸获取正确',
    'A2.01':'①onTouchStart回调触发 ②坐标正确','A2.02':'①regions数组有数据 ②id唯一','A2.03':'①点击区域内触发cb ②区域外不触发','A2.04':'①regions清空为[]',
    'A3.01':'①listeners有对应事件 ②cb被加入数组','A3.02':'①所有cb被调用 ②data正确传递','A3.03':'①cb从数组移除 ②移除后不再触发',
    'A4.01':'①scenes有对应场景 ②name唯一','A4.02':'①onExit被调用 ②onEnter被调用 ③current更新',
    'A5.01':'①矩形正确绘制 ②颜色正确','A5.02':'①文字正确绘制 ②字体大小正确','A5.03':'①红绿条按比例 ②ratio=0时全红',
    'A6.01':'①粒子数量正确 ②初始位置正确','A6.02':'①粒子位置更新 ②life<=0被过滤','A6.03':'①alpha渐变正确 ②渲染后alpha=1',
    'A7.01':'①BGM播放 ②loop=true','A7.02':'①SFX播放 ②播放完销毁',
    'B1.01':'①module.exports正确 ②常量值可引用','B2.01':'①19张卡牌数据完整 ②字段无缺失','B2.02':'①get(id)返回正确卡牌 ②不存在时console.error',
    'B3.01':'①drawPile有数据 ②hand有3张 ③洗牌后顺序随机','B3.02':'①洗牌后顺序随机 ②不遗漏不重复','B3.03':'①手牌满返回null ②牌库空返回null ③正常抽牌','B3.04':'①2秒后抽牌 ②drawTimer归零','B3.05':'①灵力够返回true ②不够返回false','B3.06':'①手牌移除 ②drawTimer=2.0 ③返回cardId',
    'B4.01':'①3套卡组各8张 ②cardId引用正确',
    'C1.01':'①所有update被调用 ②time递减 ③不报错',
    'C2.01':'①返回Unit对象 ②字段完整 ③facing正确','C2.02':'①buff被加入 ②类型正确','C2.03':'①slow减0.5 ②speed加0.5 ③stun返回0','C2.04':'①hp减少 ②hp<=0时state=dead ③返回伤害值',
    'C3.01':'①walking时移动 ②遇敌变fighting ③到大殿damageHall','C3.02':'①同列敌人优先 ②最近距离优先 ③超射程返回null','C3.03':'①距离<=attackRange返回true','C3.04':'①owner=0到顶返回true ②owner=1到底返回true',
    'C4.01':'①fighting时攻击 ②interval间隔正确 ③击杀后handleKill','C4.02':'①双方同时掉血 ②掉血量=对方atk','C4.03':'①只有目标掉血 ②攻击者不掉血','C4.04':'①范围内所有目标受伤 ②伤害值正确','C4.05':'①killer.state=walking ②killer.target=null','C4.06':'①有护盾不扣血 ②hp<=0时state=ended ③winner正确',
    'C5.01':'①冷却中返回false ②布阵成功 ③扣灵力','C5.02':'①isActive=false跳过 ②范围内敌人受伤 ③interval间隔','C5.03':'①冷却中返回false ②可布阵返回true','C5.04':'①isActive=false ②silenceTimer倒数',
    'C6.01':'①switch正确分支 ②参数传递正确','C6.02':'①全己方单位atk+1 ②speed buff 5秒','C6.03':'①范围内敌方受伤 ②伤害=4 ③含阵法','C6.04':'①目标获得speed buff ②持续5秒','C6.05':'①目标阵法isActive=false ②3秒后恢复','C6.06':'①己方hallShield=3','C6.07':'①目标y后退2格 ②受伤1点','C6.08':'①目标获得stun buff ②持续2秒','C6.09':'①所有目标受伤 ②伤害=4',
    'C7.01':'①非长老跳过 ②5秒触发 ③timer归零','C7.02':'①4分支等概率 ②对应分支被调用','C7.03':'①范围内敌人受伤 ②伤害=3','C7.04':'①长老回血3 ②附近友军回血2','C7.05':'①前方敌人受伤 ②调用castTianLei','C7.06':'①灵兽被创建 ②加入units',
    'C8.01':'①2.8秒+1 ②加时倍率1.5 ③不超上限','C8.02':'①每30秒+1 ②封顶10',
    'C9.01':'①灵力够出兵 ②单位被创建 ③扣灵力','C9.02':'①灵力够施法 ②法术生效 ③扣灵力',
    'C10.01':'①hallHp<=0返回winner ②都活着返回null','C10.02':'①time>0返回null ②平局进加时 ③加时60秒','C10.03':'①加时结束按灵力判 ②返回winner',
    'C11.01':'①dead单位被移除 ②自爆触发','C11.02':'①被毁阵法移除 ②冷却8秒记录',
    'D1.01':'①timer到间隔触发 ②think被调用','D1.02':'①能选牌出牌 ②无牌时return ③不卡死','D1.03':'①血量>60%返回0.7 ②血量<30%返回0.2','D1.04':'①过滤不可出牌 ②ratio>0.5优先unit','D1.05':'①easy随机 ②normal靠大殿 ③hard堵前方',
    'E1.01':'①model初始化 ②BattleLogic创建 ③输入注册','E1.02':'①battleLogic.update被调用','E1.03':'①分层渲染顺序正确 ②无遗漏','E1.04':'①input清空 ②model=null',
    'E2.01':'①渐变正确 ②铺满屏幕','E2.02':'①格子线均匀 ②BOARD_LENGTH条',
    'E3.01':'①大殿色块正确 ②文字显示','E3.02':'①血条按比例 ②上下各一条','E3.03':'①震动偏移 ②shakeTimer递减',
    'E4.01':'①按y排序 ②颜色区分 ③名字显示','E4.02':'①血条在单位上方 ②按比例',
    'E5.01':'①光阵透明度 ②isActive区分颜色',
    'E6.01':'①4张手牌显示 ②费用显示 ③不可出灰色','E6.02':'①点击切换选中 ②再点取消',
    'E7.01':'①背景+前景 ②按比例 ③数字显示',
    'E8.01':'①左右各半 ②红绿按比例','E8.02':'①倒计时显示 ②<=30秒变红',
    'E9.01':'①点击选中 ②灵力不够return ③再点取消','E9.02':'①有选中才执行 ②出兵/布阵/施法正确 ③执行后取消',
    'E10.01':'①胜负文字 ②颜色正确','E10.02':'①摧毁度百分比 ②双方对比','E10.03':'①切换到battle场景',
    'E11.01':'①高亮第一张牌 ②箭头指向 ③点击后step=2','E11.02':'①高亮灵力条 ②3秒后step=3','E11.03':'①高亮阵法区 ②布阵后step=4','E11.04':'①箭头指大殿 ②2秒后结束',
    'F1.01':'①wx.login成功 ②openid获取 ③存Storage','F1.02':'①数据库写入成功 ②数据完整','F2.01':'①challenge记录创建 ②返回_id','F2.02':'①对手布局加载 ②AI模拟启动','F3.01':'①开放数据域收到消息 ②排行绘制','F3.02':'①分享面板弹出 ②title/imageUrl正确',
    'G1.01':'①level+1 ②hp/atk按公式增长','G1.02':'①star+1 ②境界正确进阶','G2.01':'①广告展示 ②看完触发cb','G2.02':'①支付调起 ②成功回调',
    'H1.01':'①WebSocket连接 ②onOpen触发 ③ping心跳','H1.02':'①断线后重连 ②状态恢复','H2.01':'①action序列化发送 ②服务端收到','H2.02':'①state解析 ②model.applyState执行','H3.01':'①匹配请求发送 ②返回roomId','H3.02':'①elo增减 ②rank更新',
}

def set_w(ws,widths):
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
def big_title(ws,text,cols,row=1,color=C_INK):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=cols)
    c=ws.cell(row=row,column=1,value=text);c.font=F_BIG;c.fill=PatternFill('solid',fgColor=color);c.alignment=A_C;ws.row_dimensions[row].height=44
def sub_title(ws,text,cols,row,color=C_GOLD):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=cols)
    c=ws.cell(row=row,column=1,value=text);c.font=F_TITLE;c.fill=PatternFill('solid',fgColor=color);c.alignment=A_L;ws.row_dimensions[row].height=30
def header_row(ws,headers,row,color=C_INK):
    for i,h in enumerate(headers,1):
        c=ws.cell(row=row,column=i,value=h);c.font=F_HEADER;c.fill=PatternFill('solid',fgColor=color);c.alignment=A_C;c.border=BD
    ws.row_dimensions[row].height=30
def calc_row_height(text,col_w=40):
    if not text: return 24
    text=str(text);wrapped=sum(max(1,(len(line)+col_w-1)//col_w) for line in text.split('\n'))
    return max(24,min(120,wrapped*16+8))

# ===== Sheet 0: 开发阶段规划 =====
ws=wb.active;ws.title='0.开发阶段规划';ws.sheet_view.showGridLines=False
set_w(ws,[4,6,16,26,30,36,36,22,8])
big_title(ws,'开发阶段规划 — 先做什么后做什么，每步目的是什么、能实现什么',9)
ws.merge_cells('A2:I2')
c=ws.cell(row=2,column=1,value='按依赖关系排出8个阶段，必须从P1到P8顺序做。每个阶段做完都有"能跑起来看到的东西"。')
c.font=F_SMALL;c.alignment=A_L;ws.row_dimensions[2].height=24
header_row(ws,['','阶段','阶段名称','做什么（内容）','目的（为什么先做这个）','能实现的进展（做完能看到什么）','验证标准（怎么算做完）','对应主表编号','周期'],3)
phases=[
    {'num':'P1','name':'引擎骨架','content':'渲染循环(Director)\n输入系统(Input)\n事件总线(EventSystem)\n场景管理(SceneManager)\n渲染器(Renderer)','purpose':'这是整个游戏的地基。没有渲染循环什么也跑不起来，没有输入系统什么也点不了。必须先搭好引擎框架，后面所有功能才能挂上去。','progress':'能看到一个黑色画布在60fps跑\n能在画布上画矩形/文字\n点屏幕能触发回调\n能切换空场景A和空场景B','verify':'①Director.loop 60fps跑通\n②点击屏幕console打印坐标\n③画矩形和文字能显示\n④场景A能切换到场景B','ref':'A1~A5','period':'1周'},
    {'num':'P2','name':'数据层','content':'常量配置(Constants)\n卡牌数据表(Cards)\n卡组手牌(Deck)\n预设卡组(DeckPresets)','purpose':'引擎能跑了，但游戏需要数据。卡牌属性、灵力参数、棋盘尺寸等所有数值必须先配置好，后面的战斗逻辑才能读数据创建单位。','progress':'19张卡牌数据完整可查\n8张卡组能洗牌+抽牌\n能console打印手牌和卡牌属性\n改Cards.js数值能立即生效','verify':'①Cards.get(id)返回完整属性\n②Deck.init后hand有3张牌\n③Deck.draw能抽牌\n④canPlay能正确判断灵力','ref':'B1~B4','period':'1周'},
    {'num':'P3','name':'核心战斗逻辑','content':'战斗主循环(BattleLogic)\n单位实体(Unit)\n移动系统(Movement)\n战斗系统(Combat)\n★击杀后继续推进★','purpose':'这是整个游戏好不好玩的命门。单位能不能走、能不能打、打完能不能继续冲——这3件事决定了游戏核心循环是否成立。必须验证"攻方打得动"。','progress':'★最重要的里程碑★\nAI vs AI模拟能跑通\n能看到单位在棋盘上移动\n能看到单位互相打\n能看到单位杀掉敌人后继续冲\n能看到单位到大殿造成伤害','verify':'①单位y坐标每帧变化(在走)\n②两个单位接触后互扣血\n③单位击杀后state=walking继续走\n④单位到大殿后大殿血量减少','ref':'C1~C4','period':'2周'},
    {'num':'P4','name':'扩展战斗机制','content':'阵法系统(布阵/攻击/冷却)\n法术系统(8种法术效果)\n长老技能(4分支随机)\n灵力系统(实时回复)\n出牌执行\n胜负判定\n死亡清理','purpose':'核心循环跑通了，但只有"走和打"太单调。阵法让守方能防，法术增加策略深度，长老增加变化性，灵力系统驱动出牌节奏。这些让游戏从"能跑"变成"好玩"。','progress':'完整单局能跑通\n能布阵拦截(阵法有血量会攻击)\n能放法术(8种效果各不相同)\n长老每5秒随机放技能\n灵力实时回复能出牌\n能判定胜负(大殿破/超时)','verify':'①布阵后阵法能攻击经过敌人\n②阵法被毁后格子8秒冷却\n③8种法术效果各自正确生效\n④长老每5秒释放1个随机分支\n⑤灵力每2.8秒+1\n⑥大殿血量归零能判定胜负','ref':'C5~C11','period':'2周'},
    {'num':'P5','name':'AI对手','content':'AI决策(灵力管理/出牌/布阵)\n3档难度(简单/普通/困难)','purpose':'战斗逻辑完整了，但需要对手。没有AI玩家无法测试游戏好不好玩。AI不需要完美，但要能模拟真人节奏出牌布阵。','progress':'能和AI完整打一局(纯逻辑)\nAI会出兵也会布阵\nAI会根据血量调整攻守\n简单AI弱/困难AI强有明显差异','verify':'①AI每2.5秒(普通)决策一次\n②AI会根据大殿血量调攻守比\n③简单/普通/困难行为有差异\n④AI vs AI能完整跑完一局不卡死','ref':'D1','period':'1周'},
    {'num':'P6','name':'渲染与交互','content':'战斗场景(分层渲染)\n背景/大殿/单位/阵法渲染\n手牌栏/灵力条/HUD\n出牌交互(选牌→选目标)\n结算场景\n新手引导\n粒子特效\n音频','purpose':'到这一步为止游戏都是"纯逻辑"。P6把所有逻辑变成"能看到能操作的画面"。这是从"程序员眼里的游戏"变成"玩家眼里的游戏"的关键。','progress':'★玩家可见里程碑★\n能在微信开发者工具里看到画面\n能看到棋盘/大殿/单位/阵法\n能点手牌出牌、点格子布阵\n能看到灵力条/血条/计时器\n战斗结束有结算页\n新手引导能走通4步','verify':'①真机/开发者工具能显示画面\n②点手牌→选目标→出牌流程通\n③单位移动/交战有视觉反馈\n④大殿受击有震动+闪烁\n⑤结算页显示胜负+摧毁度\n⑥新手引导4步走通','ref':'E1~E11,A6~A7','period':'2周'},
    {'num':'P7','name':'测试与平衡调优','content':'纯逻辑模拟器(AI vs AI 10局)\n平衡性调优(参数调整)\n性能测试(60fps/内存)\n真机测试\nBug修复','purpose':'游戏能玩了，但好不好玩是另一回事。P7用模拟器跑大量对局验证平衡——攻方能不能打进去、守方能不能防下来、有没有0%死局。','progress':'★平衡达标里程碑★\n模拟10局输出摧毁度/胜率数据\n参数调到双方摧毁度20~70%\n真机60fps不卡\n无崩溃无死锁','verify':'①模拟10局:摧毁度20~70%\n②胜率40~60%\n③0%摧毁度对局<10%\n④90%对局150~210秒结束\n⑤真机60fps同屏20单位不卡','ref':'全部联调','period':'1.5周'},
    {'num':'P8','name':'提交审核上线','content':'最终Bug修复\n微信小游戏审核提交\n审核反馈处理\n正式上线','purpose':'游戏做完要上线才能让玩家玩到。修仙题材过审风险低但仍需确保无违规内容。','progress':'★V1上线里程碑★\n微信审核通过\n玩家能搜索到并玩到游戏\nV1 MVP完成','verify':'①微信审核通过\n②玩家能正常游玩\n③无严重Bug\n④新手引导完成率>80%','ref':'—','period':'1周'},
]
r=4
for p in phases:
    color=PHASE_COLORS[p['num']]
    c=ws.cell(row=r,column=2,value=p['num']);c.font=F_PHASE;c.fill=PatternFill('solid',fgColor=color);c.alignment=A_C;c.border=BD
    c=ws.cell(row=r,column=3,value=p['name']);c.font=F_PHASE;c.fill=PatternFill('solid',fgColor=color);c.alignment=A_C;c.border=BD
    c=ws.cell(row=r,column=4,value=p['content']);c.font=F_CELL;c.alignment=A_LT;c.border=BD
    c=ws.cell(row=r,column=5,value=p['purpose']);c.font=F_CELL;c.alignment=A_LT;c.border=BD
    c=ws.cell(row=r,column=6,value=p['progress']);c.font=F_RESULT;c.alignment=A_LT;c.border=BD;c.fill=PatternFill('solid',fgColor='E8F8F5')
    c=ws.cell(row=r,column=7,value=p['verify']);c.font=F_CODE;c.alignment=A_LT;c.border=BD
    c=ws.cell(row=r,column=8,value=p['ref']);c.font=F_CODE;c.alignment=A_LT;c.border=BD;c.fill=PatternFill('solid',fgColor='FDF2E9')
    c=ws.cell(row=r,column=9,value=p['period']);c.font=F_BOLD;c.alignment=A_C;c.border=BD
    ws.row_dimensions[r].height=120;r+=1
r+=1
ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=9)
c=ws.cell(row=r,column=2,value='P1→P2→P3→P4→P5→P6→P7→P8 （必须按顺序，后面依赖前面）')
c.font=Font(name='微软雅黑',size=11,bold=True,color=C_GOLD);c.alignment=A_C;c.fill=PatternFill('solid',fgColor='FEF9E7');c.border=BD;ws.row_dimensions[r].height=30

# ===== Sheet 1: 设计思路与玩法（加地图示意图） =====
ws=wb.create_sheet('1.设计思路与玩法');ws.sheet_view.showGridLines=False
set_w(ws,[4,20,20,20,20,20,4])
big_title(ws,'《宗门论道》 — 游戏设计思路与玩法',7)
r=3
sub_title(ws,'一、游戏概述',7,r,C_GOLD);r+=1
for k,v in [('游戏名称','宗门论道（暂定）'),('游戏类型','实时策略卡牌（RTS-lite + 卡牌费用制）'),('游戏题材','东方修仙 / 宗门大战'),('运行平台','微信小游戏（Canvas 2D 渲染）'),('单局时长','3~5分钟（180秒正赛+60秒加时）'),('目标用户','18~35岁，喜欢轻度策略、修仙题材、碎片化对战的玩家'),('一句话简介','皇室战争式的实时卡牌攻防，修仙皮——你与对手各据山门一座，实时派弟子出征、布阵拦截、长老施法，摧毁对方宗门大殿者胜。')]:
    ws.cell(row=r,column=2,value=k).font=F_L2;ws.cell(row=r,column=2).fill=PatternFill('solid',fgColor=L2_FILL);ws.cell(row=r,column=2).alignment=A_C;ws.cell(row=r,column=2).border=BD
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=6)
    ws.cell(row=r,column=3,value=v).font=F_CELL;ws.cell(row=r,column=3).alignment=A_LM;ws.cell(row=r,column=3).border=BD;ws.row_dimensions[r].height=28;r+=1
r+=1
sub_title(ws,'二、核心玩法 — 实时双线互推',7,r,C_BLUE);r+=1
for line in ['你和对手各据一座宗门大殿（30血），实时同时出兵互攻，谁先拆掉对方大殿谁赢，3分钟时限。','','【你做什么】','  1. 出兵：花灵力出弟子，单位自动沿山道推进，遇敌自动交战，杀完继续冲','  2. 布阵：在主路两侧阵法区布阵法（不动，拦截+攻击经过的敌人），剑阵双方通用','  3. 施法：放法术（五雷正法范围炸、万剑归宗全队加速、金钟罩大殿免疫等）','  4. 出长老：花6费出金丹长老，每5秒随机释放飞剑/丹药/符箓/御兽分支技能','','【灵力系统】','  实时回复：每2.8秒+1，开局上限5随时间涨到10。灵力不够出不了牌。','']:
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
    c=ws.cell(row=r,column=2,value=line);c.font=F_GOLD if line.startswith('【') else F_CELL;c.alignment=A_LM;c.border=BD;ws.row_dimensions[r].height=24 if line else 8;r+=1
r+=1
sub_title(ws,'三、地图布局示意图（折中方案）',7,r,C_TEAL);r+=1
map_lines=[
    '┌─────────────────────────────────────────────┐',
    '│            ★ 敌方宗门大殿 (30血) ★             │  ← P1(上方)要拆的',
    '│  ┌───────┐  ┌─────────────┐  ┌───────┐       │',
    '│  │ 阵法区 │  │   主路(9格)   │  │ 阵法区 │       │  ← 两侧布阵拦截',
    '│  │ (左侧) │  │   单位推进↑   │  │ (右侧) │       │',
    '│  └───────┘  └─────────────┘  └───────┘       │',
    '│  ┌───────┐  ┌─────────────┐  ┌───────┐       │',
    '│  │ 阵法区 │  │   主路(9格)   │  │ 阵法区 │       │  ← 剑阵双方通用',
    '│  │ (左侧) │  │   单位推进↓   │  │ (右侧) │       │',
    '│  └───────┘  └─────────────┘  └───────┘       │',
    '│            ★ 我方宗门大殿 (30血) ★             │  ← P0(下方)要守的',
    '└─────────────────────────────────────────────┘',
    '',
    '说明：逻辑上下直推(实现简单) | 视觉做成山道攻山(修仙味) | 两侧放阵法(策略空间)',
    '单位2朝向(上/下+翻转) | 镜像对称(PVP公平) | 棋盘9格(含大殿)',
]
for line in map_lines:
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
    c=ws.cell(row=r,column=2,value=line)
    if line.startswith('说明'): c.font=F_SMALL
    elif '★' in line: c.font=Font(name='Consolas',size=10,bold=True,color=C_GOLD)
    else: c.font=Font(name='Consolas',size=10,color=C_TEAL)
    c.alignment=A_C if not line.startswith('说明') else A_L
    ws.row_dimensions[r].height=18 if line and not line.startswith('说明') else 24;r+=1
r+=1
sub_title(ws,'四、设计思路 — 为什么这么设计',7,r,C_PURPLE);r+=1
thoughts=[
    ('为什么做实时不做回合制？','回合制等对方出招很无聊。实时制双方同时操作互不等待，每秒都有决策压力，紧张感是核心体验。CR验证了这一点。'),
    ('为什么"击杀后继续推进"？','★最核心的设计修复★。旧版"遇阵即停打完即没"导致攻方永远过不去，0%摧毁度。改成击杀后不消失继续冲，攻方才有突破感。'),
    ('为什么阵法有8秒冷却？','防止守方无限补阵堵路。阵被破后该格8秒不能再布，攻方有突破窗口。创造"攻守博弈"节奏。'),
    ('为什么用折中视角？','纯斜45度等距视角美术成本3~4倍，碰撞复杂。折中方案：逻辑上下直推（实现简单），视觉做成山道攻山（保留修仙味）。'),
    ('为什么金丹长老是随机技能？','固定技能会变成"最优解"。随机性增加不可预测性，每次出长老都有惊喜/惊吓，增加对局变化和观赏性。'),
    ('为什么剑阵双方通用？','阵法只能防守时攻方缺乏掩护。双方通用后攻方可铺阵掩护推进，守方布阵拦截，策略空间更大。'),
    ('为什么选修仙题材？','①国内过审零风险②修仙认知度高③"攻山"主题与推塔玩法天然契合④市场有验证。现代战争过审风险极高。'),
    ('为什么V1先做3个流派？','MVP原则：用最少的验证核心玩法。剑修+符修+阵修覆盖攻防基本循环，验证好玩后再加更多。'),
]
for q,a in thoughts:
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=2)
    ws.cell(row=r,column=2,value='问').font=F_TAG;ws.cell(row=r,column=2).fill=PatternFill('solid',fgColor=C_PURPLE);ws.cell(row=r,column=2).alignment=A_C;ws.cell(row=r,column=2).border=BD
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=6)
    ws.cell(row=r,column=3,value=q).font=F_L2;ws.cell(row=r,column=3).alignment=A_LM;ws.cell(row=r,column=3).border=BD;ws.cell(row=r,column=3).fill=PatternFill('solid',fgColor=L4_FILL);ws.row_dimensions[r].height=26;r+=1
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=2)
    ws.cell(row=r,column=2,value='答').font=F_TAG;ws.cell(row=r,column=2).fill=PatternFill('solid',fgColor=C_TEAL);ws.cell(row=r,column=2).alignment=A_C;ws.cell(row=r,column=2).border=BD
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=6)
    ws.cell(row=r,column=3,value=a).font=F_CELL;ws.cell(row=r,column=3).alignment=A_LT;ws.cell(row=r,column=3).border=BD;ws.row_dimensions[r].height=50;r+=1

# ===== Sheet 2: 开发主表（加依赖+工时+行高自适应） =====
ws=wb.create_sheet('2.开发主表');ws.sheet_view.showGridLines=False
set_w(ws,[5,8,6,12,14,18,40,22,20,18,14,7,28,6,8])
big_title(ws,'开发主表 — 从系统到函数，逐层展开（含阶段+文件路径+依赖+工时+完成标准）',15)
header_row(ws,['层级','编号','阶段','L1 系统','L2 模块','L3 子功能','L4 实现原理','L5 函数','文件路径','输入 → 输出','前置依赖','工时h','完成标准','版本','状态'],2)
ws.freeze_panes='A3'
r=3
for row_data in D:
    row_list=list(row_data)
    while len(row_list)<15: row_list.append('')
    level=row_list[0];num=row_list[1];phase=row_list[2];sys_code=row_list[3];sys_name=row_list[4];mod=row_list[5];func=row_list[6];principle=row_list[7];fn=row_list[8];filepath=row_list[9];io=row_list[10];depends=row_list[11];hours=row_list[12];ver=row_list[13];status=row_list[14]
    criteria=CRITERIA.get(num,'') if level=='L3' else ''
    vals=[level,num,phase,sys_name if level=='L1' else '',mod,func,principle,fn,filepath,io,depends,hours,criteria,ver,status]
    for i,v in enumerate(vals,1):
        c=ws.cell(row=r,column=i,value=v);c.border=BD;c.alignment=A_LT
    lc=ws.cell(row=r,column=1);lc.font=F_TAG;lc.alignment=A_C
    if level=='L1':
        bg,fg=L1_FILLS.get(sys_code,('333333','FFFFFF'));lc.fill=PatternFill('solid',fgColor=bg);lc.value='L1\n系统';ws.row_dimensions[r].height=32
    elif level=='L2':
        lc.fill=PatternFill('solid',fgColor=L2_FILL);lc.value='L2\n模块';ws.row_dimensions[r].height=24
    elif level=='L3':
        lc.fill=PatternFill('solid',fgColor=L3_FILL);lc.value='L3\n功能';ws.row_dimensions[r].height=calc_row_height(principle,40)
    ws.cell(row=r,column=2).font=F_CODE;ws.cell(row=r,column=2).alignment=A_C
    pc=ws.cell(row=r,column=3);pc.font=F_BOLD;pc.alignment=A_C
    if phase and phase in PHASE_COLORS: pc.fill=PatternFill('solid',fgColor=PHASE_COLORS[phase]);pc.font=Font(name='微软雅黑',size=9,bold=True,color='FFFFFF')
    if level=='L1':
        bg,fg=L1_FILLS.get(sys_code,('333333','FFFFFF'));c=ws.cell(row=r,column=4);c.fill=PatternFill('solid',fgColor=bg);c.font=F_L1;c.alignment=A_C
    if level=='L2':
        c=ws.cell(row=r,column=5);c.fill=PatternFill('solid',fgColor=L2_FILL);c.font=F_L2;c.alignment=A_L
    if level=='L3':
        c=ws.cell(row=r,column=6);c.fill=PatternFill('solid',fgColor=L3_FILL);c.font=F_CELL
        if '★' in (func or ''): c.font=Font(name='微软雅黑',size=10,bold=True,color=C_RED)
        ws.cell(row=r,column=7).fill=PatternFill('solid',fgColor=L4_FILL);ws.cell(row=r,column=7).font=F_CODE
        ws.cell(row=r,column=8).fill=PatternFill('solid',fgColor=L5_FILL);ws.cell(row=r,column=8).font=F_CODE
        ws.cell(row=r,column=9).font=F_CODE;ws.cell(row=r,column=10).font=F_CODE
        dc=ws.cell(row=r,column=11);dc.font=F_CODE;dc.alignment=A_C
        if depends and depends!='—': dc.fill=PatternFill('solid',fgColor='FCF3CF')
        hc=ws.cell(row=r,column=12);hc.font=F_BOLD;hc.alignment=A_C;hc.fill=PatternFill('solid',fgColor='EBF5FB')
        cc=ws.cell(row=r,column=13);cc.font=F_CODE;cc.alignment=A_LT;cc.fill=PatternFill('solid',fgColor='FEF9E7')
        vc=ws.cell(row=r,column=14);vc.alignment=A_C
        if ver=='V1': vc.fill=PatternFill('solid',fgColor='D6EAF8')
        elif ver=='V1.5': vc.fill=PatternFill('solid',fgColor='FADBD8')
        elif ver=='V2': vc.fill=PatternFill('solid',fgColor='D5D8DC')
        sc=ws.cell(row=r,column=15);sc.font=F_BOLD;sc.alignment=A_C
        if status=='待办': sc.fill=PatternFill('solid',fgColor=ST_TODO)
        elif status=='进行中': sc.fill=PatternFill('solid',fgColor=ST_DOING)
        elif status=='已完成': sc.fill=PatternFill('solid',fgColor=ST_DONE)
    r+=1
last_row=r-1
ws.conditional_formatting.add(f'O3:O{last_row}',FormulaRule(formula=['$O3="已完成"'],fill=PatternFill('solid',fgColor=ST_DONE)))
ws.conditional_formatting.add(f'O3:O{last_row}',FormulaRule(formula=['$O3="进行中"'],fill=PatternFill('solid',fgColor=ST_DOING)))
ws.conditional_formatting.add(f'O3:O{last_row}',FormulaRule(formula=['$O3="待办"'],fill=PatternFill('solid',fgColor=ST_TODO)))
r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=15)
c=ws.cell(row=r,column=1,value='层级：L1系统→L2模块→L3功能→L4原理→L5函数 | 黄色=有依赖(先做依赖项) | 工时h=预计工时 | 完成标准=验收依据 | 改状态→进度统计自动更新')
c.font=F_SMALL;c.alignment=A_L

# ===== Sheet 3: 进度统计（加当前阶段标识） =====
ws=wb.create_sheet('3.进度统计');ws.sheet_view.showGridLines=False
set_w(ws,[4,22,8,10,10,10,10,16,4])
big_title(ws,'进度统计 — 自动计算（改主表状态后刷新即更新）',9)
r=2
ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8)
c=ws.cell(row=r,column=2,value='当前应做阶段：查找第一个完成率<100%的阶段（见下方"按阶段统计"）')
c.font=F_GOLD;c.alignment=A_L;c.fill=PatternFill('solid',fgColor='FEF9E7');c.border=BD;ws.row_dimensions[r].height=26
header_row(ws,['','系统大类','阶段','总点数','已完成','进行中','待办','完成率','进度条'],3)
MAIN="'2.开发主表'"
sys_map=[('A','A.引擎框架','P1/P6'),('B','B.游戏数据','P2'),('C','C.战斗逻辑','P3/P4'),('D','D.AI系统','P5'),('E','E.渲染与UI','P6'),('F','F.社交系统','P7+'),('G','G.养成变现','P7+'),('H','H.实时PvP','V2')]
r=4
for code,name,phase_tag in sys_map:
    bg,fg=L1_FILLS.get(code,('333333','FFFFFF'))
    total_f=f'=COUNTIFS({MAIN}!B3:B{last_row},"{code}*",{MAIN}!O3:O{last_row},"待办")+COUNTIFS({MAIN}!B3:B{last_row},"{code}*",{MAIN}!O3:O{last_row},"进行中")+COUNTIFS({MAIN}!B3:B{last_row},"{code}*",{MAIN}!O3:O{last_row},"已完成")'
    done_f=f'=COUNTIFS({MAIN}!B3:B{last_row},"{code}*",{MAIN}!O3:O{last_row},"已完成")'
    doing_f=f'=COUNTIFS({MAIN}!B3:B{last_row},"{code}*",{MAIN}!O3:O{last_row},"进行中")'
    todo_f=f'=COUNTIFS({MAIN}!B3:B{last_row},"{code}*",{MAIN}!O3:O{last_row},"待办")'
    rate_f=f'=IF(D{r}=0,"0%",TEXT(E{r}/D{r},"0%"))'
    bar_f=f'=REPT("█",ROUND(E{r}/D{r}*20,0))&REPT("░",20-ROUND(E{r}/D{r}*20,0))'
    vals=['',name,phase_tag,total_f,done_f,doing_f,todo_f,rate_f,bar_f]
    for i,v in enumerate(vals,1):
        c=ws.cell(row=r,column=i,value=v);c.border=BD;c.alignment=A_C if i>2 else A_L;c.font=F_CELL
    ws.cell(row=r,column=2).fill=PatternFill('solid',fgColor=bg);ws.cell(row=r,column=2).font=F_L1;ws.cell(row=r,column=2).alignment=A_C
    ws.cell(row=r,column=9).font=Font(name='Consolas',size=11,color=C_GREEN);ws.row_dimensions[r].height=30;r+=1
total_f=f'=SUM(D4:D{r-1})';done_f=f'=SUM(E4:E{r-1})';doing_f=f'=SUM(F4:F{r-1})';todo_f=f'=SUM(G4:G{r-1})'
rate_f=f'=IF(D{r}=0,"0%",TEXT(E{r}/D{r},"0%"))';bar_f=f'=REPT("█",ROUND(E{r}/D{r}*20,0))&REPT("░",20-ROUND(E{r}/D{r}*20,0))'
vals=['','总计','',total_f,done_f,doing_f,todo_f,rate_f,bar_f]
for i,v in enumerate(vals,1):
    c=ws.cell(row=r,column=i,value=v);c.font=F_HEADER;c.fill=PatternFill('solid',fgColor=C_INK);c.alignment=A_C if i>2 else A_L;c.border=BD
ws.row_dimensions[r].height=34
r+=2
sub_title(ws,'按阶段统计（← 找第一个完成率<100%的阶段，那就是你当前该做的）',9,r,C_GOLD);r+=1
header_row(ws,['','阶段','阶段名称','总点数','已完成','进行中','待办','完成率','进度条'],r);r+=1
phase_map=[('P1','引擎骨架'),('P2','数据层'),('P3','核心战斗'),('P4','扩展战斗'),('P5','AI对手'),('P6','渲染交互'),('P7','测试平衡'),('P8','上线')]
for code,name in phase_map:
    total_f=f'=COUNTIF({MAIN}!C3:C{last_row},"{code}")'
    done_f=f'=COUNTIFS({MAIN}!C3:C{last_row},"{code}",{MAIN}!O3:O{last_row},"已完成")'
    doing_f=f'=COUNTIFS({MAIN}!C3:C{last_row},"{code}",{MAIN}!O3:O{last_row},"进行中")'
    todo_f=f'=COUNTIFS({MAIN}!C3:C{last_row},"{code}",{MAIN}!O3:O{last_row},"待办")'
    rate_f=f'=IF(D{r}=0,"0%",TEXT(E{r}/D{r},"0%"))'
    bar_f=f'=REPT("█",ROUND(E{r}/D{r}*20,0))&REPT("░",20-ROUND(E{r}/D{r}*20,0))'
    vals=['',code,name,total_f,done_f,doing_f,todo_f,rate_f,bar_f]
    for i,v in enumerate(vals,1):
        c=ws.cell(row=r,column=i,value=v);c.border=BD;c.alignment=A_C if i>2 else A_L;c.font=F_CELL
    ws.cell(row=r,column=2).fill=PatternFill('solid',fgColor=PHASE_COLORS[code]);ws.cell(row=r,column=2).font=F_L1;ws.cell(row=r,column=2).alignment=A_C
    ws.cell(row=r,column=9).font=Font(name='Consolas',size=11,color=C_GREEN);ws.row_dimensions[r].height=28;r+=1
r+=1
ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=9)
c=ws.cell(row=r,column=2,value='说明：改「2.开发主表」的"状态"列后，回到本表按Ctrl+Z再Ctrl+Y（或重新打开文件）即可刷新数字。')
c.font=F_SMALL;c.alignment=A_L

# ===== Sheet 4: 卡牌数据（加cardId+预设卡组） =====
ws=wb.create_sheet('4.卡牌数据');ws.sheet_view.showGridLines=False
set_w(ws,[14,12,10,8,8,8,8,8,28,8])
big_title(ws,'卡牌数据表 — 改数值直接在这改（cardId对应代码引用）',10)
r=3
sub_title(ws,'攻方·人物卡',10,r,C_RED);r+=1;header_row(ws,['cardId','名称','类型','费用','血量','攻击','移速','间隔','特性','稀有度'],r);r+=1
for c_data in [('body_disciple','宗门体修弟子','普通弟子',2,4,2,1.0,1.0,'普通近战士兵，基础兵','凡品'),('sword_disciple','宗门剑修弟子','普通弟子',3,3,3,0.9,1.3,'远程(射程3)，用飞剑，不掉血','凡品'),('beast_disciple','宗门御兽弟子','普通弟子',3,6,2,0.8,1.2,'控兽当肉盾，高血低速','凡品'),('golden_elder','金丹期长老','精英长老',6,10,3,0.8,1.5,'每5s随机释放4分支技能','宝品')]:
    for i,v in enumerate(c_data,1): cell=ws.cell(row=r,column=i,value=v);cell.font=F_CELL;cell.alignment=A_C;cell.border=BD
    ws.cell(row=r,column=1).font=F_CODE;ws.cell(row=r,column=2).fill=PatternFill('solid',fgColor='FADBD8');ws.row_dimensions[r].height=24;r+=1
r+=1;sub_title(ws,'守方·人物卡',10,r,C_BLUE);r+=1;header_row(ws,['cardId','名称','类型','费用','血量','攻击','射程','间隔','特性','稀有度'],r);r+=1
for c_data in [('guardian_puppet','护山傀儡','防守单位',3,6,2,1,1.2,'缓慢移动肉盾守卫','凡品'),('guardian_beast','护山灵兽','防守单位',4,8,2,1,1.0,'高血拦截，死亡自爆1伤','灵品'),('defender_elder','护法长老','精英长老',6,10,3,1,1.5,'镇守大殿前，每5s随机释放4分支','宝品')]:
    for i,v in enumerate(c_data,1): cell=ws.cell(row=r,column=i,value=v);cell.font=F_CELL;cell.alignment=A_C;cell.border=BD
    ws.cell(row=r,column=1).font=F_CODE;ws.cell(row=r,column=2).fill=PatternFill('solid',fgColor='D6EAF8');ws.row_dimensions[r].height=24;r+=1
r+=1;sub_title(ws,'阵法（双方通用）',10,r,C_TEAL);r+=1;header_row(ws,['cardId','名称','类型','费用','血量','攻击','射程','间隔','特性','稀有度'],r);r+=1
for c_data in [('jiemai_formation','截脉阵','阵法',2,4,2,1,1.0,'基础拦截，便宜','凡品'),('hanshuang_formation','寒霜阵','阵法',3,3,1,1,1.0,'命中后敌人移速-0.5(2s)','凡品'),('wanren_formation','万刃阵','阵法',4,5,3,1,1.0,'高输出拦截','灵品'),('fanzhen_formation','反震阵','阵法',3,3,0,'-',1.0,'反伤50%','灵品'),('tianluo_formation','天罗阵','阵法',5,6,2,1,1.0,'范围(打相邻所有敌人)','宝品')]:
    for i,v in enumerate(c_data,1): cell=ws.cell(row=r,column=i,value=v);cell.font=F_CELL;cell.alignment=A_C;cell.border=BD
    ws.cell(row=r,column=1).font=F_CODE;ws.cell(row=r,column=2).fill=PatternFill('solid',fgColor='E8F8F5');ws.row_dimensions[r].height=24;r+=1
r+=1;sub_title(ws,'法术',10,r,C_PURPLE);r+=1;header_row(ws,['cardId','名称','类型','费用','效果','','','','偏向','稀有度'],r);r+=1
for c_data in [('wan_jian','万剑归宗','法术',5,'全己方单位+1攻、移速+0.3(5s)','','','','攻方','宝品'),('wu_lei','五雷正法','法术',4,'区域3格内敌方受4伤','','','','通用','灵品'),('yu_feng','御风诀','法术',2,'指定己方单位移速+0.5(5s)','','','','通用','凡品'),('zhen_hun','镇魂符','法术',3,'指定敌方阵法失效3秒','','','','通用','凡品'),('jin_zhong','金钟罩','法术',3,'己方大殿免疫伤害3秒','','','','守方','灵品'),('yi_shan','移山倒海','法术',4,'区域敌人推后2格+1伤','','','','守方','灵品'),('kun_xian','困仙索','法术',2,'指定敌人定身2秒','','','','守方','凡品'),('tian_lei','天雷诀','法术',4,'范围3格内敌方受4伤(清兵)','','','','通用','灵品')]:
    for i,v in enumerate(c_data,1): cell=ws.cell(row=r,column=i,value=v);cell.font=F_CELL;cell.alignment=A_C;cell.border=BD
    ws.cell(row=r,column=1).font=F_CODE;ws.cell(row=r,column=2).fill=PatternFill('solid',fgColor='F4ECF7');ws.row_dimensions[r].height=24;r+=1
r+=1;sub_title(ws,'预设卡组（3套推荐配置，每套8张）',10,r,C_GOLD);r+=1
set_w(ws,[16,10,14,14,14,14,14,14,14,14])
header_row(ws,['卡组名','风格','卡1','卡2','卡3','卡4','卡5','卡6','卡7','卡8'],r);r+=1
for p in [('速攻流','快速铺兵','体修弟子','体修弟子','剑修弟子','剑修弟子','御风诀','御风诀','截脉阵','金丹长老'),('控制流','阵法+法术','御兽弟子','护山傀儡','截脉阵','万刃阵','反震阵','镇魂符','金钟罩','金丹长老'),('均衡流','攻守兼备','体修弟子','剑修弟子','御兽弟子','截脉阵','万刃阵','五雷正法','困仙索','金丹长老')]:
    for i,v in enumerate(p,1): cell=ws.cell(row=r,column=i,value=v);cell.font=F_CELL;cell.alignment=A_C;cell.border=BD
    ws.cell(row=r,column=1).fill=PatternFill('solid',fgColor='FEF9E7');ws.cell(row=r,column=1).font=F_BOLD;ws.row_dimensions[r].height=26;r+=1

# ===== Sheet 5: 数值配置 =====
ws=wb.create_sheet('5.数值配置');ws.sheet_view.showGridLines=False
set_w(ws,[20,12,32,14,14])
big_title(ws,'数值配置 — 所有可调参数集中管理',5)
header_row(ws,['参数名','当前值','说明','调高偏向','调低偏向'],3)
params=[('灵力回复间隔','2.8秒','每多少秒回复1点灵力','攻方(出兵多)','守方(布阵少)'),('灵力初始上限','5','开局灵力上限','攻方','守方'),('灵力上限增长间隔','30秒','每多少秒上限+1','攻方','守方'),('灵力上限封顶','10','灵力上限最大值','攻方','守方'),('大殿血量','30','大殿初始血量','守方','攻方'),('单局时长','180秒','正常对战时间','守方','攻方'),('加时时长','60秒','平局后加时时间','中立','中立'),('加时灵力倍率','1.5','加时赛灵力回复倍率','攻方','守方'),('手牌数量','4','同时持有手牌数','攻方(选择多)','守方'),('卡组数量','8','卡组总卡牌数','中立','中立'),('抽牌延迟','2秒','打出后多久抽新牌','守方','攻方'),('阵法冷却时间','8秒','阵法被毁后格子冷却','攻方','守方'),('长老技能间隔','5秒','长老多久释放一次技能','守方','攻方'),('棋盘长度','9格','大殿到大殿格子数(含大殿)','守方','攻方'),('同屏单位上限','30','最多同时存在单位数','中立(性能)','中立(性能)'),('远程单位射程','3格','剑修弟子攻击射程','攻方','守方'),('减速幅度','0.5','寒霜阵减速值','守方','攻方'),('减速持续','2秒','减速持续时间','守方','攻方'),('反震比例','50%','反震阵反弹伤害比例','守方','攻方'),('护盾持续','3秒','金钟罩大殿免疫时间','守方','攻方'),('定身持续','2秒','困仙索定身时间','守方','攻方'),('推后格数','2格','移山倒海推后距离','守方','攻方'),('加速持续','5秒','御风诀/万剑归宗加速时间','攻方','守方'),('禁阵持续','3秒','镇魂符禁阵时间','攻方','守方'),('AI简单思考间隔','3.5秒','简单AI决策间隔','AI强','AI弱'),('AI普通思考间隔','2.5秒','普通AI决策间隔','AI强','AI弱'),('AI困难思考间隔','1.5秒','困难AI决策间隔','AI强','AI弱')]
r=4
for p in params:
    for i,v in enumerate(p,1): c=ws.cell(row=r,column=i,value=v);c.font=F_CELL;c.alignment=A_C if i>1 else A_L;c.border=BD
    ws.row_dimensions[r].height=22;r+=1
r+=1;sub_title(ws,'平衡验证标准（模拟10局AI vs AI）',5,r,C_GOLD);r+=1
for v in ['双方摧毁度都在 20%~70% 区间','胜率：攻守方各 ~50%（40%~60%）','0% 摧毁度对局 < 10%','90% 对局在 150~210 秒内结束','双方大殿都被打到（有来有回）']:
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
    c=ws.cell(row=r,column=1,value='• '+v);c.font=F_CELL;c.alignment=A_L;c.border=BD;r+=1

# ===== Sheet 6: AI提问指南 =====
ws=wb.create_sheet('6.AI提问指南');ws.sheet_view.showGridLines=False
set_w(ws,[18,75])
big_title(ws,'AI提问指南 — 如何让AI替你干活',2)
r=3;sub_title(ws,'一、黄金公式',2,r,C_GOLD);r+=1
ws.cell(row=r,column=1,value='公式').font=F_BOLD;ws.cell(row=r,column=1).fill=PatternFill('solid',fgColor=L2_FILL);ws.cell(row=r,column=1).border=BD;ws.cell(row=r,column=1).alignment=A_C
ws.cell(row=r,column=2,value='【角色定位】+【具体任务】+【上下文/参考文件】+【完成标准】+【自主决策授权】').font=F_CELL;ws.cell(row=r,column=2).alignment=A_LT;ws.cell(row=r,column=2).border=BD;ws.row_dimensions[r].height=30;r+=2
sub_title(ws,'二、提示词模板（直接复制使用）',2,r,C_PURPLE);r+=1
templates=[('场景1：实现功能','你是微信小游戏开发专家。\n请实现【功能名】：【把主表"实现原理"列的内容粘过来】\n函数签名：【把主表"L5函数"列粘过来】\n文件路径：【把主表"文件路径"列粘过来】\n前置依赖：【把主表"依赖"列粘过来，确认依赖项已完成】\n完成标准：①代码能通过node --check语法检查 ②符合原理描述的逻辑\n遇到设计细节未明确的按合理默认自行决定，做完统一汇报，不要中途问我。'),('场景2：修Bug','我遇到了一个Bug：\n【现象：什么情况发生了什么】\n【错误信息：贴上报错日志】\n相关文件：【文件路径】\n请分析原因并修复。修复后说明改了什么、为什么。'),('场景3：平衡调优','请运行AI vs AI模拟10局，输出每局：双方摧毁度/胜方/时长。\n当前参数见「数值配置」表。\n如发现平衡问题（摧毁度0%或一边倒），调整参数重新模拟，直到：双方摧毁度20~70%，胜率~50%。\n每次调参记录：调了什么、从多少到多少、为什么。'),('场景4：自主连续开发','以下是连续任务（按顺序）：\n【粘贴主表中连续3~5个L3行的"编号+原理+函数+依赖"】\n请按顺序逐个完成，每个完成后：①标记完成 ②记录做了什么 ③立即开始下一个\n全部做完统一汇报。遇到细节自行决定，不要中途问我。\n每个任务必须通过完成标准验证。')]
for name,tmpl in templates:
    ws.cell(row=r,column=1,value=name).font=F_BOLD;ws.cell(row=r,column=1).fill=PatternFill('solid',fgColor=L2_FILL);ws.cell(row=r,column=1).border=BD;ws.cell(row=r,column=1).alignment=A_C
    ws.cell(row=r,column=2,value=tmpl).font=F_CELL;ws.cell(row=r,column=2).alignment=A_LT;ws.cell(row=r,column=2).border=BD;ws.row_dimensions[r].height=100;r+=1
r+=1;sub_title(ws,'三、常见错误与纠正',2,r,C_RED);r+=1
for wrong,right in [('❌太模糊','"做个战斗系统" → ✅"实现单位移动：沿y轴推进，speed×dt，遇敌停下"'),('❌太大','"把游戏做完" → ✅一次一个任务（主表一行）'),('❌没标准','"帮我写代码" → ✅给完成标准"能通过node --check + 模拟能跑"'),('❌没授权','AI每步都停下问你 → ✅"遇到细节自行决定，做完汇报，不要中途问我"'),('❌没验证','AI说做完了你就信 → ✅按完成标准自己验证（跑模拟/查语法）'),('❌没上下文','AI不知道已有代码 → ✅告诉它参考文件路径和设计文档')]:
    ws.cell(row=r,column=1,value=wrong).font=F_CELL;ws.cell(row=r,column=1).border=BD;ws.cell(row=r,column=1).alignment=A_LT
    ws.cell(row=r,column=2,value=right).font=F_CELL;ws.cell(row=r,column=2).border=BD;ws.cell(row=r,column=2).alignment=A_LT;ws.row_dimensions[r].height=30;r+=1
r+=1;sub_title(ws,'四、每日开发流程',2,r,C_TEAL);r+=1
for step in ['1. 看「0.开发阶段规划」确认当前阶段（如P3）','2. 打开「2.开发主表」筛选阶段=P3，找第一个"待办"的L3行','3. 检查"前置依赖"列，确认依赖项都已完成','4. 复制该行的"实现原理"+"函数"+"文件路径"+"依赖"','5. 用「场景1」模板粘贴发给AI','6. AI完成后你验证（node --check / 跑模拟）','7. 通过→主表状态改"已完成"→进度统计自动更新','8. 不通过→用「场景2」修Bug模板发给AI','9. 重复，每天2~4个任务']:
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
    c=ws.cell(row=r,column=1,value=step);c.font=F_CELL;c.alignment=A_L;c.border=BD;ws.row_dimensions[r].height=24;r+=1

# ===== Sheet 7: 资源清单（新增） =====
ws=wb.create_sheet('7.资源清单');ws.sheet_view.showGridLines=False
set_w(ws,[6,16,8,10,12,16,8,10,30])
big_title(ws,'美术资源清单 — 需要多少素材、什么尺寸、优先做哪些',9)
r=3;sub_title(ws,'一、角色精灵图',9,r,C_RED);r+=1;header_row(ws,['','名称','类型','尺寸px','帧数','格式','阶段','工时h','备注'],r);r+=1
for a in [('体修弟子','角色','64x64','8(行走)','PNG','P6','2','2方向(上下)+翻转'),('剑修弟子','角色','64x64','8(行走)+4(攻击)','PNG','P6','3','含飞剑特效'),('御兽弟子','角色','64x64','8(行走)','PNG','P6','2','2方向+翻转'),('金丹长老','角色','80x80','12(行走)+8(技能)','PNG','P6','4','4分支技能各1组'),('护山傀儡','角色','64x64','6(行走)','PNG','P6','2','守方单位'),('护山灵兽','角色','80x80','8(行走)+4(死亡)','PNG','P6','3','死亡自爆特效'),('护法长老','角色','80x80','12(行走)+8(技能)','PNG','P6','4','守方精英')]:
    for i,v in enumerate(a,2): c=ws.cell(row=r,column=i,value=v);c.font=F_CELL;c.alignment=A_C;c.border=BD
    ws.cell(row=r,column=2).fill=PatternFill('solid',fgColor='FADBD8');ws.row_dimensions[r].height=22;r+=1
r+=1;sub_title(ws,'二、阵法图',9,r,C_TEAL);r+=1;header_row(ws,['','名称','类型','尺寸px','帧数','格式','阶段','工时h','备注'],r);r+=1
for a in [('截脉阵','阵法','96x96','4(光效)','PNG','P6','1','绿色光阵'),('寒霜阵','阵法','96x96','4(冰效)','PNG','P6','1','蓝色冰阵'),('万刃阵','阵法','96x96','6(刀光)','PNG','P6','1.5','银色刀光'),('反震阵','阵法','96x96','4(反弹)','PNG','P6','1','紫色光罩'),('天罗阵','阵法','96x96','8(大阵)','PNG','P6','2','金色大阵+范围')]:
    for i,v in enumerate(a,2): c=ws.cell(row=r,column=i,value=v);c.font=F_CELL;c.alignment=A_C;c.border=BD
    ws.cell(row=r,column=2).fill=PatternFill('solid',fgColor='E8F8F5');ws.row_dimensions[r].height=22;r+=1
r+=1;sub_title(ws,'三、场景与UI',9,r,C_PURPLE);r+=1;header_row(ws,['','名称','类型','尺寸px','帧数','格式','阶段','工时h','备注'],r);r+=1
for a in [('山道背景','背景','375x667','1','JPG','P6','2','竖屏山道渐变'),('宗门大殿','建筑','375x60','1','PNG','P6','2','顶部+底部各1'),('大殿受击特效','特效','375x60','6','PNG','P6','1','碎裂/震动'),('手牌框','UI','80x70','2(可用/不可用)','PNG','P6','1','卡牌底框'),('灵力条','UI','335x8','1','PNG','P6','0.5','紫色进度条'),('血条','UI','可变','1','PNG','P6','0.5','红绿条'),('胜利/失败','UI','200x100','1','PNG','P6','1','结算大字'),('新手引导箭头','UI','40x60','4(动画)','PNG','P6','0.5','指示箭头')]:
    for i,v in enumerate(a,2): c=ws.cell(row=r,column=i,value=v);c.font=F_CELL;c.alignment=A_C;c.border=BD
    ws.cell(row=r,column=2).fill=PatternFill('solid',fgColor='F4ECF7');ws.row_dimensions[r].height=22;r+=1
r+=1;sub_title(ws,'四、特效',9,r,C_GOLD);r+=1;header_row(ws,['','名称','类型','尺寸px','帧数','格式','阶段','工时h','备注'],r);r+=1
for a in [('攻击命中','特效','32x32','6','PNG','P6','1','白光迸溅'),('飞剑飞行','特效','16x4','1(旋转)','PNG','P6','0.5','飞剑拖尾'),('法术释放','特效','64x64','8','PNG','P6','2','万剑/五雷/金钟各1组'),('单位死亡','特效','48x48','6','PNG','P6','1','消散粒子'),('长老技能','特效','96x96','8','PNG','P6','3','4分支各1组')]:
    for i,v in enumerate(a,2): c=ws.cell(row=r,column=i,value=v);c.font=F_CELL;c.alignment=A_C;c.border=BD
    ws.cell(row=r,column=2).fill=PatternFill('solid',fgColor='FEF9E7');ws.row_dimensions[r].height=22;r+=1
r+=1;sub_title(ws,'五、音频',9,r,C_BLUE);r+=1
set_w(ws,[6,16,8,8,8,8,8,30])
header_row(ws,['','名称','类型','时长s','格式','阶段','工时h','备注'],r);r+=1
for a in [('BGM-战斗','音乐','60','MP3','P6','2','循环，修仙风'),('SFX-出牌','音效','0.5','MP3','P6','0.5',''),('SFX-攻击','音效','0.3','MP3','P6','0.5',''),('SFX-阵法','音效','0.5','MP3','P6','0.5',''),('SFX-法术','音效','1.0','MP3','P6','1',''),('SFX-大殿受击','音效','0.5','MP3','P6','0.5',''),('SFX-胜利','音效','2.0','MP3','P6','0.5',''),('SFX-失败','音效','2.0','MP3','P6','0.5','')]:
    for i,v in enumerate(a,2): c=ws.cell(row=r,column=i,value=v);c.font=F_CELL;c.alignment=A_C;c.border=BD
    ws.cell(row=r,column=2).fill=PatternFill('solid',fgColor='D6EAF8');ws.row_dimensions[r].height=22;r+=1
r+=1;ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=9)
total_art=7+5+8+5+8
c=ws.cell(row=r,column=2,value=f'总计：{total_art}个资源。V1先用色块占位，P6阶段替换为正式素材。美术工时约33小时。')
c.font=F_GOLD;c.alignment=A_L;c.fill=PatternFill('solid',fgColor='FEF9E7');c.border=BD;ws.row_dimensions[r].height=30

# ===== Sheet 8: 决策日志与已知问题（新增） =====
ws=wb.create_sheet('8.决策日志与问题');ws.sheet_view.showGridLines=False
set_w(ws,[6,12,14,36,20,14,8])
big_title(ws,'决策日志与已知问题 — 记录每个设计决策和待解决问题',7)
r=3;sub_title(ws,'一、设计决策日志',7,r,C_PURPLE);r+=1;header_row(ws,['','日期','编号','决策内容','原因','影响范围','状态'],r);r+=1
decisions=[('2026-08-05','D001','从回合制改为实时制','回合制等对方出招无聊，实时制紧张感是核心体验','全局','已执行'),('2026-08-05','D002','单位击杀后继续推进','旧版"遇阵即停打完即没"导致0%摧毁度','C4战斗系统','已执行'),('2026-08-05','D003','阵法被毁后8秒冷却','防止守方无限补阵堵路','C5阵法系统','已执行'),('2026-08-05','D004','使用折中视角(上下直推+山道视觉)','等距视角美术成本3~4倍','E渲染系统','已执行'),('2026-08-05','D005','金丹长老随机释放4分支技能','增加不可预测性和对局变化','C7长老技能','已执行'),('2026-08-05','D006','剑阵双方通用','攻方也需要阵法掩护推进','C5阵法系统','已执行'),('2026-08-05','D007','V1只做3套预设卡组','MVP原则，先验证核心玩法','B4预设卡组','已执行'),('2026-08-05','D008','用2D Canvas不用3D','3D原生开发管线太重，2D修仙更对味','全局','已执行'),('2026-08-05','D009','纯原生JS不用引擎','V1零依赖，包体最小','全局','已执行'),('2026-08-05','D010','选用修仙题材','过审零风险，认知度高，攻山主题契合','全局','已执行')]
for d in decisions:
    for i,v in enumerate(d,2): c=ws.cell(row=r,column=i,value=v);c.font=F_CELL;c.alignment=A_LT if i==4 else A_C;c.border=BD
    ws.cell(row=r,column=2).font=F_CODE;ws.cell(row=r,column=3).font=F_BOLD
    if d[5]=='已执行': ws.cell(row=r,column=7).fill=PatternFill('solid',fgColor=ST_DONE)
    ws.row_dimensions[r].height=26;r+=1
r+=1;sub_title(ws,'二、已知问题与待解决',7,r,C_RED);r+=1;header_row(ws,['','日期','编号','问题描述','影响','优先级','状态'],r);r+=1
issues=[('2026-08-05','I001','平衡偏防守：模拟多局0%摧毁度','攻方打不进去，玩家挫败','高','已解决(D002,D003)'),('2026-08-05','I002','AI布阵位置太随机','简单AI布阵位置不合理','中','待解决'),('2026-08-05','I003','长老技能可能重复释放同一分支','随机性不够均匀','低','待解决'),('2026-08-05','I004','阵法冷却期间格子无视觉提示','玩家不知道哪个格子冷却中','中','待解决'),('2026-08-05','I005','手牌栏可能遮挡底部单位','单位在手牌栏后方不可见','中','待解决'),('2026-08-05','I006','加时赛规则不直观','玩家不理解加时赛灵力倍率','低','待解决'),('2026-08-05','I007','无断线重连机制(V2)','实时PvP断线后无法恢复','低(V2)','待解决'),('2026-08-05','I008','无防作弊机制(V2)','客户端可篡改数据','低(V2)','待解决')]
for iss in issues:
    for i,v in enumerate(iss,2): c=ws.cell(row=r,column=i,value=v);c.font=F_CELL;c.alignment=A_LT if i==4 else A_C;c.border=BD
    ws.cell(row=r,column=2).font=F_CODE;ws.cell(row=r,column=3).font=F_BOLD
    if iss[4]=='高': ws.cell(row=r,column=6).fill=PatternFill('solid',fgColor=ST_TODO)
    elif iss[4]=='中': ws.cell(row=r,column=6).fill=PatternFill('solid',fgColor=ST_DOING)
    else: ws.cell(row=r,column=6).fill=PatternFill('solid',fgColor=ST_DONE)
    if iss[5].startswith('已'): ws.cell(row=r,column=7).fill=PatternFill('solid',fgColor=ST_DONE)
    else: ws.cell(row=r,column=7).fill=PatternFill('solid',fgColor=ST_TODO)
    ws.row_dimensions[r].height=26;r+=1
r+=1
ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
c=ws.cell(row=r,column=2,value='使用方法：开发中遇到设计决策记到"决策日志"，遇到问题记到"已知问题"。已解决的标"已解决/已执行"，待解决的标"待解决"。')
c.font=F_SMALL;c.alignment=A_L;c.fill=PatternFill('solid',fgColor='FEF9E7');c.border=BD;ws.row_dimensions[r].height=30

# 保存
wb.save(PATH)
l3_count=len([d for d in D if d[0]=='L3'])
print(f'Excel已生成: {PATH}')
print(f'Sheet: {wb.sheetnames}')
print(f'L3功能点: {l3_count} (含完成标准)')
print(f'卡牌: 20张(含cardId) | 参数: 27个 | AI模板: 4个')
print(f'资源清单: 33个资源 | 决策: 10条 | 问题: 8条')
